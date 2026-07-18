"""Reconstruye mol_perf['45'] en mujer/index.html para que SOLO incluya los
competidores REALES de TRIP +45 (suplemento menopausico ORAL):

  - VIASEK MENOCARE (E6B) - ATC G02X9, Mol que contiene 'MAGNESIUM'
    -> Esto matchea SOLO la presentacion CAPS x 32 (suplemento oral con
       MAGNESIUM_PANAX GINSENG_RESVERATROL — moleculas similares a TRIP +45).
    MAT Apr 2026: ~45,479u
  - TRIP +45 (SIE) - ATC V03X0 (TABL RECUBIE x 30, suplemento oral).
    MAT Apr 2026: ~15,096u

Se EXCLUYE:
  - VIASEK MENOCARE BARRA SYNDET (G01D0, DL-LACTIC ACID): es un cleanser.
  - VIASEK MENOCARE GEL 100ML / GEL 50ML (G02X9, mol ESTROGENIC+HYALURONIC):
    son para uso VAGINAL (diferente formato que TRIP +45 oral, no son
    competidores reales).

NO se toca:
  - Otras familias en mujer mol_perf (SIN ESTROGENO, ALTA DOSIS, etc.)
  - Otras lineas
  - respPerf u otros campos
  - is_sie flags
"""
from __future__ import annotations
import json, re, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
# F4: el dato de mujer vive en data.js (window.OTC_DASHBOARD), ya no inline en
# index.html. El regex de abajo acepta ambos anchors, pero el archivo a
# leer/escribir es data.js.
DATA_FILE = REPO / 'mujer' / 'data.js'
MASTER = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs\_iqvia-master\2026-06\AR_PM_FV_Standard_Jul-2026.xlsx')

# Productos que SI deben ir en el mercado '45':
# (Product, ATC prefix, molecule_must_contain or None)
# - Si molecule_must_contain != None, SOLO se incluye la row cuyo Molecules
#   contiene ese keyword (case-insensitive). Esto sirve para distinguir
#   presentaciones del mismo Product cuando difieren por molecula.
INCLUDED = [
    # VIASEK MENOCARE: SOLO la presentacion CAPS x 32 (suplemento oral).
    # Se distingue de GEL/BARRA por la molecula MAGNESIUM_PANAX GINSENG.
    ('VIASEK MENOCARE (E6B)', 'G02X9', 'MAGNESIUM'),
    # TRIP +45: unica presentacion (TABL RECUBIE x 30, oral).
    ('TRIP +45 (SIE)',         'V03X0', None),
]

MES_INV = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,
           'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
NUM_TO_MES = {v:k for k,v in MES_INV.items()}
CIERRE = 6   # Jun 2026


def quarter_key(mk):
    parts = mk.split()
    if len(parts) != 2: return ''
    m = MES_INV.get(parts[0])
    if not m: return ''
    q = (m - 1) // 3 + 1
    return f'Q{q} {parts[1]}'


def aggregate_quarterly(monthly):
    out = defaultdict(int)
    for mk, v in monthly.items():
        qk = quarter_key(mk)
        if qk:
            try: out[qk] += int(round(float(v or 0)))
            except (TypeError, ValueError): pass
    return dict(out)


def aggregate_ytd(monthly, cierre=4):
    out = defaultdict(int)
    cierre_lbl = NUM_TO_MES[cierre]
    for mk, v in monthly.items():
        parts = mk.split()
        if len(parts) != 2: continue
        m_num = MES_INV.get(parts[0])
        if not m_num: continue
        if m_num <= cierre:
            try: out[parts[1]] += int(round(float(v or 0)))
            except (TypeError, ValueError): pass
    return {f'{cierre_lbl} {y}': v for y, v in out.items()}


def aggregate_mat(monthly, cierre=4):
    cierre_lbl = NUM_TO_MES[cierre]
    years = set()
    for mk in monthly:
        parts = mk.split()
        if len(parts) == 2 and parts[0] in MES_INV:
            try: years.add(int(parts[1]))
            except ValueError: pass
    out = {}
    for y in sorted(years):
        total = 0
        for back in range(11, -1, -1):
            tot_idx = (y * 12 + (cierre - 1)) - back
            yy, mm = divmod(tot_idx, 12)
            mk = f'{NUM_TO_MES[mm + 1]} {yy}'
            v = monthly.get(mk)
            if v is not None:
                try: total += int(round(float(v or 0)))
                except (TypeError, ValueError): pass
        out[f'{cierre_lbl} {y}'] = total
    return out


def main():
    if not MASTER.is_file():
        print(f'ERROR: master not found at {MASTER}', file=sys.stderr); return 2
    if hasattr(sys.stdout, 'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')

    print(f'Reading {MASTER.name}...')
    wb = openpyxl.load_workbook(MASTER, read_only=True, data_only=True)
    ws = wb.active
    hdrs = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    col_mfr = col_prod = col_atc = col_mol = None
    month_cols = []
    for i, h in enumerate(hdrs):
        if not h: continue
        s = str(h).strip()
        s_norm = s.replace('\n', ' ').strip().lower()
        if s_norm.startswith('manufacturer') and col_mfr is None: col_mfr = i
        elif s_norm.startswith('product') and col_prod is None: col_prod = i
        elif s_norm.startswith('molecules') and col_mol is None: col_mol = i
        # ATC: preferir 'ATC IV' (5-char, ej G02X9/V03X0 que usa el filtro '45');
        # el master 'Ateneo Total' trae tambien 'ATC III' (4-char) y last-wins lo agarraba.
        if s_norm.startswith('atc iv'): col_atc = i
        elif s_norm.startswith('atc') and col_atc is None: col_atc = i
        if s.startswith('Units') and ('\n' in s or len(s.split()) >= 2):
            after = s.split('\n', 1)[-1] if '\n' in s else s[len('Units'):].strip()
            after = after.strip()
            if after.upper().startswith('MAT') or after.upper().startswith('YTD'):
                continue
            m = re.match(r'(\w+)\s+(\d{4})$', after)
            if m and m.group(1) in MES_INV:
                month_cols.append((i, f'{m.group(1)} {m.group(2)}'))
    if col_atc is None: col_atc = 2
    if col_prod is None: col_prod = 1
    if col_mfr is None: col_mfr = 0
    if col_mol is None: col_mol = 3
    print(f'monthly columns: {len(month_cols)}, from {month_cols[0][1]} to {month_cols[-1][1]}')

    # Acumular monthly_vals por (Product, ATC prefix)
    # Cada (Product, ATC prefix) sera un producto final en mol_perf['45'].
    # Las multiples rows en master para mismo Product+ATC se suman.
    by_key = defaultdict(lambda: {'manuf': None, 'monthly': defaultdict(float), 'mols': set()})
    n_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row: continue
        prod = row[col_prod] if col_prod < len(row) else None
        atc  = row[col_atc]  if col_atc  < len(row) else None
        if not prod or not atc: continue
        prod_str = str(prod).strip()
        atc_str = str(atc).strip()
        mol_str = str(row[col_mol] or '').strip() if col_mol < len(row) else ''
        # Check if this (prod, atc, mol) matches our INCLUDED list
        matched = None
        for inc_prod, inc_atc_pref, mol_keyword in INCLUDED:
            if prod_str != inc_prod: continue
            if not atc_str.startswith(inc_atc_pref): continue
            if mol_keyword and mol_keyword.upper() not in mol_str.upper(): continue
            matched = (inc_prod, inc_atc_pref); break
        if not matched: continue
        mfr = row[col_mfr] if col_mfr < len(row) else None
        mol = row[col_mol] if col_mol < len(row) else None
        bucket = by_key[matched]
        if not bucket['manuf']:
            bucket['manuf'] = str(mfr or '').strip()
        if mol:
            bucket['mols'].add(str(mol).strip())
        for ci, mk in month_cols:
            if ci >= len(row): continue
            v = row[ci]
            if v is None: continue
            try: bucket['monthly'][mk] += float(v)
            except (TypeError, ValueError): pass
        n_rows += 1
    wb.close()
    print(f'Rows matched: {n_rows} (across {len(by_key)} (prod,atc) buckets)')

    # Construir lista de products
    # Convencion: cada (Product, ATC prefix) = 1 producto final.
    # Pero si dos buckets tienen mismo Product y diferentes ATC, los unificamos
    # bajo el mismo Product (sumando monthly_vals) para mantener shape compatible
    # con el dashboard (que muestra por Product).
    prods_by_name = defaultdict(lambda: {'manuf': None, 'monthly': defaultdict(int), 'is_sie': False})
    for (prod_name, _atc), info in by_key.items():
        agg = prods_by_name[prod_name]
        if not agg['manuf']:
            agg['manuf'] = info['manuf']
        for mk, v in info['monthly'].items():
            agg['monthly'][mk] += int(round(v))
        if 'SIE' in prod_name.upper():
            agg['is_sie'] = True

    print('\nProductos finales para mol_perf[\'45\']:')
    fam_monthly = defaultdict(int)
    products_list = []
    for prod_name, info in prods_by_name.items():
        monthly = dict(info['monthly'])
        for mk, v in monthly.items():
            fam_monthly[mk] += v
        last_apr = monthly.get('Apr 2026', 0)
        print(f'  {prod_name:40s} is_sie={info["is_sie"]} Apr 2026={last_apr:,}u')
        products_list.append({
            'prod': prod_name,
            'manuf': info['manuf'],
            'is_sie': info['is_sie'],
            'monthly_vals': monthly,
            'quarterly_vals': aggregate_quarterly(monthly),
            'ytd': aggregate_ytd(monthly, CIERRE),
            'mat': aggregate_mat(monthly, CIERRE),
        })

    fam_monthly = dict(fam_monthly)
    fam_quarterly = aggregate_quarterly(fam_monthly)
    fam_ytd = aggregate_ytd(fam_monthly, CIERRE)
    fam_mat = aggregate_mat(fam_monthly, CIERRE)
    print(f'\nFamilia 45 Apr 2026: monthly={fam_monthly.get("Apr 2026", 0):,}u  ytd={fam_ytd.get("Apr 2026", 0):,}u  mat={fam_mat.get("Apr 2026", 0):,}u')

    # ms_* per product
    def safe(num, den): return round((num or 0)/den*100, 2) if den else 0
    for p in products_list:
        mv = p['monthly_vals']; qv = p['quarterly_vals']; yv = p['ytd']; mtv = p['mat']
        p['ms_monthly']   = {mk: safe(mv.get(mk,0), fam_monthly.get(mk,0))   for mk in fam_monthly}
        p['ms_quarterly'] = {qk: safe(qv.get(qk,0), fam_quarterly.get(qk,0)) for qk in fam_quarterly}
        p['ms_ytd']       = {yk: safe(yv.get(yk,0), fam_ytd.get(yk,0))       for yk in fam_ytd}
        p['ms_mat']       = {mk: safe(mtv.get(mk,0), fam_mat.get(mk,0))      for mk in fam_mat}

    # Sort: SIE first, then by Apr 2026 units desc
    def sort_key(p):
        apr = (p.get('monthly_vals') or {}).get('Apr 2026', 0) or 0
        return (not p.get('is_sie'), -apr)
    products_list.sort(key=sort_key)

    # Patch into mujer/index.html
    t = DATA_FILE.read_text(encoding='utf-8')
    m = re.search(r'(?:const D|window\.OTC_DASHBOARD)\s*=\s*\{', t)
    ob = m.end() - 1
    D, end_idx = json.JSONDecoder().raw_decode(t[ob:])
    abs_end = ob + end_idx

    if 'mol_perf' not in D or '45' not in D['mol_perf']:
        print('ERROR: mol_perf["45"] no existe en mujer', file=sys.stderr)
        return 3

    # Preservar el campo 'family' si existe, sino default '45'
    existing_45 = D['mol_perf']['45']
    family_label = existing_45.get('family', '45')

    D['mol_perf']['45'] = {
        'family': family_label,
        'products': products_list,
        'monthly': fam_monthly,
        'quarterly': fam_quarterly,
        'ytd': fam_ytd,
        'mat': fam_mat,
    }

    new_t = t[:ob] + json.dumps(D, ensure_ascii=False) + t[abs_end:]
    DATA_FILE.write_text(new_t, encoding='utf-8', newline='')
    print(f'\nmujer/index.html actualizado: mol_perf["45"] reconstruido con {len(products_list)} productos.')
    return 0


if __name__ == '__main__':
    sys.exit(main() or 0)
