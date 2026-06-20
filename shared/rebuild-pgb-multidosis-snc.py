# -*- coding: utf-8 -*-
"""Reconstruye mol_perf['PREGABALIN'] de SNC = MERCADO MULTIDOSIS (hasta Abr-2026).

Juan pidio ver solo multidosis en el mercado de PGB. IQVIA define el mercado
"PGB Multidosis" = presentaciones en TABLETA divisible (no capsulas, no grageas).

Definicion: las 13 marcas del mercado oficial IQVIA "PGB Multidosis" (set
validado contra el regional CUP: mismas marcas en Mar-2026) en presentacion
NO capsula -> se excluye Ph.Forms 'ACA - CAPSULAS', se conservan TABLETAS y
GRAGEAS (LINPREL es gragea divisible = multidosis; por eso NO alcanza filtrar
solo TABLETAS, que lo dejaba afuera). Mar-2026 = 116k (panel PM) ~ oficial CUP
108k. Usa la col 'Ph. Forms III' para separar capsulas (mas limpio que el pack).

Fuente: AR_PM_FV_Standard_Jun-10-2026 (pack-level, llega a Abr-2026; el master
de abril cerraba en marzo). is_sie = manufacturer == SIEGFRIED. Las otras
moleculas de SNC ya estan en abril -> no se tocan. Recalcular kpis aparte
(build-kpis + build-families + sync-kpistrip).
Uso: py shared/rebuild-pgb-multidosis-snc.py [--dry-run]
"""
from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
HTML = REPO / 'SNC' / 'data.js'  # F4: dato externo
MASTER = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs\_iqvia-master\2026-04\AR_PM_FV_Standard_Jun-12-2026.xlsx')

FAM_KEY = 'PREGABALIN'
FAM_LABEL = 'Pregabalina Multidosis'
CAPSULE_FORM = 'ACA'              # Ph. Forms III: 'ACA - ORAL S.ORD.CAPSULAS' (se EXCLUYE)
# Set oficial del mercado IQVIA "PGB Multidosis" (validado contra el regional CUP,
# Mar-2026 = mismas 13 marcas). Multidosis = estas marcas en presentacion NO capsula
# (tabletas divisibles + grageas; p.ej. LINPREL es gragea pero divisible -> multidosis).
MULTI_BRANDS = {'AXUAL','DOLONEUTIN','GAVIN','KABIAN','LINPREL','LUNEL','NEURISTAN',
                'PGB','PLENICA','PREBICTAL','PREBIEN','PRINCIPIA','SERIPRAN'}

MES_INV = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,
           'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
NUM2 = {v:k for k,v in MES_INV.items()}
MONTH_RE = re.compile(r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec) \d{4}$')

# Reglas PGB multidosis: fuente real shared/close-manifest.json (seg pgb_multidosis).
# Las constantes de arriba (MULTI_BRANDS/CAPSULE_FORM/FAM_LABEL) quedan como fallback.
sys.path.insert(0, str(Path(__file__).resolve().parent))
try:
    import manifest as _mf
except Exception:
    _mf = None
def _seg(name, key, default):
    return _mf.seg_get(name, key, default) if _mf else default


def msort(mk):
    p = mk.split(); return int(p[1])*100 + MES_INV.get(p[0],0) if len(p)==2 else 0
def quarter_key(mk):
    p = mk.split()
    if len(p)!=2: return ''
    m = MES_INV.get(p[0]); return f'Q{(m-1)//3+1} {p[1]}' if m else ''
def agg_quarterly(monthly):
    o = defaultdict(int)
    for mk,v in monthly.items():
        q = quarter_key(mk)
        if q: o[q]+=v
    return dict(o)
def agg_ytd(monthly, cierre):
    by = defaultdict(int)
    for mk,v in monthly.items():
        p = mk.split()
        if len(p)!=2: continue
        mn = MES_INV.get(p[0])
        if mn and mn<=cierre: by[p[1]]+=v
    return {f'{NUM2[cierre]} {y}': v for y,v in by.items()}
def agg_mat(monthly, cierre):
    years = {int(mk.split()[1]) for mk in monthly if len(mk.split())==2 and mk.split()[0] in MES_INV}
    out = {}
    for y in sorted(years):
        tot = 0
        for back in range(11,-1,-1):
            idx = (y*12 + (cierre-1)) - back
            yy, mm = divmod(idx, 12)
            tot += int(monthly.get(f'{NUM2[mm+1]} {yy}', 0) or 0)
        out[f'{NUM2[cierre]} {y}'] = tot
    return out


def load_multidosis(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True); ws = wb.active
    r1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    ci = {}
    for i,h in enumerate(r1):
        s = str(h or '').strip().lower()
        if s.startswith('manufacturer'): ci['man'] = i
        elif s.startswith('ph. forms'): ci['phf'] = i
        elif s.startswith('product'): ci['prod'] = i
        elif s.startswith('molecules'): ci['mol'] = i
    def lab(h): return str(h).split('\n')[-1].strip()
    mcols = [(i, lab(h)) for i,h in enumerate(r1)
             if h and str(h).startswith('Units') and MONTH_RE.match(lab(h))]
    prods = defaultdict(lambda: {'manuf': None, 'monthly': defaultdict(float)})
    for row in ws.iter_rows(min_row=2, values_only=True):
        if ci['mol'] >= len(row) or str(row[ci['mol']] or '').strip().upper() != FAM_KEY: continue
        phf = str(row[ci['phf']] or '').strip() if ci['phf'] < len(row) else ''
        if phf.upper().startswith(CAPSULE_FORM): continue               # excluir capsulas
        prod = str(row[ci['prod']] or '')
        if not prod or prod.split(' ')[0].upper() not in MULTI_BRANDS: continue  # solo marcas multidosis
        b = prods[prod]; b['manuf'] = row[ci['man']] if ci['man'] < len(row) else None
        for cidx, mk in mcols:
            if cidx < len(row) and row[cidx] is not None:
                try: b['monthly'][mk] += float(row[cidx])
                except (ValueError, TypeError): pass
    wb.close()
    months = [mk for _,mk in sorted(mcols, key=lambda x: msort(x[1]))]
    return months, {p: {'manuf': i['manuf'], 'monthly': {mk: int(round(v)) for mk,v in i['monthly'].items() if v}}
                    for p,i in prods.items()}


def build_family(prods):
    fam_monthly = defaultdict(int); plist = []
    for name,info in prods.items():
        mv = info['monthly']
        for mk,v in mv.items(): fam_monthly[mk] += v
        is_sie = str(info.get('manuf') or '').strip().upper() == 'SIEGFRIED'
        plist.append({'prod':name,'manuf':info.get('manuf') or '','is_sie':is_sie,'monthly_vals':mv})
    fam_monthly = dict(fam_monthly)
    cierre = MES_INV.get(max(fam_monthly,key=msort).split()[0],12) if fam_monthly else 12
    fam_q=agg_quarterly(fam_monthly); fam_y=agg_ytd(fam_monthly,cierre); fam_m=agg_mat(fam_monthly,cierre)
    for p in plist:
        mv=p['monthly_vals']
        p['quarterly_vals']=agg_quarterly(mv); p['ytd']=agg_ytd(mv,cierre); p['mat']=agg_mat(mv,cierre)
        p['ms_monthly']  ={mk: round(mv.get(mk,0)/fv*100,2) if fv>0 else 0 for mk,fv in fam_monthly.items()}
        p['ms_quarterly']={qk: round(p['quarterly_vals'].get(qk,0)/fv*100,2) if fv>0 else 0 for qk,fv in fam_q.items()}
        p['ms_ytd']      ={y: round(p['ytd'].get(y,0)/fv*100,2) if fv>0 else 0 for y,fv in fam_y.items()}
        p['ms_mat']      ={y: round(p['mat'].get(y,0)/fv*100,2) if fv>0 else 0 for y,fv in fam_m.items()}
    plist.sort(key=lambda p: (not p['is_sie'], -(p['monthly_vals'].get(max(p['monthly_vals'],key=msort,default=''),0) if p['monthly_vals'] else 0)))
    return {'family':FAM_LABEL,'products':plist,'monthly':fam_monthly,'quarterly':fam_q,'ytd':fam_y,'mat':fam_m}


def main():
    ap=argparse.ArgumentParser()
    ap.add_argument('--master', default=str(MASTER), help='AR_PM master (default=constante actual)')
    ap.add_argument('--cierre', help='(compat orquestador; el cierre se autodetecta del master)')
    ap.add_argument('--dry-run',action='store_true')
    a=ap.parse_args()
    if hasattr(sys.stdout,'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')
    global MULTI_BRANDS, CAPSULE_FORM, FAM_LABEL
    MULTI_BRANDS = set(_seg('pgb_multidosis','brands', list(MULTI_BRANDS)))
    CAPSULE_FORM = _seg('pgb_multidosis','excludeForm', CAPSULE_FORM)
    FAM_LABEL    = _seg('pgb_multidosis','familyLabel', FAM_LABEL)
    master = Path(a.master)
    if not master.is_file(): print('ERROR master no existe:',master); return 2
    months,prods=load_multidosis(master)
    fam=build_family(prods)
    last=max(fam['monthly'],key=msort)
    print(f'Multidosis (13 marcas, no-capsula): {len(months)} meses ({months[0]}..{months[-1]}), {len(fam["products"])} productos')
    print(f'  mercado {last} = {fam["monthly"][last]:,} u')
    for p in fam['products']:
        print(f'    {"SIE " if p["is_sie"] else "    "}{p["monthly_vals"].get(last,0):>7} {p["prod"]}')

    text=HTML.read_text(encoding='utf-8',errors='replace')
    m=re.search(r'(?:const\s+D|window\.OTC_DASHBOARD)\s*=\s*',text); ob=text.index('{',m.end())
    D,end=json.JSONDecoder().raw_decode(text[ob:])
    old=D['mol_perf'].get(FAM_KEY,{})
    print(f'\nPREGABALIN actual: ultimo mes {max(old.get("monthly",{}),key=msort,default="?")}')
    D['mol_perf'][FAM_KEY]=fam
    print(f'PREGABALIN nuevo (multidosis): ultimo mes {last}')
    if a.dry_run:
        print('\nDRY RUN: nada escrito.'); return 0
    HTML.write_text(text[:ob]+json.dumps(D,ensure_ascii=False)+text[ob+end:], encoding='utf-8', newline='')
    print(f'\nEscrito SNC/index.html ({HTML.stat().st_size:,} bytes)')
    return 0


if __name__=='__main__':
    sys.exit(main())
