#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
shared/build-mercados-atc.py

Agrega a cada data.js la clave `mercadosATC`: una VISTA ALTERNATIVA del mercado por
clase terapeutica ATC III, para poder analizar cada marca contra su universo amplio
ademas de contra su molecula exacta.

Pedido del usuario: "mantene los que estan pero agrega por ejemplo para roxolan el
mercado de hipolipemeantes" / "no quiero agregar productos y demas, son formas extras
de analizar el mercado".

Ejemplo: ROXOLAN hoy se mide contra el mercado de ROSUVASTATIN (10,1M MAT, MS% 2,0%).
Con esta vista se lo puede medir tambien contra C10A - PRD REGULADORES LIPIDOS
(144 marcas, 19,6M MAT, MS% 1,05%).

POR QUE UNA CLAVE APARTE Y NO UNA FAMILIA MAS DE mol_perf
---------------------------------------------------------
build-total.py y check-total-consistency.py recorren mol_perf para armar el mercado y el
SIE de COMPANIA. Una familia con el universo ATC ancho inflaria el mercado total y
bajaria el MS% publicado de compania. Y recompute-mol-perf-aggregates.py redefine el
total de cada familia como la suma de sus productos, asi que la contaminacion seria
persistente. Por eso vive en `mercadosATC`, que ningun script de agregacion mira.
El nombre NO empieza con 'mol_perf' por higiene de grep: el triage de esta superficie se
hace buscando 'mol_perf' (519 ocurrencias en 103 archivos), y una clave hermana que caiga
en ese grep va a aparecer en toda auditoria futura del agregado de compania, invitando
justo al error de ensenarle a un agregador a incluirla. (No hay ningun script que filtre
claves por patron tipo `'mol_perf' in k` -- se verifico: 0 resultados en el repo.)

YA EXISTIA UN INTENTO DE ESTO Y ESTA ROTO: respPerf
---------------------------------------------------
cardio/ATB/respiratorio ya traen del build una clave `respPerf` con la shape
respPerf[familia] = {molecule: {all,etico,popular}, atc: {all,etico,popular}}, que es
exactamente esta idea. Pero: sus listas estan truncadas a 8 productos Y SIN fila de
residuo, asi que la suma de products da muy por debajo de su propio family.mat (en
cardio ROXOLAN: atc.all suma 9.530.488 contra family.mat 20.671.610, -54%; molecule.all
-31,5%); los seis nodos etico/popular estan VACIOS; y el corte va un mes atrasado
(May 2026 vs Jun 2026 de mol_perf). Solo respiratorio la consume, en el grafico via
getPerfData (respiratorio/index.html:643). Esto NO la arregla ni la borra: se reporta
aparte porque decidir entre repararla, repuntarla a esta clave o retirar sus controles es
del usuario.

DE DONDE SALE CADA COSA
-----------------------
  - La CLASIFICACION (que productos forman la clase, y su nombre en castellano) sale del
    Ateneo de IQVIA, que trae ATC III / ATC IV con etiqueta legible
    ('C10A - PRD REGULADORES LIPIDOS').
  - Las UNIDADES salen del master AR_PM, que tiene mensual. El Ateneo esta en MAT MOVIL
    (rolling 12 meses) y la tabla multi-periodo necesita mensual para poder armar
    MAT/YTD/TRIM/MES.
  - Los dos archivos cubren el MISMO universo (8.770 productos cada uno), asi que el
    cruce es 1 a 1 por nombre de producto.

SHAPE (deliberadamente minima)
------------------------------
    mercadosATC = {
      "corte": "Jun 2026",
      "clases": { "C10A - PRD REGULADORES LIPIDOS": {
                    "products": [ {"prod": "...", "is_sie": false,
                                   "monthly_vals": {"Feb 2024": 123, ...}}, ... ] } },
      "porFamilia": { "ROXOLAN": "C10A - PRD REGULADORES LIPIDOS", ... },
      "propios":    { "ROXOLAN": ["ROXOLAN (SIE)"], ... }
    }
Se guardan SOLO monthly_vals + is_sie porque shared/multi-period-table.js
(buildIqviaFamilies) no lee nada mas: arma el mercado sumando los monthly_vals de los
productos y calcula MS%/IE/Var pp por su cuenta con computeFamily/computeBrand. Guardar
quarterly/ytd/mat/ms_* seria ~4x de bytes que nadie lee.
Las clases se guardan UNA sola vez y `porFamilia` apunta: varias familias comparten
clase (DILATREND y DILATREND AP son las dos C07A; DIOVAN D, ENTRESTO, EXFORGE y
EXFORGE D son las cuatro C09D), asi que deduplicar evita repetir la serie entera.
`propios` es la lista EXACTA de marcas propias de cada familia, y se pasa como
budIqviaMap para que el MS% de la fila familia mida SOLO su marca y no todo lo
Siegfried que haya en la clase (en C10A estan ROXOLAN y ROXOLAN PLUS).

VALIDACION CRUZADA
------------------
El MAT de cada clase calculado desde el AR_PM (suma de los ultimos 12 meses) se compara
contra el MAT MOVIL que el Ateneo trae para esa misma clase. Son dos archivos y dos
caminos distintos: si cierran, la clase esta bien construida. Se reporta la diferencia
relativa por clase y se ABORTA si alguna se pasa de --tol.

Uso:
    py shared/build-mercados-atc.py [--ateneo <xlsx>] [--master <xlsx>]
                                    [--nivel atc3|atc4] [--tol 0.01]
                                    [--dry-run] [--report <json>]
"""
from __future__ import annotations
import argparse, importlib.util, json, re, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent

# Lista canonica de productos excluidos de TODOS los analisis
# (shared/excluded-products.py). Hay que respetarla aca tambien: la clase ATC del Ateneo
# los trae de vuelta -- 'CALCITOL D3 (SIE)' reaparecio en A11C de mujer y lo detecto
# check-syntax-and-consistency.py, que escanea cualquier data file en busca de estos
# nombres. El nombre del archivo tiene guion, asi que se carga a mano.
_spec = importlib.util.spec_from_file_location('excluded', REPO / 'shared' / 'excluded-products.py')
_excl = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_excl)
is_excluded = _excl.is_excluded
HUB = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs')
DEFAULT_MASTER = HUB / '_iqvia-master' / '2026-06' / 'AR_PM_FV_Standard_Jul-2026.xlsx'
LINES = ['cardio/data.js', 'ATB/data.js', 'OTC/data.js', 'respiratorio/data.js',
         'mujer/data.js', 'SNC/data.js', 'dermatologia/data.js']
MES = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
MI = {m: i + 1 for i, m in enumerate(MES)}
MES_LARGO = {'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june': 6,
             'july': 7, 'august': 8, 'september': 9, 'october': 10, 'november': 11,
             'december': 12}


def msort(mk):
    p = str(mk).split()
    return (int(p[1]), MI.get(p[0], 0)) if len(p) == 2 and p[1].isdigit() else (0, 0)


def es_sie(nombre, manuf):
    return '(SIE)' in str(nombre).upper() or 'SIEGFRIED' in str(manuf or '').upper()


def leer_ateneo(path, nivel):
    """producto -> {manuf, clase, mat_ateneo}. La clase es la etiqueta ATC III (o IV)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    ci, mat_cols = {}, []
    for i, h in enumerate(hdr):
        if h is None:
            continue
        s = str(h).replace('\n', ' ').strip()
        sl = s.lower()
        if sl.startswith('product'):
            ci['prod'] = i
        elif sl.startswith('manufacturer'):
            ci['manuf'] = i
        elif sl.startswith('atc iv'):
            ci['atc4'] = i
        elif sl.startswith('atc iii'):
            ci['atc3'] = i
        if sl.startswith('units') and 'mat m' in sl:
            m = re.search(r'mat m\s+(\d{4})\s+([a-z]+)', sl)
            if m and m.group(2) in MES_LARGO:
                mat_cols.append((i, int(m.group(1)), MES_LARGO[m.group(2)]))
    faltan = [k for k in ('prod', nivel) if k not in ci]
    if faltan:
        raise SystemExit('ERROR: el Ateneo no tiene columna(s) {}'.format(faltan))
    mat_cols.sort(key=lambda t: (t[1], t[2]))
    ult = mat_cols[-1] if mat_cols else None
    prods = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        j = ci['prod']
        if j >= len(row) or not row[j]:
            continue
        nom = str(row[j]).strip()
        d = prods.setdefault(nom, {'manuf': '', 'clase': '', 'mat_ateneo': 0.0})
        for k, dst in (('manuf', 'manuf'), (nivel, 'clase')):
            jj = ci.get(k)
            if jj is not None and jj < len(row) and row[jj] and not d[dst]:
                d[dst] = str(row[jj]).strip()
        if ult and ult[0] < len(row) and isinstance(row[ult[0]], (int, float)):
            d['mat_ateneo'] += row[ult[0]]
    wb.close()
    corte = '{} {}'.format(MES[ult[2] - 1], ult[1]) if ult else None
    return prods, corte


def leer_master(path):
    """producto -> {manuf, monthly}. Mensual real, es de donde salen las unidades."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    ci, mcols = {}, []
    for i, h in enumerate(hdr):
        if not h:
            continue
        s = str(h).strip()
        sl = s.replace('\n', ' ').strip().lower()
        if sl.startswith('product'):
            ci['prod'] = i
        elif sl.startswith('manufacturer'):
            ci['manuf'] = i
        if s.startswith('Units'):
            a = (s.split('\n', 1)[-1] if '\n' in s else s[len('Units'):]).strip()
            if a.upper().startswith(('MAT', 'YTD')):
                continue
            m = re.match(r'(\w+)\s+(\d{4})$', a)
            if m and m.group(1) in MI:
                mcols.append((i, '{} {}'.format(m.group(1), m.group(2))))
    prods = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        j = ci['prod']
        if j >= len(row) or not row[j]:
            continue
        nom = str(row[j]).strip()
        d = prods.setdefault(nom, {'manuf': '', 'monthly': defaultdict(float)})
        jm = ci.get('manuf')
        if jm is not None and jm < len(row) and row[jm] and not d['manuf']:
            d['manuf'] = str(row[jm]).strip()
        for c, mk in mcols:
            if c < len(row) and isinstance(row[c], (int, float)):
                d['monthly'][mk] += row[c]
    wb.close()
    for d in prods.values():
        d['monthly'] = {k: int(round(v)) for k, v in d['monthly'].items()}
    return prods, [mk for _, mk in mcols]


def cargar_data_js(rel):
    text = (REPO / rel).read_text(encoding='utf-8', errors='replace')
    m = re.search(r'window\.OTC_DASHBOARD\s*=\s*', text)
    if not m:
        return None
    s = text.index('{', m.end())
    D, end = json.JSONDecoder().raw_decode(text[s:])
    return {'text': text, 'ini': s, 'fin': s + end, 'D': D}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--ateneo', default=None, help='xlsx del Ateneo (default: el mas reciente del hub)')
    ap.add_argument('--master', default=str(DEFAULT_MASTER))
    ap.add_argument('--nivel', choices=['atc3', 'atc4'], default='atc3',
                    help='atc3 = clase terapeutica amplia (default); atc4 = subclase')
    ap.add_argument('--tol', type=float, default=0.01,
                    help='tolerancia relativa del cruce AR_PM vs Ateneo por clase (default 1%%)')
    ap.add_argument('--dry-run', action='store_true')
    ap.add_argument('--report', default=None)
    a = ap.parse_args()
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')

    at_path = Path(a.ateneo) if a.ateneo else None
    if at_path is None:
        cands = sorted(HUB.glob('Ateneo*.xlsx'), key=lambda p: p.stat().st_mtime, reverse=True)
        if not cands:
            print('ERROR: no encuentro Ateneo*.xlsx en {}'.format(HUB), file=sys.stderr)
            return 2
        at_path = cands[0]
    mp_path = Path(a.master)
    for p in (at_path, mp_path):
        if not p.is_file():
            print('ERROR: no existe {}'.format(p), file=sys.stderr)
            return 2

    print('Ateneo (clasificacion): {}'.format(at_path.name))
    AT, corte_at = leer_ateneo(at_path, a.nivel)
    print('  {} productos, MAT movil de referencia: {}'.format(len(AT), corte_at))
    print('Master  (unidades):     {}'.format(mp_path.name))
    PROD, MMONTHS = leer_master(mp_path)
    print('  {} productos, {} meses ({}..{})'.format(len(PROD), len(MMONTHS), MMONTHS[0], MMONTHS[-1]))

    # indice de la clase -> productos, salteando los excluidos forever
    por_clase = defaultdict(list)
    excluidos_vistos = set()
    for n, d in AT.items():
        if not d['clase']:
            continue
        if is_excluded(n):
            excluidos_vistos.add(n)
            continue
        por_clase[d['clase']].append(n)
    print('  {} clases {} distintas'.format(len(por_clase), a.nivel.upper()))
    if excluidos_vistos:
        print('  {} producto(s) de la lista de exclusiones salteados: {}'.format(
            len(excluidos_vistos), ', '.join(sorted(excluidos_vistos))))

    # ── cruce AR_PM vs Ateneo por clase, sobre la ventana MAT del Ateneo ──
    if corte_at:
        i_corte = MMONTHS.index(corte_at) if corte_at in MMONTHS else len(MMONTHS) - 1
    else:
        i_corte = len(MMONTHS) - 1
    w12 = MMONTHS[max(0, i_corte - 11):i_corte + 1]
    print('  ventana MAT del cruce: {} .. {}'.format(w12[0], w12[-1]))

    report = {'ateneo': at_path.name, 'master': mp_path.name, 'nivel': a.nivel,
              'corte': corte_at, 'clases': [], 'lineas': []}
    peor = (0.0, '')
    cruce_ok = 0
    for clase, nombres in sorted(por_clase.items()):
        mat_pm = sum(sum(PROD[n]['monthly'].get(mk, 0) for mk in w12) for n in nombres if n in PROD)
        mat_at = int(round(sum(AT[n]['mat_ateneo'] for n in nombres)))
        if mat_at <= 0:
            continue
        rel = abs(mat_pm - mat_at) / mat_at
        report['clases'].append({'clase': clase, 'marcas': len(nombres),
                                 'mat_master': mat_pm, 'mat_ateneo': mat_at,
                                 'dif_rel_pct': round(rel * 100, 4)})
        if rel > peor[0]:
            peor = (rel, '{} (master {:,} vs ateneo {:,})'.format(clase, mat_pm, mat_at))
        if rel <= a.tol:
            cruce_ok += 1
    print()
    print('CRUCE AR_PM vs Ateneo: {} de {} clases cierran dentro de {:.1f}%. Peor: {:.3f}% en {}'.format(
        cruce_ok, len(report['clases']), a.tol * 100, peor[0] * 100, peor[1]))
    if peor[0] > a.tol:
        print('ABORTADO: hay clases donde las dos fuentes no cierran -> la clasificacion '
              'no se puede cruzar 1 a 1 por nombre de producto.', file=sys.stderr)
        if a.report:
            Path(a.report).write_text(json.dumps(report, ensure_ascii=False, indent=1), encoding='utf-8')
        return 1

    print()
    print('{:<14} {:<22} {:<34} {:>5} {:>13} {:>7}'.format(
        'linea', 'familia', 'clase ATC (vista amplia)', 'marc', 'MAT clase', 'MS%'))
    print('-' * 104)

    planned = []
    for rel_path in LINES:
        info = cargar_data_js(rel_path)
        if not info:
            continue
        D = info['D']
        mp = D.get('mol_perf') or {}
        if not mp:
            continue
        linea = rel_path.split('/')[0]
        bim = D.get('budIqviaMap') or {}

        # ventana de meses de la linea: la union de los monthly de sus familias
        meses = set()
        for f in mp.values():
            meses |= set((f.get('monthly') or {}).keys())
        meses = [mk for mk in sorted(meses, key=msort) if mk in MMONTHS]
        if not meses:
            continue
        ult = meses[-1]
        i_u = MMONTHS.index(ult)
        w12_ln = MMONTHS[max(0, i_u - 11):i_u + 1]

        clases_out, por_fam, propios = {}, {}, {}
        for fam, f in mp.items():
            sies = [p for p in (f.get('products') or []) if p.get('is_sie')]
            if not sies:
                continue
            # marcas PROPIAS de esta familia: budIqviaMap manda; si no hay, los is_sie
            own = list(bim.get(fam) or [str(p.get('prod')) for p in sies])
            # la clase se toma de la marca propia principal (la de mayor MAT)
            ppal = max(sies, key=lambda p: (p.get('mat') or {}).get(ult, 0) or 0)
            clase = (AT.get(str(ppal.get('prod'))) or {}).get('clase')
            if not clase:
                continue
            nombres = [n for n in por_clase.get(clase, []) if n in PROD]
            if not nombres:
                continue
            if clase not in clases_out:
                prods_out = []
                for n in sorted(nombres, key=lambda x: -sum(PROD[x]['monthly'].get(mk, 0) for mk in w12_ln)):
                    monthly = {mk: int(PROD[n]['monthly'].get(mk, 0) or 0) for mk in meses}
                    if not any(monthly.values()):
                        continue        # no aporta nada a la ventana de la linea
                    prods_out.append({'prod': n,
                                      'is_sie': es_sie(n, PROD[n].get('manuf')),
                                      'monthly_vals': monthly})
                if not prods_out:
                    continue
                clases_out[clase] = {'products': prods_out}
            por_fam[fam] = clase
            propios[fam] = own
            # reporte
            ps = clases_out[clase]['products']
            mat_clase = sum(sum(p['monthly_vals'].get(mk, 0) for mk in w12_ln) for p in ps)
            mat_own = sum(sum(p['monthly_vals'].get(mk, 0) for mk in w12_ln)
                          for p in ps if p['prod'] in own)
            print('{:<14} {:<22} {:<34} {:>5} {:>13,} {:>6.2f}%'.format(
                linea[:14], fam[:22], clase[:34], len(ps), mat_clase,
                mat_own / mat_clase * 100 if mat_clase else 0))

        if not clases_out:
            continue
        D['mercadosATC'] = {'corte': ult, 'nivel': a.nivel,
                            'clases': clases_out, 'porFamilia': por_fam, 'propios': propios}
        planned.append((rel_path, info))
        report['lineas'].append({'linea': linea, 'familias': len(por_fam),
                                 'clases': len(clases_out),
                                 'productos': sum(len(c['products']) for c in clases_out.values())})

    print()
    for r in report['lineas']:
        print('  {:<14} {} familias -> {} clases, {} productos'.format(
            r['linea'], r['familias'], r['clases'], r['productos']))
    if a.report:
        Path(a.report).write_text(json.dumps(report, ensure_ascii=False, indent=1), encoding='utf-8')
        print('reporte -> {}'.format(a.report))
    if a.dry_run:
        print('\nDRY RUN: no se escribio nada.')
        return 0
    print()
    for rel_path, info in planned:
        p = REPO / rel_path
        antes = p.stat().st_size
        p.write_text(info['text'][:info['ini']] + json.dumps(info['D'], ensure_ascii=False)
                     + info['text'][info['fin']:], encoding='utf-8', newline='')
        print('-> {} ({:,} -> {:,} bytes, {:+.1f}%)'.format(
            rel_path, antes, p.stat().st_size, (p.stat().st_size / antes - 1) * 100))
    return 0


if __name__ == '__main__':
    sys.exit(main())
