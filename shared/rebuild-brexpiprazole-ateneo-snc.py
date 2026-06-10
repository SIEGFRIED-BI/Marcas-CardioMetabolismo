# -*- coding: utf-8 -*-
"""Reconstruye mol_perf['BREXPIPRAZOLE'] de SNC desde el Ateneo MAT Movil.

Problema: lo habia armado desde AR_PM (cierra Mar-2026) y BREXIL (SIE) quedo en
0 -> "no trae datos". BREXIL lanzo en ABRIL-2026; AR_PM no lo tiene, pero el
Ateneo (mismo panel/escala, validado: REXULTI Mar26=4579 en ambos; DIAZEPAM
Mar/Abr identico a lo cargado) llega a Abr-2026 y SI lo tiene.

Mercado brexpiprazol completo (Ateneo, molecula BREXPIPRAZOLE):
  REXULTI (B7I), BREMIDAN (A5U), LAPLEX (RMM), BREXIL (SIE).
BREMIDAN/LAPLEX/BREXIL son lanzamientos Abr-2026.

Fuente: Ateneo Total - MAT Movil (3) - columnas mensuales limpias c4..c63
(May 2021 -> Apr 2026; las columnas 'X to Y' son agregados trimestrales, se
ignoran). is_sie = manufacturer == SIEGFRIED. Backfill: Apr-2021 de REXULTI se
toma de la familia actual (AR_PM lo tiene) para no perder el primer mes.

Recalcular kpis aparte. Uso: py shared/rebuild-brexpiprazole-ateneo-snc.py [--dry-run]
"""
from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
HTML = REPO / 'SNC' / 'index.html'
ATENEO = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs\_iqvia-master\2026-04\Ateneo Total - MAT Movil_May-19-2026 (3).xlsx')
FAM = 'BREXPIPRAZOLE'
FAM_LABEL = 'Brexpiprazol'

MES_INV = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,
           'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
NUM2 = {v:k for k,v in MES_INV.items()}
MONTH_RE = re.compile(r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec) \d{4}$')


def msort(mk):
    p = mk.split(); return int(p[1])*100 + MES_INV.get(p[0],0) if len(p)==2 else 0
def quarter_key(mk):
    p = mk.split()
    if len(p)!=2: return ''
    m = MES_INV.get(p[0]); return f'Q{(m-1)//3+1} {p[1]}' if m else ''
def agg_quarterly(monthly):
    o = defaultdict(int)
    for mk,v in monthly.items():
        q = quarter_key(mk)
        if q: o[q]+=v
    return dict(o)
def agg_ytd(monthly, cierre):
    by = defaultdict(int)
    for mk,v in monthly.items():
        p = mk.split()
        if len(p)!=2: continue
        mn = MES_INV.get(p[0])
        if mn and mn<=cierre: by[p[1]]+=v
    return {f'{NUM2[cierre]} {y}': v for y,v in by.items()}
def agg_mat(monthly, cierre):
    years = {int(mk.split()[1]) for mk in monthly if len(mk.split())==2 and mk.split()[0] in MES_INV}
    out = {}
    for y in sorted(years):
        tot = 0
        for back in range(11,-1,-1):
            idx = (y*12 + (cierre-1)) - back
            yy, mm = divmod(idx, 12)
            tot += int(monthly.get(f'{NUM2[mm+1]} {yy}', 0) or 0)
        out[f'{NUM2[cierre]} {y}'] = tot
    return out


def load_brexpiprazole(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True); ws = wb.active
    r1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    ci = {}
    for i,h in enumerate(r1):
        s = str(h or '').strip().lower()
        if s.startswith('manufacturer'): ci['man'] = i
        elif s.startswith('product'): ci['prod'] = i
        elif s.startswith('molecules'): ci['mol'] = i
    mcols = [(i, str(h).split('\n')[-1].strip()) for i,h in enumerate(r1)
             if h and str(h).startswith('Units') and MONTH_RE.match(str(h).split('\n')[-1].strip())]
    prods = defaultdict(lambda: {'manuf': None, 'monthly': defaultdict(float)})
    for row in ws.iter_rows(min_row=2, values_only=True):
        if ci['mol'] >= len(row) or str(row[ci['mol']] or '').strip().upper() != FAM: continue
        prod = str(row[ci['prod']] or '')
        if not prod: continue
        b = prods[prod]; b['manuf'] = row[ci['man']] if ci['man'] < len(row) else None
        for cidx, mk in mcols:
            if cidx < len(row) and row[cidx] is not None:
                try: b['monthly'][mk] += float(row[cidx])
                except (ValueError, TypeError): pass
    wb.close()
    months = [mk for _,mk in sorted(mcols, key=lambda x: msort(x[1]))]
    return months, {p: {'manuf': i['manuf'], 'monthly': {mk: int(round(v)) for mk,v in i['monthly'].items() if v}}
                    for p,i in prods.items()}


def build_family(prods):
    fam_monthly = defaultdict(int); plist = []
    for name,info in prods.items():
        mv = info['monthly']
        for mk,v in mv.items(): fam_monthly[mk] += v
        is_sie = str(info.get('manuf') or '').strip().upper() == 'SIEGFRIED'
        plist.append({'prod':name,'manuf':info.get('manuf') or '','is_sie':is_sie,'monthly_vals':mv})
    fam_monthly = dict(fam_monthly)
    cierre = MES_INV.get(max(fam_monthly,key=msort).split()[0],12) if fam_monthly else 12
    fam_q=agg_quarterly(fam_monthly); fam_y=agg_ytd(fam_monthly,cierre); fam_m=agg_mat(fam_monthly,cierre)
    for p in plist:
        mv=p['monthly_vals']
        p['quarterly_vals']=agg_quarterly(mv); p['ytd']=agg_ytd(mv,cierre); p['mat']=agg_mat(mv,cierre)
        p['ms_monthly']  ={mk: round(mv.get(mk,0)/fv*100,2) if fv>0 else 0 for mk,fv in fam_monthly.items()}
        p['ms_quarterly']={qk: round(p['quarterly_vals'].get(qk,0)/fv*100,2) if fv>0 else 0 for qk,fv in fam_q.items()}
        p['ms_ytd']      ={y: round(p['ytd'].get(y,0)/fv*100,2) if fv>0 else 0 for y,fv in fam_y.items()}
        p['ms_mat']      ={y: round(p['mat'].get(y,0)/fv*100,2) if fv>0 else 0 for y,fv in fam_m.items()}
    plist.sort(key=lambda p: (not p['is_sie'], -(p['monthly_vals'].get(max(p['monthly_vals'],key=msort,default=''),0) if p['monthly_vals'] else 0)))
    return {'family':FAM_LABEL,'products':plist,'monthly':fam_monthly,'quarterly':fam_q,'ytd':fam_y,'mat':fam_m}


def main():
    ap=argparse.ArgumentParser(); ap.add_argument('--dry-run',action='store_true'); a=ap.parse_args()
    if hasattr(sys.stdout,'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')
    if not ATENEO.is_file(): print('ERROR Ateneo no existe:',ATENEO); return 2
    months,prods=load_brexpiprazole(ATENEO)

    text=HTML.read_text(encoding='utf-8',errors='replace')
    m=re.search(r'const\s+D\s*=\s*',text); ob=text.index('{',m.end())
    D,end=json.JSONDecoder().raw_decode(text[ob:])
    # backfill Apr 2021 de REXULTI desde la familia actual (Ateneo arranca May 2021)
    cur=D['mol_perf'].get(FAM,{})
    for p in cur.get('products',[]):
        if p['prod'] in prods and 'Apr 2021' in p.get('monthly_vals',{}):
            prods[p['prod']]['monthly'].setdefault('Apr 2021', p['monthly_vals']['Apr 2021'])

    fam=build_family(prods)
    last=max(fam['monthly'],key=msort)
    print(f'BREXPIPRAZOLE (Ateneo): {len(months)} meses ({months[0]}..{months[-1]}), {len(fam["products"])} productos')
    print(f'  mercado {last} = {fam["monthly"][last]:,} u')
    for p in fam['products']:
        print(f'    {"SIE " if p["is_sie"] else "    "}{p["monthly_vals"].get(last,0):>6} {p["prod"]:18} MS%_YTD={p["ms_ytd"].get(max(p["ms_ytd"],key=msort,default=""),0)}')

    D['mol_perf'][FAM]=fam
    if a.dry_run:
        print('\nDRY RUN: nada escrito.'); return 0
    HTML.write_text(text[:ob]+json.dumps(D,ensure_ascii=False)+text[ob+end:], encoding='utf-8', newline='')
    print(f'\nEscrito SNC/index.html ({HTML.stat().st_size:,} bytes)')
    return 0


if __name__=='__main__':
    sys.exit(main())
