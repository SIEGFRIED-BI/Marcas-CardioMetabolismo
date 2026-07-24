# -*- coding: utf-8 -*-
"""Rueda respPerf (el gráfico interactivo de Mercado IQVIA de respiratorio) al último
mes del master, de forma QUIRÚRGICA: APPEND del mes nuevo a cada producto de cada slice
poblado + recomputo de quarterly/ytd/mat/ms_* con cierre = mes nuevo.

Por qué existe: el roll mensual (roll-mol-perf-month.py) actualiza SOLO D.mol_perf, no
D.respPerf. respPerf es la fuente del gráfico interactivo de respi (getPerfData ->
RESP_PERF[fam][molecule|atc][all|etico|popular]). Si queda viejo (p.ej. May 2026 cuando
mol_perf/meta ya están en Jun 2026), el chart busca las keys ytd/mat del cierre nuevo,
no las encuentra y muestra "Sin datos para este filtro" en TODAS las familias.

Append por nombre exacto de producto ('MARCA (COD)') desde la col 'Product' del master,
mensual = col 'Units <Mon> <YYYY>' (NO MAT/YTD/to). Segmento 'all' = suma todos los packs
del producto (sin filtrar por Market E/OTC); etico/popular sólo si el slice está poblado
(hoy están vacíos, quedan preparados). No re-deriva el set de productos (append-only, igual
que el roll de mol_perf): un producto ausente del master ese mes queda en 0.

Uso: py shared/roll-resp-perf-month.py --master <Ateneo/AR_PM del cierre> [--dry-run]
"""
from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
DATA = REPO / 'respiratorio' / 'data.js'
MES = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
NUM = {v:k for k,v in MES.items()}
SEG_MKT = {'etico':'ETICO', 'popular':'POPULAR'}  # filtro Market (E/OTC); 'all' = sin filtro


def msort(mk):
    p = str(mk).split()
    return int(p[1])*100 + MES.get(p[0],0) if len(p)==2 and p[0] in MES else 0


def quarter_key(mk):
    p = mk.split()
    if len(p)!=2 or p[0] not in MES: return ''
    return f'Q{(MES[p[0]]-1)//3+1} {p[1]}'


def agg_q(monthly):
    out = defaultdict(int)
    for mk,v in monthly.items():
        qk = quarter_key(mk)
        if qk:
            try: out[qk]+=int(round(float(v or 0)))
            except (TypeError,ValueError): pass
    return dict(out)


def agg_ytd(monthly, cierre):
    out = defaultdict(int)
    for mk,v in monthly.items():
        p = mk.split()
        if len(p)!=2 or p[0] not in MES: continue
        if MES[p[0]] <= cierre:
            try: out[p[1]]+=int(round(float(v or 0)))
            except (TypeError,ValueError): pass
    return {f'{NUM[cierre]} {y}': v for y,v in out.items()}


def agg_mat(monthly, cierre):
    years = set()
    for mk in monthly:
        p = mk.split()
        if len(p)==2 and p[0] in MES:
            try: years.add(int(p[1]))
            except ValueError: pass
    out = {}
    for y in sorted(years):
        tot = 0
        for back in range(11,-1,-1):
            idx = (y*12+(cierre-1))-back
            yy,mm = divmod(idx,12)
            v = monthly.get(f'{NUM[mm+1]} {yy}')
            if v is not None:
                try: tot+=int(round(float(v or 0)))
                except (TypeError,ValueError): pass
        out[f'{NUM[cierre]} {y}'] = tot
    return out


def safe(num, den): return round((num or 0)/den*100, 2) if den else 0


def load_master(path):
    """Devuelve (by_prod_all, by_prod_seg, new_month). by_prod_all[prod][mes]=units (todos los packs);
    by_prod_seg[('ETICO'|'POPULAR')][prod][mes]=units."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True); ws = wb.active
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    ci_prod = ci_mkt = None
    mcols = []  # (idx, 'Mon YYYY')
    for i,h in enumerate(hdr):
        s = str(h or '').replace('\n',' ').strip()
        sl = s.lower()
        if sl == 'product': ci_prod = i
        elif sl.startswith('market'): ci_mkt = i
        m = re.match(r'^Units ([A-Z][a-z]{2}) (\d{4})$', s)  # mensual puro: NO MAT/YTD/'to'
        if m and m.group(1) in MES:
            mcols.append((i, f'{m.group(1)} {m.group(2)}'))
    if ci_prod is None:
        raise ValueError(f'no encontre columna Product en {hdr[:9]}')
    by_all = defaultdict(lambda: defaultdict(float))
    by_seg = {'ETICO': defaultdict(lambda: defaultdict(float)),
              'POPULAR': defaultdict(lambda: defaultdict(float))}
    for row in ws.iter_rows(min_row=2, values_only=True):
        prod = str(row[ci_prod] or '').strip()
        if not prod: continue
        mkt = str(row[ci_mkt] or '').strip().upper() if ci_mkt is not None else ''
        for ci,mk in mcols:
            if ci < len(row) and isinstance(row[ci],(int,float)):
                by_all[prod][mk] += row[ci]
                if mkt in by_seg: by_seg[mkt][prod][mk] += row[ci]
    wb.close()
    new_month = max((mk for _,mk in mcols), key=msort)
    return by_all, by_seg, new_month


def recompute_slice(sl, cierre):
    prods = sl.get('products', [])
    fam_m = defaultdict(int)
    for p in prods:
        for mk,v in (p.get('monthly_vals') or {}).items():
            try: fam_m[mk]+=int(round(float(v or 0)))
            except (TypeError,ValueError): pass
    fam_m = dict(fam_m); fam_q = agg_q(fam_m); fam_y = agg_ytd(fam_m,cierre); fam_mat = agg_mat(fam_m,cierre)
    for key,val in (('monthly',fam_m),('quarterly',fam_q),('ytd',fam_y),('mat',fam_mat)):
        sl[key] = val
    for p in prods:
        mv = p.get('monthly_vals') or {}
        p['quarterly_vals'] = agg_q(mv)
        p['ytd'] = agg_ytd(mv,cierre)
        p['mat'] = agg_mat(mv,cierre)
        p['ms_monthly']   = {mk: safe(mv.get(mk,0), fam_m.get(mk,0)) for mk in fam_m}
        p['ms_quarterly'] = {qk: safe(p['quarterly_vals'].get(qk,0), fam_q.get(qk,0)) for qk in fam_q}
        p['ms_ytd']       = {yk: safe(p['ytd'].get(yk,0), fam_y.get(yk,0)) for yk in fam_y}
        p['ms_mat']       = {mk: safe(p['mat'].get(mk,0), fam_mat.get(mk,0)) for mk in fam_mat}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--master', required=True)
    ap.add_argument('--dry-run', action='store_true')
    a = ap.parse_args()
    if hasattr(sys.stdout,'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')

    by_all, by_seg, new_month = load_master(Path(a.master))
    cierre = MES[new_month.split()[0]]
    t = DATA.read_text(encoding='utf-8-sig')
    m = re.search(r'window\.OTC_DASHBOARD\s*=\s*', t); ob = t.index('{', m.end())
    D, end = json.JSONDecoder().raw_decode(t[ob:])
    rp = D.get('respPerf', {})
    if not rp:
        print('  (skip) no hay respPerf en respiratorio/data.js'); return 0

    # mes actual de respPerf
    cur = max((mk for fam in rp.values() for cmp in fam.values() for seg in cmp.values()
               for p in (seg or {}).get('products',[]) for mk in (p.get('monthly_vals') or {})),
              key=msort, default='?')
    print(f'respPerf: mes actual {cur} -> nuevo {new_month} (cierre_month={cierre})')

    missing = defaultdict(list); n_slices = 0
    for fam, node in rp.items():
        for cmp in ('molecule','atc'):
            for seg in ('all','etico','popular'):
                sl = (node.get(cmp) or {}).get(seg)
                if not sl or not sl.get('products'): continue
                src = by_all if seg=='all' else by_seg.get(SEG_MKT.get(seg,''), {})
                for p in sl['products']:
                    nm = p['prod']
                    v = int(round(src.get(nm,{}).get(new_month,0))) if nm in src else 0
                    if nm not in src:
                        cur_v = float((p.get('monthly_vals') or {}).get(cur,0) or 0)
                        if cur_v != 0: missing[nm].append(f'{fam}/{cmp}/{seg} ({cur}={int(cur_v)})')
                    (p.setdefault('monthly_vals',{}))[new_month] = v
                recompute_slice(sl, cierre)
                n_slices += 1
        fj = sum(int(round(by_all.get(p['prod'],{}).get(new_month,0))) for p in node['molecule']['all']['products']) if node.get('molecule',{}).get('all',{}).get('products') else 0
        print(f'  {fam:24} slices rodados, mkt mol/all {new_month}={fj}')

    print(f'\n{n_slices} slices rodados a {new_month}.')
    if missing:
        print(f'Productos ausentes del master con {cur}>0 (quedan en 0 el mes nuevo):')
        for nm, locs in list(missing.items())[:20]:
            print(f'   {nm[:40]:40} -> {locs[0]}')

    if a.dry_run:
        print('DRY-RUN: nada escrito.'); return 0
    DATA.write_text(t[:ob] + json.dumps(D, ensure_ascii=False) + t[ob+end:], encoding='utf-8', newline='')
    print(f'Escrito {DATA}.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
