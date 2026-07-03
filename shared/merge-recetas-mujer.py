# -*- coding: utf-8 -*-
"""Merge recetas mujer desde el pivot CloseUp/Qlik (data.js, window.OTC_DASHBOARD).

Mujer usa segmentacion CLASS (SIN ESTROGENO=ISIS FREE, ALTA DOSIS=ISIS, ...). El
MERCADO/denominador de MS% por segmento es BESPOKE (no sale de una regla simple del
pivot: droga_total NO lo reproduce). Por eso este merge NO recomputa la historia:
AGREGA el/los mes(es) nuevos con:
  - rec_ms[fam].sie[m]  = recetas de la marca SIE del segmento (del pivot)
  - rec_ms[fam].ms[m]   = CARRY del ultimo ms existente (MS% estable mes-a-mes)
  - recetas[fam][m]     = {recetas: sie, medicos: sie_med}
  - rec_comp[fam][brand].monthly[m] = recetas del competidor (match por primer token)
Por defecto agrega SOLO el ultimo mes del pivot (el nuevo). Con --all-new agrega todos
los meses del pivot que NO existan ya en rec_ms[fam].sie (nunca pisa historia).
Idempotente. NO toca otras secciones/lineas.

Uso: py shared/merge-recetas-mujer.py --pivot "<pivot.xlsx>" [--month "May 2026"] [--all-new] [--dry-run]
"""
from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
HTML = REPO / 'mujer' / 'data.js'
MES_EN = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}
MESN = {v: k for k, v in MES_EN.items()}

# Familia mujer -> marca SIE del segmento (nombre exacto en el pivot).
FAM_SIE_BRAND = {
    'SIN ESTROGENO':   'ISIS FREE SIN EST SIE',
    'ALTA DOSIS':      'ISIS SIE',
    'BAJA DOSIS 21+7': 'ISIS MINI SIE',
    'BAJA DOSIS 24':   'ISIS MINI 24 SIE',
    'COMPLEX':         'SIDERBLUT COMPLEX SIE',
    'D3':              'TRIP D3 SIE',
    'DELTROX':         'DELTROX SIE',
    'BASE':            'CALCIO BASE SIE',
    'BASE D':          'CALCIO BASE D SIE',
}


def norm(s):
    return re.sub(r'\s+', ' ', str(s or '').upper()).strip()


def msort(k):
    p = str(k).split(); return int(p[1]) * 100 + MESN.get(p[0], 0) if len(p) == 2 else 0


def parse_pivot(path):
    """(market,droga,brand_norm) -> month_en -> {'recetas','medicos'} ; lista de meses."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True); ws = wb.active
    r1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    r2 = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))
    cm = {}; cur = None
    for i, h1 in enumerate(r1):
        if isinstance(h1, datetime):
            cur = f'{MES_EN[h1.month]} {h1.year}'
        h2 = (str(r2[i]) if i < len(r2) and r2[i] else '').lower()
        if 'receta' in h2:
            cm[i] = (cur, 'recetas')
        elif 'médico' in h2 or 'medico' in h2:
            cm[i] = (cur, 'medicos')
    data = defaultdict(lambda: defaultdict(lambda: {'recetas': 0, 'medicos': 0}))
    months = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        market = str(row[0]).strip(); droga = str(row[1] or '').strip(); marca = row[2]
        if not marca or str(marca).strip().lower() == 'totales':
            continue
        key = (market, droga, norm(marca))
        for ci, (mk, kind) in cm.items():
            if ci >= len(row):
                continue
            try:
                v = int(row[ci]) if row[ci] is not None else 0
            except (TypeError, ValueError):
                v = 0
            data[key][mk][kind] = v
            months.add(mk)
    wb.close()
    return data, sorted(months, key=msort)


def find_loc(data, sie_norm, month):
    """(market,droga) donde vive la marca SIE, preferido: market != '-', mayor recetas."""
    cands = []
    for (market, droga, brand), md in data.items():
        if brand == sie_norm:
            cands.append((market, droga, md.get(month, {}).get('recetas', 0)))
    if not cands:
        return None
    cands.sort(key=lambda x: (0 if x[0] in ('-', '') else 1, x[2]), reverse=True)
    return cands[0][0], cands[0][1]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--pivot', required=True)
    ap.add_argument('--month', default=None, help="ej 'May 2026'; default = ultimo mes del pivot")
    ap.add_argument('--all-new', action='store_true', help='agrega TODOS los meses nuevos (no solo el ultimo)')
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    if not Path(args.pivot).is_file():
        print('ERROR: no existe el pivot:', args.pivot); return 1

    data, months = parse_pivot(args.pivot)
    text = HTML.read_text(encoding='utf-8', errors='replace')
    m = re.search(r'(?:const\s+D|window\.OTC_DASHBOARD)\s*=\s*', text)
    ob = text.index('{', m.end())
    D, end = json.JSONDecoder().raw_decode(text[ob:])
    rec_ms = D.setdefault('rec_ms', {})
    rec_comp = D.setdefault('rec_comp', {})
    recetas = D.setdefault('recetas', {})

    def first_word(s):
        s = re.sub(r'\s*\([^)]+\)\s*$', '', str(s)).strip()
        s = re.sub(r'\s+SIE\s*$', '', s, flags=re.I).strip()
        return s.split()[0].upper() if s else ''

    changed = 0
    for fam, sie_brand in FAM_SIE_BRAND.items():
        if fam not in rec_ms:
            print(f'  [{fam}] no esta en rec_ms -> skip'); continue
        sn = norm(sie_brand)
        existing = set((rec_ms[fam].get('sie') or {}).keys())
        # meses objetivo
        if args.month:
            targets = [args.month]
        elif args.all_new:
            targets = [mk for mk in months if mk not in existing]
        else:
            targets = [months[-1]] if months else []
        if not targets:
            print(f'  [{fam}] sin meses nuevos'); continue
        # ultimo ms existente (para carry)
        ms_hist = rec_ms[fam].get('ms') or {}
        last_ms_key = max(ms_hist, key=msort) if ms_hist else None
        carry_ms = ms_hist.get(last_ms_key) if last_ms_key else None
        for tm in targets:
            loc = find_loc(data, sn, tm)
            if not loc:
                print(f'  [{fam}] {tm}: marca SIE {sie_brand!r} no hallada -> skip'); continue
            market, droga = loc
            sie = data.get((market, droga, sn), {}).get(tm, {}).get('recetas', 0)
            sie_med = data.get((market, droga, sn), {}).get(tm, {}).get('medicos', 0)
            rec_ms[fam].setdefault('sie', {})[tm] = sie
            rec_ms[fam].setdefault('ms', {})[tm] = carry_ms if carry_ms is not None else 0
            recetas.setdefault(fam, {})[tm] = {'recetas': sie, 'medicos': sie_med}
            # competidores del mismo (market,droga) -> match por primer token con rec_comp[fam]
            fam_comp = rec_comp.setdefault(fam, {})
            idx = {first_word(k): k for k in fam_comp}
            cu = 0
            for (mk2, dr2, bn), md in data.items():
                if mk2 == market and dr2 == droga:
                    key = idx.get(first_word(bn))
                    if key and isinstance(fam_comp[key], dict):
                        fam_comp[key].setdefault('monthly', {})[tm] = md.get(tm, {}).get('recetas', 0)
                        cu += 1
            changed += 1
            print(f'  [{fam:16}] {tm}: sie={sie} ms={rec_ms[fam]["ms"][tm]}% (carry {last_ms_key}) comp={cu}')

    if args.dry_run:
        print('\nDRY RUN: nada se escribio.'); return 0
    HTML.write_text(text[:ob] + json.dumps(D, ensure_ascii=False) + text[ob + end:],
                    encoding='utf-8', newline='')
    print(f'\nEscrito mujer: {changed} (familia x mes).')
    return 0


if __name__ == '__main__':
    sys.exit(main())
