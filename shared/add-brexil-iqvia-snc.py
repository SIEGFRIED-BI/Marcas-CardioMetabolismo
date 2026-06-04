# -*- coding: utf-8 -*-
"""Agrega el MERCADO IQVIA de BREXIL a SNC (targeted, sin tocar las 10 familias).

BREXIL (brexpiprazol, lanzamiento abr-2026) todavia NO figura en IQVIA
(el corte nacional/regional llega a mar-2026). Su mercado competitivo es
"Brexpiprazol / Aripiprazol": ARIPIPRAZOLE (IRAZEM, LEMIDAL, ARIZIC, ARISDAR,
SIBLIX, IBARAL, ARLEMIDE, APECUR, ARIPIPRAZOL VANNIE, ARINOVA) + BREXPIPRAZOLE
(REXULTI, el competidor originador directo).

Construye una familia mol_perf['ARIPIPRAZOLE'] (mercado BREXIL) desde el master
nacional AR_PM (mismo origen/escala que las 10 familias existentes: DIAZEPAM
Mar2026 master=60844 ~ SNC=60878). Le agrega BREXIL como producto SIE en 0
(recien lanzado), para que aparezca en el grid IQVIA y en el segmentador a 0%
-igual que en Recetas-. NO toca las 10 familias existentes (que tienen abr-2026
de un corte previo). La familia nueva llega a mar-2026: renderPerf arma los meses
POR FAMILIA, asi que muestra "MAR 2026" (su ultimo cierre), no abril en cero.

Tambien: sieMolMap['BREXIL']='ARIPIPRAZOLE', sieProds += 'BREXIL'.
(PROD_MAP -en el JS- se edita aparte.)

Idempotente. Uso: py shared/add-brexil-iqvia-snc.py [--dry-run]
"""
from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
HTML = REPO / 'SNC' / 'index.html'
MASTER = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs\_iqvia-master\2026-04\AR_PM_FV_Standard_Apr-27-2026.xlsx')

FAM_KEY = 'ARIPIPRAZOLE'
FAM_LABEL = 'Aripiprazol / Brexpiprazol'
MARKET_MOLS = {'ARIPIPRAZOLE', 'BREXPIPRAZOLE'}   # mercado competitivo de BREXIL

MES_INV = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,
           'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
NUM2 = {v:k for k,v in MES_INV.items()}


def msort(mk):
    p = mk.split()
    return int(p[1])*100 + MES_INV.get(p[0],0) if len(p)==2 else 0


def quarter_key(mk):
    p = mk.split()
    if len(p)!=2: return ''
    m = MES_INV.get(p[0])
    return f'Q{(m-1)//3+1} {p[1]}' if m else ''


def agg_quarterly(monthly):
    out = defaultdict(int)
    for mk,v in monthly.items():
        q = quarter_key(mk)
        if q: out[q]+=v
    return dict(out)


def agg_ytd(monthly, cierre):
    by = defaultdict(int)
    for mk,v in monthly.items():
        p = mk.split()
        if len(p)!=2: continue
        mn = MES_INV.get(p[0])
        if mn and mn<=cierre: by[p[1]]+=v
    return {f'{NUM2[cierre]} {y}': v for y,v in by.items()}


def agg_mat(monthly, cierre):
    years = {int(mk.split()[1]) for mk in monthly if len(mk.split())==2 and mk.split()[0] in MES_INV}
    out = {}
    for y in sorted(years):
        tot = 0
        for back in range(11,-1,-1):
            idx = (y*12 + (cierre-1)) - back
            yy, mm = divmod(idx, 12)
            tot += int(monthly.get(f'{NUM2[mm+1]} {yy}', 0) or 0)
        out[f'{NUM2[cierre]} {y}'] = tot
    return out


def load_market(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    col_manuf=col_prod=col_mol=None; month_cols=[]
    for i,h in enumerate(row1):
        if not h: continue
        s = str(h).strip(); sn = s.replace('\n',' ').strip().lower()
        if sn.startswith('manufacturer'): col_manuf=i
        elif sn.startswith('product'): col_prod=i
        elif sn.startswith('molecules'): col_mol=i
        if s.startswith('Units') and '\n' in s:
            after = s.split('\n',1)[-1].strip()
            if after.upper().startswith(('MAT','YTD')): continue
            m = re.match(r'(\w+)\s+(\d{4})$', after)
            if m and m.group(1) in MES_INV: month_cols.append((i, f'{m.group(1)} {m.group(2)}'))
    if col_manuf is None: col_manuf=0
    if col_prod is None: col_prod=1
    if col_mol is None: col_mol=5
    prods = defaultdict(lambda: {'manuf':None,'monthly':defaultdict(float)})
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row: continue
        mol = str(row[col_mol]).strip().upper() if col_mol<len(row) and row[col_mol] else ''
        if mol not in MARKET_MOLS: continue
        prod = row[col_prod] if col_prod<len(row) else None
        if not prod: continue
        b = prods[prod]; b['manuf'] = row[col_manuf] if col_manuf<len(row) else None
        for ci,mk in month_cols:
            if ci>=len(row): continue
            v = row[ci]
            if v is None: continue
            try: b['monthly'][mk]+=float(v)
            except (ValueError,TypeError): pass
    wb.close()
    months = [mk for _,mk in sorted(month_cols, key=lambda x: msort(x[1]))]
    return months, {p:{'manuf':i['manuf'],'monthly':{mk:int(round(v)) for mk,v in i['monthly'].items()}}
                    for p,i in prods.items()}


def build_family(prods):
    fam_monthly = defaultdict(int)
    plist = []
    for name,info in prods.items():
        mv = info['monthly']
        for mk,v in mv.items(): fam_monthly[mk]+=v
        plist.append({'prod':name,'manuf':info.get('manuf') or '','is_sie':False,
                      'monthly_vals':mv,'_tmp':None})
    fam_monthly = dict(fam_monthly)
    cierre = MES_INV.get(max(fam_monthly, key=msort).split()[0], 12) if fam_monthly else 12
    fam_q = agg_quarterly(fam_monthly); fam_y = agg_ytd(fam_monthly,cierre); fam_m = agg_mat(fam_monthly,cierre)
    # BREXIL SIE en 0 (recien lanzado; sin unidades IQVIA aun)
    plist.append({'prod':'BREXIL (SIE)','manuf':'SIEGFRIED','is_sie':True,'monthly_vals':{},'_tmp':None})
    for p in plist:
        mv = p.pop('_tmp', None); mv = p['monthly_vals']
        p['quarterly_vals'] = agg_quarterly(mv)
        p['ytd'] = agg_ytd(mv,cierre)
        p['mat'] = agg_mat(mv,cierre)
        p['ms_monthly']   = {mk: round(mv.get(mk,0)/fv*100,2) if fv>0 else 0 for mk,fv in fam_monthly.items()}
        p['ms_quarterly'] = {qk: round(p['quarterly_vals'].get(qk,0)/fv*100,2) if fv>0 else 0 for qk,fv in fam_q.items()}
        p['ms_ytd']       = {y: round(p['ytd'].get(y,0)/fv*100,2) if fv>0 else 0 for y,fv in fam_y.items()}
        p['ms_mat']       = {y: round(p['mat'].get(y,0)/fv*100,2) if fv>0 else 0 for y,fv in fam_m.items()}
    plist.sort(key=lambda p: (not p['is_sie'], -(p['monthly_vals'].get(max(p['monthly_vals'],key=msort,default=''),0) if p['monthly_vals'] else 0)))
    return {'family':FAM_LABEL,'products':plist,'monthly':fam_monthly,
            'quarterly':fam_q,'ytd':fam_y,'mat':fam_m}


def main():
    ap = argparse.ArgumentParser(); ap.add_argument('--dry-run',action='store_true'); a=ap.parse_args()
    if hasattr(sys.stdout,'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')
    if not MASTER.is_file(): print('ERROR master no existe:',MASTER); return 2

    months, prods = load_market(MASTER)
    print(f'Master: {len(months)} meses ({months[0]}..{months[-1]}), {len(prods)} productos del mercado')
    fam = build_family(prods)
    last = max(fam['monthly'],key=msort)
    print(f'Familia {FAM_KEY} ({FAM_LABEL}): {len(fam["products"])} productos (incl BREXIL 0)')
    print(f'  mercado {last} = {fam["monthly"][last]:,} u | productos:')
    for p in fam['products']:
        lv = p['monthly_vals'].get(last,0)
        print(f'    {"SIE " if p["is_sie"] else "    "}{lv:>7} {p["prod"]}')

    text = HTML.read_text(encoding='utf-8', errors='replace')
    m = re.search(r'const\s+D\s*=\s*', text); ob = text.index('{', m.end())
    D, end = json.JSONDecoder().raw_decode(text[ob:])

    if FAM_KEY in D.get('mol_perf',{}):
        print(f'\n[idempotente] {FAM_KEY} ya existe -> se reemplaza')
    D.setdefault('mol_perf',{})[FAM_KEY] = fam
    D.setdefault('sieMolMap',{})['BREXIL'] = FAM_KEY
    if 'BREXIL' not in D.setdefault('sieProds',[]): D['sieProds'].append('BREXIL')

    print('\nmol_perf families ahora:', list(D['mol_perf'].keys()))
    print('sieProds:', D['sieProds'])
    print("sieMolMap['BREXIL']:", D['sieMolMap']['BREXIL'])

    if a.dry_run:
        print('\nDRY RUN: nada escrito.'); return 0
    HTML.write_text(text[:ob] + json.dumps(D, ensure_ascii=False) + text[ob+end:], encoding='utf-8', newline='')
    print(f'\nEscrito SNC/index.html ({HTML.stat().st_size:,} bytes)')
    return 0


if __name__ == '__main__':
    sys.exit(main())
