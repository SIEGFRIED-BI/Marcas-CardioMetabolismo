# -*- coding: utf-8 -*-
"""Construye la vista consolidada 'Total Siegfried' agregando las 7 lineas.

FUENTE: kpis.json (ya trae, por linea y por periodo, numeradores Y denominadores
crudos: units_sie/mercado_units/recetas_sie/mercado_recetas/venta_interna con
curr/prev). Agregar = sumar los aditivos + RECALCULAR IE/MS desde las sumas
(NUNCA promediar los IE por linea — LEMA + skill evolution-index).

  IE   = (Sum propio_curr / Sum propio_prev) / (Sum mkt_curr / Sum mkt_prev) * 100
  MS%  = Sum propio_curr / Sum mkt_curr * 100
  crec = (Sum propio_curr / Sum propio_prev - 1) * 100

Escribe total/data.js (window.TOTAL_SIEGFRIED) con:
  - meta (labels, as_of), company: KPIs totales por periodo, byLine: 7 lineas.
Idempotente. Uso: py shared/build-total.py [--check]
"""
from __future__ import annotations
import argparse, json, sys
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
KPIS = REPO / 'kpis.json'
OUT = REPO / 'total' / 'data.js'
import re as _re
LINE_FILES = {'cardio':'cardio/data.js','antibio':'ATB/data.js','otx':'OTC/data.js',
              'resp':'respiratorio/data.js','mujer':'mujer/data.js','snc':'SNC/data.js',
              'derma':'dermatologia/data.js'}
# Cierre de VENTA -> indices de mes (0=Ene) para la ventana de cada periodo.
#
# ESTABA HARDCODEADO ('CIERRE_IDX = 5  # Jun') y habia que bumpearlo a mano en cada cierre.
# En Jul-2026 nadie lo movio: build-total sumaba la venta de Ene..Jun contra los KPIs de
# Ene..Jul y abortaba con 'venta budget.real=10,350,483 vs kpis=12,238,008 (dif 15.4%)'.
# Como aborta no reescribe total/data.js, y el gate check-total-consistency falla en
# cascada contra los valores viejos. Ahora sale del manifiesto (global.ventaCutoff).
def _cierre_venta():
    """(anio, idx_mes_0based) del ultimo mes COMPLETO de venta interna."""
    try:
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        import manifest
        cm = manifest._g('global', 'ventaCutoff') or manifest.cierre_month()
        y, m = str(cm).split('-')[:2]
        return int(y), int(m) - 1
    except Exception as e:
        raise SystemExit(
            f'FATAL: no resuelvo el cierre de venta desde el manifiesto ({e}). Antes esto '
            'era una constante hardcodeada y publicaba un mes viejo en silencio; abortamos.')


CIERRE_YEAR, CIERRE_IDX = _cierre_venta()


def _ultimos(n):
    """Los n meses que terminan en el cierre, como [(anio, idx0), ...] ASC.
    Misma semantica que month_range() de build-kpis.py: rolling, cruza de anio."""
    out = []
    for k in range(n - 1, -1, -1):
        t = CIERRE_YEAR * 12 + CIERRE_IDX - k
        out.append((t // 12, t % 12))
    return out


# 'semestre' era [Ene..cierre], IGUAL que ytd -- pero build-kpis.py lo define como los
# ULTIMOS 6 MESES (month_range(end, 6)). Con un cierre de junio las dos definiciones dan
# Ene..Jun y coinciden por casualidad, asi que el bug era invisible; en Jul-2026 se
# separaron (7 meses vs 6) y build-total aborto con 'semestre: venta budget.real=
# 12,238,008 vs kpis=10,749,694 (dif 13.8%)'. Ahora todas las ventanas rolling salen de
# _ultimos(), que es la misma cuenta que hace build-kpis.
PERIOD_IDX = {
    'mensual':   _ultimos(1),
    'ytd':       [(CIERRE_YEAR, i) for i in range(0, CIERRE_IDX + 1)],
    'semestre':  _ultimos(6),
    'trimestre': _ultimos(3),
    'mat':       _ultimos(12),
}


def _load_dashboard(path):
    t = (REPO / path).read_text(encoding='utf-8-sig', errors='replace')
    m = _re.search(r'window\.OTC_DASHBOARD\s*=\s*', t)
    ob = t.index('{', m.end())
    return json.JSONDecoder().raw_decode(t[ob:])[0]


# ESTIMADO = archivo de MKT vigente (hubRoot/Estimados VENTA vigentes MKT sidus.xlsx),
# hoja 'Estimados 2026': Tipo|Linea|Codigo|Producto|<meses 2026>, por SKU. Se suman las
# filas CON Linea (todos los tipos: Vigente/Lanzamiento/Sidus/Adquisicion); se EXCLUYEN
# las filas en blanco (son totales/spacers que duplican). NO usar budget-overrides.js
# (estaba viejo -> inflaba el % de cumplimiento).
MKT_LINE_MAP = {'CARDIO MET': 'cardio', 'CARDIO': 'cardio', 'ATB': 'antibio', 'OTC': 'otx',
                'RESPI': 'resp', 'RESP': 'resp', 'MUJER': 'mujer', 'NEURO': 'snc',
                'DERMATO': 'derma', 'DERMA': 'derma'}

def resolve_mkt_file():
    try:
        mani = json.loads((REPO / 'shared' / 'close-manifest.json').read_text(encoding='utf-8'))
        hub = mani['global']['hubRoot'].replace('${OneDrive}', __import__('os').environ.get('OneDrive', ''))
        import glob
        c = glob.glob(str(Path(hub) / 'Estimados*MKT*.xlsx'))
        return c[0] if c else None
    except Exception:
        return None

def mkt_estimado_by_line():
    """Estimado MKT por línea de tablero -> [12 meses 2026]. None si no está el archivo."""
    f = resolve_mkt_file()
    if not f: return None
    import openpyxl
    from collections import defaultdict
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb['Estimados 2026'] if 'Estimados 2026' in wb.sheetnames else wb.active
    r1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    lin_i = next((i for i, h in enumerate(r1) if str(h or '').strip().lower().startswith('linea')), None)
    mcols = []
    for i, h in enumerate(r1):
        s = str(h)
        if s.startswith('2026-'):
            try: mcols.append((i, int(s[5:7]) - 1))
            except Exception: pass
    if lin_i is None or not mcols: wb.close(); return None
    out = defaultdict(lambda: [0.0] * 12)
    for row in ws.iter_rows(min_row=2, values_only=True):
        lin = str(row[lin_i] or '').strip().upper()
        key = MKT_LINE_MAP.get(lin)
        if not key: continue           # blanco / Otros / Gastro -> no tablero
        for c, mi in mcols:
            if c < len(row) and isinstance(row[c], (int, float)): out[key][mi] += row[c]
    wb.close()
    return {k: [round(x) for x in v] for k, v in out.items()}


# --- Ventanas IQVIA por periodo (cierre = Jun 2026) para el recompute deduplicado ---
_M6 = ['Jan','Feb','Mar','Apr','May','Jun']
IQV_WINDOWS = {
    'mensual':   (['Jun 2026'], ['Jun 2025']),
    'ytd':       ([f'{m} 2026' for m in _M6], [f'{m} 2025' for m in _M6]),
    'semestre':  ([f'{m} 2026' for m in _M6], [f'{m} 2025' for m in _M6]),
    'trimestre': (['Apr 2026','May 2026','Jun 2026'], ['Apr 2025','May 2025','Jun 2025']),
    'mat':       ([f'{m} 2025' for m in ['Jul','Aug','Sep','Oct','Nov','Dec']] + [f'{m} 2026' for m in _M6],
                  [f'{m} 2024' for m in ['Jul','Aug','Sep','Oct','Nov','Dec']] + [f'{m} 2025' for m in _M6]),
}


def company_iqvia(master_path):
    """IQVIA sobre el REAL Siegfried (TODOS sus productos en IQVIA, no solo los de
    los tableros), replicando Market Intelligence: universo ÉTICO, Siegfried =
    manufacturer SIEGFRIED+SIDUS ('Con Sidus'), mercado = las moléculas donde
    Siegfried compite. Devuelve {'agg': {periodo: {...}}, 'products': [...todos los
    productos SIE con IE/MS por su mercado-molécula...]} o None si no hay master."""
    from pathlib import Path as _P
    if not master_path or not _P(master_path).exists():
        return None
    import openpyxl
    from collections import defaultdict
    wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True); ws = wb.active
    r1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    lab = lambda h: str(h).split('\n')[-1].strip() if h else ''
    ci = {}
    for i, h in enumerate(r1):
        z = str(h or '').strip().lower()
        if z.startswith('manufacturer'): ci['mf'] = i
        elif z.startswith('product'): ci.setdefault('prod', i)
        elif z.startswith('molecules'): ci['mol'] = i
        elif z.startswith('market (e'): ci['mkt'] = i
        if z.startswith('atc iv'): ci['atc'] = i
    col = {}
    for i, h in enumerate(r1):
        if h and str(h).startswith('Units') and _re.match(r'^[A-Z][a-z]{2} \d{4}$', lab(h)): col[lab(h)] = i
    if 'mkt' not in ci or 'prod' not in ci:
        wb.close(); return None
    M = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    WIN = {'mensual':(['Jun 2026'],['Jun 2025']),
           'ytd':([f'{m} 2026' for m in M[:6]],[f'{m} 2025' for m in M[:6]]),
           'semestre':([f'{m} 2026' for m in M[:6]],[f'{m} 2025' for m in M[:6]]),
           'trimestre':(['Apr 2026','May 2026','Jun 2026'],['Apr 2025','May 2025','Jun 2025']),
           'mat':([f'{m} 2025' for m in M[6:]]+[f'{m} 2026' for m in M[:6]],
                  [f'{m} 2024' for m in M[6:]]+[f'{m} 2025' for m in M[:6]])}
    def wsum(row, keys): return sum((row[col[k]] or 0) for k in keys if k in col and col[k] < len(row) and isinstance(row[col[k]], (int, float)))
    rows = []; sie_mols = set()
    LECHE = ('LECHE', 'MILK', 'INFANT', 'MATERNIZ', 'FORMULA INFANT', 'NUTRIC')
    n_leche = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[ci['mkt']] or '').strip().upper() != 'ETICO': continue
        mf = str(row[ci['mf']] or '').upper(); mol = str(row[ci['mol']] or '').strip().upper()
        prod = str(row[ci['prod']] or '').strip(); atc = str(row[ci.get('atc', -1)] or '').strip() if 'atc' in ci else ''
        txt = (prod + ' ' + mol + ' ' + atc).upper()
        if atc.upper().startswith('V06') or any(w in txt for w in LECHE):  # sin leches / nutrición
            n_leche += 1; continue
        issie = 'SIEGFRIED' in mf   # en el AR_PM los productos Siegfried ya vienen como SIEGFRIED; SIDUS/BIOSIDUS son otras compañías
        pv = {per: (wsum(row, cw), wsum(row, pw)) for per, (cw, pw) in WIN.items()}
        rows.append((issie, mol, prod, atc, pv))
        if issie and mol: sie_mols.add(mol)
    wb.close()
    # mercado por molécula (todos los ético sin leches)
    mol_mkt = defaultdict(lambda: {per: [0.0, 0.0] for per in WIN})
    for issie, mol, prod, atc, pv in rows:
        for per in WIN:
            mol_mkt[mol][per][0] += pv[per][0]; mol_mkt[mol][per][1] += pv[per][1]
    agg = {}
    for per in WIN:
        # MERCADO = ÉTICO SIN LECHES COMPLETO (todos los productos, no solo donde Siegfried compite)
        mc = sum(mol_mkt[m][per][0] for m in mol_mkt); mp = sum(mol_mkt[m][per][1] for m in mol_mkt)
        # secundario: mercado direccionable (moléculas donde Siegfried compite)
        adr_c = sum(mol_mkt[m][per][0] for m in sie_mols); adr_p = sum(mol_mkt[m][per][1] for m in sie_mols)
        sc = sum(pv[per][0] for issie, mol, prod, atc, pv in rows if issie)
        sp = sum(pv[per][1] for issie, mol, prod, atc, pv in rows if issie)
        agg[per] = {'sie_curr': round(sc), 'sie_prev': round(sp), 'mkt_curr': round(mc), 'mkt_prev': round(mp),
                    'ie': ie_rel(sc, sp, mc, mp), 'ms': pct(sc, mc), 'ms_prev': pct(sp, mp),
                    'growth': round((sc / sp - 1) * 100, 1) if sp else None,
                    'mkt_growth': round((mc / mp - 1) * 100, 1) if mp else None,
                    'mkt_adr_curr': round(adr_c), 'ms_adr': pct(sc, adr_c)}  # mercado direccionable
    agg['_universe'] = 'ETICO sin leches · Siegfried = manufacturer SIEGFRIED · MS% sobre ético total'
    # productos Siegfried (TODOS), agregados por nombre, con IE/MS por su mercado-molécula
    pd = defaultdict(lambda: {'mol': None, 'atc': None, 'per': {per: [0.0, 0.0] for per in WIN}})
    for issie, mol, prod, atc, pv in rows:
        if not issie or not prod: continue
        d = pd[prod]; d['mol'] = d['mol'] or mol; d['atc'] = d['atc'] or atc
        for per in WIN:
            d['per'][per][0] += pv[per][0]; d['per'][per][1] += pv[per][1]
    products = []
    for name, d in pd.items():
        mol = d['mol']; entry = {'name': name, 'mol': mol, 'atc': d['atc'], 'periods': {}}
        for per in WIN:
            c, p = d['per'][per]; mc, mp = mol_mkt[mol][per]
            entry['periods'][per] = {'units_curr': round(c), 'units_prev': round(p),
                                     'ie': ie_rel(c, p, mc, mp), 'ms': pct(c, mc)}
        products.append(entry)
    products.sort(key=lambda x: -(x['periods'].get('mat', {}).get('units_curr') or 0))
    return {'agg': agg, 'products': products}


STOCK_STATUS = ['quiebre', 'critico', 'bajo', 'alerta', 'ok']

def stock_agg():
    """Cobertura consolidada: conteo de familias por peor-estado, por linea + total,
    y lista de familias en alerta (quiebre/critico/bajo)."""
    by_line = {}; totals = {s: 0 for s in STOCK_STATUS}; alerts = []
    line_disp = {'cardio':'Cardio','antibio':'ATB','otx':'OTC','resp':'Respiratoria',
                 'mujer':'Mujer','snc':'SNC','derma':'Dermatología'}
    for key, path in LINE_FILES.items():
        try: D = _load_dashboard(path)
        except Exception: continue
        sa = D.get('stock_alerts') or {}
        cnt = {s: 0 for s in STOCK_STATUS}
        for fam, e in sa.items():
            if not (isinstance(e, dict) and isinstance(e.get('alert_indices'), list)): continue
            ws = e.get('worst_status') or 'ok'
            if ws in cnt: cnt[ws] += 1; totals[ws] += 1
            if ws in ('quiebre', 'critico', 'bajo'):
                alerts.append({'line': line_disp.get(key, key), 'familia': fam, 'status': ws,
                               'n_alerts': e.get('n_alerts')})
        by_line[key] = {'name': line_disp.get(key, key), **cnt, 'total': sum(cnt.values())}
    alerts.sort(key=lambda x: (STOCK_STATUS.index(x['status']), -(x.get('n_alerts') or 0)))
    return {'by_line': list(by_line.values()), 'totals': totals, 'alerts': alerts}


def convenios_agg():
    """Convenios consolidados por obra social: suma de unidades (periodo actual vs
    anterior) across familias + lineas. OJO: dermato usa ventana productizada
    (6m/7m) vs legacy en el resto -> el ranking de unid es robusto, el delta es
    indicativo."""
    def norm_os(os):
        s = _re.sub(r'\s*\(\d+\)\s*$', '', os)   # saca codigo (9153)
        s = _re.sub(r'\s*\|.*$', '', s)          # saca sufijo | FMLK
        return _re.sub(r'\s+', ' ', s).strip()
    os_agg = {}
    for key, path in LINE_FILES.items():
        try: D = _load_dashboard(path)
        except Exception: continue
        cv = D.get('convenios') or {}
        for fam, rows in cv.items():
            if not isinstance(rows, list): continue
            for r in rows:
                os = norm_os(r.get('os') or '')
                if not os: continue
                a = os_agg.setdefault(os, {'unid': 0, 'unid24': 0})
                a['unid'] += r.get('unid') or 0
                a['unid24'] += r.get('unid24') or 0
    out = [{'os': k, 'unid': round(v['unid']), 'unid24': round(v['unid24']),
            'delta': round((v['unid'] / v['unid24'] - 1) * 100, 1) if v['unid24'] else None}
           for k, v in os_agg.items()]
    out.sort(key=lambda x: -x['unid'])
    return out[:20]


def iqvia_dedup():
    """Total IQVIA DEDUPLICADO: cada producto SIE se cuenta 1 vez (por nombre) y cada
    mercado 1 vez (por serie mensual). Elimina el doble-conteo de mercados compartidos
    entre lineas (mometasona respi+dermato) y de sub-marcas hermanas dentro de una
    linea (DILATREND + DILATREND AP = mismo mercado carvedilol). Fuente: mol_perf.
    Devuelve tambien el 'naive' (suma sin dedup) para reportar la magnitud."""
    uniq_sie = {}   # prod_name -> monthly_vals
    uniq_mkt = {}   # serie_redondeada -> family.monthly
    naive_mkt = []  # todas las series (con duplicados) para medir magnitud
    naive_sie = {}  # prod -> lista de (linea) para medir duplicados
    for line, path in LINE_FILES.items():
        try: D = _load_dashboard(path)
        except Exception: continue
        for fam, fo in (D.get('mol_perf') or {}).items():
            mon = fo.get('monthly') or {}
            if mon:
                naive_mkt.append(mon)
                skey = tuple((m, round(v or 0)) for m, v in sorted(mon.items()))
                uniq_mkt.setdefault(skey, mon)
            for p in fo.get('products', []):
                if p.get('is_resto'): continue
                if p.get('is_sie') or '(SIE)' in (p.get('prod') or ''):
                    nm = p['prod']
                    uniq_sie.setdefault(nm, p.get('monthly_vals') or {})
                    naive_sie.setdefault(nm, set()).add((line, fam))
    def wsum(dicts, months):
        return sum((d.get(m) or 0) for d in dicts for m in months)
    out = {}
    for per, (cw, pw) in IQV_WINDOWS.items():
        sc = wsum(uniq_sie.values(), cw); sp = wsum(uniq_sie.values(), pw)
        mc = wsum(uniq_mkt.values(), cw); mp = wsum(uniq_mkt.values(), pw)
        out[per] = {'sie_curr': round(sc), 'sie_prev': round(sp),
                    'mkt_curr': round(mc), 'mkt_prev': round(mp),
                    'ie': ie_rel(sc, sp, mc, mp), 'ms': pct(sc, mc), 'ms_prev': pct(sp, mp),
                    'growth': round((sc / sp - 1) * 100, 1) if sp else None,
                    'n_sie': len(uniq_sie), 'n_mkt': len(uniq_mkt)}
    # magnitud del dedup (YTD) para reportar
    ytd_c = IQV_WINDOWS['ytd'][0]
    naive_m = sum((d.get(m) or 0) for d in naive_mkt for m in ytd_c)
    dedup_m = out['ytd']['mkt_curr']
    out['_dedup_info'] = {'naive_mkt_ytd': round(naive_m), 'dedup_mkt_ytd': dedup_m,
                          'sie_shared': [n for n, s in naive_sie.items() if len({l for l, f in s}) > 1]}
    return out


def budget_by_line():
    """Por línea: venta REAL (D.budget[fam]['2026'].real, fuente Qlik) vs ESTIMADO del
    archivo de MKT vigente (por línea). Estimado solo 2026 (el archivo no tiene 2025) ->
    MAT no tiene % (solo unidades). Devuelve {'per': {periodo:{real,estimado,pct}},
    'monthly': {real[12], est[12]}}."""
    mkt = mkt_estimado_by_line()   # {key:[12]} o None
    res = {}
    for key, path in LINE_FILES.items():
        try:
            D = _load_dashboard(path)
        except Exception:
            continue
        bud = D.get('budget') or {}
        # real mensual 2026 (suma familias del data.js = venta Qlik)
        mreal = [0.0] * 12
        for fam, yrs in bud.items():
            rv = ((yrs or {}).get('2026') or {}).get('real') or []
            for mi in range(12):
                if mi < len(rv) and isinstance(rv[mi], (int, float)): mreal[mi] += rv[mi]
        # estimado mensual 2026 = archivo de MKT (por línea); fallback al budget del data.js
        if mkt is not None and key in mkt:
            mest = list(mkt[key])
        else:
            mest = [0.0] * 12
            for fam, yrs in bud.items():
                bv = ((yrs or {}).get('2026') or {}).get('budget') or []
                for mi in range(12):
                    if mi < len(bv) and isinstance(bv[mi], (int, float)): mest[mi] += bv[mi]
        per = {}
        for pname, idxs in PERIOD_IDX.items():
            real = 0.0; est = 0.0; est_ok = True
            for (yr, mi) in idxs:
                if yr == 2026:
                    real += mreal[mi]; est += mest[mi] if mi < len(mest) else 0
                else:
                    est_ok = False   # sin estimado MKT para años previos
                    for fam, yrs2 in bud.items():
                        rv = ((yrs2 or {}).get(str(yr)) or {}).get('real') or []
                        if mi < len(rv) and isinstance(rv[mi], (int, float)): real += rv[mi]
            per[pname] = {'real': round(real), 'estimado': round(est) if est_ok else None,
                          'pct': round(real / est * 100, 1) if (est_ok and est) else None}
        res[key] = {'per': per,
                    'monthly': {'real': [round(x) for x in mreal], 'est': [round(x) for x in mest]},
                    'mkt_source': bool(mkt is not None and key in mkt)}
    return res

# metricas con mercado (para IE/MS relativo) + venta (solo YoY, sin mercado)
MKT_METRICS = [('units', 'units_sie', 'mercado_units'),
               ('recetas', 'recetas_sie', 'mercado_recetas')]


def ie_rel(sc, sp, mc, mp):
    """IE relativo al mercado desde SUMAS. None si no hay comparable o marca nueva."""
    if not sp or not mp or not mc:
        return None
    own = sc / sp
    mkt = mc / mp
    if mkt == 0:
        return None
    ie = round(own / mkt * 100, 1)
    if own > 4:  # >300% growth (marca nueva) -> sin IE fiable (tope como en el resto)
        return None
    return ie


def pct(n, d):
    return round(n / d * 100, 2) if d else None


def agg_period(lines, period):
    """Agrega un periodo (mensual/ytd/mat/trimestre) sobre las 7 lineas."""
    out = {}
    for label, sie_key, mkt_key in MKT_METRICS:
        sc = sp = mc = mp = 0.0
        n = 0
        for L in lines:
            k = (L.get('kpis') or {}).get(period)
            if not k or sie_key not in k:
                continue
            s = k[sie_key]; m = k.get(mkt_key) or {}
            if s.get('curr') is None:
                continue
            sc += s.get('curr') or 0; sp += s.get('prev') or 0
            mc += m.get('curr') or 0; mp += m.get('prev') or 0
            n += 1
        out[label] = {
            'sie_curr': round(sc), 'sie_prev': round(sp),
            'mkt_curr': round(mc), 'mkt_prev': round(mp),
            'ie': ie_rel(sc, sp, mc, mp),
            'ms': pct(sc, mc), 'ms_prev': pct(sp, mp),
            'growth': round((sc / sp - 1) * 100, 1) if sp else None,
            'n_lines': n,
        }
    # venta interna (curr/prev YoY; sin mercado)
    vc = vp = 0.0; nv = 0
    for L in lines:
        k = (L.get('kpis') or {}).get(period) or {}
        v = k.get('venta_interna')
        if v and v.get('curr') is not None:
            vc += v.get('curr') or 0; vp += v.get('prev') or 0; nv += 1
    out['venta'] = {'curr': round(vc), 'prev': round(vp),
                    'growth': round((vc / vp - 1) * 100, 1) if vp else None, 'n_lines': nv}
    return out


def resolve_master():
    """Ubica el AR_PM master del cierre corriente (hubRoot/_iqvia-master/<closeMonth>/AR_PM*.xlsx)."""
    try:
        mani = json.loads((REPO / 'shared' / 'close-manifest.json').read_text(encoding='utf-8'))
        g = mani['global']
        hub = g['hubRoot'].replace('${OneDrive}', __import__('os').environ.get('OneDrive', ''))
        cm = g['closeMonth']
        import glob
        cands = glob.glob(str(Path(hub) / '_iqvia-master' / cm / 'AR_PM*.xlsx'))
        return cands[0] if cands else None
    except Exception:
        return None


def main():
    ap = argparse.ArgumentParser(); ap.add_argument('--check', action='store_true')
    ap.add_argument('--master', help='AR_PM master para el IQVIA compañía-completa (ético). Default: autoresuelve del manifest.')
    a = ap.parse_args()
    if hasattr(sys.stdout, 'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')
    k = json.loads(KPIS.read_text(encoding='utf-8'))
    lines = k['lines']
    periods = k.get('periods') or ['mensual', 'ytd', 'mat', 'trimestre']

    company = {p: agg_period(lines, p) for p in periods}

    # ---- IQVIA DEDUPLICADO (cifra real de compania): override de 'units' ----
    # recetas y venta quedan aditivas (verificado: no duplican ni cross ni intra-linea).
    iqv = iqvia_dedup()
    dedup_info = iqv.pop('_dedup_info', {})
    for p in periods:
        if p in iqv:
            company[p]['units'] = {k: v for k, v in iqv[p].items() if k not in ('n_sie', 'n_mkt')}

    # ---- venta real vs Estimado (presupuesto), por linea + total, por periodo ----
    bud = budget_by_line()
    for p in periods:
        real = sum(((bud.get(L['key'], {}) or {}).get('per', {}).get(p, {}) or {}).get('real') or 0 for L in lines)
        est = sum(((bud.get(L['key'], {}) or {}).get('per', {}).get(p, {}) or {}).get('estimado') or 0 for L in lines)
        company[p]['venta_est'] = {'real': round(real), 'estimado': round(est),
                                   'pct': round(real / est * 100, 1) if est else None}
    # serie mensual 2026 del total (suma de lineas) para la seccion Venta Interna
    comp_mreal = [0] * 12; comp_mest = [0] * 12
    for L in lines:
        mo = (bud.get(L['key'], {}) or {}).get('monthly') or {}
        for mi in range(12):
            comp_mreal[mi] += (mo.get('real') or [0]*12)[mi]
            comp_mest[mi] += (mo.get('est') or [0]*12)[mi]
    company_venta_monthly = {'real': comp_mreal, 'est': comp_mest}

    by_line = []
    for L in lines:
        by_line.append({
            'key': L['key'], 'name': L['name'], 'icon': L.get('icon'),
            'color': L.get('color'), 'href': L.get('href'), 'owner': L.get('owner'),
            'kpis': L.get('kpis'),
            'venta_est': (bud.get(L['key']) or {}).get('per'),
            'venta_monthly': (bud.get(L['key']) or {}).get('monthly'),
            'iqvia_through': L.get('iqvia_through'), 'recetas_through': L.get('recetas_through'),
            'venta_through': L.get('venta_through'), 'has_recetas': L.get('has_recetas'),
        })

    # ---- IQVIA sobre el REAL Siegfried (ético, TODOS sus productos) — matchea Market Intelligence ----
    master = a.master or resolve_master()
    ci_full = company_iqvia(master)
    company_full = ci_full['agg'] if ci_full else None
    company_products = ci_full['products'] if ci_full else None
    cf_universe = company_full.pop('_universe', None) if company_full else None

    total = {
        'generated_at': k.get('generated_at'),
        'as_of_month': k.get('as_of_month'),
        'periods': periods,
        'period_labels': k.get('period_labels'),
        'aggregation': 'dedup',   # IQVIA cuenta cada producto/mercado 1 vez (cifra real de compania)
        'company_full': company_full,     # IQVIA real Siegfried (ético) — None si falta master
        'company_full_universe': cf_universe,
        'company_products': company_products,   # TODOS los productos SIE del master (Mercado IQVIA)
        'dedup_info': dedup_info,
        'venta_months': [f'{m} 2026' for m in ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']],
        'company_venta_monthly': company_venta_monthly,
        'company': company,
        'byLine': by_line,
        'products': k.get('products') or [],   # 63 SIE (top marcas: Mercado IQVIA + Recetas)
        'stock': stock_agg(),                  # cobertura consolidada
        'convenios': convenios_agg(),          # por obra social
    }

    # ---- consistencia ----
    errs = []
    for p in periods:
        # recetas: aditivo (no duplica) -> Total == suma lineas
        manual = sum((((L.get('kpis') or {}).get(p) or {}).get('recetas_sie') or {}).get('curr') or 0 for L in lines)
        if abs(manual - company[p]['recetas']['sie_curr']) > 1:
            errs.append(f'{p}/recetas.sie_curr: {company[p]["recetas"]["sie_curr"]} != sum {manual}')
        # IQVIA units: DEDUP -> IE/MS deben salir de las sumas deduplicadas
        u = company[p]['units']
        exp_ms = pct(u['sie_curr'], u['mkt_curr'])
        if not (exp_ms is None or abs(exp_ms - (u['ms'] or 0)) <= 0.1):
            errs.append(f'{p}/units.ms {u["ms"]} != {exp_ms} (desde sumas dedup)')
        if u['sie_prev'] and u['mkt_prev'] and u['sie_curr']/u['sie_prev'] <= 4:
            exp_ie = round((u['sie_curr']/u['sie_prev'])/(u['mkt_curr']/u['mkt_prev'])*100, 1)
            if u['ie'] is not None and abs(exp_ie - u['ie']) > 0.2:
                errs.append(f'{p}/units.ie {u["ie"]} != {exp_ie} (desde sumas dedup)')
        # venta_est.real (budget) ~ venta_interna.curr (kpis.json)
        ve = company[p].get('venta_est', {}).get('real'); vi = company[p]['venta']['curr']
        if ve is not None and vi and abs(ve - vi) / max(vi, 1) > 0.02:
            errs.append(f'{p}: venta budget.real={ve:,} vs kpis={vi:,} (dif {abs(ve-vi)/vi*100:.1f}%)')

    # print headline
    print(f"== TOTAL SIEGFRIED (as_of {total['as_of_month']}) · IQVIA DEDUPLICADO ==")
    nm = dedup_info.get('naive_mkt_ytd'); dm = dedup_info.get('dedup_mkt_ytd')
    if nm and dm:
        print(f"  dedup mercado YTD: naive {nm:,} -> {dm:,} (-{(nm-dm)/nm*100:.1f}%) · SIE compartidos: {dedup_info.get('sie_shared')}")
    for p in periods:
        c = company[p]
        u = c['units']; r = c['recetas']; ve = c.get('venta_est', {})
        print(f"\n[{p.upper()}]")
        print(f"  IQVIA  : SIE {u['sie_curr']:,}u  MS% {u['ms']}  IE {u['ie']}  crec {u['growth']}%  (mercado {u['mkt_curr']:,}u)")
        print(f"  Recetas: SIE {r['sie_curr']:,}  MS% {r['ms']}  IE {r['ie']}  crec {r['growth']}%  (mercado {r['mkt_curr']:,})")
        print(f"  Venta  : real {ve.get('real'):,}u  vs Estimado {ve.get('estimado'):,}u  =  {ve.get('pct')}%")
    if errs:
        print("\nINCONSISTENCIAS:"); [print("  ", e) for e in errs]
        return 1
    print("\nOK: IQVIA deduplicado (cada producto/mercado 1 vez); recetas/venta aditivas; IE/MS desde sumas.")

    if a.check:
        print("(--check: no se escribio)"); return 0
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text('window.TOTAL_SIEGFRIED = ' + json.dumps(total, ensure_ascii=False) + ';\n',
                   encoding='utf-8', newline='')
    print(f"\nEscrito {OUT}")
    return 0


if __name__ == '__main__':
    sys.exit(main())
