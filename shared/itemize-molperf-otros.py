#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
shared/itemize-molperf-otros.py

Desglosa el bucket 'Otros (resto del mercado)' de mol_perf para que la APERTURA
de la tabla multi-periodo (Mercado IQVIA) muestre el ranking COMPLETO del mercado:
las marcas que estan arriba de la propia y tambien las que estan abajo.

EL BUG
------
Los build-data.ps1 de cardio/ATB/OTC/respiratorio cortan en 8 productos por
mercado (`if ($selectedProductNames.Count -ge 8) { break }`) y meten todo el
resto en un unico producto sintetico 'Otros (resto del mercado)' (is_resto).
El render NO tiene tope -- shared/multi-period-table.js buildCompRows() lista
todos los products con rank #1..#N -- asi que el ranking incompleto es 100% del
dato, no de la UI.

Sintomas historicos:
  - ROXOLAN se veia #7 y en el Explorador de IQVIA es #10, porque ROSUFEN,
    ROSUSTATIN y ROSUFEC (los tres > ROXOLAN) estaban dentro de 'Otros'.
  - Aun despues de arreglar el puesto, no se veia quien estaba POR DEBAJO.
Los agregados (total de mercado y MS% propio) SIEMPRE estuvieron bien: lo unico
falso era el ranking y la lista de competidores.

MODOS
-----
  --mode full      (default) itemiza TODO el universo del mercado -> ranking
                   completo, arriba y abajo. Es lo que se ve hoy en el tablero.
  --mode outrank   comportamiento conservador original: solo las marcas que
                   SUPERAN a la marca SIE (arregla el numero de puesto y nada mas).
                   Se conserva para poder reproducir el estado previo.

COMO SE INFIERE EL UNIVERSO DEL MERCADO
---------------------------------------
Se toman los productos ya listados, se los busca por nombre en el master AR_PM, se
junta el set de sus moleculas y se consideran "ocultos" los productos del master
con esa misma molecula que no estan listados. Se usa el MISMO master del que salio
mol_perf, no competidores-data.js (ese es el panel REGIONAL de Qlik/DDD: otro
universo, mezclarlos rompe la invariante de la suma).

LA REGLA ES ASIMETRICA -- ESTO ES LO IMPORTANTE
-----------------------------------------------
La inferencia por molecula puede no coincidir con el mercado, y los dos sentidos
del error NO son equivalentes:

  SUB-CONTAR (candidatos < bucket) es SEGURO.
      El mercado tiene productos que la molecula sola no explica. Se agregan las
      marcas que si se conocen y el bucket se queda con el remanente sin explicar.
      No se inventa nada y la suma sigue cerrando.
      Caso real: respiratorio DECADRON, May 2024, bucket 6.749 vs candidatos
      5.323 -> faltan 1.426 u. Se itemiza igual; 'Otros' conserva esas 1.426.

  SOBRE-CONTAR (candidatos > bucket) es INACEPTABLE.
      Significa que la molecula abarca MAS que el mercado (mercados splitteados
      por dosis o definidos por ATC) y se estarian listando marcas de OTRO
      mercado. Se RECHAZA el mercado entero.
      Casos reales: ATB CEFALEXINA ARG (candidatos +1.953% sobre el bucket) y
      CEFALEXINA ARG DUO (+5.038%). Quedan intactos y se reportan.

La separacion medida es de tres ordenes de magnitud (peor error relativo por mes:
26 mercados <= 3,75%, DECADRON 21%, los dos CEFALEXINA 3.795% y 8.226%), asi que
el umbral de SOBRE_TOL no es un numero al voleo: cae en un hueco enorme.

COMO PRESERVA LA CONSISTENCIA
-----------------------------
NO se elimina el bucket: se lo RECALCULA como el residuo
    Otros = total_del_mercado - suma(productos listados)
Hay mas productos listados, asi que el residuo es mas chico. Por construccion
sum(products) == mol_perf[fam].{monthly,quarterly,ytd,mat} EXACTO, sin tocar ni un
valor del total del mercado. (Itemizar reemplazando el bucket por la suma cruda de
las marcas moveria el total unas decenas de unidades, porque el build redondea por
producto y el bucket absorbe ese redondeo. Eso NO se hace.)

El residuo puede quedar levemente NEGATIVO sin que nada este mal: el build redondea
producto por producto, asi que la suma de valores redondeados puede pasarse del total
por unas pocas unidades. NO se descartan marcas reales para forzar residuo >= 0 (eso
costaba 6 a 8 marcas por mercado para tapar unidades de redondeo): se deja el residuo
EXACTO y se rechaza el mercado solo si se hunde mas de RESID_TOL del total del
periodo. Medido: 26 mercados legitimos entre 0 y -85 u (<= 0,043% del mercado) contra
-33.728 u (59%) en CEFALEXINA ARG.

Los candidatos que son marcas SIEGFRIED se tratan aparte (ver el bloque es_sie): no
se pueden listar como competencia porque el nombre y el flag is_sie alimentan el
total SIE de compania.

El master cubre Jul 2021..Jun 2026 y los tableros solo necesitan ventanas desde
Feb 2024, asi que todas las ventanas MAT/YTD de los productos nuevos estan
completamente cubiertas (no hay MAT parcial).

Uso:
    py shared/itemize-molperf-otros.py [--mode full|outrank] [--master <xlsx>]
                                       [--dry-run] [--report <json>]
"""
from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
DEFAULT_MASTER = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs'
                      r'\_iqvia-master\2026-06\AR_PM_FV_Standard_Jul-2026.xlsx')
RESTO = 'Otros (resto del mercado)'
LINES = ['cardio/data.js', 'ATB/data.js', 'OTC/data.js', 'respiratorio/data.js']
MES = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
MI = {m: i + 1 for i, m in enumerate(MES)}

# Tolerancia de SOBRE-conteo: cuanto puede exceder el MAT de los candidatos al MAT
# del bucket antes de considerar que el universo por molecula no es el del mercado.
# 2% deja pasar el redondeo por-producto del build (medido: peor caso 0,45%) y
# rechaza los splits por dosis (medido: +1.953% y +5.038%).
SOBRE_TOL = 0.02
# Ademas se exige que los meses en que los candidatos exceden al bucket por mas de
# 5% no sean mas de este porcentaje del total de meses (defensa por si el MAT
# compensa un sobre-conteo mensual con un sub-conteo en otro mes).
SOBRE_MESES_MAX = 0.25
# Cota del residuo negativo, como fraccion del total del periodo. El redondeo
# por-producto del build da hasta 0,043% (medido); un universo mal inferido da 36-59%.
# 0,5% deja un margen de 10x sobre el peor caso legitimo y 70x bajo el roto.
RESID_TOL = 0.005
# Maximo que la absorcion del deficit de redondeo puede quitarle a UNA marca en UN
# mes, como fraccion de lo que esa marca vendio ese mes. Con 1%, una marca de miles
# de unidades absorbe cualquier deficit sin moverse de forma perceptible, y una de
# 16 unidades no se toca nunca (int(16*0.01) == 0).
ABSORB_MAX_FRAC = 0.01
# El cupo se calcula con int(), asi que una marca con menos de 100 unidades en el mes
# da cupo 0 y queda fuera de la absorcion por truncamiento. Para no perder por eso las
# ultimas unidades del deficit, a partir de este piso se le permite ceder 1 unidad.
# Con el piso en 100, ceder 1 unidad sigue siendo <= 1% del mes: la cota no se afloja.
ABSORB_MIN_DISP = 100


def msort(mk):
    p = str(mk).split()
    return (int(p[1]), MI.get(p[0], 0)) if len(p) == 2 and p[1].isdigit() else (0, 0)


def qkey(mk):
    p = mk.split()
    m = MI.get(p[0])
    return 'Q{} {}'.format((m - 1) // 3 + 1, p[1]) if m else None


def read_master(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    r1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    ci, mcols = {}, []
    for i, h in enumerate(r1):
        if not h:
            continue
        s = str(h).strip()
        sn = s.replace('\n', ' ').strip().lower()
        if sn.startswith('manufacturer'):
            ci['manuf'] = i
        elif sn.startswith('product'):
            ci['prod'] = i
        elif sn.startswith('molecules'):
            ci['mol'] = i
        if s.startswith('Units'):
            a = (s.split('\n', 1)[-1] if '\n' in s else s[len('Units'):]).strip()
            if a.upper().startswith(('MAT', 'YTD')):
                continue
            m = re.match(r'(\w+)\s+(\d{4})$', a)
            if m and m.group(1) in MI:
                mcols.append((i, '{} {}'.format(m.group(1), m.group(2))))
    prods = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        p = row[ci['prod']] if ci['prod'] < len(row) else None
        if not p:
            continue
        name = str(p).strip()
        d = prods.setdefault(name, {'manuf': '', 'mol': '', 'monthly': defaultdict(float)})
        if ci.get('manuf') is not None and ci['manuf'] < len(row) and row[ci['manuf']]:
            d['manuf'] = str(row[ci['manuf']]).strip()
        if ci.get('mol') is not None and ci['mol'] < len(row) and row[ci['mol']]:
            d['mol'] = str(row[ci['mol']]).strip()
        for c, mk in mcols:
            if c < len(row) and isinstance(row[c], (int, float)):
                d['monthly'][mk] += row[c]
    wb.close()
    for d in prods.values():
        d['monthly'] = {k: int(round(v)) for k, v in d['monthly'].items()}
    return prods, [mk for _, mk in mcols]


def strip_suffix(p):
    return re.sub(r'\s*\([^)]*\)\s*$', '', str(p)).strip().upper()


def agg_from_monthly(monthly, keys_ref, kind, cierre):
    """Agrega monthly a quarterly/ytd/mat usando SOLO las keys que ya existen en el
    mercado (keys_ref), para no inventar columnas nuevas."""
    out = {}
    if kind == 'quarterly':
        tmp = defaultdict(int)
        for mk, v in monthly.items():
            q = qkey(mk)
            if q:
                tmp[q] += int(v or 0)
        return {k: tmp.get(k, 0) for k in keys_ref}
    for key in keys_ref:
        p = str(key).split()
        if len(p) != 2 or not p[1].isdigit():
            continue
        cm, y = MI.get(p[0], cierre), int(p[1])
        if kind == 'ytd':
            win = ['{} {}'.format(MES[m - 1], y) for m in range(1, cm + 1)]
        else:
            win = []
            for back in range(11, -1, -1):
                idx = (y * 12 + (cm - 1)) - back
                yy, mm = divmod(idx, 12)
                win.append('{} {}'.format(MES[mm], yy))
        out[key] = sum(int(monthly.get(mk, 0) or 0) for mk in win)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--master', default=str(DEFAULT_MASTER))
    ap.add_argument('--mode', choices=['full', 'outrank'], default='full',
                    help='full = itemiza todo el mercado (default); outrank = solo los que superan al SIE')
    ap.add_argument('--dry-run', action='store_true')
    ap.add_argument('--report', default=None, help='escribe un resumen JSON en esta ruta')
    a = ap.parse_args()

    mp_path = Path(a.master)
    if not mp_path.is_file():
        print('ERROR: master no existe: {}'.format(mp_path), file=sys.stderr)
        return 2
    print('Master: {}'.format(mp_path))
    print('Modo:   {}'.format(a.mode))
    MPROD, MMONTHS = read_master(mp_path)
    print('  {} productos en el master, {} meses ({}..{})'.format(
        len(MPROD), len(MMONTHS), MMONTHS[0], MMONTHS[-1]))
    by_exact = set(MPROD)
    by_base = defaultdict(list)
    for n in MPROD:
        by_base[strip_suffix(n)].append(n)

    print()
    print('{:<12} {:<22} {:>8} {:>8} {:>7} {:>6} {:>12} {}'.format(
        'linea', 'mercado', 'rank_ant', 'rank_new', '+items', 'resid%', 'Otros nuevo', 'estado'))
    print('-' * 122)

    planned = []
    report = {'mode': a.mode, 'master': str(mp_path), 'markets': []}
    tocados = saltados = 0
    for rel in LINES:
        p = REPO / rel
        text = p.read_text(encoding='utf-8', errors='replace')
        m = re.search(r'window\.OTC_DASHBOARD\s*=\s*', text)
        s = text.index('{', m.end())
        D, end = json.JSONDecoder().raw_decode(text[s:])
        e = s + end
        cambios_linea = 0
        linea = rel.split('/')[0]
        # Nombres de producto ya presentes en CUALQUIER mercado de esta linea. Sirve
        # para el caso de los candidatos SIEGFRIED: check-total-consistency.py dedupea
        # los productos SIE por NOMBRE, asi que si el nombre ya figura en otro mercado
        # de la linea agregarlo aca no mueve el total SIE de compania. Si NO figura en
        # ninguno, agregarlo SI lo moveria -> se excluye y se reporta.
        ya_en_linea = set()
        for _f in (D.get('mol_perf') or {}).values():
            for _p in (_f.get('products') or []):
                ya_en_linea.add(str(_p.get('prod')))

        for fam, f in (D.get('mol_perf') or {}).items():
            prods = f.get('products') or []
            resto = next((x for x in prods if str(x.get('prod')) == RESTO), None)
            if not resto:
                continue
            fam_monthly = {k: int(v or 0) for k, v in (f.get('monthly') or {}).items()}
            if not fam_monthly:
                continue
            last = max(fam_monthly, key=msort)
            cierre = MI[last.split()[0]]
            listados = [x for x in prods if x is not resto]
            sie = [x for x in listados if x.get('is_sie')]
            if not sie:
                continue
            principal = max(sie, key=lambda x: (x.get('mat') or {}).get(last, 0) or 0)
            sie_mat = (principal.get('mat') or {}).get(last, 0) or 0

            # universo del mercado inferido de los listados
            mols, encontrados = set(), set()
            for x in listados:
                nm = str(x.get('prod'))
                cand = nm if nm in by_exact else (by_base.get(strip_suffix(nm)) or [None])[0]
                if cand:
                    encontrados.add(cand)
                    if MPROD[cand]['mol']:
                        mols.add(MPROD[cand]['mol'])
            if not mols:
                continue
            ocultos = [n for n, d in MPROD.items() if d['mol'] in mols and n not in encontrados]
            if not ocultos:
                continue

            orden_ant = sorted(prods, key=lambda x: -((x.get('mat') or {}).get(last, 0) or 0))
            rank_ant = [i for i, x in enumerate(orden_ant, 1) if x is principal][0]

            def fila(rank_new, n_add, n_desc, resto_new, estado):
                print('{:<12} {:<22} {:>8} {:>8} {:>7} {:>6} {:>12} {}'.format(
                    linea[:12], fam[:22], '#' + str(rank_ant), rank_new, n_add, n_desc,
                    '{:,}'.format(resto_new) if isinstance(resto_new, int) else resto_new, estado))

            # ── GATE DE SOBRE-CONTEO ──────────────────────────────────────────
            # Si el universo por molecula excede al bucket, la molecula abarca mas
            # que el mercado -> se estarian listando marcas de OTRO mercado.
            idx = MMONTHS.index(last) if last in MMONTHS else len(MMONTHS) - 1
            w12 = MMONTHS[max(0, idx - 11):idx + 1]
            cand_mat = sum(sum(MPROD[n]['monthly'].get(mk, 0) for mk in w12) for n in ocultos)
            resto_mat = int((resto.get('mat') or {}).get(last) or 0)
            rm = resto.get('monthly_vals') or {}
            meses_cmp = [mk for mk in rm if mk in MMONTHS and int(rm[mk] or 0) > 0]
            excede = 0
            for mk in meses_cmp:
                sc = sum(MPROD[n]['monthly'].get(mk, 0) for n in ocultos)
                if sc > int(rm[mk] or 0) * 1.05:
                    excede += 1
            sobre_mat = resto_mat > 0 and cand_mat > resto_mat * (1 + SOBRE_TOL)
            sobre_meses = bool(meses_cmp) and (excede / len(meses_cmp)) > SOBRE_MESES_MAX
            if sobre_mat or sobre_meses:
                saltados += 1
                exc_pct = ((cand_mat / resto_mat - 1) * 100) if resto_mat else float('inf')
                fila('-', len(ocultos), '-', '-',
                     'RECHAZADO: sobre-cuenta {:+.0f}% del bucket ({} de {} meses exceden) '
                     '-> la molecula {} abarca mas que este mercado'.format(
                         exc_pct, excede, len(meses_cmp), '/'.join(sorted(mols))[:40]))
                report['markets'].append({
                    'line': linea, 'market': fam, 'action': 'rechazado-sobreconteo',
                    'candidates': len(ocultos), 'resto_mat': resto_mat, 'cand_mat': cand_mat,
                    'exceso_pct': round(exc_pct, 2) if resto_mat else None,
                    'meses_exceden': excede, 'meses_comparados': len(meses_cmp),
                    'molecules': sorted(mols)})
                continue

            # ── candidatos SIEGFRIED: marcas PROPIAS escondidas en el bucket ───
            # Se detectan por el sufijo de laboratorio del master ('(SIE)') o por
            # manufacturer. NO se pueden agregar como si fueran competencia: el flag
            # is_sie y el nombre alimentan el total SIE de compania.
            def es_sie(n):
                return '(SIE)' in n.upper() or 'SIEGFRIED' in (MPROD[n]['manuf'] or '').upper()

            # Se excluyen TODOS, tanto los que ya figuran en otro mercado de la linea como
            # los que no. Motivo medido: check-total-consistency.py (y build-total.py)
            # arman el universo SIE de compania con sie.setdefault(p['prod'], ...), o sea
            # dedupean POR NOMBRE y se quedan con la PRIMERA copia que encuentran. Si se
            # agrega 'DILATREND AP (SIE)' como competidor del mercado DILATREND, esa copia
            # puede TAPAR a la publicada del mercado DILATREND AP segun el orden de
            # iteracion -- y si la absorcion de redondeo le toco alguna unidad, el MAT SIE
            # de compania se mueve. Medido: 20.496.830 vs 20.496.827, 3 unidades, con
            # tolerancia 2 -> el gate falla. Es una dependencia del orden de iteracion, o
            # sea fragil por diseño, y no vale la pena para mostrar en el ranking una marca
            # propia que ya tiene su propia fila de mercado.
            sie_excluidos = [n for n in ocultos if es_sie(n)]
            sie_incluidos = []
            ocultos = [n for n in ocultos if n not in sie_excluidos]
            if sie_excluidos:
                report.setdefault('sie_ocultos_excluidos', []).extend(
                    {'line': linea, 'market': fam, 'prod': n,
                     'ya_listada_en_la_linea': n in ya_en_linea,
                     'motivo': 'marca Siegfried dentro del bucket Otros; agregarla como '
                               'competidor puede mover el total SIE de compania por el dedup '
                               'por nombre'} for n in sie_excluidos)
            if not ocultos:
                continue

            # ── seleccion de candidatos segun el modo ─────────────────────────
            mat_of = {n: sum(MPROD[n]['monthly'].get(mk, 0) for mk in w12) for n in ocultos}
            if a.mode == 'outrank':
                elegidos = [n for n in ocultos if mat_of[n] > sie_mat]
                if not elegidos:
                    continue    # el rank ya era correcto
            else:
                elegidos = list(ocultos)
            elegidos.sort(key=lambda n: -mat_of[n])

            meses_mkt = sorted(fam_monthly, key=msort)
            q_ref = list((f.get('quarterly') or {}).keys())
            y_ref = list((f.get('ytd') or {}).keys())
            m_ref = list((f.get('mat') or {}).keys())

            def mk_prod(n):
                monthly = {mk: int(MPROD[n]['monthly'].get(mk, 0) or 0) for mk in meses_mkt}
                if not any(monthly.values()):
                    return None
                return {
                    'prod': n,
                    'manuf': MPROD[n]['manuf'],
                    'is_sie': n in sie_incluidos,
                    'monthly_vals': monthly,
                    'quarterly_vals': agg_from_monthly(monthly, q_ref, 'quarterly', cierre),
                    'ytd': agg_from_monthly(monthly, y_ref, 'ytd', cierre),
                    'mat': agg_from_monthly(monthly, m_ref, 'mat', cierre),
                    'ms_monthly': {}, 'ms_quarterly': {}, 'ms_ytd': {}, 'ms_mat': {},
                }

            nuevos = [x for x in (mk_prod(n) for n in elegidos) if x]
            if not nuevos:
                continue

            # ── TODO SE DERIVA DEL RESIDUO MENSUAL ────────────────────────────
            # Se verifico empiricamente que en las 4 lineas la consistencia interna es
            # perfecta: cada producto tiene quarterly == suma de sus propios meses
            # (4.005 comparaciones, 0 diferencias) y cada familia tambien
            # (702 comparaciones, 0 diferencias). Asi que el residuo se calcula UNA vez
            # a nivel mensual y sus agregados se derivan de el con agg_from_monthly,
            # igual que cualquier otro producto. Eso hace que
            #   suma(products.quarterly) == familia.quarterly
            # cierre por construccion en vez de por resta de agregados: los dos lados
            # terminan siendo sumas de los MISMOS valores mensuales.
            def suma_monthly(cands):
                acc = {mk: 0 for mk in meses_mkt}
                for x in listados + cands:
                    mv = x.get('monthly_vals') or {}
                    for mk in meses_mkt:
                        acc[mk] += int(mv.get(mk, 0) or 0)
                return acc

            r_m = {mk: fam_monthly[mk] - suma_monthly(nuevos)[mk] for mk in meses_mkt}

            # ── COTA DE CORDURA DEL RESIDUO ───────────────────────────────────
            # El residuo mensual puede quedar levemente NEGATIVO sin que nada este mal:
            # el build redondea producto por producto, asi que la suma de valores
            # redondeados puede pasarse del total por unas pocas unidades. Medido en las
            # 4 lineas: el peor residuo mensual de los mercados legitimos va de 0 a
            # -85 u, o sea <= 0,043% del mes. Los mercados con el universo mal inferido
            # estan 3 ordenes de magnitud mas abajo (CEFALEXINA ARG: -33.728 u = 59%).
            # NO se descartan marcas reales para forzar el residuo >= 0 (eso costaba 6 a
            # 8 marcas por mercado para tapar unidades de redondeo): se rechaza el
            # mercado solo si el residuo se hunde mas alla de RESID_TOL del mes.
            peor_rel, peor_k = 0.0, ''
            for mk, v in r_m.items():
                if v < 0 and fam_monthly[mk] > 0:
                    rel_err = abs(v) / fam_monthly[mk]
                    if rel_err > peor_rel:
                        peor_rel, peor_k = rel_err, 'monthly[{}]={}'.format(mk, v)
            if peor_rel > RESID_TOL:
                saltados += 1
                fila('-', len(nuevos), '-', '-',
                     'RECHAZADO: residuo {:.2f}% del total en {} -> excede la cota de redondeo '
                     '({:.1f}%), el universo inferido no es el del mercado'.format(
                         peor_rel * 100, peor_k, RESID_TOL * 100))
                report['markets'].append({
                    'line': linea, 'market': fam, 'action': 'rechazado-residuo',
                    'candidates': len(elegidos), 'peor_residuo_rel_pct': round(peor_rel * 100, 3),
                    'peor_residuo_key': peor_k})
                continue

            # ── ABSORCION DEL DEFICIT DE REDONDEO ─────────────────────────────
            # Donde el residuo mensual quedo negativo, se le descuentan esas unidades
            # al candidato NUEVO mas grande de ese mes. Reglas:
            #   - Los productos LISTADOS no se tocan NUNCA: sus valores ya estan
            #     publicados y deben seguir coincidiendo con IQVIA.
            #   - Se elige el mas grande del mes para que la distorsion relativa sea la
            #     minima posible (decenas de unidades sobre miles).
            #   - Si ni todos los candidatos juntos alcanzan a cubrir el deficit, se
            #     rechaza el mercado: eso ya no seria redondeo.
            # Esto deja el residuo >= 0 en todos los meses SIN inventar unidades: las
            # que se descuentan son exactamente las que el redondeo del build habia
            # duplicado, y la suma total no se mueve ni una unidad.
            # La cota es POR MES y RELATIVA AL VALOR DE ESE MES: a ninguna marca se le
            # quita mas de ABSORB_MAX_FRAC de lo que vendio ese mes. Consecuencias:
            #   - las marcas grandes del mes absorben el deficit entero sin moverse
            #     de manera perceptible (unas decenas sobre miles);
            #   - las marcas microscopicas NUNCA se tocan (el 1% de 16 unidades es 0),
            #     asi que no se las puede desfigurar;
            #   - si en un mes el deficit no se puede cubrir con ese criterio, el
            #     remanente queda en 'Otros' -- el residuo sigue siendo EXACTO, solo
            #     puede quedar levemente negativo en ese mes puntual.
            # (Un intento anterior comparaba las unidades quitadas en 29 meses contra el
            # MAT de los ultimos 12: mezclaba ventanas y daba falsos 56%.)
            reasignadas, meses_ajustados, meses_negativos = 0, 0, []
            peor_dist, peor_dist_det = 0.0, ''
            for mk in meses_mkt:
                falta = -r_m[mk]
                if falta <= 0:
                    continue
                for x in sorted(nuevos, key=lambda y: -(int((y['monthly_vals']).get(mk, 0) or 0))):
                    if falta <= 0:
                        break
                    disp = int(x['monthly_vals'].get(mk, 0) or 0)
                    cupo = int(disp * ABSORB_MAX_FRAC)
                    if cupo == 0 and disp >= ABSORB_MIN_DISP:
                        cupo = 1
                    quita = min(cupo, falta)
                    if quita <= 0:
                        continue
                    x['monthly_vals'][mk] = disp - quita
                    falta -= quita
                    reasignadas += quita
                    if disp > 0 and quita / disp > peor_dist:
                        peor_dist = quita / disp
                        peor_dist_det = '{} {} -{} u de {:,}'.format(x['prod'], mk, quita, disp)
                if falta > 0:
                    meses_negativos.append('{}({})'.format(mk, -falta))
                else:
                    meses_ajustados += 1

            # Recalcular el residuo y los agregados de cada nuevo DESPUES del ajuste,
            # para que cada producto siga cumpliendo agregado == suma de sus meses.
            r_m = {mk: fam_monthly[mk] - suma_monthly(nuevos)[mk] for mk in meses_mkt}
            for x in nuevos:
                x['quarterly_vals'] = agg_from_monthly(x['monthly_vals'], q_ref, 'quarterly', cierre)
                x['ytd'] = agg_from_monthly(x['monthly_vals'], y_ref, 'ytd', cierre)
                x['mat'] = agg_from_monthly(x['monthly_vals'], m_ref, 'mat', cierre)

            resto['monthly_vals'] = r_m
            resto['quarterly_vals'] = agg_from_monthly(r_m, q_ref, 'quarterly', cierre)
            resto['ytd'] = agg_from_monthly(r_m, y_ref, 'ytd', cierre)
            resto['mat'] = agg_from_monthly(r_m, m_ref, 'mat', cierre)
            r_t = resto['mat']

            finales = listados + nuevos + [resto]
            # ms_* de TODOS contra el total de familia (que no cambio)
            for x in finales:
                mv = x.get('monthly_vals') or {}
                x['ms_monthly'] = {mk: (round(int(mv.get(mk, 0) or 0) / fam_monthly[mk] * 100, 2)
                                        if fam_monthly.get(mk) else 0) for mk in meses_mkt}
                for campo, ref, dst in (('quarterly_vals', q_ref, 'ms_quarterly'),
                                        ('ytd', y_ref, 'ms_ytd'), ('mat', m_ref, 'ms_mat')):
                    src = x.get(campo) or {}
                    base = f.get('quarterly' if campo == 'quarterly_vals' else campo) or {}
                    x[dst] = {k: (round(int(src.get(k, 0) or 0) / int(base.get(k) or 0) * 100, 2)
                                  if int(base.get(k) or 0) else 0) for k in ref}
            finales.sort(key=lambda x: (not x.get('is_sie'), -((x.get('mat') or {}).get(last, 0) or 0)))
            f['products'] = finales

            orden_new = sorted(finales, key=lambda x: -((x.get('mat') or {}).get(last, 0) or 0))
            rank_new = [i for i, x in enumerate(orden_new, 1) if x is principal][0]
            tocados += 1
            cambios_linea += 1
            det = 'OK  ranking de {} marcas'.format(len(finales) - 1)
            if sie_excluidos:
                det += ' | {} SIE fuera: {}'.format(len(sie_excluidos), ', '.join(sie_excluidos))
            if sie_incluidos:
                det += ' | SIE propia visible: {}'.format(', '.join(sie_incluidos))
            if reasignadas:
                det += ' | {} u de redondeo reasignadas ({} mes), peor {:.2f}% en {}'.format(
                    reasignadas, meses_ajustados, peor_dist * 100, peor_dist_det)
            if meses_negativos:
                det += ' | Otros levemente negativo en {}: {}'.format(
                    len(meses_negativos), ', '.join(meses_negativos[:4]))
            fila('#' + str(rank_new), len(nuevos), '{:.3f}%'.format(peor_rel * 100),
                 r_t.get(last, 0), det)
            report['markets'].append({
                'line': linea, 'market': fam, 'action': 'itemizado',
                'rank_ant': rank_ant, 'rank_new': rank_new,
                'added': len(nuevos), 'n_products_final': len(finales),
                'peor_residuo_rel_pct': round(peor_rel * 100, 4),
                'peor_residuo_key': peor_k,
                'unidades_reasignadas': reasignadas, 'meses_ajustados': meses_ajustados,
                'peor_distorsion_pct': round(peor_dist * 100, 4), 'peor_distorsion': peor_dist_det,
                'meses_otros_negativo': meses_negativos,
                'sie_incluidos': sie_incluidos, 'sie_excluidos': sie_excluidos,
                'resto_mat_new': r_t.get(last, 0), 'resto_mat_old': resto_mat,
                'sie_brand': principal.get('prod')})

        if cambios_linea:
            planned.append((p, text, s, e, D))

    print()
    print('mercados itemizados: {} | rechazados/saltados: {}'.format(tocados, saltados))
    if a.report:
        Path(a.report).write_text(json.dumps(report, ensure_ascii=False, indent=1), encoding='utf-8')
        print('reporte -> {}'.format(a.report))
    if a.dry_run:
        print('\nDRY RUN: no se escribio nada.')
        return 0
    print()
    for p, text, s, e, D in planned:
        antes = p.stat().st_size
        p.write_text(text[:s] + json.dumps(D, ensure_ascii=False) + text[e:],
                     encoding='utf-8', newline='')
        print('-> {} ({:,} -> {:,} bytes, {:+.1f}%)'.format(
            p.relative_to(REPO), antes, p.stat().st_size, (p.stat().st_size / antes - 1) * 100))
    return 0


if __name__ == '__main__':
    sys.exit(main())
