# -*- coding: utf-8 -*-
"""Parsea 'Comparativa de PRECIOS_DD.MM.AAAA.xlsx' (comparativa competitiva MANUAL de
precios) e inyecta D.prec_comp + D.prec_comp_meta en mujer/index.html como SNAPSHOT fechado.

Es un dato MANUAL (no IQVIA): para refrescarlo, actualizar el Excel y re-correr:
    py shared/parse-comparativa-precios.py "<ruta al .xlsx>"
La fecha del snapshot se toma del nombre del archivo (DD.MM.AAAA).

Idempotente: reemplaza el bloque entre los marcadores /*PCOMP_START*/ ... /*PCOMP_END*/.
NO toca ninguna otra clave de const D ni ningun otro archivo.

Solo procesa las hojas de LINEA MUJER (ver SHEET_GROUPS). Las demas hojas del Excel
(Cinitral, Gastrosedol, gal, Genozym, Mailen, 'No Modificar', etc.) se ignoran.

Estructura inyectada:
  prec_comp = { "ISIS": { "pres": [ { "label","droga","rows":[
      {"prod","lab","pub","unit","gap"(frac|null),"fecha"(YYYY-MM-DD|null),"sie"(bool)} ] } ] }, ... }
  prec_comp_meta = { "fecha":"DD/MM/AAAA", "fuente":"..." }
"""
from __future__ import annotations
import sys, re, json, datetime, argparse
from types import SimpleNamespace
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent


class Grid:
    """Adaptador: hoja materializada (lista de filas) con API .cell(r,c).value /
    .max_row / .max_column, para parsear en modo read_only SIN tocar la hoja
    gigante 'No Modificar' (1M+ filas) que volveria lentisima la carga."""
    def __init__(self, rows):
        self._rows = rows
        self.max_row = len(rows)
        self.max_column = max((len(r) for r in rows), default=0)

    def cell(self, r, c):
        v = None
        if 1 <= r <= self.max_row:
            row = self._rows[r - 1]
            if 1 <= c <= len(row):
                v = row[c - 1]
        return SimpleNamespace(value=v)
HTML = REPO / 'mujer' / 'index.html'
DEFAULT_XLSX = r'C:\Users\camarinaro\AppData\Local\Temp\Comparativa de PRECIOS_07.05.2026.xlsx'

# Hoja del Excel -> nombre de grupo a mostrar (SOLO mujer)
SHEET_GROUPS = {
    'Isis': 'ISIS',
    'Trip': 'TRIP',
    'Gynoderm': 'GYNODERM',
    'Siderblut': 'SIDERBLUT',
    'Climatix': 'CLIMATIX',
    'Deltox': 'DELTROX',
    'Calcio Base - Citramar': 'CALCIO BASE',
}


def clean(s):
    if s is None:
        return ''
    return re.sub(r'\s+', ' ', str(s).replace('\xff', ' ').replace('\n', ' ')).strip()


def num(v):
    return round(float(v), 2) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def gapnum(v):
    return round(float(v), 4) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def fechastr(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%Y-%m-%d')
    return None


def troqnorm(v):
    if v is None:
        return None
    if isinstance(v, float):
        v = int(v)
    s = str(v).strip()
    return s or None


def load_prices(path):
    """Catalogo plano (Sheet1: ... Troquel ... 'PVP al DD/MM' ...) -> {troquel: PVP}.
    Usa la columna 'PVP al ...' MAS A LA DERECHA (la mas reciente)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    tcol = pcol = None
    pm = {}
    for row in ws.iter_rows(values_only=True):
        if tcol is None:
            cells = [clean(x).lower() for x in row]
            if 'troquel' in cells and any(c.startswith('pvp al') for c in cells):
                for c, txt in enumerate(cells):
                    if txt == 'troquel':
                        tcol = c
                    if txt.startswith('pvp al'):
                        pcol = c  # la ultima (mas reciente) gana
            continue
        troq = troqnorm(row[tcol]) if tcol < len(row) else None
        pv = row[pcol] if (pcol is not None and pcol < len(row)) else None
        if troq and isinstance(pv, (int, float)) and not isinstance(pv, bool) and pv > 0:
            pm[troq] = float(pv)
    wb.close()
    return pm


def find_header(ws, maxscan=14):
    """Detecta la fila de encabezado por texto y devuelve (row, {campo: col})."""
    keys = {
        'prod': ['producto'],
        'pres': ['presentaci'],
        'droga': ['droga'],
        'lab': ['laboratorio'],
        'pub': ['p. público', 'p. publico', 'p.publico', 'publico'],
        'unit': ['precio x comp'],
        'gap': ['%'],
        'fecha': ['fecha'],
        'troq': ['troquel'],
    }
    for r in range(1, maxscan + 1):
        cells = {c: clean(ws.cell(r, c).value).lower() for c in range(1, ws.max_column + 1)}
        has_prod = any(v == 'producto' for v in cells.values())
        has_pub = any(('p. p' in v) or ('publico' in v) for v in cells.values())
        if has_prod and has_pub:
            cmap = {}
            for field, kk in keys.items():
                for c, txt in cells.items():
                    if not txt:
                        continue
                    if field == 'gap':
                        ok = (txt == '%')
                    elif field == 'prod':
                        ok = (txt == 'producto')
                    elif field == 'fecha':
                        ok = (txt == 'fecha')
                    elif field == 'troq':
                        ok = (txt == 'troquel')
                    else:
                        ok = any(k in txt for k in kk)
                    if ok and field not in cmap:
                        cmap[field] = c
            return r, cmap
    return None, None


def parse_sheet(ws, pricemap=None):
    hr, cm = find_header(ws)
    if not cm or 'prod' not in cm or 'pub' not in cm:
        return []
    P, PRES, DROGA, LAB = cm.get('prod'), cm.get('pres'), cm.get('droga'), cm.get('lab')
    PUB, UNIT, GAP, FECHA = cm.get('pub'), cm.get('unit'), cm.get('gap'), cm.get('fecha')
    TROQ = cm.get('troq')
    last = min(ws.max_row, hr + 500)
    # Agrupar en bloques separados por filas no-data (vacias / titulos / headers repetidos)
    blocks, cur = [], []
    for r in range(hr + 1, last + 1):
        prod = clean(ws.cell(r, P).value)
        pub = ws.cell(r, PUB).value
        is_data = bool(prod) and isinstance(pub, (int, float)) and not isinstance(pub, bool) and pub > 0
        if is_data:
            cur.append(r)
        elif cur:
            blocks.append(cur)
            cur = []
    if cur:
        blocks.append(cur)

    pres_list = []
    for blk in blocks:
        rows = []
        sie_idx = None
        for i, r in enumerate(blk):
            lab = clean(ws.cell(r, LAB).value) if LAB else ''
            is_sie = lab.lower().startswith('siegfried')
            if is_sie and sie_idx is None:
                sie_idx = i
            rows.append({
                'prod': clean(ws.cell(r, P).value),
                'lab': lab,
                'pub': num(ws.cell(r, PUB).value),
                'unit': num(ws.cell(r, UNIT).value) if UNIT else None,
                'gap': None if is_sie else (gapnum(ws.cell(r, GAP).value) if GAP else None),
                'fecha': fechastr(ws.cell(r, FECHA).value) if FECHA else None,
                'troq': troqnorm(ws.cell(r, TROQ).value) if TROQ else None,
                'sie': is_sie,
            })
        if sie_idx is None:
            continue  # bloque sin producto SIE -> no es comparativa SIE
        if pricemap is not None:
            # Refrescar precios por Troquel desde el catalogo plano; recalcular $/unit y gap.
            for rr in rows:
                t = rr.get('troq')
                if t and t in pricemap:
                    pack = max(1, round(rr['pub'] / rr['unit'])) if (rr['pub'] and rr['unit']) else 1
                    rr['pub'] = round(float(pricemap[t]), 2)
                    rr['unit'] = round(float(pricemap[t]) / pack, 2)
            su = rows[sie_idx]['unit']
            for i2, rr in enumerate(rows):
                rr['gap'] = None if (i2 == sie_idx or not su or rr['unit'] is None) else round((rr['unit'] - su) / su, 4)
        sie = rows[sie_idx]
        pres = clean(ws.cell(blk[sie_idx], PRES).value) if PRES else ''
        droga = clean(ws.cell(blk[sie_idx], DROGA).value) if DROGA else ''
        label = (sie['prod'] + (' · ' + pres if pres else '')).strip()
        comps = [x for i, x in enumerate(rows) if i != sie_idx]
        comps.sort(key=lambda x: (x['unit'] is None, x['unit'] if x['unit'] is not None else 0))
        ordered = [sie] + comps
        pres_list.append({'label': label, 'droga': droga, 'rows': ordered})
    return pres_list


def snapshot_date(xlsx_path):
    m = re.search(r'(\d{2})[._-](\d{2})[._-](\d{4})', Path(xlsx_path).name)
    if m:
        return '%s/%s/%s' % (m.group(1), m.group(2), m.group(3))
    return ''


def main():
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx', nargs='?', default=DEFAULT_XLSX,
                    help='Excel ESTRUCTURA (hojas por marca, formato comparativa curada)')
    ap.add_argument('--prices', help='Catalogo plano con PVP actual; refresca precios por Troquel')
    ap.add_argument('--fecha', help='Fecha snapshot DD/MM/AAAA (default: del nombre del archivo)')
    args = ap.parse_args()
    xlsx = args.xlsx
    if not Path(xlsx).is_file():
        print('ERROR: no existe el Excel:', xlsx)
        return 1
    pricemap = None
    if args.prices:
        if not Path(args.prices).is_file():
            print('ERROR: no existe el catalogo de precios:', args.prices)
            return 1
        pricemap = load_prices(args.prices)
        print('Precios del catalogo: %d troqueles' % len(pricemap))
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    prec_comp, total_pres, total_rows = {}, 0, 0
    for sheet, group in SHEET_GROUPS.items():
        if sheet not in wb.sheetnames:
            print('  [aviso] hoja faltante:', sheet)
            continue
        ws = wb[sheet]
        grid = Grid([list(row) for i, row in enumerate(ws.iter_rows(values_only=True)) if i < 600])
        pres_list = parse_sheet(grid, pricemap)
        if pres_list:
            prec_comp[group] = {'pres': pres_list}
            total_pres += len(pres_list)
            total_rows += sum(len(p['rows']) for p in pres_list)
            print('  [%-12s] %2d presentaciones, %3d filas' % (group, len(pres_list),
                  sum(len(p['rows']) for p in pres_list)))
    wb.close()
    fecha = args.fecha or (snapshot_date(args.prices) if args.prices else '') or snapshot_date(xlsx)
    meta = {'fecha': fecha, 'fuente': 'Comparativa de Precios (Manual Farmacéutico)'}

    # Formato default (con espacios) para que coincida con el resto de const D y con
    # cualquier re-serializacion posterior (p.ej. merge-stock-month.py), evitando churn.
    block = ('"prec_comp":' + json.dumps(prec_comp, ensure_ascii=False)
             + ',"prec_comp_meta":' + json.dumps(meta, ensure_ascii=False) + ',')
    # Marcadores como CLAVES JSON validas (NO comentarios /* */): asi const D sigue siendo
    # JSON parseable por las herramientas (audit-full.py y los checks del pre-commit).
    payload = '"_pcomp":1,' + block + '"_pcompEnd":1,'
    assert '</script' not in block.lower() and '"_pcompEnd":1,' not in block, 'payload peligroso'

    with open(HTML, 'r', encoding='utf-8', newline='') as f:
        html = f.read()
    if 'const D = {' not in html:
        print('ERROR: no se encontro "const D = {" en', HTML)
        return 1
    # Limpiar cualquier inyeccion previa (markers viejos de comentario o sentinel keys) e insertar fresco.
    n_old = (len(re.findall(r'/\*PCOMP_START\*/.*?/\*PCOMP_END\*/', html, flags=re.DOTALL))
             + len(re.findall(r'"_pcomp"\s*:\s*1\s*,.*?"_pcompEnd"\s*:\s*1\s*,', html, flags=re.DOTALL)))
    html = re.sub(r'/\*PCOMP_START\*/.*?/\*PCOMP_END\*/', '', html, flags=re.DOTALL)
    html = re.sub(r'"_pcomp"\s*:\s*1\s*,.*?"_pcompEnd"\s*:\s*1\s*,', '', html, flags=re.DOTALL)
    html2 = html.replace('const D = {', 'const D = {' + payload, 1)
    mode = 'reemplazado' if n_old else 'insertado'
    if html2 == html:
        print('Sin cambios (payload identico).')
    else:
        with open(HTML, 'w', encoding='utf-8', newline='') as f:
            f.write(html2)
    print('\nprec_comp %s en mujer/index.html: %d grupos, %d presentaciones, %d filas. Fecha snapshot: %s'
          % (mode, len(prec_comp), total_pres, total_rows, meta['fecha']))
    return 0


if __name__ == '__main__':
    sys.exit(main())
