#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Extiende MAGNUS / MAGNUS 36 con los meses que el export curado de MKT no cubre.

EL PROBLEMA. El mercado de MAGNUS no sale del AR_PM: lo define un export curado por
Marketing ('Mercado-MAGNUS-sildenafil-tadalafil.xlsx') que separa sildenafil de
tadalafil, algo que la molecula sola no distingue bien. Ese archivo puede llegar tarde:
en el cierre de Jul-2026 tenia 60 columnas mensuales Jun-2021..May-2026, dos meses
menos que el cierre. `rebuild-otc-magnus-from-iqvia.py` reescribe los products enteros,
asi que las dos familias quedaban SIN junio (un mes ATRAS de produccion).

LA SOLUCION (decision del usuario, 2026-08-18). Hibrido: los meses que el archivo
curado cubre quedan como estan (son los publicados), y los meses faltantes se completan
desde el master AR_PM filtrando por molecula. Se midio antes de decidir:

    MAGNUS     May-2026  curado 194.484  vs master 195.119   -0,33%
    MAGNUS 36  May-2026  curado 113.479  vs master 115.692   -1,91%
    MAGNUS     Jun-2026  publicado 194.163 vs master 194.228 -0,03%  <- la costura

O sea el master reproduce la curacion de MKT con un desvio de decimas, y el mes de
empalme es practicamente identico al ya publicado.

EL MERCADO SIGUE SIENDO EL SET CURADO. Solo se completan los productos que YA estan en
la familia; las marcas de la molecula que MKT no incluyo NO se agregan (en Jul-2026 eso
deja 147 u. afuera en MAGNUS y 3.454 en MAGNUS 36, que se reportan). Cambiar la
composicion seria otra decision.

Uso:
  python shared/extend-magnus-desde-master.py --master <AR_PM.xlsx> --mes "Jun 2026" --mes "Jul 2026"
"""
from __future__ import annotations
import argparse, json, os, re, sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
TARGET = 'OTC/data.js'
# familia -> keyword de molecula que la define
FAMILIAS = {'MAGNUS': 'SILDENAFIL', 'MAGNUS 36': 'TADALAFIL'}


def leer_master(path, meses):
    """units[(molecula_grupo, PRODUCT)][mes] desde el master, columnas por HEADER."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    hdr = [str(h).replace('\n', ' ').strip() if h else ''
           for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    try:
        ci = {h[len('Units'):].strip(): i for i, h in enumerate(hdr) if h.startswith('Units')}
        pi = next(i for i, h in enumerate(hdr) if h.strip().lower() == 'product')
        mi = next(i for i, h in enumerate(hdr) if h.strip().lower().startswith('molecules'))
    except StopIteration:
        raise SystemExit(f'FATAL: no encuentro Product/Molecules por header en {path}')
    falta = [m for m in meses if m not in ci]
    if falta:
        raise SystemExit(f'FATAL: el master no trae {falta}')
    want = {m: ci[m] for m in meses}

    U, tot_mol = {}, {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or pi >= len(row) or not row[pi]:
            continue
        mol = str(row[mi] or '').upper()
        grupo = None
        for fam, kw in FAMILIAS.items():
            if kw in mol:
                grupo = fam
                break
        # TADALAFIL gana si aparecen los dos (combos): el orden del dict lo respeta
        if 'TADALAFIL' in mol:
            grupo = 'MAGNUS 36'
        if grupo is None:
            continue
        name = str(row[pi]).strip().upper()
        for m, c in want.items():
            v = row[c] if c < len(row) else None
            if isinstance(v, (int, float)):
                U[(grupo, name, m)] = U.get((grupo, name, m), 0.0) + float(v)
                tot_mol[(grupo, m)] = tot_mol.get((grupo, m), 0.0) + float(v)
    wb.close()
    return U, tot_mol


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--master', default=None)
    ap.add_argument('--mes', action='append', required=True)
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    master = args.master
    if not master:
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        import manifest
        master = manifest.resolve_source('iqvia_master')
    print(f'master: {master}')
    U, tot_mol = leer_master(master, args.mes)

    p = REPO / TARGET
    t = p.read_text(encoding='utf-8-sig', errors='replace')
    m = re.search(r'window\.OTC_DASHBOARD\s*=\s*', t)
    ob = t.index('{', m.end())
    D, end = json.JSONDecoder().raw_decode(t[ob:])

    escritos = 0
    for fam in FAMILIAS:
        o = (D.get('mol_perf') or {}).get(fam)
        if not o:
            print(f'  SKIP {fam}: no existe en mol_perf')
            continue
        for mes in args.mes:
            cubierto = 0.0
            for prod in o.get('products', []):
                name = str(prod.get('prod') or '').strip().upper()
                if 'OTROS' in name:
                    continue
                v = U.get((fam, name, mes), 0.0)
                mv = prod.setdefault('monthly_vals', {})
                if mv.get(mes) != int(round(v)):
                    if not args.dry_run:
                        mv[mes] = int(round(v))
                    escritos += 1
                cubierto += v
            total_molecula = tot_mol.get((fam, mes), 0.0)
            resto = total_molecula - cubierto
            pct = (resto / total_molecula * 100) if total_molecula else 0
            print(f'  {fam:11s} {mes}: mercado curado={cubierto:11,.0f}  '
                  f'molecula completa={total_molecula:11,.0f}  '
                  f'fuera del set curado={resto:9,.0f} ({pct:.2f}%)')

    if args.dry_run:
        print(f'\nDRY RUN: {escritos} valores a escribir.')
        return 0
    p.write_text(t[:ob] + json.dumps(D, ensure_ascii=False) + t[ob + end:],
                 encoding='utf-8', newline='')
    print(f'\n-> {TARGET}: {escritos} valores escritos.')
    print('   Correr: recompute-mol-perf-aggregates, ensure-magnus36-brandkpis, build-kpis,')
    print('   build-families, sync-kpistrip, fix-brandkpis-*, label-mercados-atrasados, build-total.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
