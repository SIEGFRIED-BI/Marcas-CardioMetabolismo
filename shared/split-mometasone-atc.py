#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
shared/split-mometasone-atc.py

Separa el "mercado" MOMETASONE en los TRES mercados terapeuticos reales que IQVIA
junta bajo la misma molecula, y le da a cada linea el que le corresponde:

    D07A0 - CORTICOST TOPICOS SOLOS      -> dermatologia/data.js   mol_perf['MOMETASONE']
    R01A1 - CORTIC NAS SIN ANTIINFECC    -> respiratorio/data.js   mol_perf['HEXALER NASAL']
    R03D1 - CORTICOIDES INHALANTES       -> respiratorio/data.js   mol_perf['HEXALER BRONQUIAL']

POR QUE (auditoria 2026-07-30). Las tres entradas tenian el MISMO total (la molecula
entera: 91.547 u. en Jun 2026 / 999.103 MAT) y cada marca calculaba su share contra
ese universo:
    - dermato mostraba MS% 72,4% para "MOMETASONE" sumando HEXALER NASAL (spray nasal)
      y HEXALER BRONQUIAL (caps p/inhalar), que son marcas de RESPIRATORIO. El share
      real de MOMETAX contra su mercado (topicos) es 58,7%.
    - respiratorio tenia HEXALER NASAL y HEXALER BRONQUIAL como dos mercados distintos
      pero con el mismo contenido (la molecula), mas un bucket 'Otros (resto del
      mercado)' que absorbia la diferencia.
Los packs no dejan dudas: MOMETAX = crema/locion; HEXALER NASAL = spray nasal;
HEXALER BRONQUIAL = capsulas para inhalar. Es la regla #2 de CLAUDE.md (una familia
de mol_perf = UN mercado) y el mismo patron que split-atb-cefalexina-duo.py, pero
partiendo por CLASE ATC en vez de por dosis.

COMO. La clasificacion sale del master AR_PM (columna ATC, molecula MOMETASONE).
Para cada mercado destino se reconstruyen los productos de SU ATC y los agregados.
Se respeta el rango de meses que ya tenia cada mercado (no se agregan ni se pierden
meses; verify-history-preserved compara la union por linea). Los meses que el mercado
tiene pero el master no cubre (dermato arranca Apr 2021 y el master en Jul 2021) se
conservan con los valores que ya estaban, quedandose solo con los productos del ATC
destino: no se inventa nada.

INVARIANTE QUE SE VERIFICA ANTES DE ESCRIBIR: para cada mes cubierto por el master,
la suma de los totales de los tres mercados nuevos == el total del mercado viejo
(es una particion de la molecula, no puede crear ni perder unidades).

Despues de correr esto hay que recomponer los KPIs (lo hace update-all.ps1):
    fix-brandkpis-market-total.py -> fix-brandkpis-ie-vs-market.py
    build-kpis.py -> build-families-perf.py -> sync-kpistrip-with-kpis-json.py

Uso:
    py shared/split-mometasone-atc.py [--master <xlsx>] [--dry-run]
"""
from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
DEFAULT_MASTER = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs'
                      r'\_iqvia-master\2026-06\AR_PM_FV_Standard_Jul-2026.xlsx')

MOLECULE = 'MOMETASONE'
MES = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
MES_INV = {m: i + 1 for i, m in enumerate(MES)}

# (archivo, clave de mol_perf, prefijo de ATC, marca SIE esperada)
TARGETS = [
    ('dermatologia/data.js', 'MOMETASONE',        'D07A0', 'MOMETAX (SIE)'),
    ('respiratorio/data.js', 'HEXALER NASAL',     'R01A1', 'HEXALER NASAL (SIE)'),
    ('respiratorio/data.js', 'HEXALER BRONQUIAL', 'R03D1', 'HEXALER BRONQUIAL (SIE)'),
]


def msort(mk):
    p = str(mk).split()
    return (int(p[1]), MES_INV.get(p[0], 0)) if len(p) == 2 and p[1].isdigit() else (0, 0)


def quarter_key(mk):
    p = mk.split()
    m = MES_INV.get(p[0]) if len(p) == 2 else None
    return 'Q{} {}'.format((m - 1) // 3 + 1, p[1]) if m else ''


def agg_quarterly(monthly):
    out = defaultdict(int)
    for mk, v in monthly.items():
        q = quarter_key(mk)
        if q:
            out[q] += int(v or 0)
    return dict(out)


def agg_ytd(monthly, cierre):
    by_year = defaultdict(int)
    for mk, v in monthly.items():
        p = mk.split()
        m = MES_INV.get(p[0]) if len(p) == 2 else None
        if m and m <= cierre:
            by_year[p[1]] += int(v or 0)
    return {'{} {}'.format(MES[cierre - 1], y): v for y, v in by_year.items()}


def agg_mat(monthly, cierre):
    years = {int(mk.split()[1]) for mk in monthly if len(mk.split()) == 2}
    out = {}
    for y in sorted(years):
        tot = 0
        for back in range(11, -1, -1):
            idx = (y * 12 + (cierre - 1)) - back
            yy, mm = divmod(idx, 12)
            tot += int(monthly.get('{} {}'.format(MES[mm], yy), 0) or 0)
        out['{} {}'.format(MES[cierre - 1], y)] = tot
    return out


def read_master(path):
    """-> {product: {'atc': str, 'manuf': str, 'monthly': {mes: u}}} para la molecula."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    r1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    ci, mcols = {}, []
    for i, h in enumerate(r1):
        if not h:
            continue
        s = str(h).strip()
        sn = s.replace('\n', ' ').strip().lower()
        if sn.startswith('manufacturer'):
            ci['manuf'] = i
        elif sn.startswith('product'):
            ci['prod'] = i
        elif sn.startswith('molecules'):
            ci['mol'] = i
        elif sn.startswith('atc'):
            ci.setdefault('atc', i)
        if s.startswith('Units'):
            a = (s.split('\n', 1)[-1] if '\n' in s else s[len('Units'):]).strip()
            if a.upper().startswith(('MAT', 'YTD')):
                continue
            m = re.match(r'(\w+)\s+(\d{4})$', a)
            if m and m.group(1) in MES_INV:
                mcols.append((i, '{} {}'.format(m.group(1), m.group(2))))
    for need in ('prod', 'mol', 'manuf', 'atc'):
        if need not in ci:
            raise RuntimeError('no se encontro la columna {} en el master'.format(need))

    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        mol = row[ci['mol']] if ci['mol'] < len(row) else None
        if not mol or str(mol).strip().upper() != MOLECULE:
            continue
        prod = str(row[ci['prod']] or '').strip()
        if not prod:
            continue
        d = out.setdefault(prod, {'atc': '', 'manuf': '', 'monthly': defaultdict(float)})
        d['manuf'] = str(row[ci['manuf']] or '').strip()
        atc = str(row[ci['atc']] or '').strip()
        if atc:
            d['atc'] = atc
        for c, mk in mcols:
            if c < len(row) and isinstance(row[c], (int, float)):
                d['monthly'][mk] += row[c]
    for d in out.values():
        d['monthly'] = {k: int(round(v)) for k, v in d['monthly'].items()}
    wb.close()
    return out, [mk for _, mk in mcols]


def load_D(rel):
    p = REPO / rel
    text = p.read_text(encoding='utf-8', errors='replace')
    m = re.search(r'window\.OTC_DASHBOARD\s*=\s*', text)
    if not m:
        raise RuntimeError('window.OTC_DASHBOARD no encontrado en {}'.format(rel))
    start = text.index('{', m.end())
    D, end = json.JSONDecoder().raw_decode(text[start:])
    return p, text, start, start + end, D


def build_market(old_fam, master, atc_prefix, master_months):
    """Devuelve (products, fam_monthly) para el ATC pedido, respetando los meses
    que el mercado ya tenia."""
    keep_months = set(old_fam.get('monthly', {}))
    mm = set(master_months)
    # productos del master con ese ATC
    names = [p for p, d in master.items() if str(d['atc']).startswith(atc_prefix)]
    # los meses del mercado que el master NO cubre: se conservan con lo que ya habia
    huerfanos = sorted(keep_months - mm, key=msort)
    old_by_prod = {p.get('prod'): p for p in old_fam.get('products', [])}

    products, fam_monthly = [], defaultdict(int)
    for name in sorted(names, key=lambda n: -sum(master[n]['monthly'].values())):
        d = master[name]
        monthly = {}
        for mk in sorted(keep_months & mm, key=msort):
            monthly[mk] = int(d['monthly'].get(mk, 0) or 0)
        # meses previos al master: valor que ya tenia ese producto en el artefacto
        op = old_by_prod.get(name)
        for mk in huerfanos:
            v = (op.get('monthly_vals', {}) or {}).get(mk) if op else None
            monthly[mk] = int(v or 0)
        if not any(monthly.values()):
            continue
        for mk, v in monthly.items():
            fam_monthly[mk] += v
        products.append({
            'prod': name,
            'manuf': d['manuf'],
            'is_sie': d['manuf'].strip().upper() == 'SIEGFRIED',
            'monthly_vals': monthly,
        })
    return products, dict(fam_monthly), huerfanos


def finish_market(products, fam_monthly):
    cierre = MES_INV[max(fam_monthly, key=msort).split()[0]]
    fam_q = agg_quarterly(fam_monthly)
    fam_y = agg_ytd(fam_monthly, cierre)
    fam_m = agg_mat(fam_monthly, cierre)
    for p in products:
        mv = p['monthly_vals']
        p['quarterly_vals'] = agg_quarterly(mv)
        p['ytd'] = agg_ytd(mv, cierre)
        p['mat'] = agg_mat(mv, cierre)
        p['ms_monthly'] = {k: round(mv.get(k, 0) / v * 100, 2) if v > 0 else 0 for k, v in fam_monthly.items()}
        p['ms_quarterly'] = {k: round(p['quarterly_vals'].get(k, 0) / v * 100, 2) if v > 0 else 0 for k, v in fam_q.items()}
        p['ms_ytd'] = {k: round(p['ytd'].get(k, 0) / v * 100, 2) if v > 0 else 0 for k, v in fam_y.items()}
        p['ms_mat'] = {k: round(p['mat'].get(k, 0) / v * 100, 2) if v > 0 else 0 for k, v in fam_m.items()}
    products.sort(key=lambda p: (not p['is_sie'], -(p['monthly_vals'].get(max(fam_monthly, key=msort), 0))))
    return {'quarterly': fam_q, 'ytd': fam_y, 'mat': fam_m}, cierre


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--master', default=str(DEFAULT_MASTER))
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    mp = Path(args.master)
    if not mp.is_file():
        print('ERROR: master no existe: {}'.format(mp), file=sys.stderr)
        return 2
    print('Master: {}'.format(mp))
    master, master_months = read_master(mp)
    if not master:
        print('ERROR: no hay filas de la molecula {} en el master'.format(MOLECULE), file=sys.stderr)
        return 3
    por_atc = defaultdict(list)
    for p, d in master.items():
        por_atc[d['atc']].append(p)
    print('  molecula {}: {} productos, {} meses ({}..{})'.format(
        MOLECULE, len(master), len(master_months), master_months[0], master_months[-1]))
    for a in sorted(por_atc):
        print('    {:<40} {} productos'.format(a[:40], len(por_atc[a])))

    # ── construir los 3 mercados ─────────────────────────────────────────
    files = {}          # rel -> (path, text, s, e, D)
    nuevos = {}         # (rel, key) -> (products, fam_monthly, aggs)
    viejos_monthly = {}
    for rel, key, atc, sie_expect in TARGETS:
        if rel not in files:
            files[rel] = load_D(rel)
        _, _, _, _, D = files[rel]
        if key not in D.get('mol_perf', {}):
            print('ERROR: {} no tiene mol_perf[{!r}]'.format(rel, key), file=sys.stderr)
            return 4
        old_fam = D['mol_perf'][key]
        viejos_monthly[(rel, key)] = {k: int(v or 0) for k, v in old_fam.get('monthly', {}).items()}
        prods, fam_monthly, huerf = build_market(old_fam, master, atc, master_months)
        if not prods:
            print('ERROR: ATC {} no dio ningun producto'.format(atc), file=sys.stderr)
            return 5
        aggs, cierre = finish_market(prods, fam_monthly)
        nuevos[(rel, key)] = (prods, fam_monthly, aggs)
        sie = [p['prod'] for p in prods if p['is_sie']]
        last = max(fam_monthly, key=msort)
        print()
        print('  [{}] {} <- ATC {}'.format(rel, key, atc))
        print('     productos: {} (SIE: {})'.format(len(prods), ', '.join(sie) or 'ninguno'))
        print('     mercado {}: {:,} u.   MAT {}: {:,} u.'.format(last, fam_monthly[last], last, aggs['mat'].get(last, 0)))
        if sie_expect not in sie:
            print('     OJO: esperaba {!r} entre los SIE'.format(sie_expect))
        if huerf:
            print('     meses previos al master conservados del artefacto: {}'.format(', '.join(huerf)))
        for p in prods[:6]:
            print('       {:<28} {:>9,} u.  MS {:>6.2f}%'.format(
                p['prod'][:28], p['monthly_vals'].get(last, 0), p['ms_monthly'].get(last, 0)))

    # ── INVARIANTES ──────────────────────────────────────────────────────
    # OJO: NO se puede exigir "suma de los 3 mercados nuevos == mercado viejo",
    # porque (a) el mercado viejo ERA la molecula entera (eso es justo lo que se
    # esta corrigiendo) y (b) los 3 mercados tienen rangos de meses distintos
    # (dermato 63 meses, los de respi 29) -> en un mes de 2021 solo existe el de
    # dermato y la suma daria solo los topicos.
    # Los dos invariantes que SI corresponden:
    #   A. la particion por ATC no pierde ni un producto de la molecula
    #   B. cada mercado nuevo == el total de SU grupo ATC en el master
    print()
    print('=' * 78)
    print('INVARIANTE A: la particion por ATC cubre la molecula completa (por mes)')
    print('=' * 78)
    atcs = [t[2] for t in TARGETS]
    malos = []
    for mk in master_months:
        tot_mol = sum(int(d['monthly'].get(mk, 0) or 0) for d in master.values())
        tot_part = 0
        for a in atcs:
            tot_part += sum(int(d['monthly'].get(mk, 0) or 0)
                            for d in master.values() if str(d['atc']).startswith(a))
        if tot_mol != tot_part:
            malos.append((mk, tot_mol, tot_part))
    if malos:
        print('FAIL: {} meses con productos fuera de los 3 ATC. Primeros:'.format(len(malos)))
        for mk, a, b in malos[:8]:
            print('   {:<10} molecula={:>10,}  particion={:>10,}  quedan afuera={:>+9,}'.format(mk, a, b, a - b))
        sueltos = sorted({p for p, d in master.items()
                          if not any(str(d['atc']).startswith(a) for a in atcs)})
        if sueltos:
            print('   productos sin ATC destino: {}'.format(sueltos))
        print('\nNo se escribio nada.')
        return 6
    print('PASS: los {} meses del master cierran exacto (ningun producto queda afuera).'.format(len(master_months)))

    print()
    print('=' * 78)
    print('INVARIANTE B: cada mercado nuevo == total de su grupo ATC en el master')
    print('=' * 78)
    malos_b = []
    for rel, key, atc, _ in TARGETS:
        fam_monthly = nuevos[(rel, key)][1]
        comunes = [mk for mk in master_months if mk in fam_monthly]
        for mk in comunes:
            esperado = sum(int(d['monthly'].get(mk, 0) or 0)
                           for d in master.values() if str(d['atc']).startswith(atc))
            if fam_monthly[mk] != esperado:
                malos_b.append((key, mk, esperado, fam_monthly[mk]))
        print('  {:<20} {:>3} meses verificados contra el master'.format(key, len(comunes)))
    if malos_b:
        print('FAIL: {} celdas no cierran. Primeras:'.format(len(malos_b)))
        for key, mk, a, b in malos_b[:8]:
            print('   {:<20} {:<10} master={:>10,}  nuevo={:>10,}'.format(key, mk, a, b))
        print('\nNo se escribio nada.')
        return 6
    print('PASS: los 3 mercados reproducen exacto el total de su ATC.')

    print()
    print('=' * 78)
    print('CAMBIO INTENCIONAL (el mercado deja de ser la molecula entera)')
    print('=' * 78)
    print('{:<20} {:<10} {:>14} {:>14} {:>9}'.format('mercado', 'mes', 'viejo', 'nuevo', 'ratio'))
    for rel, key, atc, _ in TARGETS:
        viejo = viejos_monthly[(rel, key)]
        nuevo = nuevos[(rel, key)][1]
        mk = max(nuevo, key=msort)
        a, b = viejo.get(mk, 0), nuevo.get(mk, 0)
        print('{:<20} {:<10} {:>14,} {:>14,} {:>8.2f}x'.format(key[:20], mk, a, b, (b / a) if a else 0))

    if args.dry_run:
        print('\nDRY RUN: no se escribio nada.')
        return 0

    # ── escribir ─────────────────────────────────────────────────────────
    print()
    for rel, (p, text, s, e, D) in files.items():
        for (r2, key), (prods, fam_monthly, aggs) in nuevos.items():
            if r2 != rel:
                continue
            fam = D['mol_perf'][key]
            fam['products'] = prods
            fam['monthly'] = fam_monthly
            fam['quarterly'] = aggs['quarterly']
            fam['ytd'] = aggs['ytd']
            fam['mat'] = aggs['mat']
        newtext = text[:s] + json.dumps(D, ensure_ascii=False) + text[e:]
        p.write_text(newtext, encoding='utf-8', newline='')
        print('-> {} reescrito ({:,} bytes)'.format(p.relative_to(REPO), p.stat().st_size))
    print('\nAhora recomponer KPIs: fix-brandkpis-market-total.py, fix-brandkpis-ie-vs-market.py,')
    print('build-kpis.py, build-families-perf.py, sync-kpistrip-with-kpis-json.py')
    return 0


if __name__ == '__main__':
    sys.exit(main())
