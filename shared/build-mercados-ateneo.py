#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
shared/build-mercados-ateneo.py

Agrega a cada data.js la clave `mercadosAteneo`: una VISTA ALTERNATIVA del mercado usando
los 79 MERCADOS CURADOS DEL ATENEO, para poder medir cada marca contra su universo amplio
ademas de contra su molecula exacta.

Pedido del usuario: "mantene los que estan pero agrega por ejemplo para roxolan el mercado
de hipolipemeantes" / "no quiero agregar productos y demas, son formas extras de analizar
el mercado" / "pero no me armaste como los mercados del ateneo".

REEMPLAZA A build-mercados-atc.py
--------------------------------
La primera version uso las clases ATC III ('C10A - PRD REGULADORES LIPIDOS') y el usuario
la rechazo: los mercados del Ateneo NO son clases ATC, son 79 agrupaciones curadas a mano
('Roxolan (Hipolipemeantes)', 'Betabloqueantes (Dilatrend-Nebilet)', 'Antibioticos
(Entry M.)'). Algunas coinciden con una ATC, la mayoria no.

DE DONDE SALE CADA COSA, Y POR QUE DE AHI
-----------------------------------------
  CLASIFICACION (que marcas forman cada mercado): 'Ateneo Febrero-26.xlsb', hoja DATOS.
    Es el UNICO archivo que tiene la columna `Mercado`. Se verifico que
    'Ateneo Total - MAT Movil_Jul-16-2026.xlsx' NO la tiene: ese archivo es el schema de
    AR_PM (Pack/Manufacturer/ATC IV/Product/...) y por eso el intento anterior termino
    cayendo en ATC III.
    Las filas utiles son TIPO='CUP' (por region) y TIPO='MINV' (nacional). En TIPO='ME' y
    TIPO='CUPTOTAL' la columna `Mercado` viene VACIA -- son el universo entero sin
    etiquetar, no un mercado.

  UNIDADES: el master AR_PM (mensual, al dia). El xlsb es un extracto de febrero-26 e
    IQVIA re-expresa: sobre 402 marcas contenidas enteras en algun mercado, la mediana del
    ratio xlsb/AR_PM da 0,9998-1,0002 en los 12 meses, pero solo ~60% calza al 0,1%. Esa
    diferencia es re-expresion, y se resuelve del lado del dato mas fresco.

LOS MERCADOS SE SOLAPAN A PROPOSITO -- NUNCA SUMARLOS ENTRE SI
--------------------------------------------------------------
El Ateneo tiene mercados anidados: ACANTEX esta en 'Cefalospor Iny (Acantex)' Y en
'Antibioticos (Entry M.)'; 15,6% de las marcas (529 de 3.387) cae en mas de uno. Eso es
la feature, no un bug: es la lectura ancha que se pidio. Consecuencia dura: la suma de los
79 mercados es MAYOR que el universo, asi que esta clave no se agrega a nada. Vive aparte
de mol_perf por la misma razon que la anterior (build-total.py y
check-total-consistency.py recorren mol_perf para armar el mercado y el SIE de COMPANIA;
meter aca un universo ancho inflaria el mercado total y bajaria el MS% publicado).

UNA FILA POR (PRODUCTO PROPIO, MERCADO)
---------------------------------------
La salida es `filas`: una lista de {label, mercado, propios}, NO un mapa familia->mercado.
Decision del usuario, sobre un caso concreto: ACNECLIN (50mg tabl, 16.277 u) y ACNECLIN AP
(100mg caps A.P., 124.664 u) son productos DISTINTOS pero viven en la misma familia del
tablero (dermatologia/MINOCYCLINE), que ademas no tiene budIqviaMap. Agrupando por familia
salian sumados en un 17,43% que no describe a ninguno de los dos: ACNECLIN AP es el LIDER
de 'Tetraciclinas (Acneclin)' con 15,42% y ACNECLIN esta 14no con 2,01%.
Un producto propio puede figurar en 2 familias del tablero ('MICOMAZOL (SIE)' esta en
CLOTRIMAZOLE y en CICLOPIROX), asi que las filas se DEDUPLICAN por (producto, mercado): sin
eso el toggle las colapsaba igual, por clave de objeto, pero en silencio -- `filas` decia 16
y la tabla dibujaba 15.

DECODIFICAR LAS COLUMNAS DE PERIODO DEL XLSB (para el cruce)
------------------------------------------------------------
Los headers de periodo del xlsb son indices desnudos (1..66) sin etiqueta de mes. Se
decodificaron EMPIRICAMENTE comparando el vector de cada indice contra los meses conocidos
del AR_PM sobre 402 marcas comparables:
    indices 40..51  = Mar 2025 .. Feb 2026   (mensual, ratio mediano 0,9998-1,0002)
    indices 6..17 y 35..39 = MAT movil (ratio ~12x el mes)
    indices 18..34 y 62..66 = importes en pesos (ratio ~1e4)
Buscar "el mes que mejor matchea" entre los 60 del AR_PM encuentra ratios ~1,0 por
casualidad (asi salio '51 -> Feb 2025', que es el mes correcto corrido un anio): hay que
testear cada indice contra SU mes hipotetizado y contra el anterior y el siguiente.

CRUCE (G1)
----------
Para cada mercado: suma AR_PM sobre sus marcas en Mar2025..Feb2026 vs la suma del propio
xlsb en los indices 40..51. Si un mercado no cierra dentro de --tol, se REPORTA y no se
publica ese mercado. Un mercado al que le faltan competidores por un nombre que no matchea
sale con el total corto y por lo tanto con el MS% propio INFLADO, que es el error en la
direccion peligrosa.

Uso:
    python shared/build-mercados-ateneo.py [--master <AR_PM.xlsx>] [--xlsb <Ateneo.xlsb>]
                                           [--tol 0.05] [--dry-run] [--report <json>]
"""
from __future__ import annotations
import argparse, importlib.util, json, re, sys, unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl
from pyxlsb import open_workbook as open_xlsb

REPO = Path(__file__).resolve().parent.parent

# Lista canonica de productos excluidos de TODOS los analisis (shared/excluded-products.py).
# Hay que respetarla aca tambien: el mercado del Ateneo los trae de vuelta --
# 'CALCITOL D3 (SIE)' reaparecio via la vista ATC en mujer y lo detecto
# check-syntax-and-consistency.py. El nombre del archivo tiene guion, se carga a mano.
_spec = importlib.util.spec_from_file_location('excluded', REPO / 'shared' / 'excluded-products.py')
_excl = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_excl)
is_excluded = _excl.is_excluded

HUB = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs')
DEFAULT_MASTER = HUB / '_iqvia-master' / '2026-06' / 'AR_PM_FV_Standard_Jul-2026.xlsx'
DEFAULT_XLSB = HUB / 'Ateneo Febrero-26.xlsb'
LINES = ['cardio/data.js', 'ATB/data.js', 'OTC/data.js', 'respiratorio/data.js',
         'mujer/data.js', 'SNC/data.js', 'dermatologia/data.js']

MES = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
MI = {m: i + 1 for i, m in enumerate(MES)}
# ventana mensual del xlsb, decodificada empiricamente (ver docstring)
XLSB_MESES = {40 + i: m for i, m in enumerate(
    ['Mar 2025', 'Apr 2025', 'May 2025', 'Jun 2025', 'Jul 2025', 'Aug 2025',
     'Sep 2025', 'Oct 2025', 'Nov 2025', 'Dec 2025', 'Jan 2026', 'Feb 2026'])}


def msort(mk):
    p = str(mk).split()
    return (int(p[1]), MI.get(p[0], 0)) if len(p) == 2 and p[1].isdigit() else (0, 0)


def norm(s):
    """mayusculas, sin acentos, sin puntuacion, espacios colapsados."""
    s = unicodedata.normalize('NFKD', str(s))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r'[^A-Z0-9]+', ' ', s.upper()).strip()


def sin_lab(prod):
    """'ROXOLAN (SIE)' -> 'ROXOLAN'. AR_PM sufija el laboratorio entre parentesis; el
    xlsb trae la marca pelada."""
    return re.sub(r'\s*\([^)]*\)\s*$', '', str(prod)).strip()


def es_sie(nombre, manuf):
    return '(SIE)' in str(nombre).upper() or 'SIEGFRIED' in str(manuf or '').upper()


# ── fuentes ───────────────────────────────────────────────────────────────────
def leer_xlsb(path):
    """-> (pack_norm -> {mercado}, mercado -> {pack_norm}, mercado -> {mes: unidades})

    LA LLAVE ES LA PRESENTACION, NO LA MARCA. Los mercados del Ateneo se definen pack por
    pack: una marca puede tener unas presentaciones adentro del mercado y otras afuera.
    Con llave de marca el cruce daba 53 de 79 mercados, con desvios de hasta +527%
    ('Prucalopride (Prucal)': 22 de 37 marcas encontradas y sin embargo 6x las unidades),
    o sea metia todos los packs de cada marca. norm(Presentacion) del xlsb matchea el
    norm(Pack) del AR_PM en 5.161 de 5.237 presentaciones (98,5%).
    """
    de_pack = defaultdict(set)
    del_mkt = defaultdict(set)
    mkt_mes = defaultdict(lambda: defaultdict(float))
    pres_meta = {}          # pack_norm -> {marca, u12}: lo usa el rescate de huerfanas
    with open_xlsb(str(path)) as wb:
        with wb.get_sheet('DATOS') as sh:
            it = sh.rows()
            hdr = [c.v for c in next(it)]
            iT, iM, iMarca = hdr.index('TIPO'), hdr.index('Mercado'), hdr.index('Presentación')
            iMarcaCol = hdr.index('Marca')
            # columnas de periodo: header numerico >= col 22. El indice 35 aparece dos
            # veces (col 22 y col 88); se toma la primera.
            vistos, pcols = set(), []
            for i, v in enumerate(hdr):
                if isinstance(v, float) and i >= 22 and int(v) not in vistos:
                    vistos.add(int(v))
                    if int(v) in XLSB_MESES:
                        pcols.append((i, XLSB_MESES[int(v)]))
            # SOLO TIPO='MINV'. Es el mercado a nivel NACIONAL, y es el mismo nivel del
            # que sale mkt_mes -- definir la membresia con CUP+MINV y compararla contra un
            # total que sale solo de MINV mezcla dos granularidades: las filas CUP suman
            # 4.582 presentaciones extra que no existen en AR_PM (47% sin match, contra
            # 1,5% usando MINV sola). CUP es la apertura regional de lo mismo.
            n_cup = n_minv = n_dup = 0
            vistos_par = set()
            for row in it:
                v = [c.v for c in row]
                t = v[iT]
                if t == 'CUP':
                    n_cup += 1
                    continue
                if t != 'MINV' or v[iM] is None or v[iMarca] is None:
                    continue
                n_minv += 1
                mkt = str(v[iM]).strip()
                mn = norm(v[iMarca])
                de_pack[mn].add(mkt)
                del_mkt[mkt].add(mn)
                # DEDUP por (mercado, presentacion): el xlsb repite la fila una vez por
                # molecula, asi que un combo aparece N veces con el MISMO valor. Medido:
                # 372 de 6.525 pares duplicados (5,7%), el 100% con valores identicos.
                # Sumar sin deduplicar infla la referencia y hace que el mercado real
                # parezca corto: era la causa de los 10 mercados que "no cerraban"
                # (Artro Red -49%, Tetralgin -49%, Amlodipina -45%), todos con 86%, 87%
                # y 73% de sus pares duplicados. Es la misma firma que el doble conteo
                # por droga del DDD.
                if (mkt, mn) in vistos_par:
                    n_dup += 1
                    continue
                vistos_par.add((mkt, mn))
                u12 = 0.0
                for i, mk in pcols:
                    x = v[i]
                    if isinstance(x, (int, float)):
                        mkt_mes[mkt][mk] += x
                        u12 += x
                if mn not in pres_meta:
                    pres_meta[mn] = {'marca': norm(v[iMarcaCol] or ''), 'u12': u12,
                                     'texto': str(v[iMarca] or '')}
    print('  xlsb: {} mercados, {:,} presentaciones, {:,} filas MINV (nacional) + {:,} CUP '
          '(region), {:,} filas MINV duplicadas por droga descartadas'
          .format(len(del_mkt), len(de_pack), n_minv, n_cup, n_dup))
    return de_pack, del_mkt, mkt_mes, pres_meta


def leer_master(path):
    """-> (pack_norm -> {prod, manuf, monthly{mes: u}}, meses)

    Se guarda POR PACK, no por producto: el mercado del Ateneo elige presentaciones
    sueltas, asi que agregar por marca antes de filtrar sobre-cuenta. La suma a nivel
    marca se hace despues, y solo sobre los packs que estan en el mercado.
    """
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sh = wb[wb.sheetnames[0]]
    it = sh.iter_rows(values_only=True)
    hdr = [str(h or '').replace('\n', ' ').strip() for h in next(it)]
    iPack, iProd, iMan = hdr.index('Pack'), hdr.index('Product'), hdr.index('Manufacturer')
    mes_re = re.compile(r'^Units (Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec) (\d{4})$')
    mcols = [(i, h[len('Units '):]) for i, h in enumerate(hdr) if mes_re.match(h)]
    packs, colisiones = {}, 0
    for row in it:
        if row[iPack] is None or row[iProd] is None:
            continue
        pk = norm(row[iPack])
        d = packs.get(pk)
        if d is None:
            d = packs[pk] = {'prod': str(row[iProd]).strip(),
                             'manuf': str(row[iMan] or '').strip(),
                             'monthly': defaultdict(float)}
        elif d['prod'] != str(row[iProd]).strip():
            colisiones += 1
        for i, mk in mcols:
            v = row[i]
            if isinstance(v, (int, float)):
                d['monthly'][mk] += v
    meses = sorted((mk for _, mk in mcols), key=msort)
    print('  AR_PM: {:,} packs, {} meses ({} .. {}){}'
          .format(len(packs), len(meses), meses[0], meses[-1],
                  '' if not colisiones else
                  '  OJO: {} pack(s) con mas de un Product'.format(colisiones)))
    return packs, meses


def cargar_data_js(rel):
    text = (REPO / rel).read_text(encoding='utf-8', errors='replace')
    m = re.search(r'window\.OTC_DASHBOARD\s*=\s*', text)
    if not m:
        return None
    s = text.index('{', m.end())
    D, end = json.JSONDecoder().raw_decode(text[s:])
    return {'text': text, 'ini': s, 'fin': s + end, 'D': D}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--master', default=str(DEFAULT_MASTER))
    ap.add_argument('--xlsb', default=str(DEFAULT_XLSB))
    ap.add_argument('--tol', type=float, default=0.05,
                    help='tolerancia relativa del cruce AR_PM vs xlsb por mercado (default 5%%)')
    ap.add_argument('--dry-run', action='store_true')
    ap.add_argument('--report')
    a = ap.parse_args()

    print('leyendo fuentes...')
    DE_PACK, DEL_MKT, MKT_MES, PRES = leer_xlsb(a.xlsb)
    PACKS, MMONTHS = leer_master(a.master)

    # exclusiones canonicas: se filtran por el NOMBRE DE PRODUCTO del pack
    n_excl = 0
    for pk in list(PACKS):
        if is_excluded(PACKS[pk]['prod']):
            del PACKS[pk]
            n_excl += 1
    if n_excl:
        print('  {} pack(s) de la lista de exclusiones salteados'.format(n_excl))
    # ventana comun a las dos fuentes: es la unica en la que el xlsb se puede comparar
    W12_REF = [m for m in XLSB_MESES.values() if m in MMONTHS]
    huerfanas = [pk for pk in DE_PACK if pk not in PACKS]
    print('  presentaciones del xlsb sin pack en AR_PM: {} de {} ({:.1f}%)'
          .format(len(huerfanas), len(DE_PACK), 100 * len(huerfanas) / max(1, len(DE_PACK))))

    # ── rescate de huerfanas ──────────────────────────────────────────────────
    # De las 76 huerfanas, 4 son de SIEGFRIED y valen 239.259 u -- y dos de ellas eran las
    # que hundian los dos unicos mercados que no cerraban, los DOS nombrados por la marca
    # que se estaba cayendo:
    #   Tetraciclinas (Acneclin)   ACNECLIN salia 2,38% (16.277 u) en vez de ~17%
    #   Monte Levo (Aireal Plus)   AIREAL PLUS salia 0,73% (2.993 u), ultimo del ranking
    # Dos causas distintas, las dos de nomenclatura:
    #   a) submarca en el prefijo: xlsb 'Acneclin Caps A.P 100Mg X 30'
    #                             AR_PM 'ACNECLIN AP CAPS A.P 100MG x 30'  (Product ACNECLIN AP)
    #   b) IQVIA corrigio el tamano del pack: xlsb '... 10Mg X 30 /5' -> AR_PM '... 10mg x 60 /5'
    #
    # La regla NO se queda en el parecido de nombres: EXIGE QUE LAS DOS FUENTES DEN EL
    # MISMO NUMERO en la ventana Mar2025..Feb2026 (<=1%). Eso convierte el match en algo
    # medido y no supuesto -- ACNECLIN AP cierra a -0,12% y AIREAL PLUS a +0,06%. Sin esa
    # verificacion el parecido de descriptor tambien casaba 'Qura Jbe 100Ml X 1' con
    # 'QURA PLUS JBE 100ML x 1', que son marcas DISTINTAS.
    def sin_cant(tail):
        """'TABL RECUBIE 10MG X 60 5' -> 'TABL RECUBIE 10MG X * 5' (el tamano del pack es
        justo el atributo que IQVIA re-declara)."""
        return re.sub(r'\bX \d+\b', 'X *', tail)

    por_tail = defaultdict(list)
    for pk, d in PACKS.items():
        b = norm(sin_lab(d['prod']))
        if pk.startswith(b):
            t = pk[len(b):].strip()
            if len(t) >= 8:
                por_tail[sin_cant(t)].append((pk, b))

    alias, rescatadas, rechazadas = {}, [], []
    for pk in huerfanas:
        meta = PRES.get(pk) or {}
        ma = meta.get('marca') or ''
        if not ma or not pk.startswith(ma):
            continue
        tail = sin_cant(pk[len(ma):].strip())
        if len(tail) < 8:
            continue
        u_ref = meta.get('u12') or 0.0
        cands = []
        for cpk, b in por_tail.get(tail, []):
            if not b.startswith(ma):
                continue
            u_ar = sum(PACKS[cpk]['monthly'].get(m, 0.0) for m in W12_REF)
            if u_ref > 1000 and abs(u_ar / u_ref - 1) <= 0.01:
                cands.append((cpk, u_ar))
            else:
                rechazadas.append((meta.get('texto', pk), PACKS[cpk]['prod'], u_ref, u_ar))
        # unico candidato Y el pack de AR_PM no lo reclama ya otra presentacion
        cands = [c for c in cands if c[0] not in DE_PACK]
        if len(cands) == 1:
            alias[pk] = cands[0][0]
            rescatadas.append((meta.get('texto', pk), PACKS[cands[0][0]]['prod'],
                               u_ref, cands[0][1]))
    if rescatadas:
        print('  RESCATADAS por descriptor + verificacion numerica: {}'.format(len(rescatadas)))
        for x_pres, x_prod, ur, ua in sorted(rescatadas, key=lambda x: -x[2]):
            print('    {:<44} -> {:<24} xlsb {:>9,.0f}  AR_PM {:>9,.0f}  {:+.2f}%'
                  .format(x_pres[:44], x_prod[:24], ur, ua, (ua / ur - 1) * 100 if ur else 0))
    if rechazadas:
        print('  descartadas por no dar el mismo numero: {} (p.ej. {})'.format(
            len(rechazadas), ' | '.join('{} vs {}'.format(x[0][:26], x[1][:22])
                                        for x in rechazadas[:3])))

    def packs_del_mercado(mkt):
        return [alias.get(pk, pk) for pk in DEL_MKT[mkt] if alias.get(pk, pk) in PACKS]

    # ── G1: cada mercado contra el propio xlsb, en su ventana Mar2025..Feb2026 ──
    W12 = W12_REF
    print()
    print('=' * 96)
    print('G1  cada mercado: suma AR_PM sobre sus marcas  vs  el total del propio xlsb '
          '(Mar 2025..Feb 2026)')
    print('=' * 96)
    cruce, ok_cruce = {}, 0
    for mkt, pks in DEL_MKT.items():
        ref = sum(MKT_MES[mkt].get(m, 0.0) for m in W12)
        got = sum(sum(PACKS[pk]['monthly'].get(m, 0.0) for m in W12)
                  for pk in packs_del_mercado(mkt))
        d = (got / ref - 1) if ref else None
        cruce[mkt] = {'ref_xlsb': round(ref), 'arpm': round(got),
                      'dif': None if d is None else round(d, 5),
                      'packs_xlsb': len(pks),
                      'packs_encontrados': len(packs_del_mercado(mkt))}
        if d is not None and abs(d) <= a.tol:
            ok_cruce += 1
    peor = max((v for v in cruce.values() if v['dif'] is not None),
               key=lambda v: abs(v['dif']), default=None)
    print('  {} de {} mercados cierran dentro de {:.0f}%.'
          .format(ok_cruce, len(cruce), a.tol * 100))
    malos = sorted((k for k, v in cruce.items() if v['dif'] is None or abs(v['dif']) > a.tol),
                   key=lambda k: -abs(cruce[k]['dif'] or 9))
    for k in malos[:12]:
        v = cruce[k]
        print('    {:<40} xlsb {:>12,}  AR_PM {:>12,}  {:>8}  packs {}/{}'.format(
            k[:40], v['ref_xlsb'], v['arpm'],
            'n/d' if v['dif'] is None else '{:+.1%}'.format(v['dif']),
            v['packs_encontrados'], v['packs_xlsb']))
    if len(malos) > 12:
        print('    ... y {} mas (ver --report)'.format(len(malos) - 12))
    print('  NO SE PUBLICAN los {} mercados que no cierran: un mercado corto de '
          'competidores infla el MS% propio.'.format(len(malos)))
    publicables = {k for k in cruce if k not in set(malos)}

    # ── armado por linea ──────────────────────────────────────────────────────
    print()
    print('{:<14} {:<22} {:<34} {:>5} {:>12} {:>7}'.format(
        'linea', 'producto propio', 'mercado del Ateneo', 'marc', 'MAT mercado', 'MS%'))
    print('-' * 100)
    n_fila_dup = 0
    report = {'tol': a.tol, 'cruce': cruce, 'lineas': [], 'sin_mercado': []}
    planned = []
    for rel in LINES:
        info = cargar_data_js(rel)
        if not info:
            continue
        D = info['D']
        mp = D.get('mol_perf') or {}
        if not mp:
            continue
        linea = rel.split('/')[0]
        bim = D.get('budIqviaMap') or {}

        meses = set()
        for f in mp.values():
            meses |= set((f.get('monthly') or {}).keys())
        meses = [mk for mk in sorted(meses, key=msort) if mk in MMONTHS]
        if not meses:
            continue
        ult = meses[-1]
        i_u = MMONTHS.index(ult)
        w12_ln = MMONTHS[max(0, i_u - 11):i_u + 1]

        mkts_out, filas, vistas = {}, [], set()
        for fam, f in mp.items():
            sies = [p for p in (f.get('products') or []) if p.get('is_sie')]
            if not sies:
                continue
            own = list(bim.get(fam) or [str(p.get('prod')) for p in sies])
            # los mercados de la familia = union de los mercados de los PACKS de sus
            # marcas propias (una marca propia puede tener packs en 2 mercados anidados)
            own_norm = {norm(sin_lab(p.get('prod'))) for p in sies}
            fam_mkts = set()
            for pk, mkts in DE_PACK.items():
                d = PACKS.get(pk)
                if d and norm(sin_lab(d['prod'])) in own_norm:
                    fam_mkts |= {m for m in mkts if m in publicables}
            if not fam_mkts:
                report['sin_mercado'].append({'linea': linea, 'familia': fam})
                continue
            usados = []
            for mkt in sorted(fam_mkts):
                if mkt not in mkts_out:
                    # rollup de packs -> marca, SOLO con los packs que estan en el mercado
                    agg = {}
                    for pk in packs_del_mercado(mkt):
                        d = PACKS[pk]
                        e = agg.setdefault(d['prod'], {'manuf': d['manuf'],
                                                       'monthly': defaultdict(float)})
                        for mk in meses:
                            e['monthly'][mk] += d['monthly'].get(mk, 0.0)
                    prods_out = []
                    for n in sorted(agg, key=lambda x: -sum(agg[x]['monthly'].get(mk, 0)
                                                            for mk in w12_ln)):
                        monthly = {mk: int(agg[n]['monthly'].get(mk, 0) or 0) for mk in meses}
                        if not any(monthly.values()):
                            continue          # no aporta a la ventana de esta linea
                        prods_out.append({'prod': n,
                                          'is_sie': es_sie(n, agg[n]['manuf']),
                                          'monthly_vals': monthly})
                    if not prods_out:
                        continue
                    mkts_out[mkt] = {'products': prods_out}
                if mkt not in mkts_out:
                    continue
                usados.append(mkt)
                ps = mkts_out[mkt]['products']
                mat = sum(sum(p['monthly_vals'].get(mk, 0) for mk in w12_ln) for p in ps)
                # UNA FILA POR PRODUCTO PROPIO, no una por familia.
                # Decision del usuario ("no son el mismo producto"): ACNECLIN (50mg tabl,
                # 16.277 u) y ACNECLIN AP (100mg caps A.P., 124.664 u) viven en la MISMA
                # familia del tablero (dermatologia/MINOCYCLINE) y esa familia no tiene
                # budIqviaMap, asi que una fila por familia los habria sumado en un 17,43%
                # que no describe a ninguno de los dos: ACNECLIN AP es el LIDER del mercado
                # con 15,42% y ACNECLIN esta 14no con 2,01%.
                # Es la misma logica que cardio ya aplica con DILATREND / DILATREND AP, solo
                # que alla son familias separadas y aca no.
                for prod_own in own:
                    mat_own = sum(sum(p['monthly_vals'].get(mk, 0) for mk in w12_ln)
                                  for p in ps if p['prod'] == prod_own)
                    if not mat_own:
                        continue
                    # DEDUP EXPLICITO por (producto, mercado). Un mismo producto propio
                    # puede figurar en dos familias del tablero -- 'MICOMAZOL (SIE)' esta
                    # en CLOTRIMAZOLE y en CICLOPIROX -- y generaria dos filas identicas.
                    # El toggle las colapsaria igual (misma clave de objeto), pero en
                    # silencio: `filas` decia 16 y la tabla dibujaba 15. Un descuadre que
                    # nadie mira es como se cuela el proximo bug, asi que se corta aca.
                    clave = (prod_own, mkt)
                    if clave in vistas:
                        n_fila_dup += 1
                        continue
                    vistas.add(clave)
                    filas.append({'label': sin_lab(prod_own), 'mercado': mkt,
                                  'propios': [prod_own]})
                    print('{:<14} {:<22} {:<34} {:>5} {:>12,} {:>6.2f}%'.format(
                        linea[:14], sin_lab(prod_own)[:22], mkt[:34], len(ps), mat,
                        mat_own / mat * 100 if mat else 0))

        if not mkts_out or not filas:
            continue
        D['mercadosAteneo'] = {'corte': ult, 'fuente': Path(a.xlsb).name,
                               'mercados': mkts_out, 'filas': filas}
        D.pop('mercadosATC', None)      # reemplaza a la vista ATC III que se rechazo
        D.pop('porFamilia', None)
        planned.append((rel, info))
        report['lineas'].append({'linea': linea, 'filas': len(filas),
                                 'mercados': len(mkts_out),
                                 'productos': sum(len(m['products']) for m in mkts_out.values())})

    print()
    if n_fila_dup:
        print('  {} fila(s) (producto, mercado) duplicadas descartadas: el mismo producto '
              'propio figura en mas de una familia del tablero'.format(n_fila_dup))
    for r in report['lineas']:
        print('  {:<14} {} filas -> {} mercados, {} productos'.format(
            r['linea'], r['filas'], r['mercados'], r['productos']))
    if report['sin_mercado']:
        print('  SIN mercado del Ateneo ({}): {}'.format(
            len(report['sin_mercado']),
            ', '.join('{}/{}'.format(x['linea'], x['familia'])
                      for x in report['sin_mercado'][:10])))
    if a.report:
        Path(a.report).write_text(json.dumps(report, ensure_ascii=False, indent=1), encoding='utf-8')
        print('reporte -> {}'.format(a.report))
    if a.dry_run:
        print('\nDRY RUN: no se escribio nada.')
        return 0
    print()
    for rel, info in planned:
        p = REPO / rel
        antes = p.stat().st_size
        p.write_text(info['text'][:info['ini']] + json.dumps(info['D'], ensure_ascii=False)
                     + info['text'][info['fin']:], encoding='utf-8', newline='')
        print('  {:<26} {:>10,} -> {:>10,} bytes'.format(rel, antes, p.stat().st_size))
    return 0


if __name__ == '__main__':
    sys.exit(main())
