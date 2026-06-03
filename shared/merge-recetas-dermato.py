# -*- coding: utf-8 -*-
"""Merge recetas (CloseUp) para DERMATOLOGIA (inline const D).

dermato no tenia merger de recetas. Estructura:
  rec_comp[fam][brand] = {monthly:{<MMM YYYY>:N}, quarterly:{...}, total:N}
  rec_ms[fam]          = {sie:{m:N}, ms:{m:%}, quarterly:{...}, ms_quarterly:{...}}
  recetas[fam][m]      = {recetas: <SIE>, medicos: <SIE medicos>}   (convencion dermato: SIE, no mercado)

El multi-period (tabla con IE) toma SIE de rec_ms[fam].sie y el MERCADO de la suma de
rec_comp[fam][*].monthly. Por eso este merge actualiza, por el/los mes(es) del pivot:
  - rec_comp[fam][brand].monthly[m]  (todas las marcas del mercado; agrega nuevas)
  - rec_ms[fam].sie[m]   = recetas de la marca SIE de la familia
  - rec_ms[fam].ms[m]    = sie / total_dedup_mercado * 100   (total = fila 'Totales' del pivot)
  - recetas[fam][m]      = {recetas: sie, medicos: sie_medicos}
NO toca quarterly/total (no los usa el multi-period) ni otras secciones/lineas.
Idempotente (overwrite por mes).

Mapeo explicito familia -> (substring de mercado en el pivot, marca SIE en el pivot).
La marca SIE de la familia es "<FAM> SIE" (ACNECLIN/ACNECLIN AP comparten el mercado
de tetraciclinas con SIE distinto; MICROSONA vs MICROSONA C y PALDAR vs PALDAR H se
desambiguan por el mercado).

Uso: py shared/merge-recetas-dermato.py --pivot "<pivot.xlsx>" [--dry-run]
"""
from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
HTML = REPO / 'dermatologia' / 'dermato_dashboard.html'
MES_EN = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}

# familia -> (substring del mercado en el pivot, marca SIE de la familia en el pivot)
DERMATO_FAM = {
    'ACNECLIN':    ('TETRACICLINAS (ACNECLIN)',              'ACNECLIN SIE'),
    'ACNECLIN AP': ('TETRACICLINAS (ACNECLIN)',              'ACNECLIN AP SIE'),
    'CLOBESOL':    ('CLOBETASOL (CLOBESOL)',                 'CLOBESOL SIE'),
    'MICOMAZOL':   ('ANTIM CREMA (MICOMAZOL)',               'MICOMAZOL SIE'),
    'MICROSONA':   ('HIDROCORTISO (MICROSONA)',              'MICROSONA SIE'),
    'MICROSONA C': ('TRIPLE CORT-ATB-ANTIM (MICROSONA C)',   'MICROSONA C SIE'),
    'MOMETAX':     ('MOMETAX TOTAL',                         'MOMETAX SIE'),
    'PALDAR':      ('ATB TOPICO (PALDAR)',                   'PALDAR SIE'),
    'PALDAR H':    ('ATB CORTIC TOP (PALDAR H',              'PALDAR H SIE'),
    'ROACCUTAN':   ('ISOTRETINOIN (ROACCUTAN)',              'ROACCUTAN SIE'),
}

MESN = {v: k for k, v in MES_EN.items()}


def norm(s):
    return re.sub(r'\s+', ' ', str(s or '').upper()).strip()


def norm_brand(s):
    """Para matchear marca pivot <-> rec_comp: upper, colapsa espacios. NO quita ' SIE'
    aca (lo probamos con y sin sufijo al matchear)."""
    return norm(s)


def parse_pivot(path):
    """Devuelve:
       market_brands[market][brand_norm][mk_en] = {'recetas':R,'medicos':M}
       market_total[market][mk_en] = recetas de la fila 'Totales' (dedup)
       months (lista ordenada)"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    row2 = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))
    col_map = {}
    cur = None
    for i, h1 in enumerate(row1):
        if isinstance(h1, datetime):
            cur = f'{MES_EN[h1.month]} {h1.year}'
        h2 = (str(row2[i]) if i < len(row2) and row2[i] else '').lower()
        if 'receta' in h2:
            col_map[i] = (cur, 'recetas')
        elif 'médico' in h2 or 'medico' in h2:
            col_map[i] = (cur, 'medicos')
    mb = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {'recetas': 0, 'medicos': 0})))
    mt = defaultdict(lambda: defaultdict(int))
    months = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        market = str(row[0]).strip()
        marca = row[2]
        is_tot = (str(marca or '').strip().lower() == 'totales')
        for ci, (mk, kind) in col_map.items():
            if ci >= len(row):
                continue
            try:
                v = int(row[ci]) if row[ci] is not None else 0
            except (TypeError, ValueError):
                v = 0
            months.add(mk)
            if is_tot:
                if kind == 'recetas':
                    mt[market][mk] = v
            elif marca:
                mb[market][norm_brand(marca)][mk][kind] = v
    wb.close()
    return mb, mt, sorted(months, key=lambda k: int(k.split()[1]) * 100 + MESN.get(k.split()[0], 0))


def find_market(mb, mt, target_sub):
    ts = norm(target_sub)
    # exacto primero, luego startswith / contains
    for m in list(mb.keys()) + list(mt.keys()):
        if norm(m) == ts:
            return m
    for m in list(mb.keys()) + list(mt.keys()):
        if norm(m).startswith(ts) or ts in norm(m):
            return m
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--pivot', required=True)
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    if not Path(args.pivot).is_file():
        print('ERROR: no existe el pivot:', args.pivot); return 1

    mb, mt, months = parse_pivot(args.pivot)
    print('Meses en pivot:', months)

    text = HTML.read_text(encoding='utf-8', errors='replace')
    m = re.search(r'const\s+D\s*=\s*', text)
    ob = text.index('{', m.end())
    D, end = json.JSONDecoder().raw_decode(text[ob:])
    rec_comp = D.setdefault('rec_comp', {})
    rec_ms = D.setdefault('rec_ms', {})
    recetas = D.setdefault('recetas', {})

    def last_existing(fam):
        sd = (rec_ms.get(fam, {}) or {}).get('sie', {}) or {}
        ks = [k for k in sd if len(str(k).split()) == 2]
        ks.sort(key=lambda k: int(k.split()[1]) * 100 + MESN.get(k.split()[0], 0))
        return (ks[-1], sd[ks[-1]]) if ks else (None, None)

    changed = 0
    for fam, (mk_sub, sie_brand) in DERMATO_FAM.items():
        if fam not in rec_ms:
            print(f'  [{fam}] NO esta en rec_ms -> skip'); continue
        market = find_market(mb, mt, mk_sub)
        if not market:
            print(f'  [{fam}] mercado NO encontrado en pivot ({mk_sub!r}) -> skip'); continue
        brands = mb.get(market, {})
        comp = rec_comp.setdefault(fam, {})
        # nombres rec_comp normalizados (con y sin ' SIE') -> key real
        rc_index = {}
        for k in comp:
            rc_index[norm(k)] = k
            rc_index[re.sub(r'\s+SIE$', '', norm(k))] = k
        sie_n = norm(sie_brand)
        sie_n_nos = re.sub(r'\s+SIE$', '', sie_n)
        lm, lv = last_existing(fam)
        for mk in months:
            sie_apr = (brands.get(sie_n, {}).get(mk) or {}).get('recetas', 0)
            sie_med = (brands.get(sie_n, {}).get(mk) or {}).get('medicos', 0)
            # Mercado = SUMA de las marcas YA trackeadas (set curado). Es el mismo
            # mercado que recompone el multi-period (suma de rec_comp[fam].monthly),
            # asi rec_ms.ms coincide. NO se agregan marcas nuevas del pivot (se
            # mantiene el set de competidores existente). El SIE esta dentro del set
            # -> sie <= mercado -> ms <= 100%.
            matched = skipped = 0
            market_total = 0
            for bn, monthly in brands.items():
                rec_val = (monthly.get(mk) or {}).get('recetas', 0)
                bn_nos = re.sub(r'\s+SIE$', '', bn)
                key = rc_index.get(bn) or rc_index.get(bn_nos)
                if key is None:
                    skipped += 1  # marca del pivot fuera del set curado -> no agregar
                    continue
                comp[key].setdefault('monthly', {})[mk] = rec_val
                market_total += rec_val
                matched += 1
            rms = rec_ms.setdefault(fam, {})
            rms.setdefault('sie', {})[mk] = sie_apr
            rms.setdefault('ms', {})[mk] = round(sie_apr / market_total * 100, 2) if market_total else 0
            recetas.setdefault(fam, {})[mk] = {'recetas': sie_apr, 'medicos': sie_med}
            changed += 1
            flag = ''
            if lv:
                ratio = sie_apr / lv if lv else 0
                if ratio > 3 or ratio < 0.33:
                    flag = '  <-- OJO magnitud vs %s=%s' % (lm, lv)
            print(f'  [{fam:12}] {mk}: SIE={sie_apr} (prev {lm}={lv}) mkt={market_total} ms={rms["ms"][mk]}% '
                  f'brands match={matched} skip={skipped}{flag}')

    if args.dry_run:
        print('\nDRY RUN: nada se escribio.')
        return 0
    HTML.write_text(text[:ob] + json.dumps(D, ensure_ascii=False) + text[ob + end:],
                    encoding='utf-8', newline='')
    print(f'\nEscrito dermato: {changed} (familia x mes) actualizados.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
