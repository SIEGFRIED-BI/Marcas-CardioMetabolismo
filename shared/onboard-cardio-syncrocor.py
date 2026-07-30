# -*- coding: utf-8 -*-
"""Onboarding de SYNCROCOR y SYNCROCOR D (nebivolol) a la linea Cardio.

POR QUE HACE FALTA UN SCRIPT (y no alcanza build-data.ps1):

1) IQVIA: la molecula del combo CONTIENE la del mono
   ('HYDROCHLOROTHIAZIDE_NEBIVOLOL'.Contains('NEBIVOLOL') == True) y el matcheo de
   $dashboardMarketConfig es por Contains -> el mercado de SYNCROCOR se tragaria el
   combo (mismo bug que ROXOLAN vs ROXOLAN PLUS, regla #2 de CLAUDE.md). Aca los
   mercados se definen por molecula EXACTA:
       SYNCROCOR    = NEBIVOLOL                       (ATC C07A0, 13 marcas)
       SYNCROCOR D  = HYDROCHLOROTHIAZIDE_NEBIVOLOL   (ATC C07B1,  7 marcas)

2) VENTA INTERNA: en SAP/Qlik las 5 presentaciones comparten Gran Familia = Familia =
   'SYNCROCOR' (la del combo tambien). merge-ventas-internas.py agrupa por Familia
   (col1) -> le daria las 5 a SYNCROCOR y dejaria SYNCROCOR D en 0. Aca se separan por
   Cod. Presentacion / Presentacion (mismo patron que fix-mujer-trip-venta.py y
   apply-otc-magnus-split.py).

3) STOCK: idem, la planilla 'Laboratorio - Familia - Producto' trae Familia='SYNCROCOR'
   para las 4 presentaciones; build-stock-from-laboratorio.py le asignaria el total
   (mono+combo) a SYNCROCOR. Aca se reparte por presentacion.

4) RECETAS: el mercado de CloseUp/Qlik es 'NEBIVOLOL (NEBILET)' -> build-data deriva la
   familia del PARENTESIS ('NEBILET') y no matchea ninguna familia del tablero, asi que
   la familia quedaba sin recetas en silencio (mismo caso que merge-recetas-respi-qlik.py).
   Ademas ese mercado mezcla las 2 drogas; aca se parte por Droga y el mercado de cada
   familia = suma de las marcas de SU droga (regla #3: el mercado es la suma de rec_comp).
   Verificado: mono + combo == fila 'Totales' del mercado, exacto en los 24 meses.

DATOS DISPONIBLES (relevados 2026-07-29; lo que no esta, no se inventa):
  IQVIA (master AR_PM)            Feb 2024 .. Jun 2026   -> mol_perf
  Venta interna (Qlik 'Rofina')   Jun 2026 (lanzamiento) -> budget[fam].real
  Estimado de ventas              NO HAY (SYNCROCOR no esta en 'Estimados VENTA
                                  vigentes MKT sidus'; cardio no tiene estimado) -> 0
  Recetas (CloseUp via Qlik)      Jun 2024 .. May 2026   -> recetas/rec_ms/rec_comp
  Stock (Laboratorio-Fam-Prod)    May 2026 .. Jun 2026   -> stock/stock_alerts/stock_pres
  Precios (Manual Farmaceutico)   PVP 01/03 y 07/04-2026 -> precios/prec_iqvia
  Convenios / Mostrador           NO HAY (las bases son 2024-2025; lanzo jun-2026)
  DDD regional                    NO HAY (el export por provincia no trae el mercado)

SEMANTICA que se respeta (verificada contra las familias que ya funcionan, no inventada):
  - mol_perf[fam].products = SIE propio forzado + relleno por MAT desc hasta 8 +
    'Otros (resto del mercado)' (is_resto) para que sum(products) == mercado.
  - NEBILET (SIE) / NEBILET D (SIE) quedan con is_sie=true dentro del mercado (son
    productos Siegfried). La atribucion por familia la hace budIqviaMap, que lista SOLO
    el producto propio (mismo patron que DILATREND AP, cuyo mercado incluye DILATREND
    (SIE) con is_sie=true). build-kpis desambigua por 'primary family'.
  - rec_ms[fam].sie = suma de TODAS las marcas SIE del mercado (asi lo hace el resto de
    la linea; la fila del render se rotula 'SIE', no la marca). Para SYNCROCOR eso
    incluye NEBILET SIE; el desglose por marca esta en rec_comp.
  - dias de stock = round(stock/ventas*30) (convencion del tablero, ver
    onboard-brexil-snc.py). La planilla trae su propio 'Dias' por SKU pero no es
    replicable al agregar presentaciones.

IDEMPOTENTE: re-correrlo da el mismo resultado. Cada bloque SKIPEA (sin fallar) si su
fuente no esta. Corre DESPUES de la venta y del stock en el cierre mensual
(ver shared/update-all.ps1) porque re-aplica los dos splits que esos pasos revierten.

Uso:
  py shared/onboard-cardio-syncrocor.py [--master <AR_PM.xlsx>] [--file <venta.xlsx>]
      [--cutoff YYYY-MM] [--recetas <RECETAS_qlik.xlsx>] [--stock <Laboratorio...xlsx>]
      [--precios <PRECIOS...xlsx>] [--dry-run]
"""
from __future__ import annotations
import argparse, importlib.util, json, re, sys
from collections import defaultdict
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
SHARED = REPO / 'shared'
DATA = REPO / 'cardio' / 'data.js'

sys.path.insert(0, str(SHARED))
try:
    import manifest as _mf
except Exception:
    _mf = None


def _load_py(name, path):
    spec = importlib.util.spec_from_file_location(name, path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


# recompute-mol-perf-aggregates: MISMA funcion que usa el cierre para ytd/mat/quarterly/ms_*
_recompute = _load_py('recompute_agg', SHARED / 'recompute-mol-perf-aggregates.py')

MES_EN = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
          'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}
NUM_EN = {v: k for k, v in MES_EN.items()}
MES_ES = {'ENE': 'Jan', 'FEB': 'Feb', 'MAR': 'Mar', 'ABR': 'Apr', 'MAY': 'May', 'JUN': 'Jun',
          'JUL': 'Jul', 'AGO': 'Aug', 'SEP': 'Sep', 'SET': 'Sep', 'OCT': 'Oct',
          'NOV': 'Nov', 'DIC': 'Dec'}
MES_ES_IDX = {'ene': 0, 'feb': 1, 'mar': 2, 'abr': 3, 'may': 4, 'jun': 5,
              'jul': 6, 'ago': 7, 'sep': 8, 'sept': 8, 'oct': 9, 'nov': 10, 'dic': 11}

FAM_MONO, FAM_COMBO = 'SYNCROCOR', 'SYNCROCOR D'
FAMS = (FAM_MONO, FAM_COMBO)
SIE_PROD = {FAM_MONO: 'SYNCROCOR (SIE)', FAM_COMBO: 'SYNCROCOR D (SIE)'}
IQVIA_MOL = {FAM_MONO: 'NEBIVOLOL', FAM_COMBO: 'HYDROCHLOROTHIAZIDE_NEBIVOLOL'}
IQVIA_ATC = {FAM_MONO: 'C07A', FAM_COMBO: 'C07B'}   # igual criterio que DILATREND / DILATREND D
COLORS = {FAM_MONO: '#be123c', FAM_COMBO: '#9f1239'}
# Orden en sieProds: alfabetico, entre SINTROM y TELPRES.
SIEPRODS_AFTER = 'SINTROM'

RESTO_LABEL = 'Otros (resto del mercado)'
TOP_N = 8

# --- venta interna (SAP/Qlik) ---
VENTA_FAMILIA = 'SYNCROCOR'          # Familia (col1) compartida por las 5 presentaciones
VENTA_COD_COMBO = {'3048406'}        # Cod. Presentacion del combo (sin ceros a la izquierda)
VENTA_RE_COMBO = re.compile(r'^SYNCROCOR\s+D\b')

# --- stock (Laboratorio - Familia - Producto) ---
STOCK_FAMILIA = 'SYNCROCOR'
STOCK_RE_COMBO = re.compile(r'^SYNCROCOR\s+D\b')

# --- recetas (CloseUp via Qlik) ---
REC_MARKET = 'NEBIVOLOL (NEBILET)'
REC_DROGA = {FAM_MONO: 'NEBIVOLOL', FAM_COMBO: 'HIDROCLOROTIAZIDA + NEBIVOLOL'}
REC_CONTROL = ('AMLODIPINA (TERLOC)', 'TERLOC')   # familia que YA funciona: valida el metodo

# --- precios (Manual Farmaceutico) ---
PREC_COL = {'prod': 2, 'pres': 3, 'lab': 6, 'prev': 12, 'curr': 13, 'var': 14}

is_excluded = _load_py('excluded', SHARED / 'excluded-products.py').is_excluded


# ───────────────────────────── helpers ─────────────────────────────

def msort(mk):
    p = str(mk).split()
    return int(p[1]) * 100 + MES_EN.get(p[0], 0) if len(p) == 2 and p[0] in MES_EN else 0


def qkey(mk):
    p = str(mk).split()
    if len(p) != 2 or p[0] not in MES_EN:
        return ''
    return f'Q{(MES_EN[p[0]] - 1) // 3 + 1} {p[1]}'


def qsort(qk):
    p = str(qk).split()
    return int(p[1]) * 10 + int(p[0][1]) if len(p) == 2 and p[0].startswith('Q') else 0


def to_int(v):
    if v is None or v == '' or v == '-':
        return 0
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return 0


def to_float(v):
    if v is None or v == '' or v == '-':
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace('%', '')
    try:
        return float(s)
    except ValueError:
        pass
    try:
        return float(s.replace('.', '').replace(',', '.'))
    except ValueError:
        return 0.0


def norm_ws(s):
    """Colapsa espacios (la planilla manual usa NBSP dentro de los nombres de SKU)."""
    return re.sub(r'\s+', ' ', str(s or '').replace('\xa0', ' ')).strip()


def norm_prod_key(v):
    """Puerto de Normalize-ProductKey de build-data.ps1."""
    t = norm_ws(v).upper()
    if not t:
        return ''
    t = re.sub(r'\([^)]*\)', '', t)
    t = re.sub(r'[%./,+-]', ' ', t)
    return re.sub(r'\s+', ' ', t).strip()


def prod_key_match(left, right):
    """Puerto de Test-NormalizedProductMatch (igual o uno prefijo del otro)."""
    if not left or not right:
        return False
    return left == right or left.startswith(right) or right.startswith(left)


def mat_of(monthly, window):
    """MAT = ultimos 12 meses de la ventana (criterio de orden de Convert-PerfBucket)."""
    return sum(monthly.get(mk, 0) or 0 for mk in window[-12:])


def load_data():
    text = DATA.read_text(encoding='utf-8-sig')
    m = re.search(r'window\.OTC_DASHBOARD\s*=\s*', text)
    if not m:
        raise ValueError('no encontre window.OTC_DASHBOARD en cardio/data.js')
    ob = text.index('{', m.end())
    D, end = json.JSONDecoder().raw_decode(text[ob:])
    return text, D, ob, ob + end


def newest(paths):
    paths = [p for p in paths if p and Path(p).is_file()]
    return max(paths, key=lambda p: Path(p).stat().st_mtime) if paths else None


def hub():
    if _mf:
        try:
            h = _mf.hub_root()
            if h.is_dir():
                return h
        except Exception:
            pass
    h = Path.home() / 'OneDrive - Portalcorp' / 'Documentos' / 'Hub-Marcas-Inputs'
    return h if h.is_dir() else None


def resolve(explicit, globs, source_name=None):
    """explicit > manifest source > glob recursivo en el hub (el mas reciente)."""
    if explicit:
        return Path(explicit)
    H = hub()
    cands = []
    if H:
        for g in globs:
            cands += list(H.glob(g))
    hit = newest(cands)
    if hit:
        return Path(hit)
    if source_name and _mf:
        try:
            p = _mf.resolve_source(source_name)
            if p:
                return Path(p)
        except Exception:
            pass
    return None


# ───────────────────────── 1. IQVIA (mol_perf) ─────────────────────────

def read_master(path, window):
    """-> (por_molecula, por_atc, manuf) con {clave: {producto: {mes: units}}}."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    ci, mcols = {}, []
    for i, h in enumerate(hdr):
        s = norm_ws(h)
        sl = s.lower()
        if sl == 'product':
            ci['prod'] = i
        elif sl == 'manufacturer':
            ci['mf'] = i
        elif sl == 'molecules long':
            ci['mol'] = i
        elif sl == 'atc iv':
            ci['atc'] = i
        m = re.match(r'^Units ([A-Z][a-z]{2}) (\d{4})$', s)   # mensual puro (no MAT/YTD/'to')
        if m and m.group(1) in MES_EN:
            mcols.append((i, f'{m.group(1)} {m.group(2)}'))
    for need in ('prod', 'mf', 'mol', 'atc'):
        if need not in ci:
            raise ValueError(f'falta la columna {need} en el master ({hdr[:9]})')
    have = {mk for _, mk in mcols}
    missing = [mk for mk in window if mk not in have]
    if missing:
        raise ValueError('el master no cubre la ventana de mol_perf; faltan '
                         f'{len(missing)} meses (p.ej. {missing[-3:]}). Pasa --master '
                         'con el AR_PM del cierre vigente.')
    widx = {mk: i for i, mk in mcols}

    by_mol = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    by_atc = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    manuf = {}
    wanted_mol = set(IQVIA_MOL.values())
    wanted_atc = tuple(IQVIA_ATC.values())
    for row in ws.iter_rows(min_row=2, values_only=True):
        prod = norm_ws(row[ci['prod']])
        if not prod:
            continue
        mol = norm_ws(row[ci['mol']]).upper()
        atc = norm_ws(row[ci['atc']]).upper()
        hit_mol = mol in wanted_mol
        hit_atc = atc.startswith(wanted_atc)
        if not hit_mol and not hit_atc:
            continue
        manuf.setdefault(prod, norm_ws(row[ci['mf']]))
        for mk in window:
            i = widx[mk]
            v = row[i] if i < len(row) else None
            if not isinstance(v, (int, float)):
                continue
            if hit_mol:
                by_mol[mol][prod][mk] += v
            if hit_atc:
                key = 'C07A' if atc.startswith('C07A') else 'C07B'
                by_atc[key][prod][mk] += v
    wb.close()
    return by_mol, by_atc, manuf


def select_products(prod_series, manuf, window, sie_first):
    """SIE propio forzado + relleno por MAT desc hasta TOP_N. Devuelve lista de nombres."""
    items = []
    for prod, series in prod_series.items():
        monthly = {mk: int(round(series.get(mk, 0) or 0)) for mk in window}
        if sum(monthly.values()) == 0:
            continue
        items.append((prod, monthly, mat_of(monthly, window)))
    items.sort(key=lambda x: -x[2])
    names = [n for n in sie_first if n in prod_series]
    for prod, _, _ in items:
        if len(names) >= TOP_N:
            break
        if prod not in names:
            names.append(prod)
    return names, {p: m for p, m, _ in items}


def build_family_products(prod_series, manuf, window, sie_first):
    names, monthlies = select_products(prod_series, manuf, window, sie_first)
    market = {mk: sum(int(round(prod_series[p].get(mk, 0) or 0)) for p in prod_series)
              for mk in window}
    out = []
    for prod in names:
        monthly = monthlies.get(prod) or {mk: int(round(prod_series[prod].get(mk, 0) or 0))
                                         for mk in window}
        mf = manuf.get(prod, '')
        is_sie = (prod in sie_first) or '(SIE)' in prod.upper() or 'SIEGFRIED' in mf.upper()
        out.append({'prod': prod, 'manuf': mf, 'is_sie': bool(is_sie),
                    'monthly_vals': monthly, 'quarterly_vals': {}, 'ytd': {}, 'mat': {},
                    'ms_monthly': {}, 'ms_quarterly': {}, 'ms_ytd': {}, 'ms_mat': {}})
    # RESTO: restaura sum(products) == mercado completo (invariante que asumen build-kpis,
    # build-families-perf y fix-brandkpis-*).
    resto = {mk: max(0, market[mk] - sum(p['monthly_vals'].get(mk, 0) for p in out))
             for mk in window}
    if sum(resto.values()) > 0:
        out.append({'prod': RESTO_LABEL, 'manuf': '', 'is_sie': False,
                    'monthly_vals': resto, 'is_resto': True,
                    'quarterly_vals': {}, 'ytd': {}, 'mat': {},
                    'ms_monthly': {}, 'ms_quarterly': {}, 'ms_ytd': {}, 'ms_mat': {}})
    return out, market, len(monthlies)


def step_iqvia(D, master, window, cierre_month, log):
    by_mol, by_atc, manuf = read_master(master, window)
    atc_allowed = {}
    for fam in FAMS:
        series = by_mol.get(IQVIA_MOL[fam])
        if not series:
            log(f'  (warn) {fam}: la molecula {IQVIA_MOL[fam]} no esta en el master; no toco')
            continue
        prods, market, n_all = build_family_products(series, manuf, window, [SIE_PROD[fam]])
        fam_obj = {'family': fam, 'products': prods, 'ytd': {}, 'mat': {},
                   'monthly': {}, 'quarterly': {}}
        _recompute.recompute_family(fam_obj, cierre_month=cierre_month)
        D.setdefault('mol_perf', {})[fam] = fam_obj
        last = window[-1]
        sie = next((p for p in prods if p['prod'] == SIE_PROD[fam]), None)
        sie_u = (sie or {}).get('monthly_vals', {}).get(last, 0)
        log(f'  + mol_perf[{fam}]: {len(prods)} filas de {n_all} productos | {last}: '
            f'mercado={market[last]:,} SIE={sie_u:,} '
            f'({sie_u / market[last] * 100:.2f}%)'.replace(',', '.'))
        # allowed-set para precios/prec_iqvia (mismo criterio que build-data:
        # productos EMITIDOS del bucket, normalizados -> MAT del mes de cierre)
        allowed_mol = {norm_prod_key(p['prod']): p['mat'].get(f'{NUM_EN[cierre_month]} '
                                                              f'{window[-1].split()[1]}', 0)
                       for p in prods if not p.get('is_resto')}
        a_series = by_atc.get(IQVIA_ATC[fam]) or {}
        a_names, a_monthly = select_products(a_series, manuf, window, [SIE_PROD[fam]])
        allowed_atc = {norm_prod_key(n): int(mat_of(a_monthly.get(n, {}), window))
                       for n in a_names}
        atc_allowed[fam] = {'molecule': allowed_mol, 'atc': allowed_atc}
    return atc_allowed


# ───────────────────────── 2. Venta interna (budget) ─────────────────────────

def step_venta(D, path, cutoff, log):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    col_ym = {}
    for i, h in enumerate(hdr):
        if i == 0 or not h:
            continue
        m = re.match(r'(\w+)[\s\-/](\d{4})', norm_ws(h))
        if not m:
            continue
        midx = MES_ES_IDX.get(m.group(1).lower().rstrip('.'))
        if midx is None:
            continue
        col_ym[i] = (int(m.group(2)), midx)
    if not col_ym:
        wb.close()
        log('  (skip) venta: no reconoci columnas de mes'); return
    acc = {FAM_MONO: defaultdict(int), FAM_COMBO: defaultdict(int)}
    n_sku = defaultdict(int)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        fam_src = norm_ws(row[1]) if len(row) > 1 else ''
        if fam_src.upper() != VENTA_FAMILIA:
            continue
        pres = norm_ws(row[3]) if len(row) > 3 else ''
        cod = norm_ws(row[4]).lstrip('0') if len(row) > 4 else ''
        is_combo = (cod in VENTA_COD_COMBO) or bool(VENTA_RE_COMBO.match(pres.upper()))
        fam = FAM_COMBO if is_combo else FAM_MONO
        n_sku[fam] += 1
        for ci, ym in col_ym.items():
            if ci >= len(row):
                continue
            v = row[ci]
            if v is None:
                continue
            try:
                acc[fam][ym] += int(round(float(v)))
            except (TypeError, ValueError):
                continue
    wb.close()
    if not any(acc.values()):
        log(f'  (skip) venta: la familia {VENTA_FAMILIA!r} no esta en {Path(path).name}')
        return
    years = sorted({y for d in acc.values() for (y, _) in d})
    budget = D.setdefault('budget', {})
    for fam in FAMS:
        entry = {}
        for y in years:
            real = [0] * 12
            for (yy, mi), v in acc[fam].items():
                if yy == y:
                    real[mi] = v
            if cutoff:
                cy, cm = cutoff
                for i in range(12):
                    if (y > cy) or (y == cy and i > cm):
                        real[i] = None
            entry[str(y)] = {'budget': [0] * 12, 'real': real}
        budget[fam] = entry
        shown = {f'{NUM_EN[i + 1]}-{y}': entry[str(y)]['real'][i]
                 for y in years for i in range(12) if entry[str(y)]['real'][i]}
        log(f'  + budget[{fam}].real ({n_sku[fam]} SKU): {shown or "sin venta en la ventana"} '
            f'| estimado = 0 (no hay estimado cargado para esta familia)')


# ───────────────────────── 3. Recetas ─────────────────────────

def parse_rec_month(h):
    from datetime import datetime
    if isinstance(h, datetime):
        return f'{NUM_EN[h.month]} {h.year}'
    m = re.match(r'^([A-Za-z]{3})[-/ ](\d{4})$', norm_ws(h))
    if not m:
        return None
    en = MES_ES.get(m.group(1).upper())
    return f'{en} {m.group(2)}' if en else None


def read_recetas(path):
    """-> (months, markets) con markets[mercado][droga] = {marca: {'rec':{},'med':{}}}
    y markets[mercado]['__tot__'] = {mes: {rec,med}} (fila Totales del mercado)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    r1, r2 = list(next(it)), list(next(it))
    col, months = {}, []
    for i, h in enumerate(r1):
        mk = parse_rec_month(h)
        if not mk:
            continue
        if mk not in months:
            months.append(mk)
        h2 = str(r2[i] if i < len(r2) else '').lower()
        metric = 'rec' if 'receta' in h2 else ('med' if 'dico' in h2 else None)
        if metric:
            col[i] = (mk, metric)
    months.sort(key=msort)
    markets = {}
    for row in it:
        if not row or len(row) < 3:
            continue
        merc = norm_ws(row[0])
        if not merc:
            continue
        droga, marca = norm_ws(row[1]), norm_ws(row[2])
        md = markets.setdefault(merc, {'__tot__': {}, 'drogas': {}})
        if droga == 'Totales' and not marca:
            for i, (mk, metric) in col.items():
                if i < len(row):
                    md['__tot__'].setdefault(mk, {})[metric] = to_int(row[i])
            continue
        if not marca or marca == 'Totales':
            continue
        bd = md['drogas'].setdefault(droga, {}).setdefault(marca, {'rec': {}, 'med': {}})
        for i, (mk, metric) in col.items():
            if i < len(row):
                bd[metric][mk] = bd[metric].get(mk, 0) + to_int(row[i])
    wb.close()
    return months, markets


def build_rec(brands, months, market_from_total=None):
    """(recetas, rec_ms, rec_comp) con la semantica de produccion. El mercado sale de
    market_from_total (fila Totales) si viene, si no de la suma de marcas (regla #3)."""
    brands = {n: b for n, b in brands.items() if not is_excluded(n)}
    recetas = {}
    for mk in months:
        if market_from_total is not None:
            t = market_from_total.get(mk) or {}
            recetas[mk] = {'recetas': t.get('rec', 0), 'medicos': t.get('med', 0)}
        else:
            recetas[mk] = {'recetas': sum(b['rec'].get(mk, 0) for b in brands.values()),
                           'medicos': sum(b['med'].get(mk, 0) for b in brands.values())}
    sie = {mk: sum(b['rec'].get(mk, 0) for n, b in brands.items() if 'SIE' in n.upper())
           for mk in months}
    ms = {mk: (round(sie[mk] / recetas[mk]['recetas'] * 100, 1)
               if recetas[mk]['recetas'] else 0) for mk in months}
    sq, mq = defaultdict(int), defaultdict(int)
    for mk in months:
        q = qkey(mk)
        if not q:
            continue
        sq[q] += sie[mk]; mq[q] += recetas[mk]['recetas']
    rec_ms = {'sie': sie, 'ms': ms,
              'quarterly': {q: sq[q] for q in sorted(sq, key=qsort)},
              'ms_quarterly': {q: (round(sq[q] / mq[q] * 100, 1) if mq[q] else 0)
                               for q in sorted(sq, key=qsort)}}
    rec_comp = {}
    for name, b in brands.items():
        monthly = {mk: b['rec'].get(mk, 0) for mk in months}
        bq = defaultdict(int)
        for mk in months:
            q = qkey(mk)
            if q:
                bq[q] += monthly[mk]
        rec_comp[name] = {'monthly': monthly,
                          'quarterly': {q: bq[q] for q in sorted(bq, key=qsort)},
                          'total': sum(monthly.values())}
    return recetas, rec_ms, rec_comp


def step_recetas(D, path, log):
    months, markets = read_recetas(path)
    if REC_MARKET not in markets:
        log(f'  (skip) recetas: el mercado {REC_MARKET!r} no esta en {Path(path).name}')
        return
    # control: re-derivar una familia que YA funciona y exigir igualdad
    cmkt, cfam = REC_CONTROL
    if cmkt in markets and cfam in (D.get('rec_ms') or {}):
        allb = {}
        for dr, br in markets[cmkt]['drogas'].items():
            for n, b in br.items():
                allb[n] = b
        c_rec, c_ms, _ = build_rec(allb, months, markets[cmkt]['__tot__'])
        bad = []
        for mk in months:
            if D['rec_ms'][cfam]['sie'].get(mk) != c_ms['sie'][mk]:
                bad.append(f'sie[{mk}]')
            if (D['recetas'][cfam].get(mk) or {}).get('recetas') != c_rec[mk]['recetas']:
                bad.append(f'recetas[{mk}]')
        if bad:
            log(f'  ABORTADO recetas: el control {cfam} no reproduce el dato actual '
                f'({len(bad)} difs, p.ej. {bad[:4]}). No escribo recetas.')
            return
        log(f'  control recetas OK: {cfam} re-derivado == data.js ({len(months)} meses)')
    else:
        log(f'  (warn) recetas: no pude correr el control con {cfam}')

    md = markets[REC_MARKET]
    tot_check = defaultdict(int)
    for fam in FAMS:
        brands = md['drogas'].get(REC_DROGA[fam])
        if not brands:
            log(f'  (warn) recetas: la droga {REC_DROGA[fam]!r} no esta en {REC_MARKET!r}')
            continue
        # mercado = suma de las marcas de la droga (la fila Totales del mercado mezcla
        # mono + combo). Verificado: mono + combo == Totales en todos los meses.
        recetas, rec_ms, rec_comp = build_rec(brands, months)
        for mk in months:
            tot_check[mk] += recetas[mk]['recetas']
        D.setdefault('recetas', {})[fam] = recetas
        D.setdefault('rec_ms', {})[fam] = rec_ms
        D.setdefault('rec_comp', {})[fam] = rec_comp
        last = months[-1]
        own = rec_comp.get(f'{fam} SIE', {}).get('monthly', {}).get(last, 0)
        sie_names = sorted(n for n in rec_comp if 'SIE' in n.upper())
        log(f'  + recetas[{fam}]: {last} mercado={recetas[last]["recetas"]} '
            f'SIE={rec_ms["sie"][last]} (propio {own}) MS%={rec_ms["ms"][last]} '
            f'| {len(rec_comp)} marcas, SIE: {sie_names}')
    diffs = [mk for mk in months
             if tot_check[mk] != (md['__tot__'].get(mk) or {}).get('rec', 0)]
    log(f'  recetas: mono+combo vs fila Totales del mercado -> '
        f'{"OK, identico en los %d meses" % len(months) if not diffs else "DIFIEREN en %s" % diffs[:4]}')


# ───────────────────────── 4. Stock / Cobertura ─────────────────────────

def classify(days):
    if days is None:
        return 'nd'
    if days <= 0:
        return 'quiebre'
    if days < 7:
        return 'critico'
    if days < 14:
        return 'bajo'
    if days < 20:
        return 'alerta'
    return 'ok'


WORST_ORDER = ('quiebre', 'critico', 'bajo', 'alerta', 'ok', 'nd')
ALERT_LEVELS = ('quiebre', 'critico', 'bajo', 'alerta')


def worst(statuses):
    for s in WORST_ORDER:
        if s in statuses:
            return s
    return 'nd'


def step_stock(D, path, log):
    import openpyxl, datetime as _dt
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    row0 = rows[0]
    groups, seen = [], set()
    for i, h in enumerate(row0):
        if isinstance(h, _dt.datetime):
            mes = f'{NUM_EN[h.month]} {h.year}'
            if mes in seen:
                continue
            seen.add(mes); groups.append((i, mes))
    months = [m for _, m in groups]
    if not months:
        log('  (skip) stock: no reconoci los meses de la planilla'); return

    pres = {FAM_MONO: {}, FAM_COMBO: {}}
    for r in rows[2:]:
        fam_src = norm_ws(r[1]).upper()
        prod = norm_ws(r[2])
        if fam_src != STOCK_FAMILIA or not prod or prod.lower() == 'totales':
            continue
        fam = FAM_COMBO if STOCK_RE_COMBO.match(prod.upper()) else FAM_MONO
        series = {}
        for idx, mes in groups:
            # '-' = mes sin dato (pre-lanzamiento) -> dias None, NO 0.
            # covClassify del render: null -> 'nd', 0 -> 'quiebre'. Poner 0 pintaria
            # 10 meses de quiebre inexistente en Cobertura.
            raw_d = r[idx + 3] if idx + 3 < len(r) else None
            has = isinstance(r[idx], (int, float)) or isinstance(raw_d, (int, float))
            series[mes] = {'stock': max(0, to_int(r[idx])),
                           'ventas': max(0, to_int(r[idx + 1])),
                           'facturacion': max(0, to_int(r[idx + 2])),
                           'dias': (to_int(raw_d) or None) if has else None}
        pres[fam][prod] = series
    if not any(pres.values()):
        log(f'  (skip) stock: la familia {STOCK_FAMILIA!r} no esta en {Path(path).name}')
        return

    months12 = months[-12:]
    for fam in FAMS:
        if not pres[fam]:
            log(f'  (warn) stock: sin presentaciones para {fam}'); continue
        fam_series = {}
        for mes in months:
            st = sum(s[mes]['stock'] for s in pres[fam].values())
            vt = sum(s[mes]['ventas'] for s in pres[fam].values())
            fc = sum(s[mes]['facturacion'] for s in pres[fam].values())
            # dias de la familia = stock/ventas*30 (convencion del tablero): la fila
            # 'Totales' de la planilla incluye el combo, no sirve para el mono.
            fam_series[mes] = {'stock': st, 'ventas': vt, 'facturacion': fc,
                               'dias': int(round(st / vt * 30)) if vt > 0 else None}
        D.setdefault('stock', {})[fam] = fam_series
        dias = [fam_series[m]['dias'] for m in months12]
        statuses = [classify(d) for d in dias]
        ai = [i for i, s in enumerate(statuses) if s in ALERT_LEVELS]
        D.setdefault('stock_alerts', {})[fam] = {
            'ventas': [fam_series[m]['ventas'] for m in months12],
            'dias': dias, 'statuses': statuses, 'alert_indices': ai,
            'worst_status': worst(statuses), 'n_alerts': len(ai), 'familia': fam}
        for prod, s in pres[fam].items():
            pd_ = [s[m]['dias'] for m in months12]          # dias por SKU: los de la planilla
            pst = [classify(d) for d in pd_]
            pai = [i for i, x in enumerate(pst) if x in ALERT_LEVELS]
            D.setdefault('stock_pres', {})[prod] = {
                'ventas': [s[m]['ventas'] for m in months12], 'dias': pd_,
                'statuses': pst, 'alert_indices': pai, 'worst_status': worst(pst),
                'n_alerts': len(pai), 'familia': fam}
        con = [m for m in months if fam_series[m]['stock'] or fam_series[m]['ventas']]
        log(f'  + stock[{fam}]: {len(pres[fam])} presentacion(es), con dato en '
            f'{sorted(con, key=msort)} | {months[-1]}: {fam_series[months[-1]]}')


# ───────────────────────── 5. Precios ─────────────────────────

def step_precios(D, path, allowed, log):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    rows = [r for r in it]
    wb.close()
    prev_label = norm_ws(hdr[PREC_COL['prev']])
    curr_label = norm_ws(hdr[PREC_COL['curr']])
    for fam in FAMS:
        al = (allowed or {}).get(fam)
        if not al:
            log(f'  (skip) precios: sin allowed-set IQVIA para {fam}'); continue
        prices = {'molecule': {}, 'atc': {}}
        piq = {'molecule': {}, 'atc': {}}
        for r in rows:
            name = norm_ws(r[PREC_COL['prod']])
            pres = norm_ws(r[PREC_COL['pres']])
            if not name or not pres:
                continue
            nk = norm_prod_key(name)
            lab = norm_ws(r[PREC_COL['lab']])
            for mode in ('molecule', 'atc'):
                matched = next((c for c in al[mode] if prod_key_match(c, nk)), None)
                if matched is None:
                    continue
                prices[mode].setdefault(pres, []).append({
                    'lab': lab, 'prod': name,
                    'is_sie': 'SIEGFRIED' in lab.upper(),
                    'pvp_dic25': round(to_float(r[PREC_COL['prev']]), 2),
                    'pvp_feb26': round(to_float(r[PREC_COL['curr']]), 2),
                    'var': to_float(r[PREC_COL['var']]) / 100 if to_float(r[PREC_COL['var']]) else 0.0,
                })
                piq[mode].setdefault(name.upper(), al[mode][matched])
        if not prices['molecule'] and not prices['atc']:
            log(f'  (skip) precios: 0 filas para {fam}'); continue
        D.setdefault('precios', {})[fam] = prices
        D.setdefault('prec_iqvia', {})[fam] = piq
        own = [e for lst in prices['molecule'].values() for e in lst
               if norm_prod_key(e['prod']) == norm_prod_key(fam)]
        log(f'  + precios[{fam}]: molecule={len(prices["molecule"])} pres, '
            f'atc={len(prices["atc"])} pres | propias: '
            f'{[(e["prod"], e["pvp_feb26"]) for e in own]}')
    log(f'  precios: prev={prev_label!r} curr={curr_label!r} '
        f'(meta.price_*_label ya vigente en el tablero)')


# ───────────────────────── 6. maps + brandKpis ─────────────────────────

def step_maps(D, log):
    sp = D.setdefault('sieProds', [])
    for fam in FAMS:
        if fam not in sp:
            try:
                sp.insert(sp.index(SIEPRODS_AFTER) + 1 + FAMS.index(fam), fam)
            except ValueError:
                sp.append(fam)
    for fam in FAMS:
        D.setdefault('colors', {}).setdefault(fam, COLORS[fam])
        D.setdefault('molLabels', {}).setdefault(fam, fam)
        D.setdefault('sieMolMap', {})[fam] = fam
        D.setdefault('budIqviaMap', {})[fam] = [SIE_PROD[fam]]
        # canal/conv apuntan a la propia familia: no hay convenios ni mostrador para un
        # producto que lanzo en jun-2026 (las bases son 2024-2025). El render skipea la
        # seccion cuando la key no existe (igual que ROXOLAN PLUS).
        D.setdefault('prodMap', {})[fam] = {'mol': fam, 'canal': fam, 'conv': fam,
                                            'rec': fam, 'prec': fam, 'bud': fam}
    log(f'  + maps: sieProds={[f for f in sp if f in FAMS]} '
        f'colors/molLabels/sieMolMap/prodMap/budIqviaMap OK')


def _win(end_y, end_m, kind):
    if kind == 'ytd':
        return [f'{NUM_EN[i]} {end_y}' for i in range(1, end_m + 1)]
    out = []
    for back in range(11, -1, -1):
        idx = (end_y * 12 + (end_m - 1)) - back
        yy, mm = divmod(idx, 12)
        out.append(f'{NUM_EN[mm + 1]} {yy}')
    return out


def step_brandkpis(D, cierre_y, cierre_m, venta_cut, log):
    mol = D.get('mol_perf') or {}
    for fam in FAMS:
        fo = mol.get(fam)
        if not fo:
            continue
        own = set((D.get('budIqviaMap') or {}).get(fam) or [])
        prods = fo.get('products', [])
        entry = {}
        for per in ('ytd', 'mat'):
            wc, wp = _win(cierre_y, cierre_m, per), _win(cierre_y - 1, cierre_m, per)

            def s(pl, w):
                return sum(sum((p.get('monthly_vals') or {}).get(mk, 0) or 0 for mk in w)
                           for p in pl)
            sie = [p for p in prods if p.get('is_sie') and p['prod'] in own]
            u_c, u_p = s(sie, wc), s(sie, wp)
            m_c, m_p = s(prods, wc), s(prods, wp)
            entry[per] = {
                'ie': round((u_c / u_p) / (m_c / m_p) * 100, 1) if (u_p and m_p and m_c) else None,
                'ms': round(u_c / m_c * 100, 1) if m_c else None,
                'units': u_c, 'units_prev': u_p, 'market_total': m_c,
                'growth': round((u_c / u_p - 1) * 100, 1) if u_p else None,
            }
        real = None
        b = ((D.get('budget') or {}).get(fam) or {}).get(str(venta_cut[0]))
        if b and isinstance(b.get('real'), list):
            real = b['real'][venta_cut[1]]
        entry['budget'] = {'pct': None, 'real': real, 'target': 0}
        rms = (D.get('rec_ms') or {}).get(fam) or {}
        last_rec = sorted((rms.get('ms') or {}), key=msort)
        entry['rec'] = ({'ms': rms['ms'][last_rec[-1]], 'label': last_rec[-1]}
                        if last_rec else {'ms': None, 'label': None})
        D.setdefault('brandKpis', {})[fam] = entry
        log(f'  + brandKpis[{fam}]: YTD u={entry["ytd"]["units"]} mkt={entry["ytd"]["market_total"]} '
            f'MS%={entry["ytd"]["ms"]} IE={entry["ytd"]["ie"]} | MAT MS%={entry["mat"]["ms"]} '
            f'IE={entry["mat"]["ie"]} | venta={real} | rec.ms={entry["rec"]["ms"]}')


# ───────────────────────── main ─────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--master', help='AR_PM*.xlsx (IQVIA Premium). Default: el mas reciente del hub')
    ap.add_argument('--file', help='Planilla de Ventas (manual o extracto Qlik)')
    ap.add_argument('--cutoff', help="Ultimo mes CERRADO de venta 'YYYY-MM'. Default: manifest.ventaCutoff")
    ap.add_argument('--recetas', help='RECETAS_qlik_*.xlsx')
    ap.add_argument('--stock', help='Laboratorio - Familia - Producto*.xlsx')
    ap.add_argument('--precios', help='PRECIOS HASTA*.xlsx')
    ap.add_argument('--skip', default='', help='bloques a saltear: iqvia,venta,recetas,stock,precios')
    ap.add_argument('--dry-run', action='store_true')
    a = ap.parse_args()
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    skip = {s.strip() for s in a.skip.split(',') if s.strip()}
    out = []

    def log(m):
        print(m); out.append(m)

    text, D, ob, end = load_data()

    # Ventana de meses = la de una familia ya construida (no la inventamos).
    ref = next((f for f in ('TERLOC', 'DAURAN') if f in (D.get('mol_perf') or {})), None)
    if not ref:
        print('ERROR: no hay familias en mol_perf de cardio para tomar la ventana'); return 2
    ref_sie = next((p for p in D['mol_perf'][ref]['products'] if p.get('is_sie')), None)
    window = sorted((ref_sie or {}).get('monthly_vals', {}).keys(), key=msort)
    if not window:
        print('ERROR: no pude determinar la ventana de meses de mol_perf'); return 2
    cierre_lbl = window[-1]
    cierre_m, cierre_y = MES_EN[cierre_lbl.split()[0]], int(cierre_lbl.split()[1])

    cut_str = a.cutoff or (_mf and (_mf.load().get('global', {}).get('ventaCutoff'))) or None
    cutoff = None
    if cut_str:
        m = re.match(r'(\d{4})-(\d{2})', str(cut_str))
        if m:
            cutoff = (int(m.group(1)), int(m.group(2)) - 1)
    if cutoff is None:
        cutoff = (cierre_y, cierre_m - 1)

    print(f'cardio/data.js  ventana mol_perf: {window[0]}..{cierre_lbl} ({len(window)} meses)  '
          f'cierre={cierre_y}-{cierre_m:02d}  ventaCutoff={cutoff[0]}-{cutoff[1] + 1:02d}')

    allowed = None
    if 'iqvia' not in skip:
        master = resolve(a.master, ['_iqvia-master/*/AR_PM*.xlsx', 'AR_PM*.xlsx'], 'iqvia_master')
        if not master or not master.is_file():
            log('  (skip) IQVIA: no encontre el master AR_PM*.xlsx')
        else:
            print(f'\n[IQVIA] {master.name}')
            allowed = step_iqvia(D, master, window, cierre_m, log)

    if 'venta' not in skip:
        venta = resolve(a.file, ['Planilla de Ventas*.xlsx'], 'venta_interna')
        if not venta or not venta.is_file():
            log('  (skip) venta: no encontre la Planilla de Ventas')
        else:
            print(f'\n[VENTA] {venta.name}')
            step_venta(D, venta, cutoff, log)

    if 'recetas' not in skip:
        rec = resolve(a.recetas, ['cardio/*/fuentes-originales/RECETAS_qlik_*.xlsx',
                                  'cardio/*/RECETAS_qlik_*.xlsx', 'RECETAS_qlik_*.xlsx'])
        if not rec or not rec.is_file():
            log('  (skip) recetas: no encontre RECETAS_qlik_*.xlsx de cardio')
        else:
            print(f'\n[RECETAS] {rec.name}')
            step_recetas(D, rec, log)

    if 'stock' not in skip:
        stk = resolve(a.stock, ['Laboratorio - Familia - Producto*.xlsx'])
        if not stk or not stk.is_file():
            log('  (skip) stock: no encontre Laboratorio - Familia - Producto*.xlsx')
        else:
            print(f'\n[STOCK] {stk.name}')
            step_stock(D, stk, log)

    if 'precios' not in skip:
        prc = resolve(a.precios, ['cardio/*/fuentes-originales/PRECIOS HASTA*.xlsx',
                                  'PRECIOS HASTA*.xlsx'])
        if not prc or not prc.is_file():
            log('  (skip) precios: no encontre PRECIOS HASTA*.xlsx')
        else:
            print(f'\n[PRECIOS] {prc.name}')
            step_precios(D, prc, allowed, log)

    print('\n[MAPS]')
    step_maps(D, log)
    print('\n[BRANDKPIS]')
    step_brandkpis(D, cierre_y, cierre_m, cutoff, log)

    if a.dry_run:
        print('\nDRY-RUN: nada escrito.')
        return 0
    DATA.write_text(text[:ob] + json.dumps(D, ensure_ascii=False) + text[end:],
                    encoding='utf-8', newline='')
    print(f'\nEscrito {DATA.relative_to(REPO)}.')
    print('Cascada: recompute-mol-perf-aggregates --cierre <AAAA-MM> + build-kpis + '
          'build-families-perf + sync-kpistrip + fix-brandkpis-* + build-total + '
          'finalize-labels + bump-cache-busters')
    return 0


if __name__ == '__main__':
    sys.exit(main())
