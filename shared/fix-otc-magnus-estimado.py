# -*- coding: utf-8 -*-
"""Carga el ESTIMADO de MAGNUS y MAGNUS 36 (OTC) desde la planilla por-SKU
'Estimados VENTA vigentes MKT sidus.xlsx' (en hubRoot), separando por la marca
'36' (tadalafil) vs base (sildenafil).

Por qué: el estimado del tablero sale de un panel agrupado por FAMILIA, donde
MAGNUS es una sola familia combinada -> MAGNUS 36 quedaba con budget = 0 (su
venta sí se separa via apply-otc-magnus-split, pero el estimado no). Esta fuente
'MKT sidus' SÍ trae el estimado por presentación:
  - MAGNUS 36  = SKUs cuyo Producto contiene 'MAGNUS 36'  (tadalafil)
  - MAGNUS     = resto de SKUs 'MAGNUS' (sildenafil)

Escribe D.budget['MAGNUS']['2026'].budget y D.budget['MAGNUS 36']['2026'].budget
(solo el 'budget'/estimado; la venta 'real' no se toca). Idempotente, --check.
Si falta el xlsx o openpyxl, NO falla (skip) — así no rompe update-all sin el hub.
"""
from __future__ import annotations
import re, json, sys, datetime
from pathlib import Path

SHARED = Path(__file__).resolve().parent
REPO = SHARED.parent
sys.path.insert(0, str(SHARED))

YEAR = 2026
SHEET = 'Estimados 2026'
SRC_GLOB = 'Estimados*MKT sidus*.xlsx'
PROD_COL = 3  # 'Producto'


def find_source():
    try:
        import manifest
        hub = manifest.hub_root()
    except Exception:
        hub = None
    cands = []
    if hub and hub.is_dir():
        cands += sorted(hub.glob(SRC_GLOB))
    if not cands:  # fallback: junto al repo / OneDrive comun
        for base in (REPO.parent, Path.home() / 'OneDrive - Portalcorp' / 'Documentos' / 'Hub-Marcas-Inputs'):
            if base.is_dir():
                cands += sorted(base.glob(SRC_GLOB))
    return cands[0] if cands else None


def read_magnus_estimado(xlsx):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = rows[0]
    mcols = [i for i, h in enumerate(hdr) if isinstance(h, datetime.datetime) and h.year == YEAR]
    mcols = mcols[:12]
    mag36 = [0.0] * 12
    magbase = [0.0] * 12
    for r in rows[1:]:
        prod = str(r[PROD_COL] or '').upper()
        if 'MAGNUS' not in prod:
            continue
        is36 = 'MAGNUS 36' in prod
        for j, ci in enumerate(mcols):
            v = r[ci] or 0
            try:
                v = float(v)
            except Exception:
                v = 0
            (mag36 if is36 else magbase)[j] += v
    return [round(x) for x in magbase], [round(x) for x in mag36]


def patch(check_only=False):
    xlsx = find_source()
    if xlsx is None:
        print('  (skip) no se encontro la planilla MKT sidus en hubRoot')
        return 0
    try:
        magbase, mag36 = read_magnus_estimado(xlsx)
    except ImportError:
        print('  (skip) openpyxl no disponible')
        return 0
    if not any(mag36) and not any(magbase):
        print('  (skip) sin filas MAGNUS en la planilla')
        return 0
    p = REPO / 'OTC' / 'data.js'
    t = p.read_text(encoding='utf-8-sig')
    m = re.search(r'window\.OTC_DASHBOARD\s*=\s*', t)
    ob = t.index('{', m.end())
    D, end = json.JSONDecoder().raw_decode(t[ob:])
    bud = D.get('budget', {})
    changed = 0
    for key, arr in (('MAGNUS', magbase), ('MAGNUS 36', mag36)):
        yo = (bud.setdefault(key, {})).setdefault(str(YEAR), {})
        if yo.get('budget') != arr:
            changed += 1
            if not check_only:
                yo['budget'] = arr
    if changed and not check_only:
        p.write_text(t[:ob] + json.dumps(D, ensure_ascii=False) + t[ob + end:],
                     encoding='utf-8', newline='')
    print(f'  fuente: {xlsx.name}')
    print(f'  MAGNUS estim/mes={magbase[0]}  MAGNUS 36 estim/mes={mag36[0]}  '
          f'({"a corregir" if check_only else "aplicado"}: {changed} marca(s))')
    return changed


def main():
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    check_only = '--check' in sys.argv
    n = patch(check_only)
    if check_only and n > 0:
        print('OTC-MAGNUS-ESTIMADO: budget de MAGNUS/MAGNUS 36 desalineado vs MKT sidus. '
              'Correr: py shared/fix-otc-magnus-estimado.py')
        return 1
    return 0


if __name__ == '__main__':
    sys.exit(main())
