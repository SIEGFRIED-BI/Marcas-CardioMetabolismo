#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
shared/qlik/restringir-mes-a-historico.py

Restringe el mes recien anexado, en un mercado puntual, a los productos que YA estaban en
su historia. Es para cuando la definicion de un mercado CAMBIA en la fuente entre un cierre
y el siguiente: el mes nuevo mide otro universo y el salto se lee como crecimiento.

CASO QUE LO MOTIVO (medido)
---------------------------
mujer / 'Trip +45', al anexar Jun-2026:
    May-2026: 129 productos, 1.717.150 u
    Jun-2026: 314 productos (188 nuevos), 2.355.984 u   -> +37%
Los otros 12 mercados de mujer tienen 0 a 7 productos nuevos; este solo tiene 188. Y en el
artefacto aparecian 159 marcas nuevas CON DATO UNICAMENTE EN JUNIO (PRONTAL, DERRUMAL,
AUDAX...), que no son de la linea. Entre la extraccion de julio y hoy la app se republico
(2026-08-05 16:46 UTC) y la definicion del mercado se amplio.

El +37% NO es crecimiento: es cambio de universo. Publicarlo asi rompe la comparabilidad de
la serie, que es justamente lo que el mes nuevo tiene que preservar.

QUE HACE
--------
Deja en el mes nuevo solo las filas cuyo producto ya aparecia en ese mercado en algun mes
anterior. NO toca ningun otro mercado ni ningun otro mes. Reporta los productos excluidos
para que se pueda decidir aparte si el mercado se amplia de verdad -- eso implicaria
re-expresar la historia, que es otra decision.

Uso:
  py shared/qlik/restringir-mes-a-historico.py --xlsx <archivo> --mes "Jun-2026"
     --mercado "Trip +45" [--dry-run]
"""
from __future__ import annotations
import argparse, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

C_MKT, C_MES, C_PROD, C_UNI = 1, 4, 7, 8
HEADER = ['RegionCUP', 'Mercado', 'Droga', 'Clase Terapeutica', 'AñoMes',
          'Codigo Clase Terapeutica', 'Codigo Producto', 'Producto', 'Unidades']


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--mes', required=True)
    ap.add_argument('--mercado', required=True, action='append',
                    help='mercado a restringir (se puede repetir)')
    ap.add_argument('--dry-run', action='store_true')
    a = ap.parse_args()
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')

    src = Path(a.xlsx)
    objetivo = set(a.mercado)
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # pasada 1: productos historicos por mercado (todos los meses menos el nuevo)
    hist = defaultdict(set)
    primera = True
    for row in ws.iter_rows(values_only=True):
        if primera:
            primera = False; continue
        if row is None or row[C_MKT] is None:
            continue
        mkt = str(row[C_MKT]).strip()
        if mkt not in objetivo:
            continue
        if str(row[C_MES]).strip() != a.mes:
            hist[mkt].add(str(row[C_PROD]).strip())
    wb.close()
    for m in objetivo:
        print('  {}: {} productos en la historia'.format(m, len(hist[m])))

    # pasada 2: reescribir salteando las filas del mes nuevo con productos no historicos
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    dst = src.with_name(src.stem + ' restringido.xlsx')
    wb_out = openpyxl.Workbook(write_only=True) if not a.dry_run else None
    ws_out = wb_out.create_sheet('Sheet1') if wb_out else None
    if ws_out:
        ws_out.append(HEADER)
    fuera, u_fuera, prods_fuera = 0, 0.0, defaultdict(float)
    n = 0
    primera = True
    for row in ws.iter_rows(values_only=True):
        if primera:
            primera = False; continue
        if row is None or row[C_MKT] is None:
            continue
        mkt = str(row[C_MKT]).strip()
        if mkt in objetivo and str(row[C_MES]).strip() == a.mes:
            p = str(row[C_PROD]).strip()
            if p not in hist[mkt]:
                u = row[C_UNI] if isinstance(row[C_UNI], (int, float)) else 0
                fuera += 1; u_fuera += u; prods_fuera[p] += u
                continue
        n += 1
        if ws_out:
            ws_out.append(list(row[:9]))
    wb.close()

    print()
    print('  filas del mes excluidas: {:,}  ({:,.0f} u, {} productos distintos)'.format(
        fuera, u_fuera, len(prods_fuera)))
    for p, u in sorted(prods_fuera.items(), key=lambda x: -x[1])[:8]:
        print('     {:<40} {:>10,.0f} u'.format(p[:40], u))
    if a.dry_run:
        print('\nDRY RUN: no se escribio nada.')
        return 0
    wb_out.save(dst)
    wb_out.close()
    print()
    print('  -> {} ({:,} filas, {:.0f} MB)'.format(dst.name, n, dst.stat().st_size / 1048576))
    return 0


if __name__ == '__main__':
    sys.exit(main())
