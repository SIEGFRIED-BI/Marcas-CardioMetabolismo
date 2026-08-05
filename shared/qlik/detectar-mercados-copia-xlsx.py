#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
shared/qlik/detectar-mercados-copia-xlsx.py

Detecta, dentro del xlsx regional de cada linea, los mercados que son COPIA de un slice de
otro mercado del mismo archivo. Es el bug del doble conteo del DDD.

EL BUG
------
Los mercados de la columna Mercado son una jerarquia deliberada: un contenedor y sus
sub-segmentos (Macromax ⊃ Macromax pediátr.; Antipsicóticos ⊃ Quetiapinas;
Micomazol Total ⊃ Micomazol Crema). El dato esta bien. El problema es que
build-competidores-shape-a.py re-indexa los mercados por molecula/ATC con
`units[brand][region][mes] += u` y la columna Mercado NO forma parte de la clave: si el
archivo trae el contenedor Y el contenido, esas unidades se suman DOS VECES.
Medido en el tablero de ATB, May-2026: el mercado de azitromicina publicaba 259.658 u
contra 213.664 u reales (+21,5%), y todo MS% contra ese denominador quedaba subestimado.
build-mujer-competidores-data.py NO tiene el bug: indexa por mercado, asi que cada mercado
del Ateneo es su propio bucket. Mujer no se toca (ver --sin-dedup en append-ddd-mes.py).

POR QUE SE RESUELVE CON EL ARCHIVO Y NO CON LA API
--------------------------------------------------
Un intento previo comparo mercados pidiendole a Qlik el total de cada uno y la union por
pares (`sum({<Mercado={A,B}>})` == total(A) implica B ⊆ A). No sirvio: la app devuelve
agregados A MEDIO CALCULAR y el mismo test dio resultados distintos en dos corridas
(mujer: 22 candidatos y despues 0), con contradicciones flagrantes -- marcaba
"Hexaler ⊆ Alergical" cuando Hexaler tiene 18.036.403 u y Alergical 9.456.638 u en el
propio archivo, y marcaba "Visdon ⊆ Madopar" cuando ninguno de los 5 productos de Visdon
esta en Madopar. Aplicarlo habria borrado 27,9M de unidades reales.
El archivo, en cambio, es el dato exacto que consume el builder, y el test se puede hacer
A NIVEL FILA: para cada fila de B tiene que existir una fila de A con la MISMA
(region, producto, mes) y las MISMAS unidades. Es determinístico y auditable.

CRITERIO (estricto a proposito)
-------------------------------
B se descarta solo si TODAS sus filas estan en A con identicas unidades. Si a B le sobra
una sola combinacion (region, producto, mes) que A no tiene, NO se descarta y se reporta:
podria aportar dato propio. Descartar una copia no cambia ningun valor.

Uso:
  py shared/qlik/detectar-mercados-copia-xlsx.py [--month 2026-04] --out <plan.json>
"""
from __future__ import annotations
import argparse, json, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

INPUTS = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs')
CFG = [('ATB', 'ATB'), ('respiratorio', 'respiratorio'), ('OTC', 'OTC'),
       ('dermato', 'dermato'), ('cardio', 'cardio'), ('SNC', 'PSQ'), ('mujer', 'linea-mujer')]
C_REG, C_MKT, C_MES, C_COD, C_UNI = 0, 1, 4, 6, 8


def resolver(hub, month):
    base = INPUTS / hub / month
    c = []
    for sub in ('fuentes-originales', 'ddd', ''):
        d = base / sub if sub else base
        if d.is_dir():
            c += [p for p in d.glob('Producto-Mol*provincia*.xlsx') if not p.name.startswith('~$')]
    return max(c, key=lambda p: p.stat().st_mtime) if c else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--month', default='2026-04')
    ap.add_argument('--out', required=True)
    a = ap.parse_args()
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')

    plan = {}
    for linea, hub in CFG:
        src = resolver(hub, a.month)
        if src is None:
            print('{:<13} SIN ARCHIVO'.format(linea)); continue
        wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        # celdas[mercado][(region, cod, mes)] = unidades
        celdas = defaultdict(dict)
        u_mkt = defaultdict(float)
        n = 0
        primera = True
        for row in ws.iter_rows(values_only=True):
            if primera:
                primera = False; continue
            if row is None or row[C_MKT] is None:
                continue
            n += 1
            mkt = str(row[C_MKT]).strip()
            k = (str(row[C_REG]).strip(), str(row[C_COD]).strip(), str(row[C_MES]).strip())
            u = row[C_UNI] if isinstance(row[C_UNI], (int, float)) else 0
            celdas[mkt][k] = celdas[mkt].get(k, 0) + u
            u_mkt[mkt] += u
        wb.close()

        mercados = sorted(celdas, key=lambda m: -u_mkt[m])   # de mayor a menor
        descartar, contenido_en, rechazados = [], {}, []
        for i, cont in enumerate(mercados):
            if cont in descartar:
                continue
            for dentro in mercados[i + 1:]:
                if dentro in descartar:
                    continue
                A, B = celdas[cont], celdas[dentro]
                if len(B) > len(A):
                    continue
                faltan = distintas = 0
                for k, u in B.items():
                    if k not in A:
                        faltan += 1
                    elif abs(A[k] - u) > 0.5:
                        distintas += 1
                if faltan == 0 and distintas == 0 and B:
                    descartar.append(dentro)
                    contenido_en[dentro] = cont
                elif faltan + distintas <= max(1, len(B) * 0.02) and B:
                    # casi-copia: se REPORTA y NO se descarta
                    rechazados.append({'contenido': dentro, 'contenedor': cont,
                                       'celdas_de_B': len(B), 'faltan_en_A': faltan,
                                       'con_unidades_distintas': distintas})
        u_desc = sum(u_mkt[m] for m in descartar)
        u_tot = sum(u_mkt.values())
        plan[linea] = {'archivo': src.name, 'filas': n, 'mercados': len(mercados),
                       'descartar': descartar, 'contenido_en': contenido_en,
                       'casi_copias_no_descartadas': rechazados,
                       'unidades_por_mercado': {m: round(u_mkt[m]) for m in mercados},
                       'unidades_descartadas': round(u_desc), 'unidades_totales': round(u_tot),
                       'inflacion_pct': round(u_desc / (u_tot - u_desc) * 100, 2) if u_tot > u_desc else None}
        print('{:<13} {:>9,} filas  {:>2} mercados  ->  descartar {}'.format(
            linea, n, len(mercados), len(descartar)))
        for m in descartar:
            print('     "{}"  copia dentro de  "{}"   ({:,.0f} u, {:.1f}% de la linea)'.format(
                m[:40], contenido_en[m][:40], u_mkt[m], u_mkt[m] / u_tot * 100))
        for r in rechazados:
            print('     [NO se descarta] "{}" casi-copia de "{}": le sobran {} de {} celdas'.format(
                r['contenido'][:30], r['contenedor'][:30], r['faltan_en_A'] + r['con_unidades_distintas'],
                r['celdas_de_B']))
        if plan[linea]['inflacion_pct']:
            print('     -> la linea estaba inflada {:.2f}%'.format(plan[linea]['inflacion_pct']))

    Path(a.out).write_text(json.dumps(plan, ensure_ascii=False, indent=1), encoding='utf-8')
    print()
    print('TOTAL: {} mercado(s) a descartar'.format(sum(len(p['descartar']) for p in plan.values())))
    print('plan -> {}'.format(a.out))
    return 0


if __name__ == '__main__':
    sys.exit(main())
