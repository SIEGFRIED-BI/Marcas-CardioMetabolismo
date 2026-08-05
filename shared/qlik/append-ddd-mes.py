#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
shared/qlik/append-ddd-mes.py

Agrega UN MES al xlsx regional 'Producto-Molecula-ATC-provincia' de cada linea, de forma
ESTRICTAMENTE ADITIVA: copia todas las filas del archivo vigente y le suma las filas del
mes nuevo que corresponden a los mercados de esa linea. Los meses anteriores no se
vuelven a extraer, asi que no pueden moverse.

POR QUE ASI Y NO RE-EXTRAYENDO POR LINEA
----------------------------------------
extract-ddd.mjs trae todos los meses de los mercados de una linea (~774k filas, ~5 min por
linea) y necesita un mapa mercado->linea que no existe (POC-DDD.md lo lista como
pendiente). Ademas la ventana de Qlik es movil: hoy sirve Jul-2024..Jun-2026, mientras los
tableros tienen Jun-2024..May-2026. Re-extraer PERDERIA Jun-2024. Anexando se queda con
las dos puntas: Jun-2024..Jun-2026.
El mapa mercado->linea sale de los propios archivos vigentes (los mercados que ya tienen)
y se guarda en el --mapa de salida, cerrando ese pendiente.

QUE ARCHIVO SE TOMA COMO BASE
-----------------------------
El mismo que resolveria build-competidores-shape-a.py: glob 'Producto-Mol*provincia*.xlsx'
en <hub>/<mes>/{fuentes-originales,ddd,} y el de mtime mas alto. Ojo que para cardio ese
es el de nebivolol (qlik-nebivolol 2026-07-30.xlsx), no el 'qlik 2026-07': tomar el
equivocado revertiria la reconstruccion de los mercados de nebivolol.
(La seleccion por mtime es fragil en OneDrive, que re-estampa al sincronizar; por eso el
script IMPRIME cual eligio para cada linea y aborta si el elegido ya contiene el mes.)

Uso:
  py shared/qlik/append-ddd-mes.py --mes-json <ddd_jun2026.json> --mes "Jun-2026"
     [--month 2026-04] [--sufijo jun2026] [--mapa <salida.json>] [--dry-run]
"""
from __future__ import annotations
import argparse, json, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

INPUTS = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs')
# (linea, subcarpeta del hub). Mismo mapeo que build-competidores-shape-a.py + mujer,
# que usa su propio builder (build-mujer-competidores-data.py) pero el MISMO xlsx.
CFG = [('ATB', 'ATB'), ('respiratorio', 'respiratorio'), ('OTC', 'OTC'),
       ('dermato', 'dermato'), ('cardio', 'cardio'), ('SNC', 'PSQ'), ('mujer', 'linea-mujer')]
HEADER = ['RegionCUP', 'Mercado', 'Droga', 'Clase Terapeutica', 'AñoMes',
          'Codigo Clase Terapeutica', 'Codigo Producto', 'Producto', 'Unidades']
COL_MERCADO, COL_MES, COL_UNID = 1, 4, 8


_MES_ES = {'Ene': 1, 'Feb': 2, 'Mar': 3, 'Abr': 4, 'May': 5, 'Jun': 6,
           'Jul': 7, 'Ago': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dic': 12}


def orden_mes(mk):
    """Clave cronologica para las etiquetas 'Jun-2024'. Ordenar alfabeticamente da
    'Abr-2025 .. Sep-2025' para una serie que en realidad va de Jun-2024 a May-2026."""
    p = str(mk).split('-')
    return (int(p[1]), _MES_ES.get(p[0], 0)) if len(p) == 2 and p[1].isdigit() else (0, 0)


def resolve_regional_xlsx(hub_sub, month):
    base = INPUTS / hub_sub / month
    cands = []
    for sub in ('fuentes-originales', 'ddd', ''):
        d = base / sub if sub else base
        if d.is_dir():
            cands += [p for p in d.glob('Producto-Mol*provincia*.xlsx') if not p.name.startswith('~$')]
    return max(cands, key=lambda p: p.stat().st_mtime) if cands else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--mes-json', default=None, help='JSON de extract-ddd-mes.mjs. Si se omite, '
                    'solo se aplica el dedup (--plan) sin agregar ningun mes.')
    ap.add_argument('--mes', default=None, help='etiqueta del mes, ej "Jun-2026"')
    ap.add_argument('--month', default='2026-04', help='carpeta de ciclo del hub')
    ap.add_argument('--sufijo', default=None, help='sufijo del archivo nuevo (default: el mes)')
    ap.add_argument('--mapa', default=None, help='escribe el mapa mercado->linea en este JSON')
    ap.add_argument('--plan', default=None,
                    help='JSON de detectar-mercados-copia.mjs: descarta los mercados que son '
                         'copia de otro (arregla el doble conteo del builder)')
    ap.add_argument('--sin-dedup', default='mujer',
                    help='lineas donde NO hay que aplicar el dedup, separadas por coma. '
                         'Default: mujer (ver el motivo abajo)')
    ap.add_argument('--dry-run', action='store_true')
    a = ap.parse_args()
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')

    por_mercado = defaultdict(list)
    suma_mes_total = 0
    if a.mes_json:
        if not a.mes:
            print('ERROR: con --mes-json hay que pasar --mes', file=sys.stderr)
            return 2
        filas_mes = json.loads(Path(a.mes_json).read_text(encoding='utf-8'))
        print('mes nuevo: {} | {:,} filas en el JSON de Qlik'.format(a.mes, len(filas_mes)))
        for r in filas_mes:
            if str(r[COL_MES]).strip() != a.mes:
                continue                   # defensa: el JSON tiene que ser de un solo mes
            por_mercado[str(r[COL_MERCADO]).strip()].append(r)
            suma_mes_total += (r[COL_UNID] or 0)
        print('  {} mercados distintos, {:,} unidades en total'.format(len(por_mercado), round(suma_mes_total)))
        if len(por_mercado) == 0:
            print('ERROR: el JSON no tiene filas de {}'.format(a.mes), file=sys.stderr)
            return 2
    else:
        if not a.plan:
            print('ERROR: sin --mes-json hay que pasar --plan (no habria nada que hacer)', file=sys.stderr)
            return 2
        print('SOLO DEDUP: no se agrega ningun mes, se reescriben los archivos sin las copias.')

    # Plan de descarte: mercados que son COPIA de un sub-slice de otro. Sus filas se saltean
    # tanto del historico como del mes nuevo. No cambia ningun valor -- el mercado que los
    # contiene ya tiene esas unidades -- y saca el doble conteo del builder, que acumula sin
    # mirar la columna Mercado.
    # MUJER QUEDA AFUERA DEL DEDUP A PROPOSITO.
    # El doble conteo existe porque build-competidores-shape-a.py re-indexa los mercados por
    # molecula/ATC (`units[brand][region][mes] += u`) y la columna Mercado no forma parte de
    # la clave: si el archivo trae el contenedor Y el contenido, las unidades se suman dos
    # veces. build-mujer-competidores-data.py NO hace eso: indexa por mercado
    # (`data[market][brand][region][mes] += u`, linea ~179), asi que cada mercado del Ateneo
    # es su propio bucket con su total correcto y NO hay doble conteo. Descartar ahi borraria
    # 22 mercados del tablero de mujer -- los segmentos de marketing (Isis Promocion,
    # Isis Mini 24, Siderblut Familia/IM...) son justamente lo que esa pagina muestra.
    sin_dedup = {x.strip() for x in (a.sin_dedup or '').split(',') if x.strip()}
    plan = {}
    if a.plan:
        crudo = json.loads(Path(a.plan).read_text(encoding='utf-8'))
        for ln, v in crudo.items():
            if ln in sin_dedup:
                if v.get('descartar'):
                    print('  {}: {} candidato(s) IGNORADOS (su builder indexa por mercado, '
                          'no hay doble conteo)'.format(ln, len(v['descartar'])))
                plan[ln] = set()
            else:
                plan[ln] = set(v.get('descartar') or [])
        tot = sum(len(v) for v in plan.values())
        print('plan de dedup: {} mercado(s) a descartar en {} lineas'.format(tot, len(plan)))

    sufijo = a.sufijo or (a.mes.lower().replace('-', '') if a.mes else 'dedup')
    mapa, reporte = {}, []
    for linea, hub in CFG:
        src = resolve_regional_xlsx(hub, a.month)
        if src is None:
            print('  [WARN] {}: sin regional en {}/{} -> skip'.format(linea, hub, a.month))
            continue
        print()
        print('{} <- {}'.format(linea, src.name))

        # ── pasada unica: se copia en streaming y se acumulan estadisticas ──
        dst = src.with_name('Producto-Molécula-ATC-provincia - qlik-{} {}.xlsx'.format(sufijo, a.month))
        wb_in = openpyxl.load_workbook(src, read_only=True, data_only=True)
        ws_in = wb_in[wb_in.sheetnames[0]]
        wb_out = openpyxl.Workbook(write_only=True) if not a.dry_run else None
        ws_out = wb_out.create_sheet('Sheet1') if wb_out else None
        if ws_out:
            ws_out.append(HEADER)

        descartar = plan.get(linea, set())
        mercados = set()
        por_mes_base = defaultdict(float)
        n_base = n_saltadas = 0
        u_saltadas = 0.0
        primera = True
        for row in ws_in.iter_rows(values_only=True):
            if primera:
                primera = False
                continue                   # header del origen
            if row is None or row[COL_MERCADO] is None:
                continue
            mkt = str(row[COL_MERCADO]).strip()
            u = row[COL_UNID] if isinstance(row[COL_UNID], (int, float)) else 0
            if mkt in descartar:
                n_saltadas += 1
                u_saltadas += u
                continue                   # mercado que es copia de otro: no se copia
            n_base += 1
            mercados.add(mkt)
            mk = str(row[COL_MES]).strip() if row[COL_MES] is not None else ''
            por_mes_base[mk] += u
            if ws_out:
                ws_out.append(list(row[:9]))
        wb_in.close()
        if descartar:
            print('  dedup: {} mercado(s) descartado(s) -> -{:,} filas, -{:,.0f} u del historico'.format(
                len(descartar), n_saltadas, u_saltadas))

        if a.mes and a.mes in por_mes_base:
            print('  ABORTADO: el archivo vigente YA tiene {} ({:,.0f} u). No se anexa dos veces.'.format(
                a.mes, por_mes_base[a.mes]), file=sys.stderr)
            if wb_out:
                wb_out.close()
            return 3

        # ── filas del mes nuevo que caen en los mercados de ESTA linea ──
        nuevas, suma_nuevas = 0, 0
        for m in sorted(mercados):
            for r in por_mercado.get(m, ()):
                if ws_out:
                    ws_out.append(r[:9])
                nuevas += 1
                suma_nuevas += (r[COL_UNID] or 0)
        sin_datos = [m for m in sorted(mercados) if m not in por_mercado]

        print('  base: {:,} filas, {} mercados, {} meses ({} .. {})'.format(
            n_base, len(mercados), len(por_mes_base),
            min(por_mes_base, key=orden_mes) if por_mes_base else '-',
            max(por_mes_base, key=orden_mes) if por_mes_base else '-'))
        if a.mes:
            print('  {}: +{:,} filas, {:,} unidades'.format(a.mes, nuevas, round(suma_nuevas)))
        if sin_datos and a.mes:
            print('  mercados sin filas en {}: {} -> {}'.format(
                a.mes, len(sin_datos), ', '.join(sin_datos[:4])[:90]))
        mapa[linea] = sorted(mercados)
        reporte.append({'linea': linea, 'origen': src.name, 'destino': dst.name,
                        'filas_base': n_base, 'mercados': len(mercados),
                        'mercados_descartados': sorted(descartar),
                        'filas_descartadas': n_saltadas, 'unidades_descartadas': round(u_saltadas),
                        'meses_base': len(por_mes_base),
                        'filas_nuevas': nuevas, 'unidades_nuevas': round(suma_nuevas),
                        'mercados_sin_datos_en_el_mes': sin_datos,
                        'unidades_por_mes_base': {k: round(v) for k, v in por_mes_base.items()}})
        if wb_out:
            wb_out.save(dst)
            print('  -> {} ({:.0f} MB)'.format(dst.name, dst.stat().st_size / 1048576))
            wb_out.close()

    if a.mapa:
        Path(a.mapa).write_text(json.dumps({'mes': a.mes, 'mapa_linea_mercados': mapa,
                                            'reporte': reporte}, ensure_ascii=False, indent=1),
                                encoding='utf-8')
        print('\nmapa mercado->linea -> {}'.format(a.mapa))
    print()
    tot_nuevas = sum(r['filas_nuevas'] for r in reporte)
    tot_u = sum(r['unidades_nuevas'] for r in reporte)
    print('TOTAL anexado: {:,} filas, {:,} unidades en {} lineas'.format(tot_nuevas, tot_u, len(reporte)))
    print('(Las lineas comparten mercados, asi que la suma por linea puede superar el total '
          'del mes: {:,} u del mes contra {:,} u anexadas.)'.format(round(suma_mes_total), tot_u))
    if a.dry_run:
        print('\nDRY RUN: no se escribio ningun xlsx.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
