#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
shared/fix-dermato-acneclin-split.py

Separa 'ACNECLIN (SIE)' en mol_perf.MINOCYCLINE (dermatologia/data.js) en sus
dos productos reales: ACNECLIN (tabl 50mg) y ACNECLIN AP (caps A.P 100mg).

Motivo: el AR_PM master ya reporta ambos como Product distintos desde Jul-2021
('ACNECLIN (SIE)' y 'ACNECLIN AP (SIE)'), pero mol_perf.MINOCYCLINE los tiene
mergeados en una sola entrada 'ACNECLIN (SIE)' (bug: un split hecho el 2026-05-05
(commit 2d7e43b) quedo revertido por el refresh de IQVIA del 2026-05-19, y
fix-acneclin-split.py (el script que lo arreglaba) quedo apuntando al HTML
inline viejo, no a data.js -> nunca se re-aplico).

Impacto: ACNECLIN se leia con 9.2x su volumen real (Jun 2026: 12.356 vs 1.346 u.,
MS% 42.06% vs 4.58% real) y ACNECLIN AP no tenia serie (BUD_IQVIA la resolvia a
un nombre de producto que ya no existe).

Metodo (reparticion exacta, CERO drift de familia):
  Para cada mes, el master da (a_master, b_master) = unidades reales de cada
  producto. Se reparte el valor YA PUBLICADO en mol_perf (V) proporcional a esa
  razon, con complemento exacto: a_new = round(V * a_master/s); b_new = V - a_new.
  Así SIEMPRE a_new + b_new == V (el total de familia no se mueve ni un punto,
  no dispara verify-history-preserved.py, no discute que vintage de master es
  la verdad para la familia -- eso es un tema aparte, ya explicado como ruido
  de restatement IQVIA en la auditoria).
  Meses fuera de la ventana del master (mas viejos que su primer mes, p.ej.
  Abr-Jun 2021 con este master): sin info para repartir -> 100% a ACNECLIN,
  AP=0 (mismo comportamiento que tenian antes de que IQVIA empezara a separar
  el producto).

Tambien:
  - Recomputa quarterly/ytd/mat/ms_* de los DOS productos nuevos (family
    agregada no cambia -> ms_* del resto de productos no se toca).
  - Corrige BUD_IQVIA['ACNECLIN AP'] en dermato_dashboard.html
    ('ACNECLIN (SIE)' -> 'ACNECLIN AP (SIE)'), que es lo que resuelve la
    serie del grafico de Mercado IQVIA para esa marca.
  - Recomputa brandKpis['ACNECLIN'/'ACNECLIN AP'].{ytd,mat} desde mol_perf,
    con la MISMA formula que fix-brandkpis-from-molperf.py (fuente unica),
    scopeado solo a estas 2 marcas de esta linea.

NO toca ninguna otra molecula, marca ni linea. Verifica postcondiciones antes
de escribir: la suma de productos por mes en MINOCYCLINE debe seguir dando el
family.monthly publicado (tolerancia 0).

Uso:
    py shared/fix-dermato-acneclin-split.py [--master <xlsx>] [--dry-run]
"""
from __future__ import annotations
import argparse, re, json, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
DATA = REPO / 'dermatologia' / 'data.js'
HTML = REPO / 'dermatologia' / 'dermato_dashboard.html'
DEFAULT_MASTER = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs'
                       r'\_iqvia-master\2026-06\AR_PM_FV_Standard_Jul-2026.xlsx')

MES = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
MES_INV = {m: i + 1 for i, m in enumerate(MES)}

MOL_KEY = 'MINOCYCLINE'
MERGED_NAME = 'ACNECLIN (SIE)'
SPLIT_NAMES = ('ACNECLIN (SIE)', 'ACNECLIN AP (SIE)')
BRAND_OF = {'ACNECLIN (SIE)': 'ACNECLIN', 'ACNECLIN AP (SIE)': 'ACNECLIN AP'}


def msort(mk):
    p = mk.split()
    return (int(p[1]), MES_INV.get(p[0], 0)) if len(p) == 2 else (0, 0)


def quarter_key(mk):
    p = mk.split()
    if len(p) != 2:
        return ''
    m = MES_INV.get(p[0])
    if not m:
        return ''
    return 'Q{} {}'.format((m - 1) // 3 + 1, p[1])


def aggregate_quarterly(monthly):
    out = defaultdict(int)
    for mk, v in monthly.items():
        qk = quarter_key(mk)
        if qk:
            out[qk] += v
    return dict(out)


def aggregate_ytd_per_year(monthly, cierre_month):
    by_year = defaultdict(int)
    for mk, v in monthly.items():
        p = mk.split()
        if len(p) != 2:
            continue
        m = MES_INV.get(p[0])
        if m and m <= cierre_month:
            by_year[p[1]] += v
    label = MES[cierre_month - 1]
    return {'{} {}'.format(label, y): v for y, v in by_year.items()}


def aggregate_mat(monthly, cierre_month):
    years = {int(mk.split()[1]) for mk in monthly if len(mk.split()) == 2}
    label = MES[cierre_month - 1]
    out = {}
    for y in sorted(years):
        total = 0
        for back in range(11, -1, -1):
            idx = (y * 12 + (cierre_month - 1)) - back
            yy, mm = divmod(idx, 12)
            total += int(monthly.get('{} {}'.format(MES[mm], yy), 0) or 0)
        out['{} {}'.format(label, y)] = total
    return out


def read_master_products(path):
    """-> {product_name: {mes: unidades}} para MINOCYCLINE, ambos SKU de ACNECLIN."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    col_product = col_mol = None
    month_cols = []
    for i, h in enumerate(row1):
        if not h:
            continue
        s = str(h).strip()
        sn = s.replace('\n', ' ').strip().lower()
        if sn.startswith('product'):
            col_product = i
        elif sn.startswith('molecules'):
            col_mol = i
        if s.startswith('Units'):
            after = (s.split('\n', 1)[-1] if '\n' in s else s[len('Units'):]).strip()
            if after.upper().startswith(('MAT', 'YTD')):
                continue
            m = re.match(r'(\w+)\s+(\d{4})$', after)
            if m and m.group(1) in MES_INV:
                month_cols.append((i, '{} {}'.format(m.group(1), m.group(2))))
    if col_product is None or col_mol is None:
        raise RuntimeError('No se encontraron columnas Product/Molecules Long en el master')

    out = {name: defaultdict(float) for name in SPLIT_NAMES}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        prod = row[col_product] if col_product < len(row) else None
        mol = row[col_mol] if col_mol < len(row) else None
        if not prod or not mol:
            continue
        if str(mol).strip().upper() != MOL_KEY:
            continue
        pname = str(prod).strip()
        if pname not in out:
            continue
        for ci, mk in month_cols:
            if ci >= len(row):
                continue
            v = row[ci]
            if isinstance(v, (int, float)):
                out[pname][mk] += v
    wb.close()
    return {name: dict(vals) for name, vals in out.items()}, [mk for _, mk in month_cols]


def split_monthly(merged_mv, master_by_prod):
    """Reparte merged_mv (dict mes->V publicado) en dos series que suman V exacto
    en CADA mes, proporcional a la razon real del master. Devuelve (ac_mv, ap_mv)."""
    a_master = master_by_prod['ACNECLIN (SIE)']
    b_master = master_by_prod['ACNECLIN AP (SIE)']
    ac_mv, ap_mv = {}, {}
    for mk, v in merged_mv.items():
        v = int(v or 0)
        a, b = a_master.get(mk, 0.0), b_master.get(mk, 0.0)
        s = a + b
        if s > 0:
            a_new = int(round(v * a / s))
            a_new = max(0, min(v, a_new))
        else:
            a_new = v  # sin info del master para este mes -> todo a ACNECLIN
        b_new = v - a_new
        ac_mv[mk] = a_new
        ap_mv[mk] = b_new
    return ac_mv, ap_mv


def build_product(name, monthly_vals, template, cierre_month):
    p = dict(template)
    p['prod'] = name
    p['manuf'] = 'SIEGFRIED'
    p['is_sie'] = True
    p['monthly_vals'] = monthly_vals
    p['quarterly_vals'] = aggregate_quarterly(monthly_vals)
    p['ytd'] = aggregate_ytd_per_year(monthly_vals, cierre_month)
    p['mat'] = aggregate_mat(monthly_vals, cierre_month)
    return p


def recompute_ms(p, fam_monthly, fam_quarterly, fam_ytd, fam_mat):
    mv = p['monthly_vals']
    p['ms_monthly'] = {mk: round(mv.get(mk, 0) / fv * 100, 2) if fv > 0 else 0
                       for mk, fv in fam_monthly.items()}
    qv = p['quarterly_vals']
    p['ms_quarterly'] = {qk: round(qv.get(qk, 0) / fv * 100, 2) if fv > 0 else 0
                         for qk, fv in fam_quarterly.items()}
    yv = p['ytd']
    p['ms_ytd'] = {y: round(yv.get(y, 0) / fv * 100, 2) if fv > 0 else 0
                   for y, fv in fam_ytd.items()}
    mtv = p['mat']
    p['ms_mat'] = {y: round(mtv.get(y, 0) / fv * 100, 2) if fv > 0 else 0
                   for y, fv in fam_mat.items()}


def ytd_months(year, end_m):
    return ['{} {}'.format(MES[i], year) for i in range(end_m)]


def mat_months(end_y, end_m):
    out, y, m = [], end_y, end_m
    for _ in range(12):
        out.append('{} {}'.format(MES[m - 1], y))
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    return list(reversed(out))


def sum_window(monthly, window):
    return sum(monthly.get(mk, 0) or 0 for mk in window)


def recompute_brandkpis(D, mol_obj, last_y, last_m):
    win_ytd_c = ytd_months(last_y, last_m)
    win_ytd_p = ytd_months(last_y - 1, last_m)
    win_mat_c = mat_months(last_y, last_m)
    win_mat_p = mat_months(last_y - 1, last_m)
    bk = D.get('brandKpis', {})
    for pname, brand in BRAND_OF.items():
        if brand not in bk:
            continue
        prod = next(p for p in mol_obj['products'] if p['prod'] == pname)
        for period, win_c, win_p in (('ytd', win_ytd_c, win_ytd_p), ('mat', win_mat_c, win_mat_p)):
            b_c = sum_window(prod['monthly_vals'], win_c)
            b_p = sum_window(prod['monthly_vals'], win_p)
            m_c = sum(sum_window(p['monthly_vals'], win_c) for p in mol_obj['products'])
            m_p = sum(sum_window(p['monthly_vals'], win_p) for p in mol_obj['products'])
            ms = round(b_c / m_c * 100, 1) if m_c > 0 else None
            ie = None
            if b_p > 0 and m_p > 0 and m_c > 0:
                brand_ratio, mkt_ratio = b_c / b_p, m_c / m_p
                if brand_ratio < 5 and 0.2 < mkt_ratio < 5:
                    ie = round(brand_ratio / mkt_ratio * 100, 1)
            growth = round((b_c / b_p - 1) * 100, 1) if b_p > 0 else None
            target = bk[brand].setdefault(period, {})
            target.update({'units': b_c, 'units_prev': b_p, 'market_total': m_c,
                           'ms': ms, 'ie': ie, 'growth': growth})
            print('  brandKpis[{}].{}: units={:,} ({:,} prev) market_total={:,} ms={} ie={} growth={}'.format(
                brand, period, b_c, b_p, m_c, ms, ie, growth))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--master', default=str(DEFAULT_MASTER))
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    master_path = Path(args.master)
    if not master_path.is_file():
        print('ERROR: no existe el master: {}'.format(master_path), file=sys.stderr)
        return 2

    print('Master: {}'.format(master_path))
    master_by_prod, master_months = read_master_products(master_path)
    for name in SPLIT_NAMES:
        print('  {}: {} meses en el master ({}..{})'.format(
            name, len(master_by_prod[name]),
            master_months[0] if master_months else '-', master_months[-1] if master_months else '-'))

    text = DATA.read_text(encoding='utf-8', errors='replace')
    m = re.search(r'window\.OTC_DASHBOARD\s*=\s*', text)
    if not m:
        print('ERROR: window.OTC_DASHBOARD no encontrado en {}'.format(DATA), file=sys.stderr)
        return 3
    abs_start = text.index('{', m.end())
    D, end = json.JSONDecoder().raw_decode(text[abs_start:])
    abs_end = abs_start + end

    mol_obj = D['mol_perf'][MOL_KEY]
    products = mol_obj['products']
    idx = next((i for i, p in enumerate(products) if p['prod'] == MERGED_NAME), None)
    if idx is None:
        already = any(p['prod'] == 'ACNECLIN AP (SIE)' for p in products)
        if already:
            print('ACNECLIN ya esta separado en mol_perf.MINOCYCLINE. Nada para hacer.')
            return 0
        print("ERROR: no se encontro el producto '{}' en mol_perf.{}".format(MERGED_NAME, MOL_KEY), file=sys.stderr)
        return 4

    merged = products[idx]
    merged_mv = {mk: int(v or 0) for mk, v in merged.get('monthly_vals', {}).items()}
    fam_monthly_before = {mk: int(v or 0) for mk, v in mol_obj.get('monthly', {}).items()}

    ac_mv, ap_mv = split_monthly(merged_mv, master_by_prod)

    # invariante: la reparticion no puede mover el total de ningun mes
    bad = [mk for mk in merged_mv if ac_mv[mk] + ap_mv[mk] != merged_mv[mk]]
    if bad:
        print('ERROR: la reparticion no preserva el total en {} meses: {}'.format(len(bad), bad[:5]), file=sys.stderr)
        return 5

    cierre_month = MES_INV[max(fam_monthly_before, key=msort).split()[0]]
    ac_prod = build_product('ACNECLIN (SIE)', ac_mv, merged, cierre_month)
    ap_prod = build_product('ACNECLIN AP (SIE)', ap_mv, merged, cierre_month)

    new_products = products[:idx] + [ac_prod, ap_prod] + products[idx + 1:]

    fam_monthly_after = defaultdict(int)
    for p in new_products:
        for mk, v in p['monthly_vals'].items():
            fam_monthly_after[mk] += int(v or 0)
    fam_monthly_after = dict(fam_monthly_after)
    diffs = {mk: (fam_monthly_before.get(mk, 0), fam_monthly_after.get(mk, 0))
             for mk in set(fam_monthly_before) | set(fam_monthly_after)
             if fam_monthly_before.get(mk, 0) != fam_monthly_after.get(mk, 0)}
    if diffs:
        print('ERROR: el total de familia MINOCYCLINE cambio en {} meses (no deberia): {}'.format(
            len(diffs), list(diffs.items())[:5]), file=sys.stderr)
        return 6
    print('OK: familia MINOCYCLINE.monthly identica antes/despues en los {} meses ({} u. total)'.format(
        len(fam_monthly_after), sum(fam_monthly_after.values())))

    fam_quarterly = mol_obj['quarterly']
    fam_ytd = mol_obj['ytd']
    fam_mat = mol_obj['mat']
    recompute_ms(ac_prod, fam_monthly_after, fam_quarterly, fam_ytd, fam_mat)
    recompute_ms(ap_prod, fam_monthly_after, fam_quarterly, fam_ytd, fam_mat)

    last_key = max(fam_monthly_after, key=msort)
    last_m_name, last_y = last_key.split()
    last_y, last_m = int(last_y), MES_INV[last_m_name]
    print('Ultimo mes: {} -> ventanas YTD/MAT ancladas ahi.'.format(last_key))

    mol_obj['products'] = new_products
    recompute_brandkpis(D, mol_obj, last_y, last_m)

    print('\nJun/ultimo mes - ACNECLIN: {:,} u.  ACNECLIN AP: {:,} u.  (antes mergeado: {:,} u.)'.format(
        ac_mv.get(last_key, 0), ap_mv.get(last_key, 0), merged_mv.get(last_key, 0)))

    html_text = HTML.read_text(encoding='utf-8', errors='replace')
    old_bud = '"ACNECLIN AP":"ACNECLIN (SIE)"'
    new_bud = '"ACNECLIN AP":"ACNECLIN AP (SIE)"'
    if new_bud in html_text:
        print('BUD_IQVIA ya estaba corregido.')
        bud_changed = False
    elif old_bud in html_text:
        html_text = html_text.replace(old_bud, new_bud)
        bud_changed = True
        print("BUD_IQVIA['ACNECLIN AP'] corregido -> 'ACNECLIN AP (SIE)'")
    else:
        print('ERROR: no se encontro el patron BUD_IQVIA esperado en {}'.format(HTML), file=sys.stderr)
        return 7

    if args.dry_run:
        print('\nDRY RUN: no se escribio nada.')
        return 0

    new_data_text = text[:abs_start] + json.dumps(D, ensure_ascii=False) + text[abs_end:]
    DATA.write_text(new_data_text, encoding='utf-8', newline='')
    print('-> {} reescrito ({:,} bytes)'.format(DATA, DATA.stat().st_size))
    if bud_changed:
        HTML.write_text(html_text, encoding='utf-8', newline='')
        print('-> {} reescrito ({:,} bytes)'.format(HTML, HTML.stat().st_size))
    return 0


if __name__ == '__main__':
    sys.exit(main())
