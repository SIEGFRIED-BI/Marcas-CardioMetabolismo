#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Deduplica ACNECLIN / ACNECLIN AP en dermatologia y los cierra contra el master.

QUE PASO (Jul-2026). `fix-dermato-acneclin-split.py` existe porque mol_perf.MINOCYCLINE
traia ACNECLIN y ACNECLIN AP MERGEADOS en una sola entrada, y el script los repartia
proporcionalmente segun el master. Esa premisa dejo de ser cierta: `sync-dermato-pm.py`
ahora los trae YA SEPARADOS (el AR_PM reporta ambos Product desde Jul-2021). El split
igual corrio y volvio a partir el ACNECLIN residual, creando una TERCERA fila con el
nombre repetido:

    MINOCYCLINE  ACNECLIN AP (SIE)  Jul-2026 = 10.597   <- la real (== master)
    MINOCYCLINE  ACNECLIN (SIE)     Jul-2026 =    199
    MINOCYCLINE  ACNECLIN AP (SIE)  Jul-2026 =  1.354   <- fantasma del split

Las tres suman 12.150, que es exactamente el total del master, asi que
check-molperf-suma-productos.py cerraba perfecto y nadie lo veia. Pero DOS productos con
el MISMO nombre rompen cualquier agregacion por marca: check-brandkpis-al-dia.py leia
una sola de las dos filas y daba 8 campos en desacuerdo, y eso a su vez hacia abortar a
rebuild-kpibybrand-snc.py (su validacion exige reproducir dermatologia/brandKpis campo
por campo: daba 148/156).

QUE HACE. Para cada mes, deja UNA fila por nombre con el valor del master
(ACNECLIN = 1.553, ACNECLIN AP = 10.597 en Jul-2026) y verifica que el total de las dos
no cambie respecto de lo publicado. Si el total se moviera, aborta: preferimos no tocar
antes que mover el mercado.

Uso:  python shared/fix-dermato-acneclin-dedup.py [--master <AR_PM.xlsx>] [--check]
"""
from __future__ import annotations
import argparse, json, re, sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
TARGET = 'dermatologia/data.js'
FAM = 'MINOCYCLINE'
NOMBRES = ('ACNECLIN (SIE)', 'ACNECLIN AP (SIE)')
MES = ('Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
       'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec')


def leer_master(path, nombres):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    hdr = [str(h).replace('\n', ' ').strip() if h else ''
           for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    pi = next(i for i, h in enumerate(hdr) if h.strip().lower() == 'product')
    cols = {}
    for i, h in enumerate(hdr):
        if not h.startswith('Units'):
            continue
        a = h[len('Units'):].strip()
        m = re.match(r'^(\w+)\s+(\d{4})$', a)
        if m and m.group(1) in MES:
            cols[a] = i
    out = {n: {} for n in nombres}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or pi >= len(row) or not row[pi]:
            continue
        n = str(row[pi]).strip().upper()
        if n not in out:
            continue
        for mk, ci in cols.items():
            v = row[ci] if ci < len(row) else None
            if isinstance(v, (int, float)):
                out[n][mk] = out[n].get(mk, 0.0) + float(v)
    wb.close()
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--master', default=None)
    ap.add_argument('--check', action='store_true')
    args = ap.parse_args()

    master = args.master
    if not master:
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        import manifest
        master = manifest.resolve_source('iqvia_master')

    p = REPO / TARGET
    t = p.read_text(encoding='utf-8-sig', errors='replace')
    m = re.search(r'window\.OTC_DASHBOARD\s*=\s*', t)
    ob = t.index('{', m.end())
    D, end = json.JSONDecoder().raw_decode(t[ob:])

    fam = (D.get('mol_perf') or {}).get(FAM)
    if not fam:
        print(f'  (skip) no existe mol_perf[{FAM}]')
        return 0
    filas = [x for x in fam.get('products', [])
             if str(x.get('prod', '')).strip().upper() in NOMBRES]
    dups = len(filas) - len({str(x.get('prod', '')).strip().upper() for x in filas})
    print(f'  {FAM}: {len(filas)} filas ACNECLIN*, {dups} duplicada(s) por nombre')
    if dups <= 0:
        print('  (nada que deduplicar)')
        return 0

    src = leer_master(master, [n.upper() for n in NOMBRES])

    meses = set()
    for x in filas:
        meses.update((x.get('monthly_vals') or {}).keys())
    meses = [k for k in meses if len(k.split()) == 2 and k.split()[0] in MES]

    def mkey(k):
        return (int(k.split()[1]), MES.index(k.split()[0]))

    # El master arranca despues que la serie publicada (Ago-2021 vs Abr-2021). Fuera de
    # su ventana no hay con que repartir, asi que ahi NO se reasigna: se FUSIONAN las
    # filas homonimas sumandolas, que preserva el total y el valor por nombre.
    cubiertos = set()
    for n in (x.upper() for x in NOMBRES):
        cubiertos.update(src.get(n, {}).keys())
    if not cubiertos:
        print('  ABORTA: el master no trae ninguno de los dos productos.')
        return 2
    m_min = min(cubiertos, key=mkey)
    dentro = [mk for mk in meses if mkey(mk) >= mkey(m_min)]
    fuera = [mk for mk in meses if mkey(mk) < mkey(m_min)]
    print(f'  ventana del master: desde {m_min} -> {len(dentro)} meses reasignados, '
          f'{len(fuera)} anteriores solo fusionados')

    # total publicado por mes (lo que NO se puede mover), solo donde reasignamos
    antes = {mk: sum(float((x.get('monthly_vals') or {}).get(mk, 0) or 0) for x in filas)
             for mk in dentro}
    despues = {mk: sum(src.get(n.upper(), {}).get(mk, 0.0) for n in NOMBRES) for mk in dentro}
    peor = max(((abs(despues[mk] - antes[mk]) / antes[mk], mk)
                for mk in dentro if antes[mk]), default=(0, None))
    print(f'  total de las 2 marcas: peor desvio master vs publicado = '
          f'{peor[0]*100:.3f}% ({peor[1]})')
    if peor[0] > 0.005:
        print('  ABORTA: reemplazar por el master moveria el mercado mas de 0,5%.')
        return 2

    if args.check:
        print('  --check: hay duplicados para deduplicar.')
        return 1

    # una sola fila por nombre
    base, fusion = {}, {}
    for x in filas:
        n = str(x.get('prod', '')).strip().upper()
        base.setdefault(n, x)
        acc = fusion.setdefault(n, {})
        for mk, v in (x.get('monthly_vals') or {}).items():
            acc[mk] = acc.get(mk, 0.0) + float(v or 0)
    for n, x in base.items():
        mv = {}
        for mk in sorted(meses, key=mkey):
            if mkey(mk) >= mkey(m_min):
                mv[mk] = int(round(src.get(n, {}).get(mk, 0.0)))   # master manda
            else:
                mv[mk] = int(round(fusion.get(n, {}).get(mk, 0.0)))  # solo fusion
        x['monthly_vals'] = mv
    fam['products'] = [x for x in fam.get('products', [])
                       if str(x.get('prod', '')).strip().upper() not in NOMBRES] \
                      + list(base.values())

    p.write_text(t[:ob] + json.dumps(D, ensure_ascii=False) + t[ob + end:],
                 encoding='utf-8', newline='')
    ult = max(meses, key=lambda k: (int(k.split()[1]), MES.index(k.split()[0])))
    print(f'  -> {TARGET}: {len(filas)} filas -> {len(base)}. '
          f'{ult}: ' + ', '.join(f'{n}={src.get(n,{}).get(ult,0):,.0f}' for n in
                                 (x.upper() for x in NOMBRES)))
    print('  Correr: recompute-mol-perf-aggregates, build-kpis, fix-brandkpis-*.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
