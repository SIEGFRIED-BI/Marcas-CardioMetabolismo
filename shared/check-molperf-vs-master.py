#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""G1: cierra mol_perf contra el master AR_PM leido de forma INDEPENDIENTE.

El resto de los gates son de consistencia interna (suma de productos == total de
la familia, kpiStrip == kpis.json). Todos cierran perfecto aunque el builder haya
leido la columna equivocada del xlsx: comparan el artefacto consigo mismo. Este
lee el master con openpyxl, resuelve las columnas por HEADER y compara unidades
por (Producto, mes) contra lo que quedo publicado en mol_perf.

La clave del join es exacta: mol_perf['<fam>'].products[].prod trae el mismo
string que la columna Product del master ('ROXOLAN (SIE)', 'SINLIP (GAD)').

Ratio esperado 1,000. Los dos casos que NO son 1:1 se resuelven explicando, no
aflojando la tolerancia:

  - Una marca SIE puede figurar en MAS DE UNA familia con el mismo valor (cardio:
    DILATREND (SIE) esta en 'DILATREND' y en 'DILATREND AP'). Sumar a ciegas da
    ratio 2,0000 exacto -- la firma de duplicacion estructural. Se cuenta una sola
    vez con la regla de familia PRIMARIA, la misma que usa build-kpis.py.
  - Marcas que un paso posterior al build reparte en dos claves (MAGNUS/MAGNUS 36,
    ROXOLAN/ROXOLAN PLUS, ACNECLIN/ACNECLIN AP): se comparan como GRUPO, suma
    publicada vs suma del master. Es una verificacion, no un permiso: si el split
    perdio o invento unidades, salta igual.

  'Otros (resto del mercado)' es sintetico (residuo del build) -> se excluye.

Uso:
  python shared/check-molperf-vs-master.py --master <AR_PM.xlsx> --month "Jul 2026"
  python shared/check-molperf-vs-master.py --month "Jul 2026" --tol 0.005
Exit != 0 si alguna marca SIE se desvia mas de --tol. Las marcas sin match en el
master (0 publicado y 0 en la fuente) se REPORTAN pero no bloquean: son marcas sin
venta en el mes, no un error -- que no bloqueen es lo que evita que el gate termine
desactivado por ruido.
"""
from __future__ import annotations
import argparse, json, os, re, sys
from collections import defaultdict
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent

LINES = [
    ('cardio', 'cardio/data.js'), ('antibio', 'ATB/data.js'),
    ('mujer', 'mujer/data.js'), ('snc', 'SNC/data.js'),
    ('resp', 'respiratorio/data.js'), ('otx', 'OTC/data.js'),
    ('derma', 'dermatologia/data.js'),
]

MES = ('Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
       'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec')

# Marcas que un paso POSTERIOR al build reparte entre dos claves (por presentacion
# o por molecula). Comparar cada una por separado da un desvio esperado, asi que se
# comparan como GRUPO: la suma publicada tiene que dar la suma del master. Es una
# verificacion, no un permiso -- si el split perdio o invento unidades, salta igual.
SPLIT_GROUPS = [
    ('MAGNUS', 'MAGNUS 36'),
    ('ROXOLAN', 'ROXOLAN PLUS'),
    ('TETRALGIN', 'TETRALGIN NOVO'),
    ('ACNECLIN', 'ACNECLIN AP'),
]
IN_GROUP = {b: g for g in SPLIT_GROUPS for b in g}


def parse_data_js(text):
    m = re.search(r'window\.OTC_DASHBOARD\s*=\s*', text)
    if not m:
        return None
    ob = text.index('{', m.end())
    return json.JSONDecoder().raw_decode(text[ob:])[0]


def load_master(path, months):
    """Devuelve units[product_exacto][mes] leyendo el xlsx por HEADER."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    hdr = [str(h).replace('\n', ' ').strip() if h else ''
           for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    col_product = None
    for i, h in enumerate(hdr):
        if h.strip().lower() == 'product':
            col_product = i
            break
    if col_product is None:
        raise SystemExit(f'FATAL: no encuentro la columna Product por header en {path}')

    month_cols = {}
    for i, h in enumerate(hdr):
        if not h.startswith('Units'):
            continue
        after = h[len('Units'):].strip()
        m = re.match(r'^(\w+)\s+(\d{4})$', after)
        if m and m.group(1) in MES:
            month_cols[f'{m.group(1)} {m.group(2)}'] = i

    want = {mk: c for mk, c in month_cols.items() if mk in months}
    missing = [mk for mk in months if mk not in want]
    if missing:
        raise SystemExit(f'FATAL: el master no trae {missing}. Tiene '
                         f'{min(month_cols) if month_cols else "-"}..'
                         f'{max(month_cols, key=lambda k: (int(k.split()[1]), MES.index(k.split()[0]))) if month_cols else "-"}')

    units = defaultdict(lambda: defaultdict(float))
    nrows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or col_product >= len(row):
            continue
        prod = row[col_product]
        if not prod:
            continue
        nrows += 1
        key = str(prod).strip().upper()
        for mk, ci in want.items():
            if ci < len(row):
                v = row[ci]
                if isinstance(v, (int, float)):
                    units[key][mk] += float(v)
    wb.close()
    print(f'  master: {nrows} filas leidas, {len(units)} productos distintos, '
          f'columna Product = idx {col_product}, meses {sorted(want)}')
    return units


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--master', default=None)
    ap.add_argument('--month', action='append', default=None,
                    help='month key tipo "Jul 2026" (repetible)')
    ap.add_argument('--tol', type=float, default=0.005,
                    help='desvio relativo tolerado por marca (default 0,5%%)')
    args = ap.parse_args()

    master = args.master
    if not master:
        od = os.environ.get('OneDrive', '')
        cand = Path(od) / 'Documentos' / 'Hub-Marcas-Inputs' / '_iqvia-master' / '2026-04'
        pool = sorted(cand.glob('AR_PM*.xlsx'), key=lambda p: p.stat().st_mtime)
        if not pool:
            raise SystemExit('FATAL: no resuelvo el master; pasar --master')
        master = str(pool[-1])
    months = args.month or ['Jul 2026']
    print(f'master: {master}')
    units = load_master(master, months)

    rows = []       # (linea, marca, publicado, master, ratio)
    multi = []      # productos que aparecen en mas de una familia
    for key, rel in LINES:
        p = REPO / rel
        if not p.is_file():
            continue
        D = parse_data_js(p.read_text(encoding='utf-8-sig', errors='replace'))
        if not D:
            continue
        # Una marca SIE puede figurar en MAS DE UNA familia, y hay DOS casos que se
        # ven igual en el JSON pero significan lo contrario:
        #
        #   LISTADO DUPLICADO  cardio DILATREND (SIE): 55.549 en 'DILATREND' y 55.549
        #     en 'DILATREND AP' -- el MISMO total repetido en dos mercados. Sumar da
        #     ratio 2,0000 exacto (la firma de duplicacion estructural).
        #   PARTICION REAL     derma MICOMAZOL (SIE): 15.601 en 'CLOTRIMAZOLE' y 268
        #     en 'CICLOPIROX' -- la marca tiene productos en dos moleculas y el master
        #     las trae agregadas bajo un unico Product. Aca hay que SUMAR (15.869).
        #
        # Regla: sumar los valores DISTINTOS por familia. Reproduce los dos casos.
        # Si dos familias tuvieran el mismo valor por casualidad quedaria sub-contado,
        # asi que los productos multi-familia se listan siempre (abajo), nunca en
        # silencio.
        per_fam = defaultdict(list)   # nombre -> [(familia, valor)]
        for m_key, obj in D.get('mol_perf', {}).items():
            if not isinstance(obj, dict):
                continue
            for prod in obj.get('products', []):
                if not prod.get('is_sie'):
                    continue
                name = str(prod.get('prod') or '').strip()
                if not name or 'OTROS' in name.upper():
                    continue
                mv = prod.get('monthly_vals', {})
                v = sum(float(mv.get(mk, 0) or 0) for mk in months)
                per_fam[name.upper()].append((str(m_key), v))

        pub = {}
        for name, pares in per_fam.items():
            vistos, total = set(), 0.0
            for fam, v in pares:
                if v in vistos:
                    continue
                vistos.add(v)
                total += v
            pub[name] = total
            if len(pares) > 1:
                det = ', '.join(f'{f}={v:,.0f}' for f, v in sorted(pares))
                dup = ' [valor repetido -> contado una vez]' if len(vistos) < len(pares) else ''
                multi.append(f'  {key:8s} {name:26s} {det}{dup}')
        for name, v_pub in sorted(pub.items()):
            v_src = sum(units.get(name, {}).get(mk, 0.0) for mk in months)
            ratio = (v_pub / v_src) if v_src else float('nan')
            rows.append((key, name, v_pub, v_src, ratio))

    if not rows:
        print('\nFATAL: 0 marcas SIE encontradas en mol_perf; no hay nada que conciliar.')
        return 1

    # Las marcas de un SPLIT_GROUP se comparan como grupo (suma vs suma).
    grouped = defaultdict(lambda: [0.0, 0.0, []])
    singles = []
    for r in rows:
        line, name, v_pub, v_src, _ = r
        base = re.sub(r'\s*\(.*?\)\s*$', '', name).strip().upper()
        g = IN_GROUP.get(base)
        if g:
            acc = grouped[(line, g)]
            acc[0] += v_pub
            acc[1] += v_src
            acc[2].append(base)
        else:
            singles.append(r)

    ok, bad, nomatch = [], [], []
    for r in singles:
        _, _, v_pub, v_src, ratio = r
        if v_src == 0 and v_pub == 0:
            nomatch.append(r)
        elif v_src == 0:
            bad.append(r)
        elif abs(ratio - 1.0) <= args.tol:
            ok.append(r)
        else:
            bad.append(r)

    for (line, g), (v_pub, v_src, miembros) in sorted(grouped.items()):
        label = ' + '.join(g) + f'  [grupo: {", ".join(sorted(set(miembros)))}]'
        ratio = (v_pub / v_src) if v_src else float('nan')
        row = (line, label, v_pub, v_src, ratio)
        if v_src == 0 and v_pub == 0:
            nomatch.append(row)
        elif v_src and abs(ratio - 1.0) <= args.tol:
            ok.append(row)
        else:
            bad.append(row)

    def dump(title, items):
        if not items:
            return
        print(f'\n{title}')
        for k, n, a, b, rt in items:
            rr = f'{rt:.4f}' if rt == rt else '  n/a'
            print(f'  {k:8s} {n[:34]:34s} publicado={a:12,.0f}  master={b:12,.0f}  ratio={rr}')

    print(f'\nmeses conciliados: {", ".join(months)}   tolerancia: {args.tol:.3%}')
    print(f'marcas SIE conciliadas: {len(rows)}')
    dump(f'OK  (|ratio-1| <= {args.tol:.3%}): {len(ok)}', ok[:5] if len(ok) > 5 else ok)
    if len(ok) > 5:
        print(f'  ... y {len(ok)-5} mas')

    if multi:
        print(f'\nMARCAS EN MAS DE UNA FAMILIA (se suman los valores distintos): {len(multi)}')
        for l in multi:
            print(l)
    dump(f'SIN MATCH en el master (revisar nombre): {len(nomatch)}', nomatch)
    dump(f'FUERA DE TOLERANCIA: {len(bad)}', bad)

    print('\n' + '=' * 70)
    print(f'check-molperf-vs-master: {len(ok)} OK  {len(bad)} FUERA DE TOL  '
          f'{len(nomatch)} SIN MATCH  ({len(grouped)} grupos de split verificados '
          f'como suma)')
    print('=' * 70)
    return 1 if bad else 0


if __name__ == '__main__':
    sys.exit(main())
