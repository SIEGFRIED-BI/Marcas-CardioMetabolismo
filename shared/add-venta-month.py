#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
shared/add-venta-month.py

Agrega UN mes cerrado de Venta Interna a budget[fam][anio].real en las 7 lineas,
sin tocar NINGUN otro mes.

POR QUE NO merge-ventas-internas.py
-----------------------------------
El merge general reescribe TODOS los meses que trae la planilla (_write_real hace
real_arr[midx] = v para cada mes del archivo), y agrupa por Familia (col1). Eso
tiene dos consecuencias para un cierre incremental:

  1. REVIERTE los splits de venta en TODOS los meses, no solo en el nuevo:
     familias donde varias marcas del tablero comparten la Familia SAP quedan
     colapsadas en la key padre (regla #5 de CLAUDE.md). Hay que re-aplicar
     4 correctores despues... pero 2 de ellos NO son re-aplicables hoy:
       - shared/split-cardio-roxolan.py tiene PLANILLA y CUTOFF=(2026,4)
         HARDCODEADOS -> correrlo hoy escribiria hasta May-2026 y pisaria Jun/Jul.
       - el split de venta de SYNCROCOR vive dentro de
         shared/onboard-cardio-syncrocor.py (onboarding completo, no un corrector).
  2. Cambia RETROACTIVAMENTE meses ya publicados. Ej. ROXOLAN Ene-Jun pasaria de
     105.151 (valor publicado, con split aplicado) a 121.629 (Familia SAP entera,
     que incluye ROXOLAN PLUS). Viola la regla #7 ("los merges AGREGAN meses,
     nunca reemplazan") y mueve numeros que ya se leyeron.

Este script es ADITIVO POR CONSTRUCCION: escribe exclusivamente el indice del mes
pedido. Si el valor ya existia y coincide, no hay cambio; si difiere, lo reporta.
Nunca toca los otros 11 indices ni el array 'budget' (estimado).

REGLAS DE ASIGNACION (las mismas que ya usa el pipeline, no se inventa ninguna)
------------------------------------------------------------------------------
  - generico (cardio/ATB/OTC/respi/SNC/derma): por fila (Gran Familia g, Familia f)
    se elige la budget key MAS ESPECIFICA: alias(f) -> f -> g. Identico a
    merge-ventas-internas.py:update_budget. Asi las sub-familias sin key propia
    (TETRALGIN APC, MICROSONA BB, ...) caen en la key padre, incluidos sus
    ajustes negativos (creditos).
  - mujer: por Familia (col1) via MUJER_SEGMENT_TO_FAMS (segmentos de marketing).
  - SPLITS por Presentacion (col3) / Cod. Presentacion (col4), porque varias
    marcas del tablero comparten la misma Familia SAP:
      OTC    MAGNUS / MAGNUS 36        <- 'MAGNUS 36' en la Presentacion
      mujer  45 / D3 / D3 PLUS / MAGNESIO  <- variante en la Presentacion
      cardio ROXOLAN / ROXOLAN PLUS    <- PLUS|DUO|EZ|COMBICOL|EZETIM
      cardio SYNCROCOR / SYNCROCOR D   <- Cod. Presentacion del combo
    Las reglas se importan/replican de los correctores existentes
    (apply-otc-magnus-split.py, fix-mujer-trip-venta.py, split-cardio-roxolan.py,
    onboard-cardio-syncrocor.py) para no divergir.

Uso:
    py shared/add-venta-month.py --file <planilla.xlsx> --month 2026-07 [--dry-run]
"""
from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
MES3 = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

LINES = [
    ('cardio', 'cardio/data.js'),
    ('ATB', 'ATB/data.js'),
    ('OTC', 'OTC/data.js'),
    ('respiratorio', 'respiratorio/data.js'),
    ('mujer', 'mujer/data.js'),
    ('SNC', 'SNC/data.js'),
    ('dermatologia', 'dermatologia/data.js'),
]

# ── mapeos importados del pipeline (fallbacks identicos a merge-ventas-internas) ──
sys.path.insert(0, str(Path(__file__).resolve().parent))
try:
    import manifest as _mf
except Exception:
    _mf = None


def _seg(name, key, default):
    return _mf.seg_get(name, key, default) if _mf else default


MUJER_SEGMENT_TO_FAMS = _seg('mujer_venta_segments', 'segmentToFams', {
    'SIN ESTROGENO': ['ISIS FREE'], 'ALTA DOSIS': ['ISIS'],
    'BAJA DOSIS 21+7': ['ISIS MINI'], 'BAJA DOSIS 24': ['ISIS MINI 24'],
    'COMPLEX': ['SIDERBLUT COMPLEX', 'SIDERBLUT FOLIC'],
    'SOLO': ['SIDERBLUT', 'SIDERBLUT POLI', 'FERINSOL'],
    'D3': [], 'D3 PLUS': [], '45': [], 'MAGNESIO': [],
    'DELTROX': ['DELTROX'], 'BASE': ['CALCIO BASE DUPOMAR'],
    'BASE D': ['CALCIO BASE DUPOMAR D', 'CALCIO BASE DUPOMAR D3',
               'CALCIO CITRATO DUPOMAR D3'],
})
KEY_ALIASES = _seg('mujer_venta_segments', 'keyAliases',
                   {'HEXALER BRONQUIAL DU': 'HEXALER BRONQUIAL DUO'})
MAGNUS_MARKER = _seg('magnus_split', 'presentacionMarker', 'MAGNUS 36')
MAGNUS_KEY_TADA = _seg('magnus_split', 'keyTada', 'MAGNUS 36')

# SYNCROCOR: combo por Cod. Presentacion, con fallback por nombre, igual que
# onboard-cardio-syncrocor.py:113-115 (VENTA_COD_COMBO + VENTA_RE_COMBO). El
# fallback importa si manana agregan una presentacion de combo con otro codigo:
# sin el, caeria en el mono y SYNCROCOR D quedaria corto sin que nada lo marque.
SYNCRO_FAMILIA = 'SYNCROCOR'
SYNCRO_COD_COMBO = {'3048406'}
SYNCRO_RE_COMBO = re.compile(r'^SYNCROCOR\s+D\b')
SYNCRO_KEY_COMBO = 'SYNCROCOR D'

# ROXOLAN: para VENTA la regla es 'PLUS' en la Presentacion, tal cual
# split-cardio-roxolan.py:parse_budget_split (linea 65: tgt = plus if 'PLUS' in
# pres.upper()). OJO: su is_plus() (PLUS|DUO|EZ|COMBICOL|EZETIM) es para clasificar
# productos de IQVIA, NO para el split de venta -- no confundirlas.
ROXOLAN_FAMILIA = 'ROXOLAN'
ROXOLAN_KEY_PLUS = 'ROXOLAN PLUS'

# TRIP (fix-mujer-trip-venta.py:45 classify)
TRIP_FAMILIA = 'TRIP'


def trip_classify(pres):
    pu = str(pres or '').upper()
    if '+45' in pu or re.search(r'\b45\b', pu):
        return '45'
    if 'MAGNESIO' in pu:
        return 'MAGNESIO'
    if 'D3 PLUS' in pu:
        return 'D3 PLUS'
    if 'D3' in pu:
        return 'D3'
    return None


def norm(s):
    return re.sub(r'\s+', ' ', str(s or '')).strip().upper()


def read_planilla(path, month_label):
    """-> (rows, month_col). rows = lista de dicts con g,f,pres,cod,val del mes pedido."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    idx = {}
    for i, h in enumerate(hdr):
        if not h:
            continue
        s = str(h).strip()
        sl = s.lower()
        if sl.startswith('gran familia'):
            idx['g'] = i
        elif sl == 'familia':
            idx['f'] = i
        elif sl.startswith('presentaci'):
            idx['pres'] = i
        elif sl.startswith('c') and 'presentaci' in sl:
            idx['cod'] = i
        if s == month_label:
            idx['mes'] = i
    faltan = [k for k in ('g', 'f', 'mes') if k not in idx]
    if faltan:
        raise RuntimeError('la planilla no tiene columna(s) {} (headers: {})'.format(
            faltan, [str(h)[:20] for h in hdr[:8]]))
    out = []
    for r in it:
        if not r:
            continue
        v = r[idx['mes']] if idx['mes'] < len(r) else None
        if not isinstance(v, (int, float)):
            continue
        out.append({
            'g': norm(r[idx['g']] if idx['g'] < len(r) else ''),
            'f': norm(r[idx['f']] if idx['f'] < len(r) else ''),
            'pres': str(r[idx['pres']] or '') if 'pres' in idx and idx['pres'] < len(r) else '',
            'cod': str(r[idx['cod']] or '').lstrip('0') if 'cod' in idx and idx['cod'] < len(r) else '',
            'v': float(v),
        })
    wb.close()
    return out


def compute_line(line, keys, rows):
    """-> {budget_key: unidades del mes} aplicando las reglas del pipeline."""
    acc = defaultdict(float)
    keyset = set(keys)
    alias_fam_to_key = {fam: k for k, fam in KEY_ALIASES.items() if k in keyset}

    if line == 'mujer':
        by_fam = defaultdict(float)
        for r in rows:
            by_fam[r['f']] += r['v']
        for k in keys:
            fams = MUJER_SEGMENT_TO_FAMS.get(k)
            if not fams:
                continue                      # segmento sin mapeo (TRIP) -> se resuelve abajo
            s = sum(by_fam.get(norm(tf), 0.0) for tf in fams)
            if s:
                acc[k] += s
        # TRIP: por Presentacion
        for r in rows:
            if r['f'] != TRIP_FAMILIA:
                continue
            k = trip_classify(r['pres'])
            if k and k in keyset:
                acc[k] += r['v']
        return acc

    for r in rows:
        g, f = r['g'], r['f']
        # --- splits por presentacion / codigo ---
        if line == 'OTC' and f == norm('MAGNUS'):
            k = MAGNUS_KEY_TADA if MAGNUS_MARKER.upper() in norm(r['pres']) else 'MAGNUS'
            if k in keyset:
                acc[k] += r['v']
            continue
        if line == 'cardio' and f == ROXOLAN_FAMILIA:
            k = ROXOLAN_KEY_PLUS if 'PLUS' in norm(r['pres']) else 'ROXOLAN'
            if k in keyset:
                acc[k] += r['v']
            continue
        if line == 'cardio' and f == SYNCRO_FAMILIA:
            es_combo = (r['cod'] in SYNCRO_COD_COMBO) or bool(SYNCRO_RE_COMBO.match(norm(r['pres'])))
            k = SYNCRO_KEY_COMBO if es_combo else 'SYNCROCOR'
            if k in keyset:
                acc[k] += r['v']
            continue
        # --- generico: mas especifica (alias -> f -> g) ---
        k = alias_fam_to_key.get(f) or (f if f in keyset else (g if g in keyset else None))
        if k:
            acc[k] += r['v']
    return acc


def load_D(rel):
    p = REPO / rel
    t = p.read_text(encoding='utf-8', errors='replace')
    m = re.search(r'window\.OTC_DASHBOARD\s*=\s*', t)
    if not m:
        raise RuntimeError('window.OTC_DASHBOARD no encontrado en {}'.format(rel))
    s = t.index('{', m.end())
    D, end = json.JSONDecoder().raw_decode(t[s:])
    return p, t, s, s + end, D


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--file', required=True, help='Planilla de Ventas (.xlsx)')
    ap.add_argument('--month', required=True, help="Mes a agregar, 'YYYY-MM'")
    ap.add_argument('--dry-run', action='store_true')
    a = ap.parse_args()

    m = re.match(r'(\d{4})-(\d{2})$', a.month)
    if not m:
        print('ERROR: --month debe ser YYYY-MM', file=sys.stderr)
        return 2
    year, mnum = int(m.group(1)), int(m.group(2))
    midx = mnum - 1
    label = '{}-{}'.format(MES3[midx], year)

    fp = Path(a.file)
    if not fp.is_file():
        print('ERROR: no existe {}'.format(fp), file=sys.stderr)
        return 2
    print('Planilla: {}'.format(fp))
    print('Mes a agregar: {} (indice {} del array real, anio {})'.format(label, midx, year))
    rows = read_planilla(fp, label)
    print('  filas con dato en {}: {:,}  | total: {:,.0f} u.'.format(label, len(rows), sum(r['v'] for r in rows)))
    print()

    total_escrito = 0.0
    cambios, ya_estaba, sin_dato, conflictos = [], [], [], []
    planned = []

    for line, rel in LINES:
        p, text, s, e, D = load_D(rel)
        budget = D.get('budget') or {}
        keys = list(budget.keys())
        acc = compute_line(line, keys, rows)
        n_w = 0
        suma = 0.0
        for k in keys:
            v = acc.get(k)
            if v is None:
                sin_dato.append((line, k))
                continue
            v = int(round(v))
            yo = budget[k].setdefault(str(year), {})
            real = yo.get('real')
            if not isinstance(real, list) or len(real) != 12:
                real = [None] * 12
            prev = real[midx]
            if prev is not None and int(prev or 0) != v:
                conflictos.append((line, k, prev, v))
            elif prev is not None:
                ya_estaba.append((line, k))
            real[midx] = v
            yo['real'] = real
            yo.setdefault('budget', [0] * 12)
            n_w += 1
            suma += v
        total_escrito += suma
        cambios.append((line, n_w, suma))
        planned.append((p, text, s, e, D))

    print('{:<16} {:>8} {:>16}'.format('linea', 'keys', label))
    for line, n, s_ in cambios:
        print('{:<16} {:>8} {:>16,.0f}'.format(line, n, s_))
    print('{:<16} {:>8} {:>16,.0f}'.format('TOTAL', '', total_escrito))

    if conflictos:
        print()
        print('OJO: {} key(s) ya tenian un valor DISTINTO en {} (se sobreescribe):'.format(len(conflictos), label))
        for line, k, prev, v in conflictos[:20]:
            print('   {:<14} {:<22} antes={:>10,} ahora={:>10,}'.format(line, str(k)[:22], int(prev or 0), v))
    if ya_estaba:
        print()
        print('{} key(s) ya tenian el MISMO valor (idempotente)'.format(len(ya_estaba)))
    if sin_dato:
        print()
        print('{} key(s) sin dato en la planilla para {} (se dejan como estaban):'.format(len(sin_dato), label))
        for line, k in sin_dato[:25]:
            print('   {:<14} {}'.format(line, k))

    if a.dry_run:
        print()
        print('DRY RUN: no se escribio nada.')
        return 0

    print()
    for p, text, s, e, D in planned:
        p.write_text(text[:s] + json.dumps(D, ensure_ascii=False) + text[e:],
                     encoding='utf-8', newline='')
        print('-> {} ({:,} bytes)'.format(p.relative_to(REPO), p.stat().st_size))
    return 0


if __name__ == '__main__':
    sys.exit(main())
