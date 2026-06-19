#!/usr/bin/env python3
"""
shared/merge-ventas-internas.py

Actualiza budget[fam].YYYY.real (venta interna) en todas las lineas
desde un xlsx de Planilla de Ventas con formato:

  Familia | Ene-YYYY | Feb-YYYY | ... | Dic-YYYY+1

Acepta tanto el formato corto (1 fila de header con meses) como
el largo (2 filas de header). Detecta automaticamente.

Solo actualiza venta interna (real). NO toca presupuesto (budget).
NO toca rec_ms, rec_comp, recetas, mol_perf, stock, etc.

Familias se matchean por nombre exacto contra budget keys de cada
linea. Para mujer (que usa segmentos como 'SIN ESTROGENO' en el
inline D), aplica el mapeo brand->segment definido abajo.

Uso:
    py shared/merge-ventas-internas.py [--file <path>] [--cutoff YYYY-MM] [--dry-run]

  --cutoff: ultimo mes cerrado a incluir. Meses posteriores se
            ignoran (data parcial). Ej: --cutoff 2026-04 ignora May+.
"""

from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
DEFAULT_FILE = Path(r'C:\Users\camarinaro\Downloads\Planilla de Ventas - 4 de mayo de 2026.xlsx')

# Map mes-año en Excel -> (year, month_idx 0..11)
MES_ES = {'ene':0,'feb':1,'mar':2,'abr':3,'may':4,'jun':5,
          'jul':6,'ago':7,'sep':8,'sept':8,'oct':9,'nov':10,'dic':11}

# Lineas: archivo a actualizar y donde esta budget
LINES = [
    {'key':'cardio',  'kind':'data.js', 'path':'cardio/data.js'},
    {'key':'antibio', 'kind':'data.js', 'path':'ATB/data.js'},
    {'key':'OTC',     'kind':'data.js', 'path':'OTC/data.js'},
    {'key':'respi',   'kind':'data.js', 'path':'respiratorio/data.js'},
    {'key':'SNC',     'kind':'inline',  'path':'SNC/index.html'},
    {'key':'mujer',   'kind':'inline',  'path':'mujer/index.html'},
    {'key':'dermato', 'kind':'inline',  'path':'dermatologia/dermato_dashboard.html'},
]

# Mapeo de segmento (mujer inline D) -> familia(s) del Excel
# El budget de mujer inline D suma todos los brands del segmento.
MUJER_SEGMENT_TO_FAMS = {
    'SIN ESTROGENO':   ['ISIS FREE'],
    'ALTA DOSIS':      ['ISIS'],
    'BAJA DOSIS 21+7': ['ISIS MINI'],
    'BAJA DOSIS 24':   ['ISIS MINI 24'],
    'COMPLEX':         ['SIDERBLUT COMPLEX', 'SIDERBLUT FOLIC'],
    'SOLO':            ['SIDERBLUT', 'SIDERBLUT POLI', 'FERINSOL'],
    # TRIP (D3, D3 PLUS, +45, MAGNESIO): en la planilla las 4 variantes comparten
    # Gran Familia=Familia=Producto='TRIP'; SOLO se distinguen por Presentacion (col3).
    # El merge agrupa por Familia -> NO puede separarlas. Por eso van VACIAS aca y se
    # corrigen con shared/fix-mujer-trip-venta.py (clasifica por Presentacion). Con []
    # el merge las SALTA (no las pisa), asi los valores del corrector PERSISTEN.
    # OJO: NO mapear 'D3' a 'TRIP' -> se traga todo TRIP (bug "TRIP 45 en 0", jun-2026).
    'D3':              [],
    'D3 PLUS':         [],
    '45':              [],
    'MAGNESIO':        [],
    'DELTROX':         ['DELTROX'],
    'BASE':            ['CALCIO BASE DUPOMAR'],
    'BASE D':          ['CALCIO BASE DUPOMAR D', 'CALCIO BASE DUPOMAR D3',
                        'CALCIO CITRATO DUPOMAR D3'],
    'CLIMATIX':        ['CLIMATIX'],
}


# Aliases: budget keys cuyo nombre difiere de la 'Familia' en la planilla.
# Formato: { budget_key_en_data : nombre_familia_en_planilla }
KEY_ALIASES = {
    'HEXALER BRONQUIAL DU': 'HEXALER BRONQUIAL DUO',   # a la key le falta la 'O'
}


def parse_xlsx(path, cutoff=None):
    """Devuelve dict[familia] = {(year, month_idx): value}.
    cutoff = (year, month_idx) o None. Meses posteriores se ignoran."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    # Detectar fila con headers de meses (a veces hay 2 filas, queremos
    # la que tenga "Ene-YYYY", "Feb-YYYY", etc.)
    header_row = None
    data_start = 2
    for ri in (1, 2):
        row = list(next(ws.iter_rows(min_row=ri, max_row=ri, values_only=True)))
        if not row: continue
        # ¿Tiene al menos un header tipo Mes-YYYY?
        if any(re.match(r'\w+[\s\-/]\d{4}', str(h).strip()) for h in row if h):
            header_row = row
            data_start = ri + 1
            break
    if header_row is None:
        header_row = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        data_start = 2

    col_to_ym = {}  # col_idx -> (year, month_idx 0..11)
    for i, h in enumerate(header_row):
        if not h or i == 0: continue
        s = str(h).strip()
        m = re.match(r'(\w+)[\s\-/](\d{4})', s)
        if not m: continue
        mes = m.group(1).lower().rstrip('.')
        year = int(m.group(2))
        midx = MES_ES.get(mes)
        if midx is None: continue
        # Aplicar cutoff
        if cutoff is not None:
            cy, cm = cutoff
            if (year, midx) > (cy, cm):
                continue
        col_to_ym[i] = (year, midx)

    # Detectar si hay columna 'Familia' separada de 'Gran Familia'.
    # Formato SAP nuevo: col0=Gran Familia, col1=Familia, col2=Producto, ...
    # Formato viejo: col0=Familia, col1=primer mes. Si la col 1 es un mes,
    # NO hay columna Familia separada -> f = g.
    has_familia_col = (1 not in col_to_ym)

    # by_col0: {gran_familia: {(year,midx): val}}  (suma de todas las filas)
    # pairs:   {(gran_familia, familia): {(year,midx): val}}  (granularidad fina)
    by_col0 = defaultdict(lambda: defaultdict(int))
    pairs   = defaultdict(lambda: defaultdict(int))
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row or len(row) < 1 or not row[0]: continue
        g = str(row[0]).strip()
        if not g: continue
        # Skip header-like rows
        if g.lower() in ('familia', 'familias', 'family', 'gran familia'): continue
        f = (str(row[1]).strip() if has_familia_col and len(row) > 1 and row[1] else g)
        if not f: f = g
        for ci, ym in col_to_ym.items():
            if ci >= len(row): continue
            v = row[ci]
            if v is None: continue
            try:
                iv = int(round(float(v)))
            except (ValueError, TypeError): continue
            by_col0[g][ym] += iv
            pairs[(g, f)][ym] += iv
    wb.close()
    by_col0 = {k: dict(v) for k, v in by_col0.items()}
    pairs   = {k: dict(v) for k, v in pairs.items()}
    years_seen = sorted({y for d in by_col0.values() for (y, _) in d})
    return pairs, by_col0, years_seen


def load_data_js(p):
    text = p.read_text(encoding='utf-8-sig', errors='replace')
    m1 = re.search(r'window\.OTC_DATA\s*=\s*', text)
    if not m1: raise ValueError('OTC_DATA not found in ' + str(p))
    obj_start1 = text.index('{', m1.end())
    d1, end1 = json.JSONDecoder().raw_decode(text[obj_start1:])
    abs_end1 = obj_start1 + end1
    m2 = re.search(r'window\.OTC_DASHBOARD\s*=\s*', text[abs_end1:])
    if not m2: return text, d1, None, None
    obj_start2 = abs_end1 + text[abs_end1:].index('{', m2.end())
    d2, end2 = json.JSONDecoder().raw_decode(text[obj_start2:])
    abs_end2 = obj_start2 + end2
    return text, d1, d2, (obj_start1, abs_end1, obj_start2, abs_end2)


def write_data_js(text, d1, d2, bounds):
    obj_start1, abs_end1, obj_start2, abs_end2 = bounds
    return (text[:obj_start1]
            + json.dumps(d1, ensure_ascii=False)
            + text[abs_end1:obj_start2]
            + json.dumps(d2, ensure_ascii=False)
            + text[abs_end2:])


def load_inline(p):
    text = p.read_text(encoding='utf-8', errors='replace')
    m = re.search(r'const D = (\{)', text)
    if not m: raise ValueError('const D not found in ' + str(p))
    obj_start = m.start() + len('const D = ')
    obj_start = text.index('{', obj_start)
    D, end = json.JSONDecoder().raw_decode(text[obj_start:])
    abs_end = obj_start + end
    return text, D, obj_start, abs_end


def write_inline(text, D, abs_start, abs_end):
    return text[:abs_start] + json.dumps(D, ensure_ascii=False) + text[abs_end:]


def _write_real(budget, budget_key, new_data, years_seen):
    """Escribe new_data {(year,midx):val} en budget[key].real, por año.
    NO toca budget (estimado)."""
    for year in years_seen:
        yk = str(year)
        year_obj = budget[budget_key].setdefault(yk, {})
        real_arr = year_obj.get('real')
        if not isinstance(real_arr, list) or len(real_arr) != 12:
            real_arr = [None] * 12
        for (y, midx), v in new_data.items():
            if y == year:
                real_arr[midx] = v
        year_obj['real'] = real_arr
        year_obj.setdefault('budget', [0]*12)


def update_budget(budget, pairs, by_col0, years_seen, line_key):
    """Actualiza budget[*].real desde la planilla.

    - mujer: usa by_col0 + MUJER_SEGMENT_TO_FAMS (comportamiento historico).
    - resto: asigna cada fila (Gran Familia g, Familia f) a su budget key MAS
      ESPECIFICA: si f es un budget key -> f; si no, si g es un budget key -> g.
      Asi una sub-familia (TETRALGIN NOVO, BACTRIM FORTE, DILATREND AP, ...) va
      a su propia key, y las filas restantes de la Gran Familia (incluidas
      variantes sin key propia, p.ej. TETRALGIN APC) quedan en la key padre.
    """
    updated, unmatched = 0, []

    if line_key == 'mujer':
        # IMPORTANTE: matchear por FAMILIA (col1), NO por Gran Familia (col0).
        # En la planilla, col0 'ISIS' agrupa TODAS las variantes (ISIS, ISIS FREE,
        # ISIS MINI, ISIS MINI 24). Los targets del map (ISIS, ISIS FREE, ...) son
        # nombres de Familia (col1). Si se matchea por col0, ALTA DOSIS (target
        # Familia 'ISIS') absorbe TODO el ISIS (~118k) y el resto queda en 0,
        # dando %Cumpl absurdos (705%). Por col1, ALTA DOSIS = ISIS alta dosis (~17k).
        by_col1 = defaultdict(lambda: defaultdict(int))
        for (g, f), vals in pairs.items():
            for ym, v in vals.items():
                by_col1[f][ym] += v
        for budget_key in list(budget.keys()):
            if budget_key not in MUJER_SEGMENT_TO_FAMS:
                unmatched.append(budget_key); continue
            target_fams = MUJER_SEGMENT_TO_FAMS[budget_key]
            if not target_fams:
                continue  # segmento sin mapeo, skip silencioso
            sum_data = defaultdict(int)
            had_any = False
            for tf in target_fams:
                src = by_col1.get(tf) or by_col0.get(tf)  # preferir Familia (col1)
                if src:
                    had_any = True
                    for ym, v in src.items():
                        sum_data[ym] += v
            if not had_any:
                unmatched.append(budget_key); continue
            _write_real(budget, budget_key, dict(sum_data), years_seen)
            updated += 1
        return updated, unmatched

    # Resto de lineas: asignacion mas-especifica (alias -> f -> g) por fila
    line_keys = set(budget.keys())
    # alias por familia: { familia_en_planilla : budget_key }
    alias_fam_to_key = {fam: k for k, fam in KEY_ALIASES.items() if k in line_keys}
    acc = defaultdict(lambda: defaultdict(int))
    for (g, f), vals in pairs.items():
        key = alias_fam_to_key.get(f) or (f if f in line_keys else (g if g in line_keys else None))
        if key is None:
            continue
        for ym, v in vals.items():
            acc[key][ym] += v
    for budget_key in list(budget.keys()):
        if budget_key in acc:
            _write_real(budget, budget_key, dict(acc[budget_key]), years_seen)
            updated += 1
        else:
            unmatched.append(budget_key)
    return updated, unmatched


def cap_post_cutoff(budget, cutoff):
    """Mayo cerrado: pone en None la venta real POSTERIOR al mes de corte en
    TODAS las familias (matcheen o no). El --cutoff evita ESCRIBIR meses
    parciales, pero un valor viejo (p.ej. un junio cargado antes) queda; esto
    lo limpia. Asi ninguna linea muestra venta mas alla del cierre. Devuelve
    cuantos valores limpio."""
    cy, cm = cutoff
    n = 0
    for fam, yrs in budget.items():
        if not isinstance(yrs, dict): continue
        for yk, yobj in yrs.items():
            if not isinstance(yobj, dict): continue
            try: y = int(yk)
            except (ValueError, TypeError): continue
            real = yobj.get('real')
            if not isinstance(real, list): continue
            for i in range(len(real)):
                if ((y > cy) or (y == cy and i > cm)) and real[i] not in (None, 0):
                    real[i] = None; n += 1
    return n


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--file', default=str(DEFAULT_FILE))
    ap.add_argument('--cutoff', help="Ultimo mes cerrado, formato 'YYYY-MM'. Meses posteriores se ignoran.")
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    fp = Path(args.file)
    if not fp.is_file():
        print(f'ERROR: archivo no existe: {fp}', file=sys.stderr); return 2

    cutoff = None
    if args.cutoff:
        m = re.match(r'(\d{4})-(\d{2})', args.cutoff)
        if not m:
            print(f'ERROR: --cutoff debe ser YYYY-MM, recibido: {args.cutoff}', file=sys.stderr); return 2
        cutoff = (int(m.group(1)), int(m.group(2)) - 1)
        print(f'Cutoff: incluir hasta {args.cutoff} (mes idx {cutoff[1]})')

    print(f'Leyendo: {fp}')
    pairs, by_col0, years_seen = parse_xlsx(fp, cutoff=cutoff)
    print(f'  {len(by_col0)} familias (col0), {len(pairs)} pares (gran familia, familia), años: {years_seen}')

    for line in LINES:
        path = REPO / line['path']
        if not path.is_file():
            print(f'  [{line["key"]}] SKIP: no existe {path}'); continue
        try:
            if line['kind'] == 'data.js':
                text, d1, d2, bounds = load_data_js(path)
                if d2 is None or 'budget' not in d2:
                    print(f'  [{line["key"]}] SKIP: no OTC_DASHBOARD.budget'); continue
                budget = d2['budget']
            else:
                text, D, abs_start, abs_end = load_inline(path)
                if 'budget' not in D:
                    print(f'  [{line["key"]}] SKIP: no D.budget'); continue
                budget = D['budget']

            updated, unmatched = update_budget(budget, pairs, by_col0, years_seen, line['key'])
            capped = cap_post_cutoff(budget, cutoff) if cutoff else 0

            if args.dry_run:
                print(f'  [{line["key"]}] DRY: actualizaria {updated} familias, sin match: {len(unmatched)}, limpiaria {capped} valores post-cierre'.encode('ascii','replace').decode())
                continue

            if line['kind'] == 'data.js':
                new_text = write_data_js(text, d1, d2, bounds)
            else:
                new_text = write_inline(text, D, abs_start, abs_end)
            path.write_text(new_text, encoding='utf-8', newline='')
            print(f'  [{line["key"]}] OK: {updated} familias actualizadas, {len(unmatched)} sin match, {capped} valores post-cierre limpiados'.encode('ascii','replace').decode())
            if unmatched and len(unmatched) <= 5:
                print(f'    sin match: {unmatched}')
        except Exception as e:
            print(f'  [{line["key"]}] ERROR: {e}')

    return 0


if __name__ == '__main__':
    sys.exit(main())
