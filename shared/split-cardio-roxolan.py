"""Aperturra ROXOLAN (rosuvastatina mono) y ROXOLAN PLUS (rosuvastatina+
ezetimibe) en cardio, en las 3 capas: Venta Interna, Mercado IQVIA y Recetas.

BUG corregido: el dashboard fusionaba en una sola familia 'ROXOLAN' dos
mercados que la fuente IQVIA tiene SEPARADOS:
  - 'Roxolan (Rosuvastatina)'  -> ROSUVASTATIN mono
  - 'Roxolan Plus'             -> ROSUVASTATIN + EZETIMIBE
Esto inflaba el mercado y el MS% de ROXOLAN. Todas las demas marcas con combo
(EMPAX/EMPAX MET, SILTRAN/SILTRAN MET, etc.) ya estaban separadas.

Clasificacion mono/plus verificada 100% contra cardio/DDD/competidores-data.js
(mercados 'Roxolan (Rosuvastatina)' [39] y 'Roxolan Plus' [13]).
Combo = nombre contiene PLUS / DUO / EZ / COMBICOL / EZETIM.

Preserva historia (reparte productos existentes). Idempotente.
Despues correr: recompute-mol-perf-aggregates, fix-brandkpis-from-molperf
(+esqueleto), build-kpis, build-families-perf, sync-kpistrip, audit-full.
"""
from __future__ import annotations
import json, re, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
DATA = REPO / 'cardio' / 'data.js'
PLANILLA = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs\Planilla de Ventas - 2 de junio de 2026.xlsx')
CUTOFF = (2026, 4)  # incluir hasta May 2026 (idx 4)

MES_INV = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
MES_ES = {'ene':0,'feb':1,'mar':2,'abr':3,'may':4,'jun':5,'jul':6,'ago':7,'sep':8,'sept':8,'oct':9,'nov':10,'dic':11}


def is_plus(b):
    s = ' ' + re.sub(r'\s+', ' ', re.sub(r'\([^)]*\)', '', str(b)).upper()).strip() + ' '
    if 'COMBICOL' in s or 'EZETIM' in s: return True
    return any(k in s for k in (' PLUS ', ' DUO ', ' EZ '))


def qkey(mk):
    p = mk.split()
    if len(p) != 2: return None
    m = MES_INV.get(p[0])
    return f'Q{(m-1)//3+1} {p[1]}' if m else None


def parse_budget_split():
    """De la planilla: ROXOLAN (sin PLUS) y ROXOLAN PLUS por presentacion."""
    wb = openpyxl.load_workbook(PLANILLA, read_only=True, data_only=True); ws = wb.active
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    col_ym = {}
    for i, h in enumerate(hdr):
        if not h: continue
        m = re.match(r'(\w+)[\s\-/](\d{4})', str(h).strip())
        if not m: continue
        mi = MES_ES.get(m.group(1).lower().rstrip('.'))
        if mi is None: continue
        y = int(m.group(2))
        if (y, mi) > CUTOFF: continue
        col_ym[i] = (y, mi)
    mono, plus = defaultdict(int), defaultdict(int)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or str(row[0]).strip() != 'ROXOLAN': continue
        pres = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        tgt = plus if 'PLUS' in pres.upper() else mono
        for ci, ym in col_ym.items():
            if ci < len(row) and row[ci] is not None:
                try: tgt[ym] += int(round(float(row[ci])))
                except (ValueError, TypeError): pass
    wb.close()
    return dict(mono), dict(plus)


def main():
    if hasattr(sys.stdout, 'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')
    t = DATA.read_text(encoding='utf-8-sig', errors='replace')
    m = re.search(r'window\.OTC_DASHBOARD\s*=\s*', t)
    ob = t.index('{', m.end())
    D, end = json.JSONDecoder().raw_decode(t[ob:])
    if 'ROXOLAN PLUS' in D.get('mol_perf', {}):
        print('ROXOLAN PLUS ya existe — idempotente, nada que hacer.'); return 0

    # ---- mol_perf ----
    src = D['mol_perf']['ROXOLAN']
    mono_p = [p for p in src['products'] if not is_plus(p['prod'])]
    plus_p = [p for p in src['products'] if is_plus(p['prod'])]
    D['mol_perf']['ROXOLAN'] = {'family':'ROXOLAN','products':mono_p,'monthly':{},'quarterly':{},'ytd':{},'mat':{}}
    D['mol_perf']['ROXOLAN PLUS'] = {'family':'ROXOLAN PLUS','products':plus_p,'monthly':{},'quarterly':{},'ytd':{},'mat':{}}
    print(f'mol_perf: ROXOLAN {len(mono_p)} | ROXOLAN PLUS {len(plus_p)}')

    # ---- maps ----
    for d in ('sieMolMap','molLabels'):
        if isinstance(D.get(d), dict): D[d]['ROXOLAN PLUS'] = 'ROXOLAN PLUS'
    if 'colors' in D: D['colors'].setdefault('ROXOLAN PLUS', '#4d7c0f')  # verde oliva oscuro (ROXOLAN=#65a30d)
    if 'prodMap' in D and 'ROXOLAN' in D['prodMap']:
        D['prodMap']['ROXOLAN PLUS'] = {'mol':'ROXOLAN PLUS','canal':'ROXOLAN','conv':'ROXOLAN','rec':'ROXOLAN PLUS','prec':'ROXOLAN','bud':'ROXOLAN PLUS'}
    if isinstance(D.get('budIqviaMap'), dict):
        D['budIqviaMap']['ROXOLAN'] = ['ROXOLAN (SIE)']
        D['budIqviaMap']['ROXOLAN PLUS'] = ['ROXOLAN PLUS (SIE)']

    # ---- recetas ----
    rc = D['rec_comp']['ROXOLAN']
    rc_mono = {b: bd for b, bd in rc.items() if not is_plus(b)}
    rc_plus = {b: bd for b, bd in rc.items() if is_plus(b)}
    D['rec_comp']['ROXOLAN'] = rc_mono
    D['rec_comp']['ROXOLAN PLUS'] = rc_plus
    print(f'rec_comp: ROXOLAN {len(rc_mono)} | ROXOLAN PLUS {len(rc_plus)}')

    def build_recms(comp, sie_brand):
        mkt = defaultdict(int)
        for b, bd in comp.items():
            for mk, v in (bd.get('monthly') or {}).items(): mkt[mk] += int(v or 0)
        sie = {mk: int(v or 0) for mk, v in ((comp.get(sie_brand) or {}).get('monthly') or {}).items()}
        ms = {mk: round(sie.get(mk,0)/mkt[mk]*100, 2) for mk in mkt if mkt[mk]}
        qs, qm = defaultdict(int), defaultdict(int)
        for mk, v in sie.items():
            k = qkey(mk); qs[k] += v if k else 0
        for mk, v in mkt.items():
            k = qkey(mk); qm[k] += v if k else 0
        msq = {k: round(qs[k]/qm[k]*100, 2) for k in qm if qm[k]}
        return {'sie':sie,'ms':ms,'quarterly':dict(qs),'ms_quarterly':msq,'mkt':dict(mkt)}, dict(mkt)

    rms_mono, mkt_mono = build_recms(rc_mono, 'ROXOLAN SIE')
    rms_plus, mkt_plus = build_recms(rc_plus, 'ROXOLAN PLUS SIE')
    D['rec_ms']['ROXOLAN'] = rms_mono
    D['rec_ms']['ROXOLAN PLUS'] = rms_plus

    orig = D['recetas'].get('ROXOLAN', {})
    rec_mono, rec_plus = {}, {}
    for mk, obj in orig.items():
        tot = mkt_mono.get(mk,0) + mkt_plus.get(mk,0)
        med = (obj or {}).get('medicos',0) or 0
        a, b = mkt_mono.get(mk,0), mkt_plus.get(mk,0)
        rec_mono[mk] = {'recetas':a, 'medicos': int(round(med*a/tot)) if tot else 0}
        rec_plus[mk] = {'recetas':b, 'medicos': int(round(med*b/tot)) if tot else 0}
    D['recetas']['ROXOLAN'] = rec_mono
    D['recetas']['ROXOLAN PLUS'] = rec_plus

    # ---- budget (Venta Interna) ----
    bmono, bplus = parse_budget_split()
    years = sorted({y for (y, _) in list(bmono) + list(bplus)})
    def wbud(key, data):
        D['budget'].setdefault(key, {})
        for year in years:
            yo = D['budget'][key].setdefault(str(year), {})
            arr = yo.get('real')
            if not isinstance(arr, list) or len(arr) != 12: arr = [None]*12
            for (y, mi), v in data.items():
                if y == year: arr[mi] = v
            yo['real'] = arr; yo.setdefault('budget', [0]*12)
    wbud('ROXOLAN', bmono)
    wbud('ROXOLAN PLUS', bplus)
    print(f'budget 2026: ROXOLAN {[bmono.get((2026,i)) for i in range(5)]} | PLUS {[bplus.get((2026,i)) for i in range(5)]}')

    DATA.write_text(t[:ob] + json.dumps(D, ensure_ascii=False) + t[ob+end:], encoding='utf-8', newline='')
    print('cardio/data.js actualizado: ROXOLAN / ROXOLAN PLUS separados (IQVIA+Recetas+VentaInterna).')
    return 0


if __name__ == '__main__':
    sys.exit(main() or 0)
