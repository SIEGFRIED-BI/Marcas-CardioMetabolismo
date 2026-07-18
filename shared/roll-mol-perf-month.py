# -*- coding: utf-8 -*-
"""Rueda el mol_perf (Mercado IQVIA) de una linea al ultimo mes del master, de forma
QUIRURGICA: APPEND del mes nuevo a cada producto (no reescribe la historia) + recomputo
del RESTO ("Otros") = mercado_mes − emitidos. Preserva ventana, split de familias, y todo
lo demas (venta/recetas/DDD/budget intactos). No usa PowerShell (evita el serializer).

A diferencia de update-iqvia-month.py: (1) recomputa el RESTO, (2) hace APPEND del mes
nuevo (no mete los 61 meses del master ni expande la ventana), (3) verifica el mercado
derivado contra el actual por familia (red de seguridad).

Mercado por familia = productos del master cuya MOLECULA esta en el set de moleculas de los
productos EMITIDOS (no-RESTO) de esa familia. Esto separa limpio ROXOLAN (ROSUVASTATIN) de
ROXOLAN PLUS (EZETIMIBE_ROSUVASTATIN) y funciona igual en lineas keyeadas por marca
(cardio/ATB/OTC/respi/mujer) o por molecula (SNC/dermato).

Despues correr: recompute-mol-perf-aggregates --cierre <YYYY-MM>.
Uso: py shared/roll-mol-perf-month.py --line cardio --master <AR_PM junio> [--dry-run]
"""
from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
LINE_FILE = {'cardio':'cardio/data.js','ATB':'ATB/data.js','OTC':'OTC/data.js',
             'respiratorio':'respiratorio/data.js','mujer':'mujer/data.js',
             'SNC':'SNC/data.js','dermato':'dermatologia/data.js'}
ANCHORS=[r'window\.OTC_DASHBOARD\s*=\s*', r'window\.MUJER_DATA\s*=\s*', r'const\s+D\s*=\s*']
MES={'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
def msort(mk):
    p=str(mk).split(); return int(p[1])*100+MES.get(p[0],0) if len(p)==2 and p[0] in MES else 0


def load_master(path):
    wb=openpyxl.load_workbook(path,read_only=True,data_only=True); ws=wb.active
    r1=list(next(ws.iter_rows(min_row=1,max_row=1,values_only=True)))
    ci={}
    for i,h in enumerate(r1):
        s=str(h or '').strip().lower()
        if s.startswith('product'): ci['prod']=i
        elif s.startswith('molecules'): ci['mol']=i
    lab=lambda h:str(h).split('\n')[-1].strip() if h else ''
    mcols=[(i,lab(h)) for i,h in enumerate(r1) if h and str(h).startswith('Units') and re.match(r'^[A-Z][a-z]{2} \d{4}$',lab(h))]
    prod_month=defaultdict(lambda:defaultdict(float)); prod_mol={}
    mol_month=defaultdict(lambda:defaultdict(float))
    for row in ws.iter_rows(min_row=2,values_only=True):
        prod=str(row[ci['prod']] or '').strip(); mol=str(row[ci['mol']] or '').strip().upper()
        if not prod: continue
        prod_mol[prod]=mol
        for cidx,mk in mcols:
            if cidx<len(row) and isinstance(row[cidx],(int,float)):
                prod_month[prod][mk]+=row[cidx]; mol_month[mol][mk]+=row[cidx]
    wb.close()
    last=max((mk for _,mk in mcols),key=msort)
    return prod_month, prod_mol, mol_month, last


def is_resto(p): return bool(p.get('is_resto'))


def main():
    ap=argparse.ArgumentParser(); ap.add_argument('--line',required=True); ap.add_argument('--master',required=True)
    ap.add_argument('--dry-run',action='store_true')
    ap.add_argument('--allow-warn',action='store_true',help='escribir aunque haya familias con warning de derivacion (mercado segmentado)')
    a=ap.parse_args()
    if hasattr(sys.stdout,'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')
    f=REPO/LINE_FILE[a.line]; text=f.read_text(encoding='utf-8-sig',errors='replace')
    anc=None
    for pat in ANCHORS:
        for m in re.finditer(pat,text):
            ob=text.index('{',m.end())
            try: obj,_=json.JSONDecoder().raw_decode(text[ob:])
            except Exception: continue
            if isinstance(obj,dict) and obj.get('mol_perf'): anc=(m,ob); break
        if anc: break
    m,ob=anc; D,end=json.JSONDecoder().raw_decode(text[ob:])
    prod_month,prod_mol,mol_month,new_month=load_master(Path(a.master))
    mp=D['mol_perf']
    # mes actual (ultimo de la linea)
    cur_last=max((mk for fo in mp.values() for p in fo.get('products',[]) for mk in (p.get('monthly_vals') or {})),key=msort,default='?')
    print(f'{a.line}: mol_perf {len(mp)} fam | mes actual {cur_last} -> nuevo {new_month}')
    if new_month==cur_last:
        print('  master no es mas nuevo que la linea; nada que rodar.'); return 0
    warns=0; dropped_all=[]
    for fam,fo in mp.items():
        prods=fo.get('products') or []
        emitted=[p for p in prods if not is_resto(p)]
        resto=next((p for p in prods if is_resto(p)),None)
        # APPEND nuevo mes a cada producto nombrado, por nombre exacto en el master.
        # Si el producto NO esta en el master: no escribo (lo dejo ausente ese mes) y lo
        # cuento como "dropeado" solo si tenia valor el mes actual (drift de nombre a revisar).
        emit_new=0; dropped=[]
        for p in emitted:
            if p['prod'] in prod_month:
                v=int(round(prod_month[p['prod']].get(new_month,0)))
                (p.setdefault('monthly_vals',{}))[new_month]=v; emit_new+=v
            else:
                cur=float((p.get('monthly_vals') or {}).get(cur_last,0) or 0)
                if cur!=0: dropped.append(p['prod']); dropped_all.append((fam,p['prod'],int(round(cur))))
        # RESTO ("Otros") solo si la familia lo tiene: mercado derivado por molecula − emitidos.
        mkt=defaultdict(float); ratio=None; der=cur_mkt=0
        if resto is not None:
            mols={prod_mol[p['prod']] for p in emitted if prod_mol.get(p['prod'])}
            for mol in mols:
                for mk,v in mol_month.get(mol,{}).items(): mkt[mk]+=v
            cur_mkt=(fo.get('monthly') or {}).get(cur_last,0); der=mkt.get(cur_last,0)
            ratio=(der/cur_mkt) if cur_mkt else 0
            if not(cur_mkt and 0.9<=ratio<=1.1): warns+=1
            resto.setdefault('monthly_vals',{})[new_month]=max(int(round(mkt.get(new_month,0)))-emit_new,0)
        if dropped: warns+=1
        # resumen por familia
        if resto is not None:
            rflag='' if (ratio is not None and 0.9<=ratio<=1.1) else '  <-- REVISAR mkt'
            print(f'  {fam:18s} [resto] mkt {cur_last}={int(round(der))} (act {int(round(cur_mkt))}, r={ratio:.2f}) | {new_month}: mkt={int(round(mkt.get(new_month,0)))} emit={emit_new} resto={int(round(mkt.get(new_month,0)))-emit_new}{rflag}')
        else:
            dflag=f'  <-- {len(dropped)} DROPEADO(S)' if dropped else ''
            print(f'  {fam:18s} [enum {len(emitted)}p] {new_month}: total={emit_new}{dflag}')
    if dropped_all:
        print(f'\nProductos con {cur_last}>0 ausentes del master {new_month} (se pierden si no se resuelve el nombre):')
        for fam,nm,v in sorted(dropped_all,key=lambda x:-x[2])[:20]:
            print(f'   {fam:16s} {nm[:44]:44s} {cur_last}={v}')
    print(f'\nWarnings totales (mkt segmentado + productos dropeados): {warns}')
    if a.dry_run:
        print('DRY-RUN: nada escrito.'); return 0
    if warns>0 and not a.allow_warn:
        print(f'ABORTADO: {warns} warning(s) (mercado segmentado con RESTO mal derivado, o '
              'productos nombrados que se dropearian por drift de nombre). No escribo para no '
              'corromper. Resolver esas familias aparte (p.ej. sync-<linea>-pm.py) o --allow-warn '
              'si estas seguro.'); return 2
    f.write_text(text[:ob]+json.dumps(D,ensure_ascii=False)+text[ob+end:],encoding='utf-8',newline='')
    print(f'Escrito {f}. Ahora: py shared/recompute-mol-perf-aggregates.py --cierre <YYYY-MM>')
    return 0


if __name__=='__main__':
    sys.exit(main())
