"""Aperturra MAGNUS en MAGNUS (sildenafil) y MAGNUS 36 (tadalafil) en la
VENTA INTERNA de OTC.

La planilla SAP agrupa todo bajo Gran Familia 'MAGNUS'; el "36" solo aparece
en la columna Presentacion. Este script lee la planilla, separa por
presentacion ('MAGNUS 36' -> MAGNUS 36; el resto -> MAGNUS) y escribe en
OTC/data.js:
  - budget['MAGNUS']     = filas SIN '36'   (sildenafil)
  - budget['MAGNUS 36']  = filas CON '36'   (tadalafil)  [se crea si no existe]

Para los meses que NO estan en la planilla (anteriores a Jun-2025), MAGNUS
queda con su valor combinado historico y MAGNUS 36 = null (no hay detalle
por SKU para separarlos).

Idempotente. Correr DESPUES de merge-ventas-internas.py.
Uso: py shared/apply-otc-magnus-split.py [--file <xlsx>] [--cutoff YYYY-MM] [--dry-run]
"""
from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
DATA = REPO / 'OTC' / 'data.js'
DEFAULT_FILE = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs\Planilla de Ventas - 2 de junio de 2026.xlsx')

MES_ES = {'ene':0,'feb':1,'mar':2,'abr':3,'may':4,'jun':5,
          'jul':6,'ago':7,'sep':8,'sept':8,'oct':9,'nov':10,'dic':11}


def parse_magnus(path, cutoff=None):
    """Devuelve (sin36, con36) cada uno {(year, midx): val} sumado por presentacion."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    col_to_ym = {}
    for i, h in enumerate(hdr):
        if not h: continue
        m = re.match(r'(\w+)[\s\-/](\d{4})', str(h).strip())
        if not m: continue
        midx = MES_ES.get(m.group(1).lower().rstrip('.'))
        if midx is None: continue
        y = int(m.group(2))
        if cutoff and (y, midx) > cutoff: continue
        col_to_ym[i] = (y, midx)
    # col0=Gran Familia, col3=Presentacion (formato SAP)
    sin36 = defaultdict(int); con36 = defaultdict(int)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]: continue
        if str(row[0]).strip() != 'MAGNUS': continue
        pres = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        tgt = con36 if 'MAGNUS 36' in pres.upper() else sin36
        for ci, ym in col_to_ym.items():
            if ci >= len(row): continue
            v = row[ci]
            if v is None: continue
            try: tgt[ym] += int(round(float(v)))
            except (ValueError, TypeError): pass
    wb.close()
    return dict(sin36), dict(con36)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--file', default=str(DEFAULT_FILE))
    ap.add_argument('--cutoff', help='YYYY-MM ultimo mes cerrado')
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()
    if hasattr(sys.stdout, 'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')

    fp = Path(args.file)
    if not fp.is_file():
        print(f'ERROR: no existe {fp}', file=sys.stderr); return 2
    cutoff = None
    if args.cutoff:
        m = re.match(r'(\d{4})-(\d{2})', args.cutoff)
        cutoff = (int(m.group(1)), int(m.group(2)) - 1)

    sin36, con36 = parse_magnus(fp, cutoff)
    NUM = {0:'Ene',1:'Feb',2:'Mar',3:'Abr',4:'May',5:'Jun',6:'Jul',7:'Ago',8:'Sep',9:'Oct',10:'Nov',11:'Dic'}
    print('MAGNUS (sin 36) 2026:', {NUM[mi]: v for (y,mi), v in sorted(sin36.items()) if y==2026})
    print('MAGNUS 36       2026:', {NUM[mi]: v for (y,mi), v in sorted(con36.items()) if y==2026})

    # Cargar OTC/data.js
    t = DATA.read_text(encoding='utf-8-sig', errors='replace')
    m = re.search(r'window\.OTC_DASHBOARD\s*=\s*', t)
    ob = t.index('{', m.end())
    D, end = json.JSONDecoder().raw_decode(t[ob:])
    budget = D['budget']

    years = sorted({y for (y, _) in list(sin36) + list(con36)})

    def write_key(key, data):
        if key not in budget:
            budget[key] = {}
        for year in years:
            yk = str(year)
            yo = budget[key].setdefault(yk, {})
            arr = yo.get('real')
            if not isinstance(arr, list) or len(arr) != 12:
                arr = [None] * 12
            for (y, midx), v in data.items():
                if y == year:
                    arr[midx] = v
            yo['real'] = arr
            yo.setdefault('budget', [0]*12)

    # MAGNUS = sin36 (reemplaza el combinado en los meses de la planilla);
    # los meses fuera de la planilla quedan con su valor historico.
    write_key('MAGNUS', sin36)
    write_key('MAGNUS 36', con36)

    if args.dry_run:
        print('DRY RUN: no se escribio.'); return 0

    new_t = t[:ob] + json.dumps(D, ensure_ascii=False) + t[ob + end:]
    DATA.write_text(new_t, encoding='utf-8', newline='')
    print(f'OTC/data.js actualizado: budget MAGNUS / MAGNUS 36 separados.')
    return 0


if __name__ == '__main__':
    sys.exit(main() or 0)
