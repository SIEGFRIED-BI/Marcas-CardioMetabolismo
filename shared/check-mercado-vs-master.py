#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""G1 del MERCADO: cierra el total de cada familia de mol_perf contra el master.

POR QUE EXISTE. `check-molperf-vs-master.py` concilia las marcas SIEGFRIED contra el
AR_PM, y eso deja un hueco: el DENOMINADOR. En Jul-2026 el mercado de 11 familias de
cardio quedo inflado (DIOVAN x1,90, TERLOC x1,98, SILTRAN x1,84, ROXOLAN x1,16) porque
build-data.ps1 matcheaba la molecula por SUBSTRING y los combos entraban al mercado
mono: 'HYDROCHLOROTHIAZIDE_VALSARTAN'.Contains('VALSARTAN') = True. Las unidades SIE
estaban perfectas -- se movia solo el mercado, o sea el MS% de toda la linea.

Ningun gate lo veia:
  - sum(productos) == total de familia cerraba EXACTO (sobre el universo equivocado);
  - verify-history-preserved solo mira que no falten meses;
  - check-molperf-vs-master solo miraba el numerador.
Lo caza este: compara el total publicado de la familia contra la suma del master
filtrando por IGUALDAD de molecula, que es lo que el config declara.

De donde sale el mapa familia -> molecula: del propio config de cada build-data.ps1
(`'DIOVAN' = @{ ... filters = @(@{ molecules = @('VALSARTAN') ... `). Asi el gate no
duplica la definicion: si alguien cambia el config, el gate lo sigue.

Familias que comparten molecula (DILATREND / DILATREND AP, BACTRIM / BACTRIM FORTE,
CEFALEXINA ARG / ARG DUO) publican las dos el mismo mercado a proposito: cada una da
ratio ~1 contra el mismo total y se listan aparte, no fallan.

Uso:  python shared/check-mercado-vs-master.py [--month "Jul 2026"] [--tol 0.02]
Exit != 0 si alguna familia se desvia mas de --tol.
"""
from __future__ import annotations
import argparse, io, json, os, re, sys
from collections import defaultdict
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
# Solo las lineas cuyo build filtra por `molecules` (Test-TextEqualsAny).
BUILDERS = {'cardio': ('cardio/build-data.ps1', 'cardio/data.js'),
            'ATB': ('ATB/build-data.ps1', 'ATB/data.js'),
            'respiratorio': ('respiratorio/build-data.ps1', 'respiratorio/data.js')}
MES = ('Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
       'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec')


def cfg_molecules(ps1):
    """familia -> [moleculas] leidas del literal de config del build."""
    txt = io.open(REPO / ps1, encoding='utf-8-sig', errors='replace').read()
    out = {}
    for m in re.finditer(r"^\s*'([^']+)'\s*=\s*@\{\s*label\s*=.*$", txt, re.M):
        vals = []
        for g in re.findall(r"molecules\s*=\s*@\(([^)]*)\)", m.group(0)):
            vals += [x.strip().strip("'").upper() for x in g.split(',') if x.strip()]
        if vals:
            out[m.group(1)] = sorted(set(vals))
    return out


def parse_data_js(path):
    t = (REPO / path).read_text(encoding='utf-8-sig', errors='replace')
    m = re.search(r'window\.OTC_DASHBOARD\s*=\s*', t)
    if not m:
        return None
    ob = t.index('{', m.end())
    return json.JSONDecoder().raw_decode(t[ob:])[0]


def master_por_molecula(path, mes):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    hdr = [str(h).replace('\n', ' ').strip() if h else ''
           for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    try:
        mi = next(i for i, h in enumerate(hdr) if h.strip().lower().startswith('molecules'))
        ci = next(i for i, h in enumerate(hdr) if h.strip() == f'Units {mes}')
    except StopIteration:
        raise SystemExit(f'FATAL: no encuentro Molecules o "Units {mes}" en {path}')
    tot = defaultdict(float)
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or mi >= len(r):
            continue
        v = r[ci] if ci < len(r) else None
        if isinstance(v, (int, float)):
            tot[str(r[mi] or '').strip().upper()] += float(v)
    wb.close()
    return tot


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--master', default=None)
    ap.add_argument('--month', default=None)
    ap.add_argument('--tol', type=float, default=0.02)
    args = ap.parse_args()

    sys.path.insert(0, str(Path(__file__).resolve().parent))
    import manifest
    master = args.master or manifest.resolve_source('iqvia_master')
    mes = args.month
    if not mes:
        cm = manifest.cierre_month()
        y, mm = str(cm).split('-')[:2]
        mes = f'{MES[int(mm) - 1]} {y}'
    print(f'master: {master}\nmes: {mes}   tolerancia: {args.tol:.1%}\n')
    src = master_por_molecula(master, mes)

    ok, bad, compartidas, particiones = [], [], [], []
    for linea, (ps1, djs) in BUILDERS.items():
        cfg = cfg_molecules(ps1)
        D = parse_data_js(djs)
        if not D:
            continue
        por_mol = defaultdict(list)
        for fam, mols in cfg.items():
            por_mol[tuple(mols)].append(fam)

        # Dos familias pueden compartir la MISMA molecula por dos motivos opuestos, y
        # se distinguen por como suman sus ratios:
        #   LISTADO DUPLICADO  cada una publica el mercado ENTERO (ratio ~1 c/u).
        #                      cardio DILATREND / DILATREND AP.
        #   PARTICION          se reparten la molecula (los ratios SUMAN ~1).
        #                      ATB CEFALEXINA ARG 0,421 + ARG DUO 0,579 = 1,000,
        #                      que hace split-atb-cefalexina-duo.py por dosis.
        # Las dos son correctas; lo que NO puede pasar es que sumen mas que la molecula.
        for mols, fams in sorted(por_mol.items()):
            exp = sum(src.get(m, 0.0) for m in mols)
            if not exp:
                continue
            filas = []
            for fam in sorted(fams):
                o = (D.get('mol_perf') or {}).get(fam)
                if not o:
                    continue
                pub = sum(float((p.get('monthly_vals') or {}).get(mes, 0) or 0)
                          for p in o.get('products', []))
                filas.append((linea, fam, pub, exp, pub / exp, len(fams)))
            if not filas:
                continue
            if all(abs(r[4] - 1.0) <= args.tol for r in filas):
                (compartidas if len(filas) > 1 else ok).extend(filas)
            elif len(filas) > 1 and abs(sum(r[4] for r in filas) - 1.0) <= args.tol:
                particiones.extend(filas)
            else:
                bad.extend(filas)

    def dump(t, items):
        if not items:
            return
        print(f'\n{t}')
        for l, f, p, e, r, n in items:
            print(f'  {l:13s} {f:20s} publicado={p:12,.0f}  master={e:12,.0f}  ratio={r:6.3f}')

    print(f'familias conciliadas: {len(ok) + len(bad) + len(compartidas)}')
    dump(f'OK: {len(ok)}', ok[:6])
    if len(ok) > 6:
        print(f'  ... y {len(ok)-6} mas')
    dump(f'MOLECULA COMPARTIDA (cada familia publica el mercado entero): {len(compartidas)}',
         compartidas)
    if particiones:
        tot = sum(r[4] for r in particiones)
        dump(f'PARTICION de la molecula (los ratios suman {tot:.3f}): {len(particiones)}',
             particiones)
    dump(f'FUERA DE TOLERANCIA: {len(bad)}', bad)
    if bad:
        print('\n  Un ratio ~2 en un mercado "mono" suele ser el COMBO colandose:')
        print('  revisar que el filtro de molecula compare por IGUALDAD, no por substring')
        print('  (shared/../build-data.ps1 -> Test-TextEqualsAny). Regla #2 de CLAUDE.md.')

    print('\n' + '=' * 70)
    print(f'check-mercado-vs-master: {len(ok)+len(compartidas)+len(particiones)} OK  '
          f'{len(bad)} FUERA DE TOL')
    print('=' * 70)
    return 1 if bad else 0


if __name__ == '__main__':
    sys.exit(main())
