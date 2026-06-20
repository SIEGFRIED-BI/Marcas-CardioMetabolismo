#!/usr/bin/env python3
"""
shared/build-competidores-shape-a.py

Generaliza el builder Shape A de competidores DDD para cualquier linea.
Lee un xlsx en formato 'Producto-Mol-ATC-provincia' (cols:
RegionCUP, Mercado, Droga, Clase Terapeutica, AñoMes, Codigo Clase,
Codigo Producto, Producto, Unidades) y produce un competidores-data.js
con shape A (brand_monthly[brand][region][N months]).

Aplica a:
  - ATB       <- Hub-Marcas-Inputs/ATB/2026-04/fuentes-originales/DDD ATB.xlsx
  - respi     <- Hub-Marcas-Inputs/respiratorio/2026-04/ddd/Producto-Mol...xlsx
  - mujer     <- ya hecho separadamente (build-mujer-competidores-data.py)
  - OTC, cardio, SNC, dermato: NO aplica (cardio/SNC/dermato ya son
    Shape A nativos; OTC no tiene xlsx fuente)
"""
from __future__ import annotations
import re, json, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
INPUTS = Path(r'C:\Users\camarinaro\OneDrive - Portalcorp\Documentos\Hub-Marcas-Inputs')

# (line, hub_subfolder, out_data_js). El xlsx regional se resuelve por glob del
# mes (Producto-Molecula-ATC-provincia*.xlsx, el mas reciente) -> month-agnostico.
LINE_CFG = [
    ('ATB',          'ATB',         'ATB/DDD/competidores-data.js'),
    ('respiratorio', 'respiratorio','respiratorio/DDD/competidores-data.js'),
    ('OTC',          'OTC',         'OTC/DDD/competidores-data.js'),
    ('dermato',      'dermato',     'dermatologia/competidores-data.js'),
    ('cardio',       'cardio',      'cardio/DDD/competidores-data.js'),
    ('SNC',          'PSQ',         'SNC/DDD/competidores-data.js'),
    # mujer NO va aca: usa mercados PERSONALIZADOS (segmentos de marketing:
    # Isis/Mini/Nat, Trip, Siderblut...) via build-mujer-competidores-data.py.
    # NO regenerar mujer con este builder generico molecula/ATC (colapsa los
    # segmentos custom). Ver memoria iqvia-base-unica-arquitectura.
]

def resolve_regional_xlsx(hub_sub, month):
    """Glob del Producto-Molecula-ATC-provincia*.xlsx mas reciente del mes."""
    base = INPUTS / hub_sub / month
    cands = []
    for sub in ('fuentes-originales', 'ddd', ''):
        d = base / sub if sub else base
        if d.is_dir():
            cands += [p for p in d.glob('Producto-Mol*provincia*.xlsx') if not p.name.startswith('~$')]
    return max(cands, key=lambda p: p.stat().st_mtime) if cands else None

def resolve_lines(month):
    out = []
    for line, hub, rel in LINE_CFG:
        xlsx = resolve_regional_xlsx(hub, month)
        if xlsx is None:
            print(f'  [WARN] {line}: sin regional Producto-Mol*provincia en {hub}/{month} -> skip')
            continue
        out.append((line, xlsx, REPO / rel))
    return out

MONTH_ORDER = {'Ene':1,'Feb':2,'Mar':3,'Abr':4,'May':5,'Jun':6,
               'Jul':7,'Ago':8,'Sep':9,'Oct':10,'Nov':11,'Dic':12}

PACKAGE_TOKENS = {
    'TABL','TABLE','TABLET','TABLETA','TABLETAS',
    'COMP','COMPR','COMPS','COMPRIMIDO','COMPRIMIDOS',
    'CAP','CAPS','CAPSULA','CAPSULAS',
    'AMP','AMPOLLA','AMPOLLAS',
    'JBE','JARABE',
    'INY','INYECTABLE','V.IM','V.IV',
    'SUSP','SUSPENSION',
    'POL','POLVO',
    'LIQ','LIQUIDO',
    'GTAS','GOTAS',
    'SOL','SOLUC','SOLUCION','SOLUCIONES',
    'CR','CREMA','CREMAS',
    'UNG','UNGUENTO',
    'SOB','SOBRE','SOBRES',
    'BOLS','BOLSA',
    'EFER','EFERV','EFERVES','EFERVESCENTE','EFERVESCENTES',
    'DESLEI','DESLEIBLES','DESLEÍBLES',
    'REC','RECUB','RECUBIE','RECUBIERTO','RECUBIERTOS','RECUBIERTAS',
    'IM','PED','AD','ADULT','INF',
    'SPRAY','PVO','PLV',
    'DISP','DISPERS','DISPERSABLE','DISPERSABLES',
    'MAST','GRAG','GRAGEAS',
    'INHAL','INHALACION',
    'OFT','GINEC','VAGINAL','OVUL','OVULO','OVULOS','PESARIO',
    'GEL','EMUL','LOC','LOCION',
    'COLIRIO','COLIR','PARCHE','APOSITO',
    'AERO','AEROSOL',
    'ANILLO','ANILLOS',
}


def extract_brand(producto: str) -> str:
    if not producto: return 'UNKNOWN'
    s = str(producto).strip().upper().replace('\xa0', ' ')
    s = re.sub(r'\s+', ' ', s)
    parts = s.split(' ')
    brand = []
    for i, p in enumerate(parts):
        clean = re.sub(r'^[.,;:()\[\]/\\-]+|[.,;:()\[\]/\\-]+$', '', p)
        if not clean: continue
        if clean == 'X' and i > 0: break
        if clean in PACKAGE_TOKENS: break
        if re.match(r'^[\d.,]+([A-Z%]+)?$', clean):
            if len(brand) >= 1 and re.match(r'^\d{1,2}$', clean):
                brand.append(clean); continue
            break
        if re.match(r'^\d+(\.\d+)?(MG|ML|G|MCG|UI|%).*', clean): break
        brand.append(clean)
    return ' '.join(brand) if brand else (parts[0] if parts else 'UNKNOWN')


# SIE detection patterns per linea (regex case-insensitive)
SIE_PATTERNS_BY_LINE = {
    'ATB': [r'^ACANTEX\b', r'^BACTRIM\b', r'^CEFALEXINA ARG', r'^MACROMAX\b'],
    'respiratorio': [r'^ACEMUK\b', r'^AIREAL\b', r'^ALIDIAL\b', r'^DECADRON\b',
                     r'^DUO-DECADRON\b', r'^HEXALER\b'],
    'OTC': [r'^ACERPES\b', r'^ACI-TIP\b', r'^ALUMPAK\b', r'^ARTRO RED\b',
            r'^FLEXINA\b', r'^MAGNUS\b', r'^TETRALGIN\b'],
    'dermato': [r'^ACNECLIN\b', r'^CLOBESOL\b', r'^MICOMAZOL\b',
                r'^MICROSONA\b', r'^MOMETAX\b', r'^PALDAR\b', r'^ROACCUTAN\b'],
    'cardio': [r'^DAURAN\b', r'^DILATREND\b', r'^DIOVAN\b', r'^EMPAX\b',
               r'^ENTRESTO\b', r'^EXFORGE\b', r'^METGLUCON\b', r'^PIXABAN\b',
               r'^ROXOLAN\b', r'^SILTRAN\b', r'^SINTROM\b', r'^TELPRES\b',
               r'^TERLOC\b'],
    'mujer': [r'^ISIS\b', r'^SIDERBLUT\b', r'^SIDER\b', r'^TRIP\b',
              r'^CALCIO BASE\b', r'^CALCIO CITRATO\b', r'^CLIMATIX\b',
              r'^DELTROX\b', r'^GYNODERM\b', r'^ROXOLAN\b', r'^ALUMPAK\b'],
    'SNC': [r'^VALIUM\b', r'^MADOPAR\b', r'^QTP\b', r'^PGB\b', r'^EMERAL\b',
            r'^LURAP\b', r'^VALQUIR\b', r'^MELERIL\b', r'^LEVITAL\b', r'^VISDON\b'],
}


def is_sie_brand(brand: str, line: str) -> bool:
    pats = SIE_PATTERNS_BY_LINE.get(line, [])
    s = str(brand).upper()
    return any(re.match(p, s) for p in pats)


def month_sort_key(mk: str) -> int:
    parts = mk.split('-')
    return int(parts[1]) * 100 + MONTH_ORDER.get(parts[0], 0)


def _title(s: str) -> str:
    return ' '.join(w.capitalize() for w in str(s).split())


# Overrides curados (semilla del parametrizador): corrigen mislabels del panel
# IQVIA que el agrupamiento automatico molecula/ATC no puede detectar solo.
_OVR_FILE = REPO / 'shared' / 'ddd-competidores-overrides.json'
try:
    OVERRIDES = json.loads(_OVR_FILE.read_text(encoding='utf-8')) if _OVR_FILE.is_file() else {}
except (ValueError, OSError):
    OVERRIDES = {}


def apply_overrides(line: str, molecules: set, atc: str, key: str):
    """Devuelve (key, corrected_atc). Si una regla molecule_regroup de la linea
    matchea (la marca contiene if_molecule y su ATC == in_atc), re-asigna el
    mercado a new_key y reporta corrected_atc para la metadata. Sin match -> sin
    cambios. Cada regla esta justificada por una inconsistencia interna del dato."""
    for rule in OVERRIDES.get('molecule_regroup', {}).get(line, []):
        if rule.get('if_molecule') in molecules:
            in_atc = rule.get('in_atc')
            if not in_atc or atc == in_atc:
                return rule['new_key'], rule.get('corrected_atc')
    return key, None


def build_one(line: str, xlsx: Path, out: Path) -> str:
    """Agrupa competidores HIBRIDO molecula/ATC (no por la columna 'Mercado', que
    mezcla moleculas). Mono-producto -> mercado = su MOLECULA (Droga). Combo (>1
    molecula) -> mercado = su ATC (Codigo Clase Terapeutica), que el panel asigna
    una por combo. Solo se exportan los mercados que contienen una marca SIE; sus
    competidores = todas las marcas del mismo grupo (misma molecula/ATC)."""
    if not xlsx.is_file():
        return f'  [{line}] SKIP: xlsx no existe ({xlsx})'

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    cols = {str(h or '').strip(): i for i, h in enumerate(hdr)}
    c_region = cols.get('RegionCUP', 0)
    c_droga  = cols.get('Droga', 2)
    c_atc    = cols.get('Codigo Clase Terapeutica',
               cols.get('Codigo Clase Terapeutica '.strip(),
               cols.get('Código Clase Terapeutica', None)))
    c_mes    = cols.get('AñoMes', 4)
    c_prod   = cols.get('Producto', 7)
    c_unid   = cols.get('Unidades', 8)

    months_set, regions_set = set(), set()
    units = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))   # brand->region->mes->u
    brand_drogas = defaultdict(set)                                      # brand->{droga}
    brand_atc = defaultdict(lambda: defaultdict(int))                    # brand->atc->count
    brand_is_sie = {}

    n_rows = n_kept = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        n_rows += 1
        if not row: continue
        region = row[c_region] if c_region < len(row) else None
        mes    = row[c_mes]    if c_mes    < len(row) else None
        prod   = row[c_prod]   if c_prod   < len(row) else None
        unid   = row[c_unid]   if c_unid   < len(row) else None
        if not region or not mes or not prod: continue
        region = str(region).strip()
        if region in ('Totales', '-'): continue
        mes = str(mes).strip()
        try: u = int(round(float(unid or 0)))
        except (TypeError, ValueError): u = 0
        if u <= 0: continue
        brand = extract_brand(prod)
        if not brand or brand == 'UNKNOWN': continue
        droga = (str(row[c_droga]).strip().upper() if c_droga is not None and c_droga < len(row) and row[c_droga] else '')
        atc   = (str(row[c_atc]).strip().upper()   if c_atc   is not None and c_atc   < len(row) and row[c_atc]   else '')
        units[brand][region][mes] += u
        if droga: brand_drogas[brand].add(droga)
        if atc:   brand_atc[brand][atc] += 1
        brand_is_sie.setdefault(brand, is_sie_brand(brand, line))
        regions_set.add(region); months_set.add(mes)
        n_kept += 1
    wb.close()

    def main_atc(b):
        d = brand_atc.get(b)
        return max(d, key=d.get) if d else ''

    # Mercado por marca: mono -> molecula ; combo -> ATC (clave con namespace para no colisionar)
    brand_market, meta = {}, {}
    for b in units:
        dz = brand_drogas.get(b, set())
        atc = main_atc(b)
        if len(dz) <= 1:
            key = 'MOL:' + (next(iter(dz)) if dz else b)
        else:
            key = 'ATC:' + (atc if atc else '+'.join(sorted(dz)))
        # Override curado (parametrizador): corrige mislabels del panel que el
        # agrupamiento automatico no detecta (ej: montelukast etiquetado R01B0).
        key, corr_atc = apply_overrides(line, dz, atc, key)
        brand_market[b] = key
        m = meta.setdefault(key, {'mol': set(), 'atc': set()})
        m['mol'] |= dz
        eff_atc = corr_atc or atc
        if eff_atc: m['atc'].add(eff_atc)

    mkt_brands = defaultdict(list)
    for b, k in brand_market.items():
        mkt_brands[k].append(b)
    # solo mercados con al menos una marca SIE
    keep = [k for k in mkt_brands if any(brand_is_sie.get(b) for b in mkt_brands[k])]

    months = sorted(months_set, key=month_sort_key)
    regions = sorted(regions_set, key=lambda r: (r.startswith('_'), r))

    out_obj = {'months': months, 'regions': regions, 'markets': {}}
    used_names = {}
    for key in keep:
        brands = sorted(mkt_brands[key])
        m = meta[key]
        mols = sorted(m['mol']); atcs = sorted(m['atc'])
        sie_in = sorted([b for b in brands if brand_is_sie.get(b)])
        # Nombre del mercado = molecula(s) de la marca SIE PRINCIPAL (la de mas
        # unidades), no la union del grupo (que en clases ATC anchas de combos da
        # nombres monstruosos). Asi el nombre refleja el producto SIE real.
        def _btot(b): return sum(units[b].get(r, {}).get(mk, 0) for r in units[b] for mk in months)
        primary = sorted(sie_in, key=lambda b: -_btot(b))[0] if sie_in else brands[0]
        label_mols = sorted(brand_drogas.get(primary, set()))
        label = '+'.join(_title(x) for x in label_mols) if label_mols else key.split(':', 1)[1]
        if sie_in:
            label = f'{label} ({_title(primary)})'
        # dedup de nombres (raro): apendar ATC
        name = label
        if name in used_names:
            name = f'{label} [{atcs[0] if atcs else used_names[label]}]'
        used_names[label] = (atcs[0] if atcs else '')
        bm = {}
        for b in brands:
            bm[b] = {}
            for region in regions:
                arr = [units[b].get(region, {}).get(mk, 0) for mk in months]
                if any(arr): bm[b][region] = arr
        tm = {}
        for region in regions:
            arr = [sum(units[b].get(region, {}).get(mk, 0) for b in brands) for mk in months]
            if any(arr): tm[region] = arr
        out_obj['markets'][name] = {
            'brands': sie_in,
            'molecules': mols,
            'atc': atcs,
            'brand_monthly': bm,
            'total_monthly': tm,
        }

    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(
        f'window.SFG_COMP_DATA = {json.dumps(out_obj, ensure_ascii=False)};\n',
        encoding='utf-8', newline=''
    )
    return (f'  [{line}] OK rows={n_rows:,} kept={n_kept:,} '
            f'months={len(months)} ({months[0] if months else "-"}..{months[-1] if months else "-"}) '
            f'regions={len(regions)} mercados(SIE)={len(out_obj["markets"])} '
            f'-> {out.name} ({out.stat().st_size:,} bytes)')


def patch_html_loader(line: str, html_path: Path):
    """Make the competidores.html load competidores-data.js instead of ../data.js."""
    if not html_path.is_file():
        return f'  [{line}/html] SKIP (no existe)'
    text = html_path.read_text(encoding='utf-8', errors='replace')
    if '<script src="./competidores-data.js"></script>' in text:
        return f'  [{line}/html] ya carga competidores-data.js'
    new_text = text.replace('<script src="../data.js"></script>',
                             '<script src="./competidores-data.js"></script>', 1)
    if new_text == text:
        return f'  [{line}/html] WARN: no match for ../data.js'
    html_path.write_text(new_text, encoding='utf-8', newline='')
    return f'  [{line}/html] OK (loader updated)'


def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument('--month', default='2026-04', help='Carpeta de mes en Hub-Marcas-Inputs (YYYY-MM)')
    ap.add_argument('--lines', help='Coma-separadas para limitar (ej: cardio,SNC,ATB,respiratorio). Default: todas.')
    a = ap.parse_args()
    LINES = resolve_lines(a.month)
    if a.lines:
        want = {x.strip() for x in a.lines.split(',') if x.strip()}
        LINES = [t for t in LINES if t[0] in want]
    print(f'Build Shape A competidores-data.js (mes {a.month}, {len(LINES)} lineas)...\n')
    for line, xlsx, out in LINES:
        print(build_one(line, xlsx, out))

    print('\nPatch competidores.html loaders...\n')
    for line, _, out in LINES:
        html = out.parent / 'competidores.html'
        print(patch_html_loader(line, html))

    print('\nListo.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
