# -*- coding: utf-8 -*-
"""Refresca D.prec_comp de mujer (seccion "Comparativa de Precios", #s-pcomp) con los
precios del Manual Farmaceutico, matcheando por TROQUEL.

QUE ES ESTA SECCION: un snapshot MANUAL, exclusivo de la linea mujer -- ninguna otra
linea la tiene. Compara el precio publico y por unidad de cada producto SIE contra sus
competidores, presentacion por presentacion. La arma parse-comparativa-precios.py desde
un Excel 'Comparativa de PRECIOS_DD.MM.AAAA.xlsx' con una hoja por grupo. Este script es
distinto: NO rearma el snapshot, solo le actualiza los PRECIOS desde el dump crudo del
Manual Farmaceutico (Registro/Troquel/Producto/.../PVP al <fecha>), que es la misma
fuente pero en el formato en que llega mes a mes.

POR QUE POR TROQUEL: es el identificador estable del SKU y ya esta guardado en cada fila
del snapshot (campo 'troq'). Matchear por (producto, presentacion) fallaria con los
nombres del Manual, que difieren.

LA CANTIDAD NO SE RECALCULA. 'unit' = pub / cantidad-de-la-presentacion, y esa cantidad
NO es la columna 'Q Pres' del Manual (que trae 1 para un ISIS de 28 comprimidos): sale
del texto de la presentacion. Para no inventarla parseando texto, se REUSA la que el
snapshot ya tiene implicita: q = round(pub_viejo / unit_viejo). Asi el precio se actualiza
y la unidad de medida queda exactamente como fue validada cuando se armo el snapshot.

FILAS SIN MATCH: conservan su precio y su fecha viejos (no se borran ni se estiman). Se
reportan al final, y como el render muestra la fecha por fila, quedan visibles como lo
que son: mas viejas que el resto.

Idempotente. --check para ver si hace falta. --dry-run para no escribir.
Uso: py shared/update-mujer-preccomp.py [--file <manual.xlsx>] [--check|--dry-run]
"""
from __future__ import annotations
import argparse, datetime, json, re, sys
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
TARGET = REPO / 'mujer' / 'data.js'


def hub_root():
    try:
        sys.path.insert(0, str(REPO / 'shared'))
        import manifest
        h = manifest.hub_root()
        if h.is_dir():
            return h
    except Exception:
        pass
    h = Path.home() / 'OneDrive - Portalcorp' / 'Documentos' / 'Hub-Marcas-Inputs'
    return h if h.is_dir() else None


def resolver(explicit):
    """El dump del Manual mas reciente. Los nombres traen tildes ('linea-mujer',
    'Sin titulo...'), asi que se resuelve por glob y no por ruta armada a mano."""
    if explicit:
        return Path(explicit)
    H = hub_root()
    if not H:
        return None
    cands = []
    for d in list(H.iterdir()) + [H]:
        try:
            if d.is_dir():
                cands += [p for p in d.glob('*.xlsx') if not p.name.startswith('~$')]
        except OSError:
            continue
    # solo los que tienen pinta de dump del Manual Farmaceutico
    import openpyxl
    ok = []
    for p in sorted(cands, key=lambda x: -x.stat().st_mtime)[:12]:
        try:
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
            hdr = [str(c or '').strip() for c in next(wb.worksheets[0].iter_rows(values_only=True))]
            wb.close()
            if 'Troquel' in hdr and any(h.startswith('PVP al ') for h in hdr):
                ok.append(p)
        except Exception:
            continue
    return ok[0] if ok else None


def leer_manual(path):
    """-> (por_troquel, label_prev, label_curr)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    hdr = [str(c or '').strip() for c in next(it)]
    pvp = [(i, h) for i, h in enumerate(hdr) if h.startswith('PVP al ')]
    if len(pvp) < 2:
        wb.close()
        raise ValueError(f'esperaba 2 columnas "PVP al ...", encontre {len(pvp)}: {hdr}')
    (i_prev, l_prev), (i_curr, l_curr) = pvp[-2], pvp[-1]
    I = {n: i for i, n in enumerate(hdr)}
    out = {}
    for r in it:
        tq = str(r[I['Troquel']] or '').strip()
        if not tq:
            continue
        def f(i):
            try:
                return float(r[i])
            except (TypeError, ValueError):
                return None
        fecha = r[I.get('Fecha Vigencia', -1)] if 'Fecha Vigencia' in I else None
        if isinstance(fecha, datetime.datetime):
            fecha = fecha.date().isoformat()
        elif fecha:
            fecha = str(fecha)[:10]
        out[tq] = {'pub': f(i_curr), 'pub_prev': f(i_prev),
                   'fecha': fecha or None,
                   'lab': str(r[I['Laboratorio']] or '').strip() if 'Laboratorio' in I else '',
                   'prod': str(r[I['Producto']] or '').strip() if 'Producto' in I else ''}
    wb.close()
    return out, l_prev, l_curr


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--file')
    ap.add_argument('--check', action='store_true')
    ap.add_argument('--dry-run', action='store_true')
    a = ap.parse_args()
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')

    src = resolver(a.file)
    if not src or not src.is_file():
        print('  (skip) no encontre un dump del Manual Farmaceutico (Troquel + "PVP al ...")')
        return 0
    print(f'  fuente: {src.name}')
    try:
        manual, l_prev, l_curr = leer_manual(src)
    except ImportError:
        print('  (skip) openpyxl no disponible')
        return 0
    print(f'  columnas de precio: {l_prev!r} -> {l_curr!r}   ({len(manual):,} troqueles)')

    text = TARGET.read_text(encoding='utf-8-sig')
    m = re.search(r'window\.OTC_DASHBOARD\s*=\s*', text)
    ob = text.index('{', m.end())
    D, end = json.JSONDecoder().raw_decode(text[ob:])
    pc = D.get('prec_comp')
    if not pc:
        print('  (skip) mujer no tiene prec_comp')
        return 0
    meta = D.get('prec_comp_meta') or {}
    print(f'  snapshot actual: {meta.get("fecha")}   {len(pc)} grupos')

    # fecha del nuevo snapshot: la del label de la columna actual (PVP al DD/MM/AAAA)
    mm = re.search(r'(\d{2}/\d{2}/\d{4})', l_curr)
    nueva_fecha = mm.group(1) if mm else meta.get('fecha')

    tot = act = sin_match = sin_q = 0
    faltantes = []
    for grupo, g in pc.items():
        for pres in g.get('pres', []):
            for row in pres.get('rows', []):
                tot += 1
                tq = str(row.get('troq') or '').strip()
                src_row = manual.get(tq) if tq else None
                if not src_row or src_row['pub'] is None:
                    sin_match += 1
                    faltantes.append(f'{grupo}/{row.get("prod")} (troq={tq or "sin troquel"})')
                    continue
                # cantidad IMPLICITA en el snapshot: no se recalcula desde el texto
                pub_v, unit_v = row.get('pub'), row.get('unit')
                if not pub_v or not unit_v:
                    sin_q += 1
                    faltantes.append(f'{grupo}/{row.get("prod")} (sin unit previo)')
                    continue
                q = round(pub_v / unit_v)
                if q <= 0:
                    sin_q += 1
                    continue
                row['pub'] = round(src_row['pub'], 2)
                row['unit'] = round(src_row['pub'] / q, 2)
                if src_row['fecha']:
                    row['fecha'] = src_row['fecha']
                act += 1
            # gap = unit del competidor / unit del SIE - 1, dentro de la presentacion
            sie = next((r for r in pres.get('rows', []) if r.get('sie')), None)
            base = (sie or {}).get('unit')
            for row in pres.get('rows', []):
                if row.get('sie'):
                    row['gap'] = None
                elif base and row.get('unit'):
                    row['gap'] = round(row['unit'] / base - 1, 4)

    fecha_vieja = meta.get('fecha')
    meta['fecha'] = nueva_fecha
    meta['fuente'] = 'Comparativa de Precios (Manual Farmacéutico)'
    D['prec_comp_meta'] = meta

    print(f'  filas: {tot}  actualizadas: {act}  sin match: {sin_match}  sin unidad previa: {sin_q}')
    print(f'  snapshot: {fecha_vieja} -> {nueva_fecha}')
    if faltantes:
        print(f'  conservan su precio y fecha viejos ({len(faltantes)}):')
        for f in faltantes[:8]:
            print(f'      {f}')
        if len(faltantes) > 8:
            print(f'      ... y {len(faltantes) - 8} mas')

    nuevo = text[:ob] + json.dumps(D, ensure_ascii=False) + text[ob + end:]
    if nuevo == text:
        print('  sin cambios (ya estaba al dia)')
        return 0
    if a.check:
        print(f'PRECCOMP: hay que actualizar. Correr: py shared/update-mujer-preccomp.py')
        return 1
    if a.dry_run:
        print('DRY-RUN: nada escrito.')
        return 0
    TARGET.write_text(nuevo, encoding='utf-8', newline='')
    print(f'Escrito {TARGET.relative_to(REPO)}.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
