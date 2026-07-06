# -*- coding: utf-8 -*-
"""Actualiza SOLO el mol_perf (Mercado IQVIA) de una linea al ultimo mes del master.

Camino confiable que evita el serializer de PowerShell (lento/cuelga en 5.1 con
data grande). Reusa la membresia de mercados EXISTENTE en el data.js (cada
mercado -> lista de productos), y recalcula desde el master nuevo:
  - monthly_vals de cada producto = suma de las filas del master con ese Product
  - family monthly/quarterly/ytd/mat + ms_* por producto (vs total mercado)
NO toca venta/stock/recetas/canales/etc. (esas no cambian). Mantiene el orden y
los demas campos. as_of = ultimo mes del master.

LIMITACION: reusa los productos ya presentes (no agrega competidores nuevos que
hayan entrado este mes). Para un avance de 1 mes es despreciable; el build-data
completo los capturaria.

Uso: py shared/update-iqvia-month.py --line OTC --master <ruta> [--dry-run]
     (line: cardio|ATB|OTC|respiratorio|mujer ; objeto window.OTC_DASHBOARD / const D / window.MUJER_DATA)
"""
from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
LINE_FILE = {
    'cardio': 'cardio/data.js', 'ATB': 'ATB/data.js', 'OTC': 'OTC/data.js',
    'respiratorio': 'respiratorio/data.js', 'mujer': 'mujer/index.html',
}
ANCHORS = [r'window\.OTC_DASHBOARD\s*=\s*', r'window\.MUJER_DATA\s*=\s*', r'const\s+D\s*=\s*']
MES_INV = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
NUM2 = {v:k for k,v in MES_INV.items()}
MONTH_RE = re.compile(r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec) \d{4}$')


def msort(mk):
    p = str(mk).split(); return int(p[1])*100 + MES_INV.get(p[0],0) if len(p)==2 and p[0] in MES_INV else 0
def quarter_key(mk):
    p = mk.split()
    if len(p)!=2: return ''
    m = MES_INV.get(p[0]); return f'Q{(m-1)//3+1} {p[1]}' if m else ''
def agg_quarterly(monthly):
    o=defaultdict(int)
    for mk,v in monthly.items():
        q=quarter_key(mk)
        if q:o[q]+=v
    return dict(o)
def agg_ytd(monthly,cierre):
    by=defaultdict(int)
    for mk,v in monthly.items():
        p=mk.split()
        if len(p)!=2:continue
        mn=MES_INV.get(p[0])
        if mn and mn<=cierre:by[p[1]]+=v
    return {f'{NUM2[cierre]} {y}':v for y,v in by.items()}
def agg_mat(monthly,cierre):
    years={int(mk.split()[1]) for mk in monthly if len(mk.split())==2 and mk.split()[0] in MES_INV}
    out={}
    for y in sorted(years):
        tot=0
        for back in range(11,-1,-1):
            idx=(y*12+(cierre-1))-back; yy,mm=divmod(idx,12)
            tot+=int(monthly.get(f'{NUM2[mm+1]} {yy}',0) or 0)
        out[f'{NUM2[cierre]} {y}']=tot
    return out


def load_master_by_product(path):
    """{product_name: {month: units}} sumando filas (packs) por Product."""
    wb=openpyxl.load_workbook(path, read_only=True, data_only=True); ws=wb.active
    r1=list(next(ws.iter_rows(min_row=1,max_row=1,values_only=True)))
    ci={}
    for i,h in enumerate(r1):
        if str(h or '').strip().lower().startswith('product'): ci['prod']=i
    lab=lambda h:str(h).split('\n')[-1].strip()
    mcols=[(i,lab(h)) for i,h in enumerate(r1) if h and str(h).startswith('Units') and MONTH_RE.match(lab(h))]
    prods=defaultdict(lambda: defaultdict(float))
    for row in ws.iter_rows(min_row=2,values_only=True):
        prod=str(row[ci['prod']] or '').strip() if ci['prod']<len(row) else ''
        if not prod: continue
        d=prods[prod]
        for cidx,mk in mcols:
            if cidx<len(row) and row[cidx] is not None:
                try: d[mk]+=float(row[cidx])
                except (ValueError,TypeError): pass
    wb.close()
    months=[mk for _,mk in sorted(mcols,key=lambda x:msort(x[1]))]
    return months, {p:{mk:int(round(v)) for mk,v in d.items() if v} for p,d in prods.items()}


def rebuild_family(products, master, cierre):
    fam_monthly=defaultdict(int)
    for p in products:
        mv = master.get(p['prod'], p.get('monthly_vals') or {})   # si no esta en master, conserva
        p['monthly_vals']=dict(mv)
        for mk,v in mv.items(): fam_monthly[mk]+=v
    fam_monthly=dict(fam_monthly)
    fam_q=agg_quarterly(fam_monthly); fam_y=agg_ytd(fam_monthly,cierre); fam_m=agg_mat(fam_monthly,cierre)
    for p in products:
        mv=p['monthly_vals']
        p['quarterly_vals']=agg_quarterly(mv); p['ytd']=agg_ytd(mv,cierre); p['mat']=agg_mat(mv,cierre)
        p['ms_monthly']  ={mk: round(mv.get(mk,0)/fv*100,2) if fv>0 else 0 for mk,fv in fam_monthly.items()}
        p['ms_quarterly']={qk: round(p['quarterly_vals'].get(qk,0)/fv*100,2) if fv>0 else 0 for qk,fv in fam_q.items()}
        p['ms_ytd']      ={y: round(p['ytd'].get(y,0)/fv*100,2) if fv>0 else 0 for y,fv in fam_y.items()}
        p['ms_mat']      ={y: round(p['mat'].get(y,0)/fv*100,2) if fv>0 else 0 for y,fv in fam_m.items()}
    return fam_monthly, fam_q, fam_y, fam_m


def main():
    ap=argparse.ArgumentParser(); ap.add_argument('--line',required=True); ap.add_argument('--master',required=True); ap.add_argument('--dry-run',action='store_true')
    a=ap.parse_args()
    if hasattr(sys.stdout,'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')
    f=REPO/LINE_FILE[a.line]
    text=f.read_text(encoding='utf-8',errors='replace')
    anc=next((m for pat in ANCHORS for m in [re.search(pat,text)] if m and 'mol_perf' in text[m.end():m.end()+200000]), None)
    if not anc:
        # fallback: primer objeto con mol_perf
        for pat in ANCHORS:
            for m in re.finditer(pat,text):
                ob=text.index('{',m.end())
                try: obj,_=json.JSONDecoder().raw_decode(text[ob:])
                except Exception: continue
                if isinstance(obj,dict) and obj.get('mol_perf'): anc=m; break
            if anc: break
    ob=text.index('{',anc.end()); D,end=json.JSONDecoder().raw_decode(text[ob:])
    months,master=load_master_by_product(Path(a.master))
    last=months[-1]; cierre=MES_INV[last.split()[0]]
    mp=D['mol_perf']; before_last=None
    for fam,fobj in mp.items():
        prods=fobj.get('products') or []
        if before_last is None and prods:
            before_last=max((max(p.get('monthly_vals',{}),key=msort,default='') for p in prods),key=msort,default='?')
        fm,fq,fy,fmat=rebuild_family(prods,master,cierre)
        fobj['monthly']=fm; fobj['quarterly']=fq; fobj['ytd']=fy; fobj['mat']=fmat
    new_last=max((max(p.get('monthly_vals',{}),key=msort,default='') for fobj in mp.values() for p in (fobj.get('products') or [])),key=msort,default='?')
    print(f'{a.line}: mol_perf {len(mp)} mercados | ultimo mes {before_last} -> {new_last}')
    if a.dry_run:
        print('DRY RUN: nada escrito.'); return 0
    f.write_text(text[:ob]+json.dumps(D,ensure_ascii=False)+text[ob+end:], encoding='utf-8', newline='')
    print(f'Escrito {f} ({f.stat().st_size:,} bytes)')
    return 0


if __name__=='__main__':
    sys.exit(main())
