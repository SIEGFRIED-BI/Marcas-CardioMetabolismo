# -*- coding: utf-8 -*-
"""Separa el mercado de CEFALEXINA ARG (comun) y CEFALEXINA ARG DUO en ATB/data.js,
splitteando por DOSIS a nivel PACK desde el master IQVIA.

Por que: en IQVIA la molecula (CEFALEXIN) y el ATC (J01D1) son los MISMOS para comun y
DUO, y Siegfried figura con UN SOLO producto 'CEFALEXINA ARGENTI (SIE)' que engloba las
dos presentaciones. Resultado: las dos familias del tablero apuntaban al mismo producto y
al mismo mercado -> en "Mercado IQVIA" se veian DUPLICADAS (identicas: mismas unidades
SIE, mismo MS%, mismo IE) y el mercado de la linea contaba cefalexina dos veces.

Lo que SI distingue comun de DUO es el PACK (la dosis):
  DUO   = pack con 'DUO' | '1.00G' | '1000MG' | '750MG'   (1g comprimidos / 750mg susp)
  COMUN = el resto (500MG / 250MG)
Ej: 'CEFALEXINA ARGENTI TABL DUO 1.00G x 14' vs 'CEFALEXINA ARGENTI TABL 500MG x 16'.

El split reconcilia con la venta interna (YTD jun-2026): IQVIA comun 233.505 vs venta
229.388; IQVIA DUO 250.623 vs venta 240.630. Y comun+DUO == el mercado unico anterior.

Cada familia queda con SU mercado (disjunto) y sus competidores reales: en DUO aparecen
CEPOREXIN DUO / CEFADUO / BUTEFINA DUO, en comun CEPOREXIN / FECOFAR / etc. Marcas con
las dos presentaciones (RICHET, KEFORAL, SEPTILISIN, SINURIT) figuran en ambos con el
volumen que corresponde a cada dosis.

Despues correr la cascada estandar: recompute-mol-perf-aggregates --cierre <M> +
build-kpis + sync-kpistrip + fix-brandkpis-* + build-total. Idempotente.

Uso: py shared/split-atb-cefalexina-duo.py --master <master IQVIA> [--dry-run]
"""
from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
DATA = REPO / 'ATB' / 'data.js'
MOLECULE = 'CEFALEXIN'
FAM_COMUN, FAM_DUO = 'CEFALEXINA ARG', 'CEFALEXINA ARG DUO'
RESTO_LABEL = 'Otros (resto del mercado)'
TOP_N = 8  # productos nombrados por sub-mercado; la cola va al RESTO
DUO_PACK = re.compile(r'DUO|1\.00G|1000MG|750MG')
MES = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}


def msort(mk):
    p = str(mk).split()
    return int(p[1])*100 + MES.get(p[0],0) if len(p)==2 and p[0] in MES else 0


def read_master(path, window):
    """Devuelve {'COMUN': {prod: {mes: units}}, 'DUO': ...} + manuf por producto."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True); ws = wb.active
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    ci = {}
    mcols = []
    for i, h in enumerate(hdr):
        s = str(h or '').replace('\n', ' ').strip()
        sl = s.lower()
        if sl == 'product': ci['prod'] = i
        elif sl == 'pack': ci['pack'] = i
        elif sl == 'manufacturer': ci['mf'] = i
        elif sl == 'molecules long': ci['mol'] = i
        m = re.match(r'^Units ([A-Z][a-z]{2}) (\d{4})$', s)   # mensual puro (no MAT/YTD/'to')
        if m and m.group(1) in MES:
            mcols.append((i, f'{m.group(1)} {m.group(2)}'))
    for need in ('prod', 'pack', 'mf', 'mol'):
        if need not in ci:
            raise ValueError(f'falta la columna {need} en el master: {hdr[:9]}')
    win = [mk for mk in window if mk in {mk2 for _, mk2 in mcols}]
    if not win:
        raise ValueError('la ventana del mol_perf no interseca con los meses del master')

    seg = {'COMUN': defaultdict(lambda: defaultdict(float)),
           'DUO':   defaultdict(lambda: defaultdict(float))}
    manuf = {}
    widx = {mk: i for i, mk in mcols}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[ci['mol']] or '').strip().upper() != MOLECULE:
            continue
        prod = str(row[ci['prod']] or '').strip()
        if not prod:
            continue
        manuf.setdefault(prod, str(row[ci['mf']] or '').strip())
        key = 'DUO' if DUO_PACK.search(str(row[ci['pack']] or '').upper()) else 'COMUN'
        for mk in win:
            i = widx[mk]
            if i < len(row) and isinstance(row[i], (int, float)):
                seg[key][prod][mk] += row[i]
    wb.close()
    return seg, manuf, win


def build_products(prod_series, manuf, win):
    """top-N nombrados (SIE primero, resto por unidades del ultimo mes) + RESTO agregado."""
    last = win[-1]
    items = []
    for prod, series in prod_series.items():
        monthly = {mk: int(round(series.get(mk, 0))) for mk in win}
        if sum(monthly.values()) == 0:
            continue
        items.append((prod, monthly, 'SIEG' in manuf.get(prod, '').upper()))
    items.sort(key=lambda x: (0 if x[2] else 1, -x[1].get(last, 0)))
    named, tail = items[:TOP_N], items[TOP_N:]

    def blank():
        return {'ytd': {}, 'mat': {}, 'quarterly_vals': {},
                'ms_monthly': {}, 'ms_ytd': {}, 'ms_mat': {}, 'ms_quarterly': {}}

    out = []
    for prod, monthly, is_sie in named:
        out.append({'prod': prod, 'manuf': manuf.get(prod, ''), 'is_sie': is_sie,
                    'monthly_vals': monthly, **blank()})
    if tail:
        resto = {mk: sum(m.get(mk, 0) for _, m, _ in tail) for mk in win}
        if sum(resto.values()) > 0:
            out.append({'prod': RESTO_LABEL, 'manuf': '', 'is_sie': False, 'is_resto': True,
                        'monthly_vals': resto, **blank()})
    return out, len(items)


def _win(end_y, end_m, kind):
    """Ventana de meses 'Mon YYYY' para ytd|mat terminando en end_y/end_m."""
    if kind == 'ytd':
        return [f'{k} {end_y}' for k, v in MES.items() if v <= end_m]
    out = []
    for back in range(11, -1, -1):
        idx = (end_y*12 + (end_m-1)) - back
        yy, mm = divmod(idx, 12)
        out.append(f'{[k for k,v in MES.items() if v==mm+1][0]} {yy}')
    return out


def fix_brandkpis(D, fams, last):
    """Recompone brandKpis[fam].{ytd,mat} desde mol_perf[fam] con la atribucion SIE de
    budIqviaMap (cada familia suma SOLO sus productos propios). Arregla dos cosas:
    (a) el split comun/DUO, y (b) el desfasaje de mes que traia brandKpis (units venia en
    YTD Ene-May contra un mercado ya a Jun -> MS% subestimado ~9 pp)."""
    bk = D.get('brandKpis') or {}
    mol = D.get('mol_perf') or {}
    own_map = D.get('budIqviaMap') or {}
    end_m, end_y = MES[last.split()[0]], int(last.split()[1])
    n = 0
    for fam in fams:
        kobj, fo = bk.get(fam), mol.get(fam)
        if not isinstance(kobj, dict) or not fo: continue
        own = own_map.get(fam)
        prods = fo.get('products', [])
        for per in ('ytd', 'mat'):
            if not isinstance(kobj.get(per), dict): continue
            wc = _win(end_y, end_m, per); wp = _win(end_y-1, end_m, per)
            def s(pl, w): return sum(sum((p.get('monthly_vals') or {}).get(mk, 0) or 0 for mk in w) for p in pl)
            sie = [p for p in prods if p.get('is_sie') and (not own or p['prod'] in own)]
            u_c, u_p = s(sie, wc), s(sie, wp)
            m_c, m_p = s(prods, wc), s(prods, wp)
            st = kobj[per]
            st['units'], st['units_prev'], st['market_total'] = u_c, u_p, m_c
            st['ms'] = round(u_c/m_c*100, 1) if m_c else None
            st['growth'] = round((u_c/u_p - 1)*100, 1) if u_p else None
            st['ie'] = round((u_c/u_p)/(m_c/m_p)*100, 1) if (u_p and m_p and m_c) else None
            n += 1
    return n


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--master', required=True)
    ap.add_argument('--dry-run', action='store_true')
    a = ap.parse_args()
    if hasattr(sys.stdout, 'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')

    t = DATA.read_text(encoding='utf-8-sig')
    m = re.search(r'window\.OTC_DASHBOARD\s*=\s*', t); ob = t.index('{', m.end())
    D, end = json.JSONDecoder().raw_decode(t[ob:])
    mol = D.get('mol_perf', {})
    if FAM_COMUN not in mol or FAM_DUO not in mol:
        print(f'  (skip) faltan las familias {FAM_COMUN}/{FAM_DUO} en mol_perf'); return 0

    sie_prod = next((p for p in mol[FAM_COMUN]['products'] if p.get('is_sie')), None)
    window = sorted((sie_prod or {}).get('monthly_vals', {}).keys(), key=msort)
    if not window:
        print('  (skip) no pude determinar la ventana de meses'); return 0

    seg, manuf, win = read_master(Path(a.master), window)
    last = win[-1]
    prev_mkt = sum((p.get('monthly_vals') or {}).get(last, 0) for p in mol[FAM_COMUN]['products'])

    total_new = 0
    for fam, key in ((FAM_COMUN, 'COMUN'), (FAM_DUO, 'DUO')):
        prods, n_all = build_products(seg[key], manuf, win)
        if not prods:
            print(f'  (warn) {fam}: 0 productos en el sub-mercado {key}; no toco'); continue
        mkt = sum(p['monthly_vals'].get(last, 0) for p in prods)
        total_new += mkt
        sie = [p['prod'] for p in prods if p['is_sie']]
        sie_u = sum(p['monthly_vals'].get(last, 0) for p in prods if p['is_sie'])
        mol[fam]['products'] = prods
        print(f'  {fam:20} [{key:5}] {len(prods)} filas de {n_all} prods | {last}: mkt={mkt} '
              f'SIE={sie_u} ({sie_u/mkt*100:.1f}%) {sie}')

    print(f'\n  mercado {last}: antes {prev_mkt} (contado 2 veces = {prev_mkt*2}) -> '
          f'ahora comun+DUO = {total_new} (disjuntos, 1 vez)')

    n = fix_brandkpis(D, (FAM_COMUN, FAM_DUO), last)
    print(f'  brandKpis recompuestos desde mol_perf: {n} periodo(s)')
    for fam in (FAM_COMUN, FAM_DUO):
        st = ((D.get('brandKpis') or {}).get(fam) or {}).get('ytd') or {}
        print(f'    {fam:20} YTD units={st.get("units")} mkt={st.get("market_total")} '
              f'ms={st.get("ms")} ie={st.get("ie")}')
    if a.dry_run:
        print('DRY-RUN: nada escrito.'); return 0
    DATA.write_text(t[:ob] + json.dumps(D, ensure_ascii=False) + t[ob+end:], encoding='utf-8', newline='')
    print(f'Escrito {DATA}. Ahora: recompute-mol-perf-aggregates + build-kpis + sync-kpistrip + fix-brandkpis-* + build-total')
    return 0


if __name__ == '__main__':
    sys.exit(main())
