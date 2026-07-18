#!/usr/bin/env python3
"""
shared/slice-iqvia-master.py

Lee el AR_PM Premium master de IQVIA y produce versiones recortadas por linea
que cada build-data.ps1 puede consumir mucho mas rapido (decenas de KB en
lugar de 18 MB).

Hoy slicea para:
- mujer: filtra por ATC-4 codes que mujer compite (extraidos de IQUVIA_VENTAS),
  reformatea columnas para que coincidan con el layout que el parser de mujer
  espera (col 4 = ATC, col 5 = Molecules, headers data como 'MMM YYYY\\nUnits').

Uso:
    py shared/slice-iqvia-master.py \\
        --master "<ruta>/AR_PM_FV_Standard_<fecha>.xlsx" \\
        --out-dir "<ruta>/_iqvia-master/<mes>/sliced" \\
        --lines mujer

El orquestador shared/build-all.ps1 invoca este script automaticamente cuando
detecta que mujer esta en el set de lineas a procesar y existe el AR_PM
centralizado.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl


# ATC-4 codes en los que mujer compite. Extraidos de IQUVIA_VENTAS.xlsx
# corte 2026-04 (los 13 codes que cubren todos los productos SIE de mujer
# y sus competidores). Si mujer agrega un mercado nuevo, sumar el ATC-4 aca.
MUJER_ATCS: set[str] = {
    'V03X0',  # TOD.DEMAS.PROD.TERAPEUT          (TRIP +45)
    'A12C1',  # SUPLEMENTOS DE MAGNESIO          (TRIP MAGNESIO)
    'A11C2',  # VITAMINA D SOLA                  (TRIP D3, CALCITOL D3)
    'G02X9',  # OTROS GINECOLOGICOS              (GYNODERM CARE GEL)
    'G03A1',  # PREPAR MONOF C/<50 MC EST        (ISIS, ISIS MINI, ISIS MINI 24)
    'A12A0',  # CALCIO                           (CALCIO BASE, CALCIO CIT, CITRAMAR)
    'A11X9',  # OTR VIT, SOLAS Y COMB            (TRIP D3 PLUS)
    'M05B3',  # BIFOSFONA D OSTEOPOROSIS         (ALENATO, DELTROX NF)
    'G01D0',  # ANTISEPT GINECOLOGICOS           (GYNODERM, GYNODERM HYALU)
    'B03A2',  # PROD COMB DE HIERRO              (SIDERBLUT COMPLEX, SIDERBLUT FOLICO)
    'B03A1',  # HIERRO SOLO                      (FER-IN-SOL, SIDERBLUT, SIDERBLUT POLI)
    'G03X0',  # OTR.HORM.SEXUAL&PR.SIMIL         (CLIMATIX)
    'G03A5',  # PREPAR SOLAS PROGEST ORAL        (ISIS FREE S/ESTROG)
}


LINE_RULES: dict[str, dict] = {
    'mujer': {
        'atc4': MUJER_ATCS,
        'output_name': 'AR_PM_mujer.xlsx',
    },
}


def _extract_atc4(atc_full: str | None) -> str | None:
    """Extrae el codigo ATC-4 de un string como 'A12A0 - CALCIO' -> 'A12A0'."""
    if not atc_full:
        return None
    m = re.match(r'^\s*([A-Z]\d{2}[A-Z0-9]\d)', str(atc_full))
    return m.group(1) if m else None


def _normalize_data_header(raw: str | None) -> str | None:
    """Convierte un header de data column del AR_PM a formato mujer.

    AR_PM Premium real:  'Units\\nApr 2021'   (Units arriba, mes abajo)
    Mujer parser regex:  ^MMM YYYY Units$    (despues de reemplazar whitespace)

    Output deseado: 'Apr 2021\\nUnits' (cuando se hace replace de '\\s+' a ' '
    queda 'Apr 2021 Units' que matchea el regex de mujer).
    """
    if not raw:
        return None
    s = str(raw)
    m = re.search(r'([A-Za-z]{3})\s+(\d{4})', s)
    if not m:
        return None
    return f'{m.group(1)} {m.group(2)}\nUnits'


def slice_for_line(
    master_path: Path,
    output_path: Path,
    atc4_set: set[str],
    line_name: str,
) -> None:
    print(f'[{line_name}] Lee master: {master_path}')
    wb_in = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
    ws_in = wb_in.active

    # Detectar columnas de entrada POR HEADER (no por posicion): IQVIA cambia el
    # orden entre entregas. El 'AR_PM Premium' estandar viene
    # Manuf|Product|Pack|ATC IV|Ph.Forms|Molecules|..., pero el export
    # 'Ateneo Total - MAT Movil' viene Pack|Manuf|ATC IV|Ph.Forms|Product|Molecules|...
    # y ademas trae DOS columnas ATC ('ATC IV' 5-char + 'ATC III' 4-char). Detectar
    # por nombre + preferir 'ATC IV' (el nivel de 5 chars, ej. G03A1, que usa mujer)
    # lo hace robusto a cualquiera de los dos layouts.
    headers_in = list(next(ws_in.iter_rows(min_row=1, max_row=1, values_only=True)))
    ci = {'manuf': None, 'product': None, 'pack': None, 'atc': None, 'mol': None}
    ci_atc_iv = None
    for idx, raw in enumerate(headers_in):
        h = str(raw or '').strip().lower()
        if h.startswith('manufacturer') and ci['manuf'] is None: ci['manuf'] = idx
        elif h.startswith('product') and ci['product'] is None: ci['product'] = idx
        elif h.startswith('pack') and ci['pack'] is None: ci['pack'] = idx
        elif h.startswith('molecules') and ci['mol'] is None: ci['mol'] = idx
        if h.startswith('atc iv'): ci_atc_iv = idx
        elif h.startswith('atc') and ci['atc'] is None: ci['atc'] = idx
    if ci_atc_iv is not None: ci['atc'] = ci_atc_iv
    faltan = [k for k, v in ci.items() if v is None]
    if faltan:
        raise SystemExit(f'[{line_name}] No encontre columnas por header: {faltan}. '
                         f'Headers: {[str(x) for x in headers_in[:8]]}')

    # Columnas de data mensual (0-based): header 'Units' cuyo label sea 'MMM YYYY'
    # exacto (excluye MAT/YTD). Output normalizado 'MMM YYYY\\nUnits' que el parser espera.
    data_cols: list[tuple[int, str]] = []
    for idx, raw in enumerate(headers_in):
        s = str(raw or '')
        if not s.startswith('Units'):
            continue
        lastline = s.split('\n')[-1].strip()
        if re.match(r'^[A-Z][a-z]{2} \d{4}$', lastline):
            data_cols.append((idx, f'{lastline}\nUnits'))

    if not data_cols:
        print(f'[{line_name}] WARN: no se detectaron columnas de data en el master.')

    # Filtrar rows por ATC-4 (usando la col ATC detectada por header)
    keepers: list[tuple] = []
    total_rows = 0
    ci_atc = ci['atc']
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        total_rows += 1
        if not row:
            continue
        atc4 = _extract_atc4(row[ci_atc] if ci_atc < len(row) else None)
        if atc4 and atc4 in atc4_set:
            keepers.append(row)
    wb_in.close()

    print(f'[{line_name}] Master rows: {total_rows}, kept: {len(keepers)} ({len(data_cols)} data cols)')

    # Layout output (imita IQUVIA_VENTAS, que es lo que mujer parser espera):
    # col 1: Manufacturer
    # col 2: Product
    # col 3: Pack
    # col 4: ATC-4         (el ATC IV crudo del AR_PM, ej. 'A12A0 - CALCIO')
    # col 5: Molecules Long
    # col 6: Market Type   (placeholder: el parser de mujer no usa esta col)
    # col 7+: 'MMM YYYY\\nUnits'
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = 'PM ARGENTINA Premium'

    new_headers = ['Manufacturer', 'Product', 'Pack', 'ATC-4', 'Molecules Long', 'Market Type']
    new_headers.extend(h for _, h in data_cols)
    ws_out.append(new_headers)

    def _g(row, i):
        return row[i] if i is not None and i < len(row) else None
    for row in keepers:
        new_row = [
            _g(row, ci['manuf']),   # Manufacturer
            _g(row, ci['product']), # Product
            _g(row, ci['pack']),    # Pack
            _g(row, ci['atc']),     # ATC IV crudo -> col ATC-4
            _g(row, ci['mol']),     # Molecules Long
            'POPULAR',              # placeholder Market Type
        ]
        for zero_idx, _ in data_cols:
            new_row.append(row[zero_idx] if zero_idx < len(row) else None)
        ws_out.append(new_row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(output_path)
    size_kb = output_path.stat().st_size // 1024
    print(f'[{line_name}] OK -> {output_path}  ({size_kb} KB)')


def main() -> int:
    ap = argparse.ArgumentParser(
        description='Slicea el AR_PM Premium en cuts mas chicos por linea.',
    )
    ap.add_argument('--master', required=True, help='Path al AR_PM master .xlsx')
    ap.add_argument('--out-dir', required=True, help='Carpeta donde escribir los slices')
    ap.add_argument(
        '--lines',
        nargs='+',
        default=['mujer'],
        help='Lineas a slicear (default: mujer)',
    )
    args = ap.parse_args()

    master = Path(args.master)
    if not master.is_file():
        print(f'ERROR: master no existe: {master}', file=sys.stderr)
        return 2

    out_dir = Path(args.out_dir)

    failed = 0
    for line in args.lines:
        if line not in LINE_RULES:
            print(f'WARN: linea desconocida {line!r}, saltando.')
            continue
        rules = LINE_RULES[line]
        try:
            slice_for_line(
                master_path=master,
                output_path=out_dir / rules['output_name'],
                atc4_set=rules['atc4'],
                line_name=line,
            )
        except Exception as e:
            print(f'[{line}] ERROR: {e}', file=sys.stderr)
            failed += 1

    return 1 if failed else 0


if __name__ == '__main__':
    sys.exit(main())
